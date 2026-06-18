from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import hashlib
import json
import zipfile

from openpyxl import load_workbook
import pandas as pd

from app.services.accounting_cogs import (
    COGS_ESTIMATE_SOURCES,
    COGS_REVIEW_SOURCES,
    cogs_basis_bucket,
)
from app.utils.time import utc_today


COLORADO_SUTS_ACCOUNT_NUMBER = "080390"
COLORADO_SUTS_GOLDEN_STATE_ACCOUNT_NUMBER = "970074130001"
COLORADO_SUTS_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2]
    / "assets"
    / "templates"
    / "CO-SUTS-Excel-Template-127596.xlsx"
)


def tax_review_packet_prefixes() -> set[str]:
    return {
        "tax_summary_estimated",
        "tax_by_marketplace_estimated",
        "tax_detail_estimated",
        "tax_exceptions_advisor_review",
        "tax_reporting_signoffs",
        "tax_reporting_signoff_review",
        "quarterly_estimated_tax_summary",
        "quarterly_estimated_tax_payments",
        "quarterly_estimated_tax_sales_detail",
        "quarterly_estimated_tax_fee_detail",
        "quarterly_estimated_tax_fee_summary",
        "quarterly_estimated_tax_cogs_detail",
        "quarterly_estimated_tax_returns",
        "quarterly_estimated_tax_local_tax",
        "quarterly_estimated_tax_payment_evidence",
        "quarterly_estimated_tax_payment_review",
    }


def render_taxes_workspace_intro(*, st, render_help_panel) -> None:
    st.subheader("Taxes")
    st.caption(
        "Dedicated tax workspace for Colorado SUTS, sales-tax review packets, and quarterly estimated-tax planning."
    )
    render_help_panel(
        section_title="Taxes",
        goal="Prepare tax-advisor review packets and filing support without mixing them into day-to-day reporting.",
        steps=[
            "Set the Reports date range for the tax evidence packet and sales-tax review context.",
            "Use Tax Reporting Scope to keep marketplace-facilitator channels such as eBay out of normal SUTS remittance unless advisor-confirmed.",
            "Use Quarterly Estimated Tax Planning for federal/Colorado estimated income-tax planning; spouse-owned LLC assumptions are review-only until confirmed by your tax advisor.",
            "Use Colorado SUTS Upload for monthly SUTS workbooks, normally scoped to direct/local sales rather than eBay facilitator sales.",
            "Record payment/sign-off evidence after advisor review so packet hashes, confirmations, and evidence links stay audit-ready.",
        ],
        roadmap_phase="GS Tax Reporting + Accounting Hardening",
    )
    st.info(
        "Tax outputs are planning estimates. Confirm LLC partnership treatment, member allocations, safe-harbor rules, "
        "deductions, credits, marketplace facilitator handling, bullion/coin exemptions, and final filing/payment amounts "
        "with your tax advisor."
    )
    with st.expander("Taxes Workspace Instructions", expanded=True):
        st.markdown(
            "- **SUTS monthly filing:** keep eBay/facilitator channels unselected for normal Colorado SUTS remittance unless your advisor tells you to include them.\n"
            "- **Golden rows:** use Golden State `110042` with account `970074130001`; Golden Local/Self-Collected may be blank only for advisor/SUTS-accepted zero filing.\n"
            "- **Quarterly estimated taxes:** Golden Stackers is modeled as a 50/50 spouse-owned partnership LLC by default; update each spouse's W-2 Social Security wages before relying on SE-tax estimates.\n"
            "- **Evidence:** download the Tax Review Packet and Quarterly Estimated Tax Packet before filing/payment, then record payment evidence with confirmation/reference and packet hash.\n"
        )


def _tax_safe_float(value: object) -> float:
    try:
        return float(value or 0.0)
    except Exception:
        return 0.0


def _tax_audit_changes(row) -> dict:
    try:
        payload = json.loads(str(getattr(row, "changes_json", "") or "{}"))
    except Exception:
        payload = {}
    return payload if isinstance(payload, dict) else {}


def _tax_truthy(value) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on"}
    return bool(value)


def _parse_csv_set(value: str) -> set[str]:
    return {str(part).strip().lower() for part in str(value or "").split(",") if str(part).strip()}


def _default_tax_marketplace_scope(
    sales_marketplace_options: list[str],
    facilitator_channels: set[str],
) -> list[str]:
    options = [str(v or "").strip().lower() for v in (sales_marketplace_options or []) if str(v or "").strip()]
    unique_options: list[str] = []
    seen: set[str] = set()
    for value in options:
        if value in seen:
            continue
        seen.add(value)
        unique_options.append(value)
    filtered = [m for m in unique_options if m not in set(facilitator_channels or set())]
    return filtered


def render_tax_reporting_scope_controls(
    *,
    st,
    tax_workspace: bool,
    sales,
    tax_default_jurisdiction: str,
    tax_default_rate: float,
    tax_shipping_taxable_default: bool,
    tax_exempt_categories_default_csv: str,
    facilitator_channels_default_csv: str,
    tax_profile_rows: list[dict[str, object]],
) -> dict[str, object]:
    if "reports_tax_exempt_categories_csv" not in st.session_state:
        st.session_state["reports_tax_exempt_categories_csv"] = str(
            tax_exempt_categories_default_csv or "bullion,coins"
        )
    if "reports_tax_facilitator_channels_csv" not in st.session_state:
        st.session_state["reports_tax_facilitator_channels_csv"] = str(
            facilitator_channels_default_csv or "ebay"
        )
    tax_exempt_categories = _parse_csv_set(
        str(st.session_state.get("reports_tax_exempt_categories_csv") or "")
    )
    sales_marketplace_options = sorted(
        {
            str((getattr(s, "marketplace", "") or "")).strip().lower()
            for s in sales
            if str((getattr(s, "marketplace", "") or "")).strip()
        }
    )
    facilitator_channels = _parse_csv_set(
        str(st.session_state.get("reports_tax_facilitator_channels_csv") or "")
    )
    default_tax_marketplaces = _default_tax_marketplace_scope(
        sales_marketplace_options=sales_marketplace_options,
        facilitator_channels=facilitator_channels,
    )
    if "reports_tax_jurisdiction" not in st.session_state:
        st.session_state["reports_tax_jurisdiction"] = str(
            tax_default_jurisdiction or "Golden, Colorado"
        )
    if "reports_tax_rate_percent" not in st.session_state:
        st.session_state["reports_tax_rate_percent"] = float(max(0.0, tax_default_rate))
    if "reports_tax_shipping_taxable" not in st.session_state:
        st.session_state["reports_tax_shipping_taxable"] = bool(tax_shipping_taxable_default)
    if "reports_tax_marketplaces" not in st.session_state:
        st.session_state["reports_tax_marketplaces"] = list(default_tax_marketplaces)

    if tax_workspace:
        st.markdown("### Tax Reporting Scope")
        tr1, tr2, tr3 = st.columns(3)
        with tr1:
            tax_jurisdiction = st.text_input(
                "Tax Jurisdiction Context",
                key="reports_tax_jurisdiction",
            ).strip()
        with tr2:
            tax_rate_percent = st.number_input(
                "Estimated Tax Rate (%)",
                min_value=0.0,
                step=0.01,
                key="reports_tax_rate_percent",
            )
        with tr3:
            tax_shipping_taxable = st.checkbox(
                "Treat shipping as taxable",
                key="reports_tax_shipping_taxable",
            )
        if facilitator_channels:
            st.caption(
                "Marketplace facilitator channels are excluded by default in local tax scope: "
                + ", ".join(sorted(facilitator_channels))
                + "."
            )
    else:
        tax_jurisdiction = str(
            st.session_state.get("reports_tax_jurisdiction")
            or tax_default_jurisdiction
            or "Golden, Colorado"
        )
        tax_rate_percent = float(
            st.session_state.get("reports_tax_rate_percent") or max(0.0, tax_default_rate)
        )
        tax_shipping_taxable = bool(
            st.session_state.get("reports_tax_shipping_taxable", tax_shipping_taxable_default)
        )

    selected_tax_profile_context: dict[str, object] = {}
    if tax_workspace and tax_profile_rows:
        profile_options = {
            f"{row.get('profile_name') or row.get('profile_key')} [{row.get('profile_key')}]": row
            for row in tax_profile_rows
        }
        sp1, sp2 = st.columns([2, 1])
        with sp1:
            selected_tax_profile_label = st.selectbox(
                "Saved Tax Profile",
                options=list(profile_options.keys()),
                key="reports_saved_tax_profile",
            )
        with sp2:
            if st.button("Apply Saved Tax Profile", key="reports_apply_saved_tax_profile_btn"):
                profile = profile_options.get(selected_tax_profile_label) or {}
                st.session_state["reports_tax_jurisdiction"] = str(
                    profile.get("jurisdiction") or tax_default_jurisdiction or "Golden, Colorado"
                )
                st.session_state["reports_tax_rate_percent"] = float(
                    max(0.0, _tax_safe_float(profile.get("tax_rate_percent")))
                )
                st.session_state["reports_tax_shipping_taxable"] = bool(profile.get("shipping_taxable"))
                st.session_state["reports_tax_facilitator_channels_csv"] = str(
                    profile.get("facilitator_channels") or ""
                )
                st.session_state["reports_tax_exempt_categories_csv"] = str(
                    profile.get("tax_exempt_categories") or ""
                )
                st.session_state["reports_tax_marketplaces"] = _default_tax_marketplace_scope(
                    sales_marketplace_options=sales_marketplace_options,
                    facilitator_channels=_parse_csv_set(str(profile.get("facilitator_channels") or "")),
                )
                st.success(f"Applied saved tax profile `{profile.get('profile_key')}`.")
                st.rerun()
        selected_profile = profile_options.get(str(st.session_state.get("reports_saved_tax_profile") or "")) or {}
        selected_tax_profile_context = dict(selected_profile)
        validation_status = str(selected_profile.get("human_validation_status") or "").strip()
        if validation_status and validation_status != "advisor_validated":
            st.warning(
                "Selected tax profile is not advisor-validated "
                f"(`{validation_status}`). Treat outputs as review-only estimates."
            )
        elif selected_profile:
            st.caption("Selected saved tax profile is advisor-validated.")

    if tax_workspace:
        preset_map = _tax_report_presets(
            default_jurisdiction=tax_default_jurisdiction,
            default_tax_rate_percent=float(max(0.0, tax_default_rate)),
            default_shipping_taxable=bool(tax_shipping_taxable_default),
        )
        tp1, tp2 = st.columns([2, 1])
        with tp1:
            tax_preset_name = st.selectbox(
                "Tax Report Preset",
                options=list(preset_map.keys()),
                key="reports_tax_preset_name",
            )
        with tp2:
            if st.button("Apply Tax Report Preset", key="reports_apply_tax_preset_btn"):
                preset = preset_map.get(tax_preset_name) or {}
                st.session_state["reports_tax_jurisdiction"] = str(
                    preset.get("jurisdiction") or tax_default_jurisdiction or "Golden, Colorado"
                )
                st.session_state["reports_tax_rate_percent"] = float(
                    max(0.0, _tax_safe_float(preset.get("tax_rate_percent")))
                )
                st.session_state["reports_tax_shipping_taxable"] = bool(
                    preset.get("shipping_taxable", False)
                )
                marketplace_mode = str(preset.get("marketplace_mode") or "all").strip().lower()
                if marketplace_mode == "local_only":
                    local_candidates = [
                        m for m in sales_marketplace_options if m in {"local", "in_person", "pos"}
                    ]
                    st.session_state["reports_tax_marketplaces"] = local_candidates
                else:
                    st.session_state["reports_tax_marketplaces"] = list(sales_marketplace_options)
                st.success(f"Applied tax report preset `{tax_preset_name}`.")
                st.rerun()

    current_tax_marketplaces = st.session_state.get("reports_tax_marketplaces")
    if current_tax_marketplaces is None:
        current_tax_marketplaces = list(default_tax_marketplaces)
    st.session_state["reports_tax_marketplaces"] = [
        m for m in current_tax_marketplaces if m in sales_marketplace_options
    ]
    if tax_workspace:
        selected_tax_marketplaces = st.multiselect(
            "Tax Marketplace Filter",
            options=sales_marketplace_options,
            key="reports_tax_marketplaces",
            help=(
                "Choose marketplaces you are responsible for reporting/remitting. "
                "Leave eBay/facilitator channels unselected for normal SUTS filing unless your advisor tells you otherwise."
            ),
        )
    else:
        selected_tax_marketplaces = list(
            st.session_state.get("reports_tax_marketplaces") or default_tax_marketplaces
        )
    selected_tax_marketplace_set = {
        str(v).strip().lower() for v in selected_tax_marketplaces if str(v).strip()
    }
    tax_query_marketplace_set = (
        set(selected_tax_marketplace_set)
        if selected_tax_marketplace_set
        else {"__no_tax_marketplace_selected__"}
    )
    tax_marketplace_scope_label = (
        ",".join(sorted(selected_tax_marketplace_set)) if selected_tax_marketplace_set else "none"
    )
    if tax_workspace and sales_marketplace_options and not selected_tax_marketplace_set:
        st.warning(
            "No non-facilitator tax marketplaces are selected. Tax/SUTS outputs for this scope will be zero "
            "unless you intentionally select a marketplace for advisor-reviewed reporting."
        )
    selected_facilitator_marketplaces = selected_tax_marketplace_set.intersection(facilitator_channels)
    if tax_workspace and selected_facilitator_marketplaces:
        st.warning(
            "Marketplace facilitator channel(s) selected for tax/SUTS scope: "
            + ", ".join(sorted(selected_facilitator_marketplaces))
            + ". For normal SUTS remittance, keep eBay/facilitator channels unselected because the facilitator "
            "generally collects/remits those taxes. Select only with advisor-confirmed reporting instructions."
        )
    elif tax_workspace:
        st.caption(
            "For SUTS filing, keep eBay/facilitator channels unselected unless your advisor says to report "
            "marketplace-facilitator gross sales. Direct/local channels are the normal SUTS remittance scope."
        )
        st.caption(
            "Tax-exempt categories (runtime): "
            + (", ".join(sorted(tax_exempt_categories)) if tax_exempt_categories else "(none)")
        )
        st.info(
            "Tax outputs in this report are estimates for operational planning. "
            "Validate local/state tax treatment (including bullion/coin exemptions) with your tax advisor."
        )

    return {
        "tax_exempt_categories": tax_exempt_categories,
        "sales_marketplace_options": sales_marketplace_options,
        "facilitator_channels": facilitator_channels,
        "default_tax_marketplaces": default_tax_marketplaces,
        "tax_jurisdiction": tax_jurisdiction,
        "tax_rate_percent": tax_rate_percent,
        "tax_shipping_taxable": tax_shipping_taxable,
        "selected_tax_profile_context": selected_tax_profile_context,
        "selected_tax_marketplaces": selected_tax_marketplaces,
        "selected_tax_marketplace_set": selected_tax_marketplace_set,
        "tax_query_marketplace_set": tax_query_marketplace_set,
        "tax_marketplace_scope_label": tax_marketplace_scope_label,
    }


def _tax_profile_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    profiles: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows or []:
        payload = _tax_audit_changes(row)
        profile_key = str(payload.get("profile_key") or "").strip().lower()
        if not profile_key or profile_key in seen:
            continue
        seen.add(profile_key)
        if not _tax_truthy(payload.get("is_active", True)):
            continue
        profiles.append(
            {
                "profile_key": profile_key,
                "profile_name": str(payload.get("profile_name") or profile_key).strip(),
                "jurisdiction": str(payload.get("jurisdiction") or "").strip(),
                "tax_rate_percent": _tax_safe_float(payload.get("tax_rate_percent")),
                "shipping_taxable": _tax_truthy(payload.get("shipping_taxable")),
                "facilitator_channels": str(payload.get("facilitator_channels") or "").strip(),
                "tax_exempt_categories": str(payload.get("tax_exempt_categories") or "").strip(),
                "effective_from": str(payload.get("effective_from") or "").strip(),
                "effective_to": str(payload.get("effective_to") or "").strip(),
                "human_validation_status": str(payload.get("human_validation_status") or "").strip().lower(),
                "advisor_evidence_link": str(payload.get("advisor_evidence_link") or "").strip(),
            }
        )
    return profiles


def _tax_signoff_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for row in rows or []:
        payload = _tax_audit_changes(row)
        output.append(
            {
                "recorded_at_utc": (
                    getattr(row, "created_at").isoformat(timespec="seconds")
                    if getattr(row, "created_at", None)
                    else ""
                ),
                "actor": str(getattr(row, "actor", "") or ""),
                "target_env": str(payload.get("target_env") or ""),
                "tax_period": str(payload.get("tax_period") or ""),
                "jurisdiction": str(payload.get("jurisdiction") or ""),
                "profile_key": str(payload.get("profile_key") or ""),
                "status": str(payload.get("status") or "").strip().lower(),
                "owner": str(payload.get("owner") or ""),
                "signoff_date": str(payload.get("signoff_date") or ""),
                "tax_packet_ref": str(payload.get("tax_packet_ref") or ""),
                "tax_packet_hash": str(
                    payload.get("tax_packet_hash")
                    or payload.get("tax_packet_evidence_hash_sha256")
                    or ""
                ).strip(),
                "advisor_evidence_link": str(payload.get("advisor_evidence_link") or ""),
                "tax_exception_count": int(_tax_safe_float(payload.get("tax_exception_count"))),
                "notes": str(payload.get("notes") or "")[:220],
            }
        )
    return output


def _quarterly_estimated_tax_payment_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for row in rows or []:
        payload = _tax_audit_changes(row)
        output.append(
            {
                "recorded_at_utc": (
                    getattr(row, "created_at").isoformat(timespec="seconds")
                    if getattr(row, "created_at", None)
                    else ""
                ),
                "actor": str(getattr(row, "actor", "") or ""),
                "target_env": str(payload.get("target_env") or ""),
                "tax_year": int(_tax_safe_float(payload.get("tax_year"))),
                "quarter": str(payload.get("quarter") or "").strip().upper(),
                "jurisdiction": str(payload.get("jurisdiction") or "").strip(),
                "payment_type": str(payload.get("payment_type") or "").strip(),
                "status": str(payload.get("status") or "").strip().lower(),
                "payment_date": str(payload.get("payment_date") or "").strip(),
                "amount": round(_tax_safe_float(payload.get("amount")), 2),
                "confirmation_ref": str(payload.get("confirmation_ref") or "").strip(),
                "evidence_link": str(payload.get("evidence_link") or "").strip(),
                "packet_ref": str(payload.get("packet_ref") or "").strip(),
                "packet_hash": str(payload.get("packet_hash") or "").strip(),
                "notes": str(payload.get("notes") or "")[:220],
            }
        )
    return output


def _load_colorado_suts_jurisdiction_options(
    template_path: str = str(COLORADO_SUTS_TEMPLATE_PATH),
) -> list[dict[str, object]]:
    path = Path(template_path)
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=False)
    if "Upload Data" not in wb.sheetnames:
        return []
    ws = wb["Upload Data"]
    rows: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        account_type = str(ws.cell(row_idx, 2).value or "").strip()
        jurisdiction_code = str(ws.cell(row_idx, 3).value or "").strip()
        jurisdiction_name = str(ws.cell(row_idx, 4).value or "").strip()
        if not jurisdiction_code or not jurisdiction_name:
            continue
        rows.append(
            {
                "row": row_idx,
                "account_type": account_type,
                "jurisdiction_code": jurisdiction_code,
                "jurisdiction_name": jurisdiction_name,
                "label": f"{jurisdiction_code} | {jurisdiction_name} | {account_type}",
            }
        )
    return rows


def _format_suts_gross_text(value: object) -> str:
    amount = round(_tax_safe_float(value), 2)
    if abs(amount) < 0.005:
        return "0"
    return f"{amount:.2f}"


def _month_bounds_from_yyyy_mm(month_value: str) -> tuple[datetime, datetime]:
    raw = str(month_value or "").strip()
    try:
        month_start = datetime.strptime(raw, "%Y-%m")
    except Exception as exc:
        raise ValueError("Month must be in YYYY-MM format.") from exc
    if month_start.month == 12:
        next_month = datetime(month_start.year + 1, 1, 1)
    else:
        next_month = datetime(month_start.year, month_start.month + 1, 1)
    return month_start, next_month - timedelta(microseconds=1)


def _filter_tax_detail_for_month(tax_detail_df: pd.DataFrame, *, month_value: str) -> pd.DataFrame:
    if tax_detail_df is None or tax_detail_df.empty:
        return pd.DataFrame()
    month_start, month_end = _month_bounds_from_yyyy_mm(month_value)
    if "sold_at" not in tax_detail_df.columns:
        return tax_detail_df.copy()
    filtered = tax_detail_df.copy()
    sold_at = pd.to_datetime(filtered["sold_at"], errors="coerce")
    filtered = filtered[(sold_at >= month_start) & (sold_at <= month_end)].copy()
    return filtered


_load_colorado_suts_jurisdiction_options.clear = lambda: None


def _estimated_tax_period_rows(tax_year: int) -> list[dict[str, object]]:
    year = int(tax_year or utc_today().year)
    return [
        {
            "quarter": "Q1",
            "income_period": f"{year}-01-01 to {year}-03-31",
            "start": datetime(year, 1, 1),
            "end": datetime(year, 3, 31, 23, 59, 59, 999999),
            "federal_due_date": datetime(year, 4, 15),
            "colorado_due_date": datetime(year, 4, 15),
        },
        {
            "quarter": "Q2",
            "income_period": f"{year}-04-01 to {year}-05-31",
            "start": datetime(year, 4, 1),
            "end": datetime(year, 5, 31, 23, 59, 59, 999999),
            "federal_due_date": datetime(year, 6, 15),
            "colorado_due_date": datetime(year, 6, 15),
        },
        {
            "quarter": "Q3",
            "income_period": f"{year}-06-01 to {year}-08-31",
            "start": datetime(year, 6, 1),
            "end": datetime(year, 8, 31, 23, 59, 59, 999999),
            "federal_due_date": datetime(year, 9, 15),
            "colorado_due_date": datetime(year, 9, 15),
        },
        {
            "quarter": "Q4",
            "income_period": f"{year}-09-01 to {year}-12-31",
            "start": datetime(year, 9, 1),
            "end": datetime(year, 12, 31, 23, 59, 59, 999999),
            "federal_due_date": datetime(year + 1, 1, 15),
            "colorado_due_date": datetime(year + 1, 1, 15),
        },
    ]


def _estimated_tax_period_for_quarter(tax_year: int, quarter: str) -> dict[str, object]:
    normalized = str(quarter or "").strip().upper()
    for row in _estimated_tax_period_rows(tax_year):
        if str(row.get("quarter") or "").upper() == normalized:
            return row
    raise ValueError("Quarter must be Q1, Q2, Q3, or Q4.")


def _quarterly_estimated_tax_scope_notice(
    *,
    report_from_date,
    report_to_date,
    period: dict[str, object],
) -> dict[str, object]:
    def _date_only(value):
        if isinstance(value, datetime):
            return value.date()
        return value

    start_dt = period.get("start")
    end_dt = period.get("end")
    quarter = str(period.get("quarter") or "").strip()
    income_period = str(period.get("income_period") or "").strip()
    quarter_start = _date_only(start_dt)
    quarter_end = _date_only(end_dt)
    report_start = _date_only(report_from_date)
    report_end = _date_only(report_to_date)
    matches = bool(report_start == quarter_start and report_end == quarter_end)
    if matches:
        return {
            "status": "match",
            "message": f"Quarterly estimated-tax worksheet and Reports page are both scoped to {income_period}.",
        }
    return {
        "status": "mismatch",
        "message": (
            f"Quarterly estimated-tax worksheet uses {quarter} income period {income_period}. "
            f"Other report sections on this page use {report_start} to {report_end}, so totals can differ."
        ),
    }


def _filter_df_by_datetime_window(
    df: pd.DataFrame,
    *,
    column: str,
    start_dt: datetime,
    end_dt: datetime,
) -> pd.DataFrame:
    if df is None or df.empty or column not in df.columns:
        return pd.DataFrame()
    filtered = df.copy()
    timestamps = pd.to_datetime(filtered[column], errors="coerce")
    filtered = filtered[(timestamps >= start_dt) & (timestamps <= end_dt)].copy()
    return filtered


def _build_quarterly_estimated_tax_payload(
    *,
    tax_year: int,
    quarter: str,
    sales_df: pd.DataFrame,
    cogs_margin_df: pd.DataFrame,
    qbo_adjustments_df: pd.DataFrame,
    tax_detail_df: pd.DataFrame,
    federal_income_tax_rate_percent: float,
    colorado_income_tax_rate_percent: float,
    self_employment_tax_rate_percent: float,
    self_employment_net_earnings_multiplier_percent: float = 92.35,
    include_self_employment_tax: bool = True,
    w2_social_security_wages: float = 0.0,
    spouse_w2_social_security_wages: float = 0.0,
    social_security_wage_base: float = 184500.0,
    owner_allocation_percent: float = 50.0,
    prior_estimated_payments: float = 0.0,
    other_income_adjustments: float = 0.0,
    deductible_adjustments: float = 0.0,
) -> dict[str, object]:
    period = _estimated_tax_period_for_quarter(tax_year, quarter)
    start_dt = period["start"]
    end_dt = period["end"]
    assert isinstance(start_dt, datetime)
    assert isinstance(end_dt, datetime)

    sales_period_df = _filter_df_by_datetime_window(
        sales_df,
        column="sold_at",
        start_dt=start_dt,
        end_dt=end_dt,
    )
    cogs_period_df = _filter_df_by_datetime_window(
        cogs_margin_df,
        column="sold_at",
        start_dt=start_dt,
        end_dt=end_dt,
    )
    adjustments_period_df = _filter_df_by_datetime_window(
        qbo_adjustments_df,
        column="txn_date",
        start_dt=start_dt,
        end_dt=end_dt,
    )
    tax_period_df = _filter_df_by_datetime_window(
        tax_detail_df,
        column="sold_at",
        start_dt=start_dt,
        end_dt=end_dt,
    )

    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df[column], errors="coerce").fillna(0.0).sum())

    def _to_bool(value: object) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return False
        raw = str(value).strip().lower()
        return raw in {"1", "true", "yes", "y", "on"}

    def _bool_series(df: pd.DataFrame, column: str, default: bool = False) -> pd.Series:
        if df is None or df.empty or column not in df.columns:
            return pd.Series([default] * (0 if df is None else len(df)), index=getattr(df, "index", None))
        return df[column].map(_to_bool).fillna(default).astype(bool)

    gross_sales = _sum(sales_period_df, "gross_sales")
    shipping_charged = _sum(sales_period_df, "actual_shipping_charged") or _sum(sales_period_df, "shipping_cost")
    fees = _sum(sales_period_df, "actual_fee") or _sum(sales_period_df, "fees")
    label_spend = _sum(sales_period_df, "actual_shipping_label_cost") or _sum(sales_period_df, "shipping_label_cost")
    net_before_cogs = _sum(cogs_period_df, "net_before_cogs") or _sum(sales_period_df, "actual_net_before_cogs")
    fifo_cogs = _sum(cogs_period_df, "fifo_cogs")
    profit_before_returns = net_before_cogs - fifo_cogs
    returns_refund_total = abs(_sum(adjustments_period_df, "net_adjustment"))
    returns_cogs_reversal = _sum(adjustments_period_df, "cogs_reversal_estimate")
    returns_profit_impact = -returns_refund_total + returns_cogs_reversal
    net_business_income = (
        profit_before_returns
        + returns_profit_impact
        + float(other_income_adjustments or 0.0)
        - float(deductible_adjustments or 0.0)
    )
    positive_business_income = max(0.0, net_business_income)
    self_employment_net_earnings = (
        positive_business_income
        * (max(0.0, float(self_employment_net_earnings_multiplier_percent or 0.0)) / 100.0)
        if include_self_employment_tax
        else 0.0
    )
    federal_income_tax_estimate = positive_business_income * (
        max(0.0, float(federal_income_tax_rate_percent or 0.0)) / 100.0
    )
    colorado_income_tax_estimate = positive_business_income * (
        max(0.0, float(colorado_income_tax_rate_percent or 0.0)) / 100.0
    )
    owner_allocation = max(0.0, min(100.0, float(owner_allocation_percent or 0.0)))
    spouse_allocation = max(0.0, 100.0 - owner_allocation)
    owner_se_net_earnings = self_employment_net_earnings * (owner_allocation / 100.0)
    spouse_se_net_earnings = self_employment_net_earnings * (spouse_allocation / 100.0)
    owner_social_security_wage_base_remaining = max(
        0.0,
        float(social_security_wage_base or 0.0) - max(0.0, float(w2_social_security_wages or 0.0)),
    )
    spouse_social_security_wage_base_remaining = max(
        0.0,
        float(social_security_wage_base or 0.0) - max(0.0, float(spouse_w2_social_security_wages or 0.0)),
    )
    self_employment_tax_rate = max(0.0, float(self_employment_tax_rate_percent or 0.0)) / 100.0
    se_social_security_rate = min(0.124, self_employment_tax_rate)
    se_medicare_rate = max(0.0, self_employment_tax_rate - se_social_security_rate)
    owner_se_social_security_taxable_earnings = min(
        owner_se_net_earnings,
        owner_social_security_wage_base_remaining,
    )
    spouse_se_social_security_taxable_earnings = min(
        spouse_se_net_earnings,
        spouse_social_security_wage_base_remaining,
    )
    se_social_security_taxable_earnings = (
        owner_se_social_security_taxable_earnings + spouse_se_social_security_taxable_earnings
    )
    se_social_security_tax_estimate = se_social_security_taxable_earnings * se_social_security_rate
    se_medicare_tax_estimate = self_employment_net_earnings * se_medicare_rate
    self_employment_tax_estimate = (
        se_social_security_tax_estimate + se_medicare_tax_estimate
        if include_self_employment_tax
        else 0.0
    )
    total_estimated_tax = federal_income_tax_estimate + colorado_income_tax_estimate + self_employment_tax_estimate
    suggested_payment_after_prior = max(0.0, total_estimated_tax - float(prior_estimated_payments or 0.0))

    cogs_review_count = 0
    cogs_estimate_count = 0
    missing_cogs_count = 0
    cogs_review_amount = 0.0
    cogs_estimate_amount = 0.0
    missing_cogs_amount = 0.0
    if cogs_period_df is not None and not cogs_period_df.empty:
        source_series = (
            cogs_period_df["fifo_cost_source"].fillna("").astype(str)
            if "fifo_cost_source" in cogs_period_df.columns
            else pd.Series([""] * len(cogs_period_df), index=cogs_period_df.index)
        )
        if "cogs_basis_bucket" in cogs_period_df.columns:
            bucket_series = cogs_period_df["cogs_basis_bucket"].fillna("").astype(str).str.lower()
        else:
            bucket_series = source_series.map(cogs_basis_bucket)
        if "basis_review_required" in cogs_period_df.columns:
            review_series = _bool_series(cogs_period_df, "basis_review_required")
        else:
            review_series = bucket_series.eq("review")
        if "basis_is_estimate" in cogs_period_df.columns:
            estimate_series = _bool_series(cogs_period_df, "basis_is_estimate")
        else:
            estimate_series = bucket_series.eq("estimate")
        review_mask = review_series | bucket_series.eq("review") | source_series.isin(COGS_REVIEW_SOURCES)
        estimate_mask = estimate_series | bucket_series.eq("estimate") | source_series.isin(COGS_ESTIMATE_SOURCES)
        missing_mask = source_series.isin({"missing_cost_basis", "unknown", ""})
        cogs_amount_series = (
            pd.to_numeric(cogs_period_df["fifo_cogs"], errors="coerce").fillna(0.0)
            if "fifo_cogs" in cogs_period_df.columns
            else pd.Series([0.0] * len(cogs_period_df), index=cogs_period_df.index)
        )
        cogs_review_count = int(review_mask.sum())
        cogs_estimate_count = int(estimate_mask.sum())
        missing_cogs_count = int(missing_mask.sum())
        cogs_review_amount = float(cogs_amount_series[review_mask].sum())
        cogs_estimate_amount = float(cogs_amount_series[estimate_mask].sum())
        missing_cogs_amount = float(cogs_amount_series[missing_mask].sum())

    local_tax_estimated_collected = _sum(tax_period_df, "estimated_tax_collected")
    local_taxable_sales = _sum(tax_period_df, "taxable_subtotal")
    local_gross_sales = _sum(tax_period_df, "gross_sales")
    federal_due = period["federal_due_date"]
    colorado_due = period["colorado_due_date"]
    assert isinstance(federal_due, datetime)
    assert isinstance(colorado_due, datetime)
    summary_rows = [
        {"field": "Tax year", "value": int(tax_year)},
        {"field": "Estimated tax quarter", "value": str(period["quarter"])},
        {"field": "Income period", "value": str(period["income_period"])},
        {"field": "Federal estimated payment due", "value": federal_due.date().isoformat()},
        {"field": "Colorado estimated payment due", "value": colorado_due.date().isoformat()},
        {"field": "Sale rows", "value": int(len(sales_period_df))},
        {"field": "Gross sales", "value": round(gross_sales, 2)},
        {"field": "Shipping charged", "value": round(shipping_charged, 2)},
        {"field": "Marketplace fees", "value": round(fees, 2)},
        {"field": "Shipping label spend", "value": round(label_spend, 2)},
        {"field": "Net before COGS", "value": round(net_before_cogs, 2)},
        {"field": "FIFO COGS", "value": round(fifo_cogs, 2)},
        {"field": "Profit before returns", "value": round(profit_before_returns, 2)},
        {"field": "Returns refund impact", "value": round(-returns_refund_total, 2)},
        {"field": "Returns COGS reversal", "value": round(returns_cogs_reversal, 2)},
        {"field": "Other income adjustments", "value": round(float(other_income_adjustments or 0.0), 2)},
        {"field": "Deductible adjustments", "value": round(float(deductible_adjustments or 0.0), 2)},
        {"field": "Estimated business income", "value": round(net_business_income, 2)},
        {"field": "Federal income tax reserve", "value": round(federal_income_tax_estimate, 2)},
        {"field": "Colorado income tax reserve", "value": round(colorado_income_tax_estimate, 2)},
        {"field": "Self-employment tax included", "value": "yes" if include_self_employment_tax else "no"},
        {"field": "Owner/member allocation to you %", "value": round(owner_allocation, 2)},
        {"field": "Spouse/other member allocation %", "value": round(spouse_allocation, 2)},
        {"field": "Your SE net earnings share", "value": round(owner_se_net_earnings, 2)},
        {"field": "Spouse SE net earnings share", "value": round(spouse_se_net_earnings, 2)},
        {"field": "Your W-2 Social Security wages entered", "value": round(float(w2_social_security_wages or 0.0), 2)},
        {
            "field": "Spouse W-2 Social Security wages entered",
            "value": round(float(spouse_w2_social_security_wages or 0.0), 2),
        },
        {"field": "Social Security wage base", "value": round(float(social_security_wage_base or 0.0), 2)},
        {"field": "Your Social Security wage base remaining", "value": round(owner_social_security_wage_base_remaining, 2)},
        {
            "field": "Spouse Social Security wage base remaining",
            "value": round(spouse_social_security_wage_base_remaining, 2),
        },
        {"field": "Self-employment net earnings estimate", "value": round(self_employment_net_earnings, 2)},
        {"field": "Your SE Social Security taxable earnings", "value": round(owner_se_social_security_taxable_earnings, 2)},
        {
            "field": "Spouse SE Social Security taxable earnings",
            "value": round(spouse_se_social_security_taxable_earnings, 2),
        },
        {"field": "SE Social Security taxable earnings", "value": round(se_social_security_taxable_earnings, 2)},
        {"field": "SE Social Security tax reserve", "value": round(se_social_security_tax_estimate, 2)},
        {"field": "SE Medicare tax reserve", "value": round(se_medicare_tax_estimate, 2)},
        {"field": "Self-employment tax reserve", "value": round(self_employment_tax_estimate, 2)},
        {"field": "Total estimated tax reserve", "value": round(total_estimated_tax, 2)},
        {"field": "Prior estimated payments entered", "value": round(float(prior_estimated_payments or 0.0), 2)},
        {"field": "Suggested payment after prior payments", "value": round(suggested_payment_after_prior, 2)},
        {"field": "COGS review-needed sale rows", "value": cogs_review_count},
        {"field": "COGS review-needed amount", "value": round(cogs_review_amount, 2)},
        {"field": "COGS estimate sale rows", "value": cogs_estimate_count},
        {"field": "COGS estimate amount", "value": round(cogs_estimate_amount, 2)},
        {"field": "Missing/unknown COGS sale rows", "value": missing_cogs_count},
        {"field": "Missing/unknown COGS amount", "value": round(missing_cogs_amount, 2)},
        {"field": "Local/SUTS gross in selected tax scope", "value": round(local_gross_sales, 2)},
        {"field": "Local/SUTS taxable subtotal in selected tax scope", "value": round(local_taxable_sales, 2)},
        {"field": "Local/SUTS tax estimate in selected tax scope", "value": round(local_tax_estimated_collected, 2)},
    ]
    payment_rows = [
        {
            "jurisdiction": "Federal IRS",
            "form_or_payment": "1040-ES / IRS Direct Pay estimated tax",
            "income_period": str(period["income_period"]),
            "due_date": federal_due.date().isoformat(),
            "estimated_amount": round(federal_income_tax_estimate + self_employment_tax_estimate, 2),
            "notes": "Planning estimate only; reconcile with Form 1040-ES worksheet and prior-year safe-harbor rules.",
        },
        {
            "jurisdiction": "Colorado",
            "form_or_payment": "DR 0104EP estimated income tax",
            "income_period": str(period["income_period"]),
            "due_date": colorado_due.date().isoformat(),
            "estimated_amount": round(colorado_income_tax_estimate, 2),
            "notes": "Planning estimate only; confirm entity/taxpayer treatment and credits with advisor.",
        },
    ]
    advisor_checklist_rows = [
        {
            "area": "Entity/taxpayer treatment",
            "status": "advisor_review",
            "detail": (
                "Confirm LLC federal tax treatment. A spouse-owned LLC may be taxed as a partnership unless "
                "a corporate election or advisor-confirmed exception applies; confirm whether self-employment "
                "tax applies to each spouse/member."
            ),
            "amount": "",
        },
        {
            "area": "Spouse/member allocation",
            "status": "advisor_review",
            "detail": (
                f"Worksheet currently notes {owner_allocation:.0f}% allocation to you and "
                f"{spouse_allocation:.0f}% to spouse/other member. Confirm Form 1065/K-1 "
                "or advisor-approved allocation and whether SE tax should be computed separately by spouse."
            ),
            "amount": round(net_business_income * (owner_allocation / 100.0), 2),
        },
        {
            "area": "W-2 Social Security wage base",
            "status": "advisor_review" if include_self_employment_tax and w2_social_security_wages else "ok",
            "detail": (
                "W-2 wages can reduce the Social Security portion of self-employment tax; "
                "Medicare tax has no wage-base cap."
            ),
            "amount": round(se_social_security_tax_estimate, 2),
        },
        {
            "area": "Safe harbor",
            "status": "advisor_review",
            "detail": "Confirm prior-year safe-harbor target, withholding, credits, and annualized-income method before payment.",
            "amount": "",
        },
        {
            "area": "COGS basis",
            "status": "review_needed" if cogs_review_count or missing_cogs_count else "ok",
            "detail": (
                f"{cogs_review_count} review-needed row(s), {missing_cogs_count} missing/unknown row(s), "
                f"{cogs_estimate_count} estimate row(s)."
            ),
            "amount": round(cogs_review_amount + missing_cogs_amount, 2),
        },
        {
            "area": "Marketplace fees",
            "status": "advisor_review" if fees else "ok",
            "detail": (
                "eBay/marketplace selling fees are tracked separately as potential commissions/fees expense evidence. "
                "Confirm Schedule C/partnership return treatment, capitalization rules, and account mapping with your tax advisor."
            ),
            "amount": round(fees, 2),
        },
        {
            "area": "Manual adjustments",
            "status": "advisor_review" if other_income_adjustments or deductible_adjustments else "ok",
            "detail": "Confirm manual other-income and deductible adjustments are complete, supported, and not double-counted.",
            "amount": round(float(other_income_adjustments or 0.0) - float(deductible_adjustments or 0.0), 2),
        },
        {
            "area": "Local/SUTS tax scope",
            "status": "advisor_review" if local_gross_sales or local_tax_estimated_collected else "ok",
            "detail": "Confirm marketplace-facilitator exclusions, local/direct sales scope, and SUTS filing treatment.",
            "amount": round(local_tax_estimated_collected, 2),
        },
        {
            "area": "Payment evidence",
            "status": "required_after_payment",
            "detail": "Record payment confirmation/reference, evidence link, packet reference, and packet hash after payment.",
            "amount": round(suggested_payment_after_prior, 2),
        },
    ]
    fee_detail_columns = [
        "sale_id",
        "sold_at",
        "marketplace",
        "sku",
        "title",
        "gross_sales",
        "actual_fee",
        "fees",
        "actual_fee_source",
        "cogs_basis_bucket",
    ]
    fee_detail_df = pd.DataFrame()
    if sales_period_df is not None and not sales_period_df.empty:
        available_columns = [column for column in fee_detail_columns if column in sales_period_df.columns]
        fee_detail_df = sales_period_df[available_columns].copy() if available_columns else pd.DataFrame()
        if not fee_detail_df.empty:
            if "actual_fee" not in fee_detail_df.columns:
                fee_detail_df["actual_fee"] = 0.0
            if "fees" not in fee_detail_df.columns:
                fee_detail_df["fees"] = 0.0
            fee_detail_df["actual_fee"] = pd.to_numeric(fee_detail_df["actual_fee"], errors="coerce").fillna(0.0)
            fee_detail_df["fees"] = pd.to_numeric(fee_detail_df["fees"], errors="coerce").fillna(0.0)
            fee_detail_df["deductible_fee_planning_amount"] = fee_detail_df["actual_fee"]
            missing_actual_mask = fee_detail_df["deductible_fee_planning_amount"].abs() < 0.005
            fee_detail_df.loc[missing_actual_mask, "deductible_fee_planning_amount"] = fee_detail_df.loc[
                missing_actual_mask, "fees"
            ]
            fee_detail_df["tax_planning_category"] = "marketplace_commissions_and_fees"
            fee_detail_df["advisor_review_required"] = True
            fee_detail_df["advisor_review_note"] = (
                "Planning evidence only. Confirm deductibility, capitalization, and return line mapping with tax advisor."
            )
            fee_detail_df = fee_detail_df[
                pd.to_numeric(fee_detail_df["deductible_fee_planning_amount"], errors="coerce").fillna(0.0) > 0
            ].copy()
    fee_summary_df = pd.DataFrame()
    if fee_detail_df is not None and not fee_detail_df.empty:
        summary = fee_detail_df.copy()
        if "marketplace" not in summary.columns:
            summary["marketplace"] = ""
        if "actual_fee_source" not in summary.columns:
            summary["actual_fee_source"] = ""
        summary["marketplace"] = summary["marketplace"].fillna("").astype(str).str.strip().replace("", "unknown")
        summary["actual_fee_source"] = (
            summary["actual_fee_source"].fillna("").astype(str).str.strip().replace("", "unknown")
        )
        fee_summary_df = (
            summary.groupby(["marketplace", "actual_fee_source", "tax_planning_category"], as_index=False)
            .agg(
                sale_rows=("sale_id", "count") if "sale_id" in summary.columns else ("deductible_fee_planning_amount", "count"),
                gross_sales=("gross_sales", "sum") if "gross_sales" in summary.columns else ("deductible_fee_planning_amount", "sum"),
                fee_planning_amount=("deductible_fee_planning_amount", "sum"),
            )
            .sort_values(["fee_planning_amount", "marketplace"], ascending=[False, True])
        )
        fee_summary_df["fee_planning_amount"] = fee_summary_df["fee_planning_amount"].round(2)
        fee_summary_df["gross_sales"] = fee_summary_df["gross_sales"].round(2)
        fee_summary_df["advisor_review_required"] = True
        fee_summary_df["advisor_review_note"] = (
            "Planning rollup only. Confirm fee deductibility, capitalization, and tax return line mapping with advisor."
        )

    return {
        "period": period,
        "summary_rows": summary_rows,
        "payment_rows": payment_rows,
        "advisor_checklist_rows": advisor_checklist_rows,
        "sales_detail_df": sales_period_df,
        "fee_detail_df": fee_detail_df,
        "fee_summary_df": fee_summary_df,
        "cogs_detail_df": cogs_period_df,
        "return_adjustments_df": adjustments_period_df,
        "tax_detail_df": tax_period_df,
        "review": {
            "cogs_review_needed_sale_rows": cogs_review_count,
            "cogs_review_needed_amount": round(cogs_review_amount, 2),
            "cogs_estimate_sale_rows": cogs_estimate_count,
            "cogs_estimate_amount": round(cogs_estimate_amount, 2),
            "missing_cogs_sale_rows": missing_cogs_count,
            "missing_cogs_amount": round(missing_cogs_amount, 2),
            "requires_advisor_review": True,
        },
    }


def _build_quarterly_estimated_tax_payment_payload(
    *,
    target_env: str,
    tax_year: int,
    quarter: str,
    jurisdiction: str,
    payment_type: str,
    status: str,
    payment_date,
    amount: float,
    confirmation_ref: str,
    evidence_link: str,
    packet_ref: str,
    packet_hash: str,
    notes: str = "",
) -> dict[str, object]:
    return {
        "target_env": str(target_env or "").strip().lower(),
        "tax_year": int(tax_year or 0),
        "quarter": str(quarter or "").strip().upper(),
        "jurisdiction": str(jurisdiction or "").strip(),
        "payment_type": str(payment_type or "").strip(),
        "status": str(status or "").strip().lower(),
        "payment_date": payment_date.isoformat() if hasattr(payment_date, "isoformat") else str(payment_date or ""),
        "amount": round(float(amount or 0.0), 2),
        "confirmation_ref": str(confirmation_ref or "").strip(),
        "evidence_link": str(evidence_link or "").strip(),
        "packet_ref": str(packet_ref or "").strip(),
        "packet_hash": str(packet_hash or "").strip(),
        "notes": str(notes or "").strip(),
    }


def _build_quarterly_estimated_tax_payment_review(
    *,
    payment_df: pd.DataFrame,
    payment_worksheet_df: pd.DataFrame,
    tax_year: int,
    quarter: str,
) -> pd.DataFrame:
    if payment_worksheet_df is None or payment_worksheet_df.empty:
        return pd.DataFrame()
    year = int(tax_year or 0)
    q = str(quarter or "").strip().upper()
    current_payments = pd.DataFrame()
    if payment_df is not None and not payment_df.empty:
        df = payment_df.copy()
        df["tax_year"] = pd.to_numeric(df.get("tax_year", 0), errors="coerce").fillna(0).astype(int)
        df["quarter"] = df.get("quarter", pd.Series(dtype=str)).fillna("").astype(str).str.upper().str.strip()
        df["status"] = df.get("status", pd.Series(dtype=str)).fillna("").astype(str).str.lower().str.strip()
        df["jurisdiction"] = df.get("jurisdiction", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        df["amount"] = pd.to_numeric(df.get("amount", 0), errors="coerce").fillna(0.0)
        current_payments = df.loc[(df["tax_year"] == year) & (df["quarter"] == q) & (df["status"] == "paid")].copy()

    rows: list[dict[str, object]] = []
    for worksheet_row in payment_worksheet_df.to_dict("records"):
        jurisdiction = str(worksheet_row.get("jurisdiction") or "").strip()
        expected = round(_tax_safe_float(worksheet_row.get("estimated_amount")), 2)
        paid_df = pd.DataFrame()
        if not current_payments.empty:
            paid_df = current_payments.loc[
                current_payments["jurisdiction"].str.lower() == jurisdiction.lower()
            ].copy()
        paid_amount = round(float(paid_df["amount"].sum()) if not paid_df.empty else 0.0, 2)
        confirmation_count = (
            int(
                paid_df.get("confirmation_ref", pd.Series(dtype=str))
                .fillna("")
                .astype(str)
                .str.strip()
                .astype(bool)
                .sum()
            )
            if not paid_df.empty
            else 0
        )
        evidence_link_count = (
            int(
                paid_df.get("evidence_link", pd.Series(dtype=str))
                .fillna("")
                .astype(str)
                .str.strip()
                .astype(bool)
                .sum()
            )
            if not paid_df.empty
            else 0
        )
        packet_hash_count = (
            int(
                paid_df.get("packet_hash", pd.Series(dtype=str))
                .fillna("")
                .astype(str)
                .str.strip()
                .astype(bool)
                .sum()
            )
            if not paid_df.empty
            else 0
        )
        delta = round(paid_amount - expected, 2)
        remaining = round(max(0.0, expected - paid_amount), 2)
        status = "pass" if expected <= 0 or paid_amount + 0.005 >= expected else "warn"
        rows.append(
            {
                "jurisdiction": jurisdiction,
                "tax_year": year,
                "quarter": q,
                "expected_amount": expected,
                "recorded_paid_amount": paid_amount,
                "remaining_amount": remaining,
                "delta_paid_minus_expected": delta,
                "paid_records": int(len(paid_df)) if not paid_df.empty else 0,
                "confirmation_refs_present": confirmation_count,
                "evidence_links_present": evidence_link_count,
                "packet_hashes_present": packet_hash_count,
                "status": status,
                "details": (
                    "Recorded paid amount meets or exceeds the planning worksheet amount."
                    if status == "pass"
                    else "Recorded paid amount is below the planning worksheet amount or payment has not been recorded."
                ),
            }
        )
    return pd.DataFrame(rows)


def _filter_tax_drilldown_rows(
    tax_detail_df: pd.DataFrame,
    *,
    marketplace: str,
    taxability: str,
) -> pd.DataFrame:
    if tax_detail_df is None or tax_detail_df.empty:
        return pd.DataFrame()
    filtered = tax_detail_df.copy()
    marketplace_norm = str(marketplace or "all").strip().lower()
    taxability_norm = str(taxability or "all").strip().lower()
    if marketplace_norm != "all":
        filtered = filtered[
            filtered["marketplace"].astype(str).str.strip().str.lower() == marketplace_norm
        ]
    if taxability_norm == "taxable_only":
        filtered = filtered[filtered["taxable_subtotal"].astype(float) > 0.0]
    elif taxability_norm == "exempt_only":
        filtered = filtered[filtered["is_tax_exempt_category"].astype(bool)]
    return filtered


def _build_tax_drilldown_sale_option_rows(filtered_tax_detail: pd.DataFrame) -> list[dict]:
    if filtered_tax_detail is None or filtered_tax_detail.empty:
        return []
    rows: list[dict] = []
    for row in filtered_tax_detail.itertuples(index=False):
        sale_id = int(getattr(row, "sale_id", 0) or 0)
        if sale_id <= 0:
            continue
        label = (
            f"sale#{sale_id} | {str(getattr(row, 'marketplace', '') or '').strip()} | "
            f"{str(getattr(row, 'sku', '') or '').strip()} | "
            f"tax={float(getattr(row, 'estimated_tax_collected', 0.0) or 0.0):,.2f}"
        )
        rows.append({"label": label, "sale_id": sale_id})
    return rows


def _tax_drilldown_kpis(filtered_tax_detail: pd.DataFrame) -> dict[str, float | int]:
    if filtered_tax_detail is None or filtered_tax_detail.empty:
        return {
            "rows": 0,
            "taxable_subtotal": 0.0,
            "estimated_tax": 0.0,
        }
    return {
        "rows": int(len(filtered_tax_detail)),
        "taxable_subtotal": float(filtered_tax_detail["taxable_subtotal"].sum()),
        "estimated_tax": float(filtered_tax_detail["estimated_tax_collected"].sum()),
    }


def _build_tax_exception_rows(
    tax_detail_df: pd.DataFrame,
    *,
    tax_jurisdiction: str,
    tax_rate_percent: float,
    shipping_taxable: bool,
    facilitator_channels: set[str] | None,
    tax_exempt_categories: set[str] | None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    jurisdiction = str(tax_jurisdiction or "").strip()
    facilitator_set = {
        str(v).strip().lower()
        for v in (facilitator_channels or set())
        if str(v).strip()
    }
    exempt_set = {
        str(v).strip().lower()
        for v in (tax_exempt_categories or set())
        if str(v).strip()
    }
    tax_rate = _tax_safe_float(tax_rate_percent)

    def _append(
        *,
        exception_type: str,
        severity: str,
        message: str,
        sale_id: int | None = None,
        marketplace: str = "",
        category: str = "",
        evidence: str = "",
    ) -> None:
        rows.append(
            {
                "severity": severity,
                "exception_type": exception_type,
                "sale_id": int(sale_id or 0) if sale_id else None,
                "marketplace": marketplace,
                "category": category,
                "message": message,
                "evidence": evidence,
            }
        )

    if not jurisdiction:
        _append(
            exception_type="missing_tax_jurisdiction",
            severity="P1",
            message="Tax report has no jurisdiction configured for the selected scope.",
            evidence="tax_jurisdiction is blank",
        )
    if tax_rate <= 0 and tax_detail_df is not None and not tax_detail_df.empty:
        taxable_total = float(tax_detail_df.get("taxable_subtotal", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if taxable_total > 0:
            _append(
                exception_type="missing_or_zero_tax_rate",
                severity="P1",
                message="Taxable sales exist but the configured tax rate is zero or missing.",
                evidence=f"taxable_subtotal={taxable_total:.2f}; tax_rate_percent={tax_rate:.4f}",
            )
    if tax_detail_df is None or tax_detail_df.empty:
        return rows

    for raw in tax_detail_df.to_dict(orient="records"):
        sale_id = int(raw.get("sale_id") or 0)
        marketplace = str(raw.get("marketplace") or "").strip().lower()
        category = str(raw.get("category") or "").strip().lower()
        gross_sales = _tax_safe_float(raw.get("gross_sales"))
        shipping_cost = _tax_safe_float(raw.get("shipping_cost"))
        taxable_subtotal = _tax_safe_float(raw.get("taxable_subtotal"))
        taxable_shipping = _tax_safe_float(raw.get("taxable_shipping_subtotal"))
        estimated_tax = _tax_safe_float(raw.get("estimated_tax_collected"))
        is_exempt = bool(raw.get("is_tax_exempt_category"))

        if marketplace in facilitator_set and taxable_subtotal > 0:
            _append(
                exception_type="facilitator_channel_in_tax_scope",
                severity="P2",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Marketplace facilitator channel is included in local tax liability scope.",
                evidence=f"taxable_subtotal={taxable_subtotal:.2f}; estimated_tax={estimated_tax:.2f}",
            )
        if gross_sales > 0 and not category:
            _append(
                exception_type="missing_tax_category",
                severity="P2",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Sale has no product category for taxable/exempt classification.",
                evidence=f"gross_sales={gross_sales:.2f}",
            )
        if is_exempt:
            _append(
                exception_type="exempt_category_review_needed",
                severity="P3",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Sale uses configured tax-exempt category treatment; confirm exemption basis with advisor.",
                evidence=f"category={category}; configured_exempt_categories={','.join(sorted(exempt_set))}",
            )
        if is_exempt and bool(shipping_taxable) and shipping_cost > 0 and taxable_shipping > 0:
            _append(
                exception_type="exempt_item_taxable_shipping_review_needed",
                severity="P2",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Exempt item has taxable shipping under current report settings.",
                evidence=f"shipping_cost={shipping_cost:.2f}; taxable_shipping={taxable_shipping:.2f}",
            )
    return rows


def _tax_report_presets(
    *,
    default_jurisdiction: str,
    default_tax_rate_percent: float,
    default_shipping_taxable: bool,
) -> dict[str, dict]:
    return {
        "Golden Local Retail": {
            "jurisdiction": default_jurisdiction or "Golden, Colorado",
            "tax_rate_percent": float(default_tax_rate_percent),
            "shipping_taxable": bool(default_shipping_taxable),
            "marketplace_mode": "local_only",
        },
        "Marketplace Shipped": {
            "jurisdiction": default_jurisdiction or "Golden, Colorado",
            "tax_rate_percent": float(default_tax_rate_percent),
            "shipping_taxable": False,
            "marketplace_mode": "all",
        },
        "Bullion Exempt Focus": {
            "jurisdiction": default_jurisdiction or "Golden, Colorado",
            "tax_rate_percent": float(default_tax_rate_percent),
            "shipping_taxable": False,
            "marketplace_mode": "all",
        },
    }


def _build_tax_reporting_signoff_payload(
    *,
    target_env: str,
    tax_period: str,
    jurisdiction: str,
    profile_key: str,
    status: str,
    owner: str,
    signoff_date,
    tax_packet_ref: str,
    tax_packet_hash: str,
    advisor_evidence_link: str,
    tax_exception_count: int,
    notes: str = "",
) -> dict[str, object]:
    return {
        "target_env": str(target_env or "").strip().lower(),
        "tax_period": str(tax_period or "").strip(),
        "jurisdiction": str(jurisdiction or "").strip(),
        "profile_key": str(profile_key or "").strip().lower(),
        "status": str(status or "").strip().lower(),
        "owner": str(owner or "").strip(),
        "signoff_date": signoff_date.isoformat() if hasattr(signoff_date, "isoformat") else str(signoff_date or ""),
        "tax_packet_ref": str(tax_packet_ref or "").strip(),
        "tax_packet_hash": str(tax_packet_hash or "").strip(),
        "tax_packet_evidence_hash_sha256": str(tax_packet_hash or "").strip(),
        "advisor_evidence_link": str(advisor_evidence_link or "").strip(),
        "tax_exception_count": int(tax_exception_count or 0),
        "notes": str(notes or "").strip(),
    }


def _build_tax_reporting_signoff_review(
    *,
    signoff_df: pd.DataFrame,
    tax_period: str,
    jurisdiction: str,
    profile_key: str,
    tax_exception_count: int,
    current_packet_hash: str = "",
    to_date=None,
) -> pd.DataFrame:
    period = str(tax_period or "").strip()
    jurisdiction_norm = str(jurisdiction or "").strip().lower()
    profile_key_norm = str(profile_key or "").strip().lower()

    def _col(name: str) -> pd.Series:
        if signoff_df is None or signoff_df.empty or name not in signoff_df.columns:
            return pd.Series(dtype=str)
        return signoff_df[name].fillna("").astype(str)

    matching_df = pd.DataFrame()
    if signoff_df is not None and not signoff_df.empty:
        period_series = _col("tax_period").str.strip()
        matching_df = signoff_df.loc[period_series == period].copy()
    approved_df = pd.DataFrame()
    if not matching_df.empty and "status" in matching_df.columns:
        approved_df = matching_df.loc[matching_df["status"].fillna("").astype(str).str.lower() == "approved"].copy()
        if not approved_df.empty:
            approved_df["_signoff_sort_date"] = pd.to_datetime(
                approved_df.get("signoff_date", pd.Series(dtype=str)),
                errors="coerce",
                utc=True,
            )
            approved_df["_recorded_sort_date"] = pd.to_datetime(
                approved_df.get("recorded_at_utc", pd.Series(dtype=str)),
                errors="coerce",
                utc=True,
            )
            approved_df = approved_df.sort_values(
                ["_signoff_sort_date", "_recorded_sort_date"],
                ascending=[False, False],
                na_position="last",
                kind="mergesort",
            ).drop(columns=["_signoff_sort_date", "_recorded_sort_date"])

    rows: list[dict[str, object]] = [
        {
            "check": "Tax Sign-Off Evidence Present",
            "status": "pass" if not approved_df.empty else ("warn" if int(tax_exception_count or 0) == 0 else "info"),
            "expected": "approved tax reporting sign-off for selected period",
            "observed": "approved" if not approved_df.empty else "missing",
            "details": (
                "Approved tax reporting sign-off evidence found for the selected period."
                if not approved_df.empty
                else "No approved tax reporting sign-off evidence found for the selected period."
            ),
        }
    ]
    if approved_df.empty:
        return pd.DataFrame(rows)

    latest = approved_df.iloc[0].to_dict()
    signoff_jurisdiction = str(latest.get("jurisdiction") or "").strip().lower()
    signoff_profile_key = str(latest.get("profile_key") or "").strip().lower()
    signoff_exception_count = int(_tax_safe_float(latest.get("tax_exception_count")))
    signoff_packet_ref = str(latest.get("tax_packet_ref") or "").strip()
    signoff_packet_hash = str(latest.get("tax_packet_hash") or "").strip().lower()
    current_packet_hash = str(current_packet_hash or "").strip().lower()
    advisor_evidence = str(latest.get("advisor_evidence_link") or "").strip()
    signoff_owner = str(latest.get("owner") or "").strip()
    signoff_date = str(latest.get("signoff_date") or "").strip()

    def _coerce_review_date(value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if all(hasattr(value, attr) for attr in ("year", "month", "day")):
            try:
                return datetime(int(value.year), int(value.month), int(value.day)).date()
            except Exception:
                return None
        raw = str(value).strip()
        if not raw:
            return None
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
        except Exception:
            try:
                return datetime.fromisoformat(raw[:10]).date()
            except Exception:
                return None

    signoff_review_date = _coerce_review_date(signoff_date)
    period_end_date = _coerce_review_date(to_date)
    today = utc_today()
    signoff_date_valid = bool(
        signoff_review_date
        and (period_end_date is None or signoff_review_date >= period_end_date)
        and signoff_review_date <= today
    )
    if not signoff_date:
        signoff_date_observed = "missing"
    elif signoff_review_date is None:
        signoff_date_observed = f"{signoff_date} (unparseable)"
    elif period_end_date is not None and signoff_review_date < period_end_date:
        signoff_date_observed = f"{signoff_review_date.isoformat()} before {period_end_date.isoformat()}"
    elif signoff_review_date > today:
        signoff_date_observed = f"{signoff_review_date.isoformat()} after {today.isoformat()}"
    else:
        signoff_date_observed = signoff_review_date.isoformat()

    rows.extend(
        [
            {
                "check": "Approved Tax Sign-Off Jurisdiction Match",
                "status": "pass" if signoff_jurisdiction == jurisdiction_norm else "warn",
                "expected": jurisdiction_norm or "missing",
                "observed": signoff_jurisdiction or "missing",
                "details": "Approved tax sign-off jurisdiction should match the selected tax review assumptions.",
            },
            {
                "check": "Approved Tax Sign-Off Profile Match",
                "status": "pass" if signoff_profile_key == profile_key_norm else ("info" if not profile_key_norm else "warn"),
                "expected": profile_key_norm or "no selected profile",
                "observed": signoff_profile_key or "missing",
                "details": "Approved tax sign-off profile key should match the selected saved tax profile when one is used.",
            },
            {
                "check": "Approved Tax Sign-Off Exception Count",
                "status": "pass" if signoff_exception_count == int(tax_exception_count or 0) else "warn",
                "expected": int(tax_exception_count or 0),
                "observed": signoff_exception_count,
                "details": "Approved tax sign-off exception count should match recalculated Tax Exceptions / Advisor Review rows.",
            },
            {
                "check": "Approved Tax Sign-Off Owner Present",
                "status": "pass" if signoff_owner else "warn",
                "expected": "review owner",
                "observed": signoff_owner or "missing",
                "details": "Approved tax sign-off should identify the reviewer or owner.",
            },
            {
                "check": "Approved Tax Sign-Off Date Present",
                "status": "pass" if signoff_date else "warn",
                "expected": "sign-off date",
                "observed": signoff_date or "missing",
                "details": "Approved tax sign-off should include the approval date.",
            },
            {
                "check": "Approved Tax Sign-Off Date Validity",
                "status": "pass" if signoff_date_valid else ("info" if not signoff_date else "warn"),
                "expected": "parseable date from period end through today",
                "observed": signoff_date_observed,
                "details": "Approved tax sign-off date should be parseable, on or after the reviewed period end, and not future-dated.",
            },
            {
                "check": "Approved Tax Sign-Off Advisor Evidence",
                "status": "pass" if advisor_evidence else "warn",
                "expected": "advisor/evidence link",
                "observed": advisor_evidence or "missing",
                "details": "Approved tax sign-off should reference advisor review evidence or workpaper context.",
            },
            {
                "check": "Approved Tax Sign-Off Packet Evidence",
                "status": "pass" if signoff_packet_ref or advisor_evidence else "warn",
                "expected": "tax packet reference or advisor evidence link",
                "observed": signoff_packet_ref or advisor_evidence or "missing",
                "details": "Approved tax sign-off should reference the Tax Review Packet or external evidence reviewed.",
            },
            {
                "check": "Approved Tax Sign-Off Packet Hash",
                "status": (
                    "pass"
                    if signoff_packet_hash and current_packet_hash and signoff_packet_hash == current_packet_hash
                    else ("info" if not current_packet_hash else "warn")
                ),
                "expected": current_packet_hash or "current packet hash unavailable",
                "observed": signoff_packet_hash or "missing",
                "details": "Approved tax sign-off packet hash should match the recalculated Tax Review Packet evidence hash.",
            },
        ]
    )
    return pd.DataFrame(rows)



def _build_colorado_suts_scope_summary_rows(
    reportable_tax_detail_df: pd.DataFrame,
    excluded_facilitator_tax_detail_df: pd.DataFrame,
    *,
    selected_marketplaces: set[str] | None,
    facilitator_channels: set[str] | None,
) -> list[dict[str, object]]:
    selected = {str(v or "").strip().lower() for v in (selected_marketplaces or set()) if str(v or "").strip()}
    facilitators = {str(v or "").strip().lower() for v in (facilitator_channels or set()) if str(v or "").strip()}

    def _summary_row(
        df: pd.DataFrame,
        *,
        scope: str,
        marketplace_scope: set[str],
        suts_treatment: str,
        note: str,
    ) -> dict[str, object]:
        if df is None or df.empty:
            sales_count = 0
            gross_sales = 0.0
            marketplaces_seen: list[str] = []
        else:
            sales_count = int(len(df))
            gross_sales = round(float(df.get("gross_sales", pd.Series(dtype=float)).fillna(0).astype(float).sum()), 2)
            marketplaces_seen = sorted(
                {
                    str(v or "").strip().lower()
                    for v in df.get("marketplace", pd.Series(dtype=str)).dropna().tolist()
                    if str(v or "").strip()
                }
            )
        return {
            "scope": scope,
            "marketplace_scope": ", ".join(sorted(marketplace_scope)) if marketplace_scope else "(none)",
            "marketplaces_seen": ", ".join(marketplaces_seen) if marketplaces_seen else "(none)",
            "sales_count": sales_count,
            "gross_sales": gross_sales,
            "suts_treatment": suts_treatment,
            "note": note,
        }

    excluded_facilitators = facilitators - selected
    rows = [
        _summary_row(
            reportable_tax_detail_df,
            scope="Reportable SUTS upload gross",
            marketplace_scope=selected,
            suts_treatment="included",
            note="Direct/local marketplace scope selected above; this is what the SUTS gross-sales rows use.",
        )
    ]
    if excluded_facilitators or (excluded_facilitator_tax_detail_df is not None and not excluded_facilitator_tax_detail_df.empty):
        rows.append(
            _summary_row(
                excluded_facilitator_tax_detail_df,
                scope="Marketplace facilitator gross",
                marketplace_scope=excluded_facilitators,
                suts_treatment="excluded by default",
                note="eBay/facilitator sales are kept out of normal SUTS remittance unless advisor-confirmed.",
            )
        )
    if selected.intersection(facilitators):
        rows.append(
            _summary_row(
                pd.DataFrame(),
                scope="Advisor override",
                marketplace_scope=selected.intersection(facilitators),
                suts_treatment="facilitator selected",
                note="A marketplace facilitator channel is selected; only submit that scope with advisor-confirmed reporting instructions.",
            )
        )
    return rows


def _suts_jurisdiction_key(jurisdiction_code: object, account_type: object) -> str:
    return (
        f"{str(jurisdiction_code or '').strip()}|"
        f"{str(account_type or '').strip().upper()}"
    )


def _build_colorado_suts_upload_workbook(
    tax_detail_df: pd.DataFrame,
    *,
    account_number: str = COLORADO_SUTS_ACCOUNT_NUMBER,
    gross_jurisdiction_code: str = "",
    gross_jurisdiction_key: str = "",
    gross_jurisdiction_codes: list[str] | None = None,
    gross_jurisdiction_keys: list[str] | None = None,
    zero_filing_jurisdiction_codes: list[str] | None = None,
    zero_filing_jurisdiction_keys: list[str] | None = None,
    account_number_by_jurisdiction_code: dict[str, str] | None = None,
    account_number_by_jurisdiction_key: dict[str, str] | None = None,
    allow_blank_account_jurisdiction_keys: set[str] | None = None,
    custom_jurisdictions: list[dict[str, object]] | None = None,
    remove_unselected_template_rows: bool = True,
    template_path: str = str(COLORADO_SUTS_TEMPLATE_PATH),
) -> tuple[bytes, pd.DataFrame]:
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Colorado SUTS template not found: {path}")
    wb = load_workbook(path, data_only=False)
    if "Upload Data" not in wb.sheetnames:
        raise ValueError("Colorado SUTS template must contain an `Upload Data` sheet.")
    ws = wb["Upload Data"]
    ws["A1"].value = None

    account_digits = "".join(ch for ch in str(account_number or "").strip() if ch.isdigit())
    if not account_digits:
        raise ValueError("Colorado SUTS account number is required.")
    gross_code = str(gross_jurisdiction_code or "").strip()
    gross_key = str(gross_jurisdiction_key or "").strip()
    gross_codes = {str(code).strip() for code in (gross_jurisdiction_codes or []) if str(code).strip()}
    gross_keys = {str(key).strip() for key in (gross_jurisdiction_keys or []) if str(key).strip()}
    if gross_code:
        gross_codes.add(gross_code)
    if gross_key:
        gross_keys.add(gross_key)
    zero_codes = {str(code).strip() for code in (zero_filing_jurisdiction_codes or []) if str(code).strip()}
    zero_keys = {str(key).strip() for key in (zero_filing_jurisdiction_keys or []) if str(key).strip()}
    zero_codes -= gross_codes
    zero_keys -= gross_keys
    account_override_digits_by_code = {
        str(code or "").strip(): "".join(ch for ch in str(value or "").strip() if ch.isdigit())
        for code, value in (account_number_by_jurisdiction_code or {}).items()
        if str(code or "").strip()
    }
    account_override_digits_by_key = {
        str(key or "").strip(): "".join(ch for ch in str(value or "").strip() if ch.isdigit())
        for key, value in (account_number_by_jurisdiction_key or {}).items()
        if str(key or "").strip()
    }
    allow_blank_keys = {str(key or "").strip() for key in (allow_blank_account_jurisdiction_keys or set()) if str(key or "").strip()}

    gross_sales_total = 0.0
    if tax_detail_df is not None and not tax_detail_df.empty and "gross_sales" in tax_detail_df.columns:
        gross_sales_total = float(tax_detail_df["gross_sales"].fillna(0).astype(float).sum())

    selected_codes = {*gross_codes, *zero_codes}
    selected_codes.discard("")
    selected_keys = {*gross_keys, *zero_keys}
    selected_keys.discard("")
    existing_keys = {
        _suts_jurisdiction_key(ws.cell(row_idx, 3).value, ws.cell(row_idx, 2).value)
        for row_idx in range(4, ws.max_row + 1)
    }
    for custom in custom_jurisdictions or []:
        code = str(custom.get("jurisdiction_code") or "").strip()
        account_type = str(custom.get("account_type") or "LOCAL").strip().upper() or "LOCAL"
        key = _suts_jurisdiction_key(code, account_type)
        if not code or key in existing_keys or (code not in selected_codes and key not in selected_keys):
            continue
        row_idx = ws.max_row + 1
        ws.cell(row_idx, 2).value = account_type
        ws.cell(row_idx, 3).value = int(code) if code.isdigit() else code
        ws.cell(row_idx, 4).value = str(custom.get("jurisdiction_name") or "Custom Jurisdiction").strip()
        existing_keys.add(key)

    summary_rows: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        jurisdiction_code = str(ws.cell(row_idx, 3).value or "").strip()
        account_type = str(ws.cell(row_idx, 2).value or "").strip().upper()
        jurisdiction_key = _suts_jurisdiction_key(jurisdiction_code, account_type)
        if (
            not jurisdiction_code
            or (jurisdiction_code not in selected_codes and jurisdiction_key not in selected_keys)
        ):
            continue
        if jurisdiction_key in account_override_digits_by_key:
            row_account_digits = account_override_digits_by_key.get(jurisdiction_key) or ""
        elif jurisdiction_code in account_override_digits_by_code:
            row_account_digits = account_override_digits_by_code.get(jurisdiction_code) or ""
        elif jurisdiction_key in allow_blank_keys:
            row_account_digits = ""
        else:
            row_account_digits = account_digits
        if not row_account_digits and jurisdiction_key not in allow_blank_keys:
            raise ValueError(f"Account number is required for Colorado SUTS jurisdiction {jurisdiction_code}.")
        account_cell = ws.cell(row_idx, 1)
        if row_account_digits:
            account_cell.value = int(row_account_digits)
            account_cell.number_format = "0" * max(1, len(row_account_digits))
        else:
            account_cell.value = None
            account_cell.number_format = "General"

        is_gross_row = (
            jurisdiction_key in gross_keys
            if gross_keys
            else jurisdiction_code in gross_codes
        )
        gross_text = _format_suts_gross_text(gross_sales_total if is_gross_row else 0.0)
        gross_cell = ws.cell(row_idx, 5)
        gross_cell.value = gross_text
        gross_cell.number_format = "@"
        summary_rows.append(
            {
                "account_number": row_account_digits,
                "account_type": account_type,
                "jurisdiction_code": jurisdiction_code,
                "jurisdiction_key": jurisdiction_key,
                "jurisdiction_name": str(ws.cell(row_idx, 4).value or "").strip(),
                "gross_amount": gross_text,
                "filing_type": "gross_sales" if is_gross_row else "zero_filing",
                "worksheet_row": row_idx,
            }
        )

    if (selected_codes or selected_keys) and not summary_rows:
        raise ValueError("No selected Colorado SUTS jurisdiction codes were found in the template.")
    if remove_unselected_template_rows:
        selected_row_indexes = {
            int(row["worksheet_row"])
            for row in summary_rows
            if int(row.get("worksheet_row") or 0) >= 4
        }
        for row_idx in range(ws.max_row, 3, -1):
            if row_idx not in selected_row_indexes:
                ws.delete_rows(row_idx)

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), pd.DataFrame(summary_rows)


def _colorado_suts_summary_warnings(summary_df: pd.DataFrame) -> list[str]:
    if summary_df is None or summary_df.empty:
        return ["No Colorado SUTS rows are selected for export."]
    warnings: list[str] = []
    rows = summary_df.to_dict(orient="records")
    gross_rows = [row for row in rows if str(row.get("filing_type") or "") == "gross_sales"]
    zero_rows = [row for row in rows if str(row.get("filing_type") or "") == "zero_filing"]
    if not gross_rows and not zero_rows:
        warnings.append("Selected SUTS rows do not contain a gross-sales row or zero-filing row.")
    for row in rows:
        account_type = str(row.get("account_type") or "").strip().upper()
        jurisdiction_code = str(row.get("jurisdiction_code") or "").strip()
        account_number = str(row.get("account_number") or "").strip()
        filing_type = str(row.get("filing_type") or "").strip()
        if account_type == "STATE" and jurisdiction_code == "110042" and not account_number:
            warnings.append("Golden STATE row is missing SUTS account 970074130001.")
        if filing_type == "gross_sales" and not account_number:
            warnings.append(
                f"{str(row.get('jurisdiction_name') or jurisdiction_code).strip()} "
                f"{account_type} gross-sales row is missing an account number."
            )
        if (
            account_type == "LOCAL"
            and jurisdiction_code == "110042"
            and filing_type == "gross_sales"
            and not account_number
        ):
            warnings.append(
                "Golden LOCAL gross-sales row has a blank account number. SUTS accepted blank account only "
                "for zero filing in observed testing; confirm whether nonzero self-collected Golden filing "
                "requires a local account before submitting."
            )
    duplicate_codes = {
        str(code)
        for code, count in summary_df["jurisdiction_code"].astype(str).value_counts().items()
        if int(count) > 1
    } if "jurisdiction_code" in summary_df.columns else set()
    for code in sorted(duplicate_codes):
        account_types = sorted(
            {
                str(row.get("account_type") or "").strip().upper()
                for row in rows
                if str(row.get("jurisdiction_code") or "").strip() == code
            }
        )
        if len(account_types) != len(
            [
                row
                for row in rows
                if str(row.get("jurisdiction_code") or "").strip() == code
            ]
        ):
            warnings.append(f"Jurisdiction {code} has duplicate rows with the same account type.")
    return warnings

def _build_tax_review_export_packet(
    *,
    reports: list[tuple[str, pd.DataFrame, str]],
    from_date,
    to_date,
    tax_jurisdiction: str,
    tax_rate_percent: float,
    shipping_taxable: bool,
    marketplace_scope: str,
    facilitator_channels: set[str] | None,
    tax_exempt_categories: set[str] | None,
    tax_profile: dict[str, object] | None = None,
    extra_artifacts: list[tuple[str, bytes]] | None = None,
) -> bytes:
    extra_artifact_hashes = {
        str(name or "").strip(): hashlib.sha256(payload or b"").hexdigest()
        for name, payload in (extra_artifacts or [])
        if str(name or "").strip()
    }
    evidence_hash = _tax_review_packet_evidence_hash_from_reports(
        reports=reports,
        from_date=from_date,
        to_date=to_date,
        tax_jurisdiction=tax_jurisdiction,
        tax_rate_percent=tax_rate_percent,
        shipping_taxable=shipping_taxable,
        marketplace_scope=marketplace_scope,
        facilitator_channels=facilitator_channels,
        tax_exempt_categories=tax_exempt_categories,
        tax_profile=tax_profile,
        extra_artifact_hashes=extra_artifact_hashes,
    )
    tax_report_prefixes = tax_review_packet_prefixes()
    selected_reports = [
        (label, df, prefix)
        for label, df, prefix in reports
        if str(prefix or "").strip() in tax_report_prefixes
    ]
    report_csv_by_prefix: dict[str, str] = {}
    for _label, df, prefix in selected_reports:
        export_df = df if df is not None else pd.DataFrame()
        report_csv_by_prefix[str(prefix)] = export_df.to_csv(index=False)
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    manifest_rows = [
        {"key": "generated_at", "value": generated_at},
        {"key": "tax_packet_evidence_hash_sha256", "value": evidence_hash},
        {"key": "from_date", "value": str(from_date)},
        {"key": "to_date", "value": str(to_date)},
        {"key": "tax_jurisdiction", "value": str(tax_jurisdiction or "")},
        {"key": "tax_rate_percent", "value": str(float(tax_rate_percent or 0.0))},
        {"key": "shipping_taxable", "value": str(bool(shipping_taxable)).lower()},
        {"key": "marketplace_scope", "value": str(marketplace_scope or "all")},
        {"key": "facilitator_channels", "value": ",".join(sorted(facilitator_channels or set()))},
        {"key": "tax_exempt_categories", "value": ",".join(sorted(tax_exempt_categories or set()))},
    ]
    if tax_profile:
        for key in [
            "profile_key",
            "profile_name",
            "effective_from",
            "effective_to",
            "human_validation_status",
            "advisor_evidence_link",
        ]:
            manifest_rows.append({"key": f"tax_profile_{key}", "value": str(tax_profile.get(key) or "")})
    for label, df, prefix in selected_reports:
        manifest_rows.append(
            {
                "key": f"row_count_{prefix}",
                "value": str(0 if df is None else int(len(df))),
            }
        )
    for name, payload in extra_artifacts or []:
        safe_name = str(name or "").strip()
        if not safe_name:
            continue
        manifest_rows.append({"key": f"artifact_{safe_name}", "value": str(len(payload or b""))})
        manifest_rows.append(
            {
                "key": f"sha256_{safe_name}",
                "value": hashlib.sha256(payload or b"").hexdigest(),
            }
        )

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as packet:
        packet.writestr("manifest.csv", pd.DataFrame(manifest_rows).to_csv(index=False))
        packet.writestr(
            "README.txt",
            (
                "GoldenStackers tax review export packet.\n"
                "Tax outputs are estimates for operational planning and advisor review.\n"
                "Validate local/state treatment, marketplace facilitator rules, shipping taxability, "
                "and bullion/coin exemptions with your tax advisor before filing or remittance decisions.\n"
            ),
        )
        for label, df, prefix in selected_reports:
            packet.writestr(f"{prefix}.csv", report_csv_by_prefix.get(str(prefix), ""))
        for name, payload in extra_artifacts or []:
            safe_name = str(name or "").strip()
            if safe_name:
                packet.writestr(safe_name, payload or b"")
    buffer.seek(0)
    return buffer.getvalue()


def _tax_review_packet_evidence_hash_from_reports(
    *,
    reports: list[tuple[str, pd.DataFrame, str]],
    from_date,
    to_date,
    tax_jurisdiction: str,
    tax_rate_percent: float,
    shipping_taxable: bool,
    marketplace_scope: str,
    facilitator_channels: set[str] | None,
    tax_exempt_categories: set[str] | None,
    tax_profile: dict[str, object] | None = None,
    extra_artifact_hashes: dict[str, str] | None = None,
) -> str:
    tax_report_prefixes = tax_review_packet_prefixes()
    report_csv_by_prefix: dict[str, str] = {}
    for _label, df, prefix in reports:
        if str(prefix or "").strip() not in tax_report_prefixes:
            continue
        export_df = df if df is not None else pd.DataFrame()
        report_csv_by_prefix[str(prefix)] = export_df.to_csv(index=False)
    hash_payload = {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "tax_jurisdiction": str(tax_jurisdiction or ""),
        "tax_rate_percent": float(tax_rate_percent or 0.0),
        "shipping_taxable": bool(shipping_taxable),
        "marketplace_scope": str(marketplace_scope or "all"),
        "facilitator_channels": sorted(facilitator_channels or set()),
        "tax_exempt_categories": sorted(tax_exempt_categories or set()),
        "tax_profile": tax_profile or {},
        "extra_artifact_hashes": extra_artifact_hashes or {},
        "reports": report_csv_by_prefix,
    }
    return hashlib.sha256(json.dumps(hash_payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()
