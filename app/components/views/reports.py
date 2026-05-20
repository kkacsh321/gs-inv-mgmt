from datetime import datetime, timedelta, timezone
from collections import defaultdict, deque
from io import BytesIO
import hashlib
import json
from pathlib import Path
import time
from types import SimpleNamespace
import zipfile

from openpyxl import load_workbook
import pandas as pd
import streamlit as st
from sqlalchemy import select

from app.auth import current_user, ensure_permission, has_permission
from app.db.models import AuditLog, OrderFinanceEntry
from app.config import settings
from app.components.ui_helpers import iso_or_none
from app.components.views.shared import (
    dataframe_to_xlsx_bytes,
    handoff_to_documents_draft,
    render_help_panel,
)
from app.components.views.workspace_shell import render_workspace_feedback
from app.repository import InventoryRepository
from app.services.ai_orchestration import execute_comp_summary
from app.services.fee_calibration import build_final_value_rate_calibration
from app.services.fee_reconciliation import build_ebay_fee_reconciliation_rows
from app.services.runtime_settings import get_runtime_bool, get_runtime_float, get_runtime_str
from app.utils.time import utc_today

COLORADO_SUTS_ACCOUNT_NUMBER = "080390"
COLORADO_SUTS_GOLDEN_STATE_ACCOUNT_NUMBER = "970074130001"
COLORADO_SUTS_TEMPLATE_PATH = (
    Path(__file__).resolve().parents[2]
    / "assets"
    / "templates"
    / "CO-SUTS-Excel-Template-127596.xlsx"
)

ACCOUNTING_FIELD_SEMANTICS_ROWS = [
    {
        "field": "products.acquisition_cost",
        "meaning": "Per-unit product acquisition basis before optional tax, shipping, and handling landed components.",
        "accounting_use": "Fallback unit COGS/inventory value when no product-lot cost history is available.",
    },
    {
        "field": "products.acquisition_tax_paid / shipping_paid / handling_paid",
        "meaning": "Per-unit landed cost components attached directly to the product record.",
        "accounting_use": "Added to acquisition cost when product-level landed cost is used.",
    },
    {
        "field": "products.product_cost",
        "meaning": "Fallback per-unit internal/product basis for converted/intake items when acquisition cost is blank.",
        "accounting_use": "Last-resort product-level COGS/inventory value fallback after lot and acquisition landed cost.",
    },
    {
        "field": "purchase_lots.total_cost / tax / shipping / handling",
        "meaning": "Whole-lot landed cost for one purchase lot that may contain multiple product SKUs.",
        "accounting_use": "Allocated only to blank-cost product-lot assignment quantity after explicit assignment costs are honored; if expected lot quantity is set, partial check-ins use that denominator.",
    },
    {
        "field": "purchase_lots.expected_total_quantity",
        "meaning": "Optional expected unit/item count for a purchase lot whose full contents have not been checked in yet.",
        "accounting_use": "Prevents early blank-cost assignments from absorbing the full lot cost before all products are assigned.",
    },
    {
        "field": "product_lot_assignments.allocation_weight",
        "meaning": "Optional proportional share used for mixed lots with different product values when per-product dollar cost is not known.",
        "accounting_use": "Splits remaining whole-lot landed cost across blank-cost assignments by weight before falling back to equal quantity allocation.",
    },
    {
        "field": "product_lot_assignments.unit_cost / unit_tax / unit_shipping / unit_handling",
        "meaning": "Per-unit landed components for a specific product quantity acquired in a lot.",
        "accounting_use": "Highest-priority COGS source for that lot quantity in time-aware FIFO sale costing.",
    },
    {
        "field": "product_lot_assignments.allocated_cost / allocated_tax / allocated_shipping / allocated_handling",
        "meaning": "Total allocated landed components for that assignment quantity.",
        "accounting_use": "Used as assignment unit cost when per-unit assignment components are blank.",
    },
    {
        "field": "FIFO remaining lot cost",
        "meaning": "Remaining unsold cost from lots already available by acquisition date.",
        "accounting_use": "Inventory value source when product has lot history; prevents current inventory from defaulting to stale weighted product cost.",
    },
]


def _safe_float(value) -> float:
    try:
        if value is None:
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _sale_listing_bundle_summary(sale: object) -> dict[str, object]:
    listing = getattr(sale, "listing", None)
    bundle_payload = InventoryRepository._listing_bundle_payload(listing)
    components = InventoryRepository._bundle_components_from_payload(
        bundle_payload,
        int(getattr(sale, "quantity_sold", 0) or 0),
    )
    return {
        "listing_is_bundle": bool(components),
        "listing_bundle_kind": str(bundle_payload.get("kind") or "").strip(),
        "listing_bundle_component_count": len(components),
        "listing_bundle_units_per_listing": sum(
            max(1, int(component.get("quantity_per_listing") or 1))
            for component in components
        ),
        "listing_bundle_inventory_units_sold": sum(
            max(1, int(component.get("quantity_total") or 1))
            for component in components
        ),
    }


def _return_listing_bundle_summary(ret: object) -> dict[str, object]:
    if isinstance(ret, dict):
        return {
            "listing_is_bundle": bool(ret.get("listing_is_bundle")),
            "listing_bundle_kind": str(ret.get("listing_bundle_kind") or "").strip(),
            "listing_bundle_component_count": int(_safe_float(ret.get("listing_bundle_component_count"))),
            "listing_bundle_units_per_return": int(_safe_float(ret.get("listing_bundle_units_per_return"))),
            "listing_bundle_inventory_units_returned": int(
                _safe_float(ret.get("listing_bundle_inventory_units_returned"))
            ),
        }
    sale = getattr(ret, "sale", None)
    listing = getattr(sale, "listing", None)
    bundle_payload = InventoryRepository._listing_bundle_payload(listing)
    components = InventoryRepository._bundle_components_from_payload(
        bundle_payload,
        int(getattr(ret, "quantity", 0) or 0),
    )
    return {
        "listing_is_bundle": bool(components),
        "listing_bundle_kind": str(bundle_payload.get("kind") or "").strip(),
        "listing_bundle_component_count": len(components),
        "listing_bundle_units_per_return": sum(
            max(1, int(component.get("quantity_per_listing") or 1))
            for component in components
        ),
        "listing_bundle_inventory_units_returned": sum(
            max(1, int(component.get("quantity_total") or 1))
            for component in components
        ),
    }


def _audit_changes(row) -> dict:
    try:
        payload = json.loads(str(getattr(row, "changes_json", "") or "{}"))
    except Exception:
        payload = {}
    return payload if isinstance(payload, dict) else {}


def _truthy(value) -> bool:
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y", "on"}
    return bool(value)


def _tax_profile_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    profiles: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows or []:
        payload = _audit_changes(row)
        profile_key = str(payload.get("profile_key") or "").strip().lower()
        if not profile_key or profile_key in seen:
            continue
        seen.add(profile_key)
        if not _truthy(payload.get("is_active", True)):
            continue
        profiles.append(
            {
                "profile_key": profile_key,
                "profile_name": str(payload.get("profile_name") or profile_key).strip(),
                "jurisdiction": str(payload.get("jurisdiction") or "").strip(),
                "tax_rate_percent": _safe_float(payload.get("tax_rate_percent")),
                "shipping_taxable": _truthy(payload.get("shipping_taxable")),
                "facilitator_channels": str(payload.get("facilitator_channels") or "").strip(),
                "tax_exempt_categories": str(payload.get("tax_exempt_categories") or "").strip(),
                "effective_from": str(payload.get("effective_from") or "").strip(),
                "effective_to": str(payload.get("effective_to") or "").strip(),
                "human_validation_status": str(payload.get("human_validation_status") or "").strip().lower(),
                "advisor_evidence_link": str(payload.get("advisor_evidence_link") or "").strip(),
            }
        )
    return profiles


def _latest_tax_profile_rows(repo: InventoryRepository, *, limit: int = 100) -> list[dict[str, object]]:
    db = getattr(repo, "db", None)
    if db is None or not hasattr(db, "scalars"):
        return []
    try:
        rows = db.scalars(
            select(AuditLog)
            .where(AuditLog.entity_type == "tax_profile")
            .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
            .limit(max(1, int(limit)))
        ).all()
        return _tax_profile_rows_from_audit_logs(rows)
    except Exception:
        if hasattr(db, "rollback"):
            db.rollback()
        return []


def _tax_signoff_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for row in rows or []:
        payload = _audit_changes(row)
        output.append(
            {
                "recorded_at_utc": (
                    getattr(row, "created_at").isoformat(timespec="seconds")
                    if getattr(row, "created_at", None)
                    else ""
                ),
                "actor": str(getattr(row, "actor", "") or ""),
                "target_env": str(payload.get("target_env") or ""),
                "tax_period": str(payload.get("tax_period") or ""),
                "jurisdiction": str(payload.get("jurisdiction") or ""),
                "profile_key": str(payload.get("profile_key") or ""),
                "status": str(payload.get("status") or "").strip().lower(),
                "owner": str(payload.get("owner") or ""),
                "signoff_date": str(payload.get("signoff_date") or ""),
                "tax_packet_ref": str(payload.get("tax_packet_ref") or ""),
                "tax_packet_hash": str(
                    payload.get("tax_packet_hash")
                    or payload.get("tax_packet_evidence_hash_sha256")
                    or ""
                ).strip(),
                "advisor_evidence_link": str(payload.get("advisor_evidence_link") or ""),
                "tax_exception_count": int(_safe_float(payload.get("tax_exception_count"))),
                "notes": str(payload.get("notes") or "")[:220],
            }
        )
    return output


def _latest_tax_signoff_rows(repo: InventoryRepository, *, limit: int = 100) -> list[dict[str, object]]:
    db = getattr(repo, "db", None)
    if db is None or not hasattr(db, "scalars"):
        return []
    try:
        rows = db.scalars(
            select(AuditLog)
            .where(AuditLog.entity_type == "tax_reporting_signoff")
            .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
            .limit(max(1, int(limit)))
        ).all()
        return _tax_signoff_rows_from_audit_logs(rows)
    except Exception:
        if hasattr(db, "rollback"):
            db.rollback()
        return []


def _accounting_close_signoff_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for row in rows or []:
        payload = _audit_changes(row)
        output.append(
            {
                "recorded_at_utc": (
                    getattr(row, "created_at").isoformat(timespec="seconds")
                    if getattr(row, "created_at", None)
                    else ""
                ),
                "actor": str(getattr(row, "actor", "") or ""),
                "target_env": str(payload.get("target_env") or "").strip().lower(),
                "signoff_type": str(payload.get("signoff_type") or "").strip().lower(),
                "close_period": str(payload.get("close_period") or "").strip(),
                "status": str(payload.get("status") or "").strip().lower(),
                "owner": str(payload.get("owner") or "").strip(),
                "signoff_date": str(payload.get("signoff_date") or "").strip(),
                "close_readiness_status": str(payload.get("close_readiness_status") or "").strip(),
                "exception_count": int(_safe_float(payload.get("exception_count"))),
                "unresolved_blocker_count": int(_safe_float(payload.get("unresolved_blocker_count"))),
                "period_drift_warn_count": int(_safe_float(payload.get("period_drift_warn_count"))),
                "ai_review_followup_count": int(_safe_float(payload.get("ai_review_followup_count"))),
                "accounting_packet_ref": str(payload.get("accounting_packet_ref") or "").strip(),
                "accounting_packet_hash": str(
                    payload.get("accounting_packet_hash")
                    or payload.get("accounting_close_packet_evidence_hash_sha256")
                    or ""
                ).strip(),
                "evidence_link": str(payload.get("evidence_link") or "").strip(),
                "notes": str(payload.get("notes") or "")[:220],
            }
        )
    return output


def _latest_accounting_close_signoff_rows(repo: InventoryRepository, *, limit: int = 100) -> list[dict[str, object]]:
    db = getattr(repo, "db", None)
    if db is None or not hasattr(db, "scalars"):
        return []
    try:
        rows = db.scalars(
            select(AuditLog)
            .where(AuditLog.entity_type == "accounting_close_signoff")
            .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
            .limit(max(1, int(limit)))
        ).all()
        return _accounting_close_signoff_rows_from_audit_logs(rows)
    except Exception:
        if hasattr(db, "rollback"):
            db.rollback()
        return []


def _ai_review_outcome_rows_from_audit_logs(rows) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for row in rows or []:
        payload = _audit_changes(row)
        after = payload.get("after", {}) if isinstance(payload, dict) else {}
        if not isinstance(after, dict):
            continue
        metadata = after.get("metadata", {})
        if not isinstance(metadata, dict):
            metadata = {}
        event_type = str(metadata.get("event_type") or "").strip().lower()
        if not event_type.endswith("_outcome"):
            continue
        output.append(
            {
                "recorded_at_utc": (
                    getattr(row, "created_at").isoformat(timespec="seconds")
                    if getattr(row, "created_at", None)
                    else ""
                ),
                "actor": str(getattr(row, "actor", "") or ""),
                "review_type": str(metadata.get("review_type") or event_type.removesuffix("_outcome")).strip(),
                "outcome": str(metadata.get("outcome") or "").strip().lower(),
                "prompt_hash_sha256": str(metadata.get("prompt_hash_sha256") or "").strip(),
                "data_scope_hash_sha256": str(metadata.get("data_scope_hash_sha256") or "").strip(),
                "answer_hash_sha256": str(metadata.get("answer_hash_sha256") or "").strip(),
                "accounting_close_packet_evidence_hash": str(
                    (metadata.get("data_scope") or {}).get("accounting_close_packet_evidence_hash")
                    if isinstance(metadata.get("data_scope"), dict)
                    else ""
                ).strip(),
                "tax_packet_evidence_hash": str(
                    (metadata.get("data_scope") or {}).get("tax_packet_evidence_hash")
                    if isinstance(metadata.get("data_scope"), dict)
                    else ""
                ).strip(),
                "intent": str(after.get("intent") or "").strip(),
                "answer_preview": str(after.get("answer_preview") or "")[:220],
            }
        )
    return output


def _latest_ai_review_outcome_rows(repo: InventoryRepository, *, limit: int = 100) -> list[dict[str, object]]:
    db = getattr(repo, "db", None)
    if db is None or not hasattr(db, "scalars"):
        return []
    try:
        rows = db.scalars(
            select(AuditLog)
            .where(AuditLog.entity_type == "ai_chat")
            .order_by(AuditLog.created_at.desc(), AuditLog.id.desc())
            .limit(max(1, int(limit)))
        ).all()
        return _ai_review_outcome_rows_from_audit_logs(rows)
    except Exception:
        if hasattr(db, "rollback"):
            db.rollback()
        return []


def _default_accounting_close_period(from_date, to_date) -> str:
    try:
        if getattr(from_date, "year", None) == getattr(to_date, "year", None) and getattr(
            from_date, "month", None
        ) == getattr(to_date, "month", None):
            return f"{int(from_date.year):04d}-{int(from_date.month):02d}"
    except Exception:
        pass
    return f"{from_date}..{to_date}"


def _build_accounting_close_signoff_payload(
    *,
    target_env: str,
    close_period: str,
    status: str,
    owner: str,
    signoff_date,
    close_summary: dict[str, float | int | str],
    accounting_packet_ref: str,
    accounting_packet_hash: str,
    evidence_link: str = "",
    notes: str = "",
    signoff_type: str = "monthly_close_review",
) -> dict[str, object]:
    summary = close_summary or {}
    return {
        "target_env": str(target_env or "").strip().lower(),
        "signoff_type": str(signoff_type or "monthly_close_review").strip().lower(),
        "close_period": str(close_period or "").strip(),
        "status": str(status or "").strip().lower(),
        "owner": str(owner or "").strip(),
        "signoff_date": signoff_date.isoformat() if hasattr(signoff_date, "isoformat") else str(signoff_date or ""),
        "close_readiness_status": str(summary.get("readiness_status") or "").strip().lower(),
        "exception_count": int(_safe_float(summary.get("total_exceptions"))),
        "unresolved_blocker_count": int(_safe_float(summary.get("blocker_count"))),
        "period_drift_warn_count": int(_safe_float(summary.get("period_drift_warn_count"))),
        "ai_review_followup_count": int(_safe_float(summary.get("ai_review_followup_count"))),
        "accounting_packet_ref": str(accounting_packet_ref or "").strip(),
        "accounting_packet_hash": str(accounting_packet_hash or "").strip(),
        "accounting_close_packet_evidence_hash_sha256": str(accounting_packet_hash or "").strip(),
        "evidence_link": str(evidence_link or "").strip(),
        "notes": str(notes or "").strip(),
    }


def _build_tax_reporting_signoff_payload(
    *,
    target_env: str,
    tax_period: str,
    jurisdiction: str,
    profile_key: str,
    status: str,
    owner: str,
    signoff_date,
    tax_packet_ref: str,
    tax_packet_hash: str,
    advisor_evidence_link: str,
    tax_exception_count: int,
    notes: str = "",
) -> dict[str, object]:
    return {
        "target_env": str(target_env or "").strip().lower(),
        "tax_period": str(tax_period or "").strip(),
        "jurisdiction": str(jurisdiction or "").strip(),
        "profile_key": str(profile_key or "").strip().lower(),
        "status": str(status or "").strip().lower(),
        "owner": str(owner or "").strip(),
        "signoff_date": signoff_date.isoformat() if hasattr(signoff_date, "isoformat") else str(signoff_date or ""),
        "tax_packet_ref": str(tax_packet_ref or "").strip(),
        "tax_packet_hash": str(tax_packet_hash or "").strip(),
        "tax_packet_evidence_hash_sha256": str(tax_packet_hash or "").strip(),
        "advisor_evidence_link": str(advisor_evidence_link or "").strip(),
        "tax_exception_count": int(tax_exception_count or 0),
        "notes": str(notes or "").strip(),
    }


def _build_tax_reporting_signoff_review(
    *,
    signoff_df: pd.DataFrame,
    tax_period: str,
    jurisdiction: str,
    profile_key: str,
    tax_exception_count: int,
    current_packet_hash: str = "",
    to_date=None,
) -> pd.DataFrame:
    period = str(tax_period or "").strip()
    jurisdiction_norm = str(jurisdiction or "").strip().lower()
    profile_key_norm = str(profile_key or "").strip().lower()

    def _col(name: str) -> pd.Series:
        if signoff_df is None or signoff_df.empty or name not in signoff_df.columns:
            return pd.Series(dtype=str)
        return signoff_df[name].fillna("").astype(str)

    matching_df = pd.DataFrame()
    if signoff_df is not None and not signoff_df.empty:
        period_series = _col("tax_period").str.strip()
        matching_df = signoff_df.loc[period_series == period].copy()
    approved_df = pd.DataFrame()
    if not matching_df.empty and "status" in matching_df.columns:
        approved_df = matching_df.loc[matching_df["status"].fillna("").astype(str).str.lower() == "approved"].copy()
        if not approved_df.empty:
            approved_df["_signoff_sort_date"] = pd.to_datetime(
                approved_df.get("signoff_date", pd.Series(dtype=str)),
                errors="coerce",
                utc=True,
            )
            approved_df["_recorded_sort_date"] = pd.to_datetime(
                approved_df.get("recorded_at_utc", pd.Series(dtype=str)),
                errors="coerce",
                utc=True,
            )
            approved_df = approved_df.sort_values(
                ["_signoff_sort_date", "_recorded_sort_date"],
                ascending=[False, False],
                na_position="last",
                kind="mergesort",
            ).drop(columns=["_signoff_sort_date", "_recorded_sort_date"])

    rows: list[dict[str, object]] = [
        {
            "check": "Tax Sign-Off Evidence Present",
            "status": "pass" if not approved_df.empty else ("warn" if int(tax_exception_count or 0) == 0 else "info"),
            "expected": "approved tax reporting sign-off for selected period",
            "observed": "approved" if not approved_df.empty else "missing",
            "details": (
                "Approved tax reporting sign-off evidence found for the selected period."
                if not approved_df.empty
                else "No approved tax reporting sign-off evidence found for the selected period."
            ),
        }
    ]
    if approved_df.empty:
        return pd.DataFrame(rows)

    latest = approved_df.iloc[0].to_dict()
    signoff_jurisdiction = str(latest.get("jurisdiction") or "").strip().lower()
    signoff_profile_key = str(latest.get("profile_key") or "").strip().lower()
    signoff_exception_count = int(_safe_float(latest.get("tax_exception_count")))
    signoff_packet_ref = str(latest.get("tax_packet_ref") or "").strip()
    signoff_packet_hash = str(latest.get("tax_packet_hash") or "").strip().lower()
    current_packet_hash = str(current_packet_hash or "").strip().lower()
    advisor_evidence = str(latest.get("advisor_evidence_link") or "").strip()
    signoff_owner = str(latest.get("owner") or "").strip()
    signoff_date = str(latest.get("signoff_date") or "").strip()

    def _coerce_review_date(value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if all(hasattr(value, attr) for attr in ("year", "month", "day")):
            try:
                return datetime(int(value.year), int(value.month), int(value.day)).date()
            except Exception:
                return None
        raw = str(value).strip()
        if not raw:
            return None
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
        except Exception:
            try:
                return datetime.fromisoformat(raw[:10]).date()
            except Exception:
                return None

    signoff_review_date = _coerce_review_date(signoff_date)
    period_end_date = _coerce_review_date(to_date)
    today = utc_today()
    signoff_date_valid = bool(
        signoff_review_date
        and (period_end_date is None or signoff_review_date >= period_end_date)
        and signoff_review_date <= today
    )
    if not signoff_date:
        signoff_date_observed = "missing"
    elif signoff_review_date is None:
        signoff_date_observed = f"{signoff_date} (unparseable)"
    elif period_end_date is not None and signoff_review_date < period_end_date:
        signoff_date_observed = f"{signoff_review_date.isoformat()} before {period_end_date.isoformat()}"
    elif signoff_review_date > today:
        signoff_date_observed = f"{signoff_review_date.isoformat()} after {today.isoformat()}"
    else:
        signoff_date_observed = signoff_review_date.isoformat()

    rows.extend(
        [
            {
                "check": "Approved Tax Sign-Off Jurisdiction Match",
                "status": "pass" if signoff_jurisdiction == jurisdiction_norm else "warn",
                "expected": jurisdiction_norm or "missing",
                "observed": signoff_jurisdiction or "missing",
                "details": "Approved tax sign-off jurisdiction should match the selected tax review assumptions.",
            },
            {
                "check": "Approved Tax Sign-Off Profile Match",
                "status": "pass" if signoff_profile_key == profile_key_norm else ("info" if not profile_key_norm else "warn"),
                "expected": profile_key_norm or "no selected profile",
                "observed": signoff_profile_key or "missing",
                "details": "Approved tax sign-off profile key should match the selected saved tax profile when one is used.",
            },
            {
                "check": "Approved Tax Sign-Off Exception Count",
                "status": "pass" if signoff_exception_count == int(tax_exception_count or 0) else "warn",
                "expected": int(tax_exception_count or 0),
                "observed": signoff_exception_count,
                "details": "Approved tax sign-off exception count should match recalculated Tax Exceptions / Advisor Review rows.",
            },
            {
                "check": "Approved Tax Sign-Off Owner Present",
                "status": "pass" if signoff_owner else "warn",
                "expected": "review owner",
                "observed": signoff_owner or "missing",
                "details": "Approved tax sign-off should identify the reviewer or owner.",
            },
            {
                "check": "Approved Tax Sign-Off Date Present",
                "status": "pass" if signoff_date else "warn",
                "expected": "sign-off date",
                "observed": signoff_date or "missing",
                "details": "Approved tax sign-off should include the approval date.",
            },
            {
                "check": "Approved Tax Sign-Off Date Validity",
                "status": "pass" if signoff_date_valid else ("info" if not signoff_date else "warn"),
                "expected": "parseable date from period end through today",
                "observed": signoff_date_observed,
                "details": "Approved tax sign-off date should be parseable, on or after the reviewed period end, and not future-dated.",
            },
            {
                "check": "Approved Tax Sign-Off Advisor Evidence",
                "status": "pass" if advisor_evidence else "warn",
                "expected": "advisor/evidence link",
                "observed": advisor_evidence or "missing",
                "details": "Approved tax sign-off should reference advisor review evidence or workpaper context.",
            },
            {
                "check": "Approved Tax Sign-Off Packet Evidence",
                "status": "pass" if signoff_packet_ref or advisor_evidence else "warn",
                "expected": "tax packet reference or advisor evidence link",
                "observed": signoff_packet_ref or advisor_evidence or "missing",
                "details": "Approved tax sign-off should reference the Tax Review Packet or external evidence reviewed.",
            },
            {
                "check": "Approved Tax Sign-Off Packet Hash",
                "status": (
                    "pass"
                    if signoff_packet_hash and current_packet_hash and signoff_packet_hash == current_packet_hash
                    else ("info" if not current_packet_hash else "warn")
                ),
                "expected": current_packet_hash or "current packet hash unavailable",
                "observed": signoff_packet_hash or "missing",
                "details": "Approved tax sign-off packet hash should match the recalculated Tax Review Packet evidence hash.",
            },
        ]
    )
    return pd.DataFrame(rows)


def _sale_net_before_cogs_from_fields(sale) -> float:
    return (
        _safe_float(getattr(sale, "sold_price", None))
        + _safe_float(getattr(sale, "shipping_cost", None))
        - _safe_float(getattr(sale, "fees", None))
        - _safe_float(getattr(sale, "shipping_label_cost", None))
    )


def _build_qbo_sales_export_rows(
    sales,
    fifo_unit_cost_by_sale: dict[int, float],
    fifo_unit_cost_source_by_sale: dict[int, str] | None = None,
    actual_econ_by_sale_id: dict[int, dict[str, float | str]] | None = None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    actual_by_sale = actual_econ_by_sale_id or {}
    cogs_source_by_sale = fifo_unit_cost_source_by_sale or {}
    for s in sales:
        sale_id = int(getattr(s, "id", 0) or 0)
        actual = actual_by_sale.get(sale_id) or {}
        bundle_summary = _sale_listing_bundle_summary(s)
        cogs_estimate = _safe_float(fifo_unit_cost_by_sale.get(sale_id)) * int(getattr(s, "quantity_sold", 0) or 0)
        cogs_source = str(cogs_source_by_sale.get(sale_id) or "missing_cost_basis").strip() or "missing_cost_basis"
        fees = (
            _safe_float(actual.get("allocated_fee_actual"))
            if actual
            else _safe_float(getattr(s, "fees", None))
        )
        shipping_charged = (
            _safe_float(actual.get("allocated_shipping_charged"))
            if actual and "allocated_shipping_charged" in actual
            else _safe_float(getattr(s, "shipping_cost", None))
        )
        shipping_label = (
            _safe_float(actual.get("allocated_shipping_actual"))
            if actual
            else _safe_float(getattr(s, "shipping_label_cost", None))
        )
        net_amount = (
            _safe_float(actual.get("net_before_cogs_actual"))
            if actual
            else (
                _safe_float(getattr(s, "sold_price", None))
                + shipping_charged
                - fees
                - shipping_label
            )
        )
        quantity = int(getattr(s, "quantity_sold", 0) or 0)
        sold_price = _safe_float(getattr(s, "sold_price", None))
        product = getattr(s, "product", None)
        listing = getattr(s, "listing", None)
        listing_product = getattr(listing, "product", None) if listing is not None else None
        item_product = product or listing_product
        item_product_source = (
            "sale_product"
            if product is not None
            else ("listing_product" if listing_product is not None else "missing_product")
        )
        profit_before_returns_estimate = round(float(net_amount - cogs_estimate), 2)
        rows.append(
            {
                "txn_date": s.sold_at.date().isoformat() if getattr(s, "sold_at", None) else "",
                "doc_number": getattr(s, "external_order_id", "") or f"SALE-{sale_id}",
                "customer_ref": str(getattr(s, "marketplace", "") or "").upper(),
                "item_sku": getattr(item_product, "sku", "") if item_product else "",
                "item_description": getattr(item_product, "title", "") if item_product else "",
                "item_product_source": item_product_source,
                "quantity": quantity,
                "rate": float(sold_price / quantity) if quantity else float(sold_price),
                "amount": float(sold_price),
                "fees": round(float(fees), 2),
                "fee_source": str(actual.get("actual_fee_source") or "sale_fees_field"),
                "shipping_cost": round(float(shipping_charged), 2),
                "shipping_label_cost": round(float(shipping_label), 2),
                "shipping_label_source": str(actual.get("actual_shipping_source") or "sale_shipping_label_field"),
                "tracking_number": getattr(s, "tracking_number", ""),
                "tracking_status": getattr(s, "tracking_status", ""),
                "cogs_input_estimate": round(float(cogs_estimate), 2),
                "cogs_source": cogs_source,
                **bundle_summary,
                "gross_margin_estimate": profit_before_returns_estimate,
                "profit_before_returns_estimate": profit_before_returns_estimate,
                "net_amount": round(float(net_amount), 2),
                "marketplace": getattr(s, "marketplace", ""),
            }
        )
    return rows


def _build_qbo_adjustment_export_rows(
    returns,
    fifo_unit_cost_by_sale: dict[int, float],
    fifo_unit_cost_source_by_sale: dict[int, str] | None = None,
) -> list[dict[str, object]]:
    cogs_source_by_sale = fifo_unit_cost_source_by_sale or {}
    rows: list[dict[str, object]] = []
    for ret in returns:
        is_dict = isinstance(ret, dict)
        getter = ret.get if is_dict else lambda key, default=None: getattr(ret, key, default)
        bundle_summary = _return_listing_bundle_summary(ret)
        sale_id = int(getter("sale_id", 0) or 0)
        quantity = max(0, int(getter("quantity", 0) or 0))
        refund_amount = _safe_float(getter("refund_amount", 0.0))
        refund_fees = _safe_float(getter("refund_fees", 0.0))
        refund_shipping = _safe_float(getter("refund_shipping", 0.0))
        refund_total = refund_amount + refund_fees + refund_shipping
        cogs_per_returned_listing = _safe_float(fifo_unit_cost_by_sale.get(sale_id))
        returned_cogs = cogs_per_returned_listing * quantity
        restocked = bool(getter("restocked", False))
        cogs_reversal = returned_cogs if restocked else 0.0
        returned_inventory_units = int(bundle_summary.get("listing_bundle_inventory_units_returned") or 0)
        if returned_inventory_units <= 0:
            returned_inventory_units = quantity
        returned_at = getter("returned_at", None)
        if hasattr(returned_at, "date"):
            txn_date = returned_at.date().isoformat()
        else:
            txn_date = str(returned_at or "")[:10]
        return_id = int(getter("return_id", 0) or getter("id", 0) or 0)
        doc_number = str(getter("external_return_id", "") or "").strip() or f"RETURN-{return_id}"
        source_order = str(getter("source_order", "") or "").strip()
        if not source_order and not is_dict:
            sale = getattr(ret, "sale", None)
            source_order = str(getattr(sale, "external_order_id", "") or "").strip() if sale else ""
        sku = str(getter("sku", "") or "").strip()
        sku_source = "return_product" if sku else "missing_product"
        if not sku and not is_dict:
            product = getattr(ret, "product", None)
            sale = getattr(ret, "sale", None)
            listing = getattr(sale, "listing", None) if sale is not None else None
            listing_product = getattr(listing, "product", None) if listing is not None else None
            sku_product = product or listing_product
            sku = str(getattr(sku_product, "sku", "") or "").strip() if sku_product else ""
            sku_source = (
                "return_product"
                if product is not None
                else ("listing_product" if listing_product is not None else "missing_product")
            )
        elif sku:
            sku_source = "return_row"
        description = str(getter("reason", "") or getter("notes", "") or "Return/Refund").strip()
        rows.append(
            {
                "txn_date": txn_date,
                "doc_number": doc_number,
                "source_order": source_order,
                "marketplace": str(getter("marketplace", "") or "").strip(),
                "sku": sku,
                "sku_source": sku_source,
                "description": description,
                "adjustment_type": "refund",
                "returned_listing_units": quantity,
                "returned_inventory_units": returned_inventory_units,
                "refund_amount": refund_amount,
                "refund_fees": refund_fees,
                "refund_shipping": refund_shipping,
                "net_adjustment": -refund_total,
                "cogs_per_returned_listing": round(float(cogs_per_returned_listing), 2),
                "returned_cogs_estimate": round(float(returned_cogs), 2),
                "cogs_reversal_estimate": round(float(cogs_reversal), 2),
                "cogs_source": str(cogs_source_by_sale.get(sale_id) or "missing_cost_basis"),
                **bundle_summary,
                "estimated_profit_impact": round(float(-refund_total + cogs_reversal), 2),
                "return_status": str(getter("status", "") or getter("return_status", "") or "").strip(),
                "restocked": restocked,
            }
        )
    return rows


def _build_marketplace_reconciliation_fallback_rows(
    sales,
    orders,
    returns_df: pd.DataFrame,
    actual_econ_by_sale_id: dict[int, dict[str, float | str]] | None = None,
) -> list[dict[str, object]]:
    actual_by_sale = actual_econ_by_sale_id or {}
    rows: list[dict[str, object]] = []
    marketplaces = sorted(
        {
            (s.marketplace or "").strip().lower()
            for s in sales
            if (s.marketplace or "").strip()
        }
        | {
            (o.marketplace or "").strip().lower()
            for o in orders
            if (o.marketplace or "").strip()
        }
    )
    for mp in marketplaces:
        mp_sales = [s for s in sales if (s.marketplace or "").strip().lower() == mp]
        mp_orders = [o for o in orders if (o.marketplace or "").strip().lower() == mp]
        mp_returns_df = (
            returns_df[returns_df["marketplace"].fillna("").astype(str).str.strip().str.lower() == mp]
            if not returns_df.empty and "marketplace" in returns_df.columns
            else pd.DataFrame()
        )
        sales_gross = 0.0
        sales_fees = 0.0
        sales_shipping = 0.0
        sales_label_spend = 0.0
        sales_net = 0.0
        for sale in mp_sales:
            sale_id = int(getattr(sale, "id", 0) or 0)
            actual = actual_by_sale.get(sale_id) or {}
            gross = _safe_float(getattr(sale, "sold_price", None))
            fee = _safe_float(actual.get("allocated_fee_actual")) if actual else _safe_float(getattr(sale, "fees", None))
            shipping = (
                _safe_float(actual.get("allocated_shipping_charged"))
                if actual
                else _safe_float(getattr(sale, "shipping_cost", None))
            )
            label = (
                _safe_float(actual.get("allocated_shipping_actual"))
                if actual
                else _safe_float(getattr(sale, "shipping_label_cost", None))
            )
            net = _safe_float(actual.get("net_before_cogs_actual")) if actual else gross + shipping - fee - label
            sales_gross += gross
            sales_fees += fee
            sales_shipping += shipping
            sales_label_spend += label
            sales_net += net
        returns_total = (
            float(mp_returns_df.get("refund_amount", pd.Series(dtype=float)).fillna(0).astype(float).sum())
            + float(mp_returns_df.get("refund_fees", pd.Series(dtype=float)).fillna(0).astype(float).sum())
            + float(mp_returns_df.get("refund_shipping", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        )
        order_totals = sum(_safe_float(o.total_amount) for o in mp_orders)
        delta = order_totals - sales_gross
        rows.append(
            {
                "marketplace": mp,
                "sales_count": len(mp_sales),
                "orders_count": len(mp_orders),
                "returns_count": int(len(mp_returns_df)),
                "sales_gross": round(sales_gross, 2),
                "sales_fees": round(sales_fees, 2),
                "sales_shipping_cost": round(sales_shipping, 2),
                "sales_shipping_label_cost": round(sales_label_spend, 2),
                "sales_net_before_returns": round(sales_net, 2),
                "returns_refund_total": round(returns_total, 2),
                "net_after_returns": round(sales_net - returns_total, 2),
                "order_total_sum": round(order_totals, 2),
                "delta_order_total_vs_sales_gross": round(delta, 2),
                "reconcile_flag": abs(delta) > 0.01,
            }
        )
    return rows


def _safe_product_id(row) -> int | None:
    raw = getattr(row, "product_id", None)
    if raw is None:
        product = getattr(row, "product", None)
        raw = getattr(product, "id", None)
    if raw is None:
        return None
    try:
        return int(raw)
    except Exception:
        return None


def _landed_unit_cost_from_product(product) -> float:
    landed = (
        _safe_float(getattr(product, "acquisition_cost", 0))
        + _safe_float(getattr(product, "acquisition_tax_paid", 0))
        + _safe_float(getattr(product, "acquisition_shipping_paid", 0))
        + _safe_float(getattr(product, "acquisition_handling_paid", 0))
    )
    if landed > 0:
        return landed
    return _safe_float(getattr(product, "product_cost", 0))


def _assignment_lot_id(assignment) -> int | None:
    raw = getattr(assignment, "lot_id", None)
    if raw is None:
        lot = getattr(assignment, "lot", None)
        raw = getattr(lot, "id", None)
    if raw is None:
        return None
    try:
        return int(raw)
    except Exception:
        return None


def _lot_landed_total_from_assignment(assignment) -> float:
    lot = getattr(assignment, "lot", None)
    direct_values = [
        getattr(assignment, "lot_total_cost", None),
        getattr(assignment, "lot_total_tax_paid", None),
        getattr(assignment, "lot_total_shipping_paid", None),
        getattr(assignment, "lot_total_handling_paid", None),
    ]
    if any(value is not None for value in direct_values):
        return sum(_safe_float(value) for value in direct_values)
    return (
        _safe_float(getattr(lot, "total_cost", None))
        + _safe_float(getattr(lot, "total_tax_paid", None))
        + _safe_float(getattr(lot, "total_shipping_paid", None))
        + _safe_float(getattr(lot, "total_handling_paid", None))
    )


def _explicit_landed_unit_cost_from_assignment(assignment) -> float:
    qty = float(max(0, int(getattr(assignment, "quantity_acquired", 0) or 0)))
    unit_cost = (
        _safe_float(getattr(assignment, "unit_cost", None))
        + _safe_float(getattr(assignment, "unit_tax_paid", None))
        + _safe_float(getattr(assignment, "unit_shipping_paid", None))
        + _safe_float(getattr(assignment, "unit_handling_paid", None))
    )
    if unit_cost > 0:
        return unit_cost

    allocated_landed = (
        _safe_float(getattr(assignment, "allocated_cost", None))
        + _safe_float(getattr(assignment, "allocated_tax_paid", None))
        + _safe_float(getattr(assignment, "allocated_shipping_paid", None))
        + _safe_float(getattr(assignment, "allocated_handling_paid", None))
    )
    if allocated_landed > 0 and qty > 0:
        return allocated_landed / qty
    return 0.0


def _lot_fallback_unit_costs_by_lot(assignments) -> dict[int, float]:
    lot_fallbacks, _assignment_fallbacks = _lot_fallback_unit_cost_maps(assignments)
    return lot_fallbacks


def _lot_fallback_unit_cost_maps(assignments) -> tuple[dict[int, float], dict[int, float]]:
    explicit_cost_by_lot: dict[int, float] = defaultdict(float)
    explicit_qty_by_lot: dict[int, float] = defaultdict(float)
    blank_qty_by_lot: dict[int, float] = defaultdict(float)
    lot_total_by_lot: dict[int, float] = {}
    expected_qty_by_lot: dict[int, float] = {}
    blank_rows_by_lot: dict[int, list[dict[str, float | int]]] = defaultdict(list)
    for assignment in assignments:
        lot_id = _assignment_lot_id(assignment)
        if lot_id is None:
            continue
        qty = float(max(0, int(getattr(assignment, "quantity_acquired", 0) or 0)))
        if qty <= 0:
            continue
        lot_total_by_lot[int(lot_id)] = max(
            lot_total_by_lot.get(int(lot_id), 0.0),
            _lot_landed_total_from_assignment(assignment),
        )
        expected_qty = _safe_float(
            getattr(assignment, "lot_expected_total_quantity", None)
            if getattr(assignment, "lot_expected_total_quantity", None) is not None
            else getattr(getattr(assignment, "lot", None), "expected_total_quantity", None)
        )
        if expected_qty > 0:
            expected_qty_by_lot[int(lot_id)] = max(expected_qty_by_lot.get(int(lot_id), 0.0), expected_qty)
        explicit_unit = _explicit_landed_unit_cost_from_assignment(assignment)
        if explicit_unit > 0:
            explicit_cost_by_lot[int(lot_id)] += explicit_unit * qty
            explicit_qty_by_lot[int(lot_id)] += qty
        else:
            blank_qty_by_lot[int(lot_id)] += qty
            blank_rows_by_lot[int(lot_id)].append(
                {
                    "assignment_id": int(getattr(assignment, "assignment_id", 0) or getattr(assignment, "id", 0) or 0),
                    "qty": qty,
                    "allocation_weight": _safe_float(getattr(assignment, "allocation_weight", None)),
                }
            )

    fallback: dict[int, float] = {}
    assignment_fallback: dict[int, float] = {}
    for lot_id, blank_qty in blank_qty_by_lot.items():
        if blank_qty <= 0:
            continue
        remaining_landed = max(0.0, lot_total_by_lot.get(lot_id, 0.0) - explicit_cost_by_lot.get(lot_id, 0.0))
        if remaining_landed > 0:
            weighted_rows = [
                row
                for row in blank_rows_by_lot.get(lot_id, [])
                if _safe_float(row.get("allocation_weight")) > 0
                and int(row.get("assignment_id") or 0) > 0
                and _safe_float(row.get("qty")) > 0
            ]
            total_weight = sum(_safe_float(row.get("allocation_weight")) for row in weighted_rows)
            if total_weight > 0:
                for row in weighted_rows:
                    assignment_id = int(row.get("assignment_id") or 0)
                    qty = _safe_float(row.get("qty"))
                    weight = _safe_float(row.get("allocation_weight"))
                    assignment_fallback[assignment_id] = (remaining_landed * (weight / total_weight)) / qty
                continue
            expected_remaining_qty = max(0.0, expected_qty_by_lot.get(lot_id, 0.0) - explicit_qty_by_lot.get(lot_id, 0.0))
            fallback[lot_id] = remaining_landed / max(blank_qty, expected_remaining_qty)
    return fallback, assignment_fallback


def _landed_unit_cost_from_assignment(
    assignment,
    *,
    lot_fallback_unit_costs: dict[int, float] | None = None,
    assignment_fallback_unit_costs: dict[int, float] | None = None,
) -> float:
    explicit_unit = _explicit_landed_unit_cost_from_assignment(assignment)
    if explicit_unit > 0:
        return explicit_unit
    assignment_id = int(getattr(assignment, "assignment_id", 0) or getattr(assignment, "id", 0) or 0)
    if assignment_id > 0 and assignment_fallback_unit_costs:
        assignment_fallback = _safe_float(assignment_fallback_unit_costs.get(assignment_id, 0.0))
        if assignment_fallback > 0:
            return assignment_fallback
    lot_id = _assignment_lot_id(assignment)
    return _safe_float((lot_fallback_unit_costs or {}).get(int(lot_id or 0), 0.0))


def _parse_csv_set(value: str) -> set[str]:
    return {str(part).strip().lower() for part in str(value or "").split(",") if str(part).strip()}


def _add_margin_pct_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    gross = df["gross_sales"].replace(0, pd.NA)
    df["fifo_margin_pct"] = (df["fifo_margin"] / gross).fillna(0.0)
    df["lot_margin_pct"] = (df["lot_margin"] / gross).fillna(0.0)
    df["fifo_margin_actual_pct"] = (df["fifo_margin_actual"] / gross).fillna(0.0)
    df["lot_margin_actual_pct"] = (df["lot_margin_actual"] / gross).fillna(0.0)
    return df


def _build_economics_intelligence_drilldowns(
    economics_intel_df: pd.DataFrame,
    *,
    min_margin_alert_pct: float,
    max_fee_variance_alert_usd: float,
    min_group_sales_for_alert: int,
) -> dict[str, pd.DataFrame]:
    empty = {
        "by_sku": pd.DataFrame(),
        "by_marketplace": pd.DataFrame(),
        "alerts": pd.DataFrame(),
    }
    if economics_intel_df.empty:
        return empty

    working = economics_intel_df.copy()
    for col in [
        "sold_price",
        "estimated_fee_alloc",
        "expected_shipping_alloc",
        "estimated_net_before_cogs",
        "actual_fee_alloc",
        "actual_shipping_alloc",
        "actual_net_before_cogs",
        "fee_variance_actual_minus_estimated",
        "net_variance_actual_minus_estimated",
    ]:
        if col not in working.columns:
            working[col] = 0.0
        working[col] = pd.to_numeric(working[col], errors="coerce").fillna(0.0)

    if "estimate_available" not in working.columns:
        working["estimate_available"] = False
    working["estimate_available"] = working["estimate_available"].astype(bool)
    working["estimate_available_count"] = working["estimate_available"].astype(int)
    working["abs_fee_variance"] = working["fee_variance_actual_minus_estimated"].abs()
    working["abs_net_variance"] = working["net_variance_actual_minus_estimated"].abs()
    gross = working["sold_price"].replace(0, pd.NA)
    working["actual_margin_pct_of_gross"] = ((working["actual_net_before_cogs"] / gross) * 100.0).fillna(0.0)

    def _aggregate(group_cols: list[str]) -> pd.DataFrame:
        grouped = (
            working.groupby(group_cols, dropna=False, as_index=False)
            .agg(
                sales_count=("sale_id", "count"),
                estimate_covered_sales=("estimate_available_count", "sum"),
                gross_sales=("sold_price", "sum"),
                estimated_fee_total=("estimated_fee_alloc", "sum"),
                actual_fee_total=("actual_fee_alloc", "sum"),
                fee_variance_total=("fee_variance_actual_minus_estimated", "sum"),
                avg_abs_fee_variance=("abs_fee_variance", "mean"),
                expected_net_total=("estimated_net_before_cogs", "sum"),
                actual_net_total=("actual_net_before_cogs", "sum"),
                net_variance_total=("net_variance_actual_minus_estimated", "sum"),
                avg_abs_net_variance=("abs_net_variance", "mean"),
            )
            .sort_values(["sales_count"], ascending=[False])
        )
        gross_group = grouped["gross_sales"].replace(0, pd.NA)
        grouped["expected_margin_pct_of_gross"] = ((grouped["expected_net_total"] / gross_group) * 100.0).fillna(0.0)
        grouped["actual_margin_pct_of_gross"] = ((grouped["actual_net_total"] / gross_group) * 100.0).fillna(0.0)
        grouped["estimate_coverage_pct"] = (
            (grouped["estimate_covered_sales"] / grouped["sales_count"].replace(0, pd.NA)) * 100.0
        ).fillna(0.0)
        min_samples = max(1, int(min_group_sales_for_alert or 1))
        margin_floor = float(min_margin_alert_pct or 0.0)
        fee_threshold = max(0.0, float(max_fee_variance_alert_usd or 0.0))
        grouped["alert_margin_below_floor"] = (
            (grouped["sales_count"] >= min_samples) & (grouped["actual_margin_pct_of_gross"] < margin_floor)
        )
        grouped["alert_fee_variance_high"] = (
            (grouped["estimate_covered_sales"] >= min_samples) & (grouped["avg_abs_fee_variance"] > fee_threshold)
        )
        grouped["alert_any"] = grouped["alert_margin_below_floor"] | grouped["alert_fee_variance_high"]
        return grouped.sort_values(["alert_any", "sales_count"], ascending=[False, False])

    by_sku_df = _aggregate(["sku", "product_title"])
    by_marketplace_df = _aggregate(["marketplace"])

    alerts_df = working.copy()
    margin_floor = float(min_margin_alert_pct or 0.0)
    fee_threshold = max(0.0, float(max_fee_variance_alert_usd or 0.0))
    alerts_df["alert_margin_below_floor"] = alerts_df["actual_margin_pct_of_gross"] < margin_floor
    alerts_df["alert_fee_variance_high"] = (
        alerts_df["estimate_available"] & (alerts_df["abs_fee_variance"] > fee_threshold)
    )
    alerts_df["alert_any"] = alerts_df["alert_margin_below_floor"] | alerts_df["alert_fee_variance_high"]
    alerts_df = alerts_df[alerts_df["alert_any"]].copy()
    if not alerts_df.empty:
        alerts_df = alerts_df.sort_values(["sold_at", "abs_net_variance"], ascending=[False, False])
        keep_cols = [
            "sold_at",
            "sale_id",
            "marketplace",
            "sku",
            "product_title",
            "sold_price",
            "estimated_net_before_cogs",
            "actual_net_before_cogs",
            "net_variance_actual_minus_estimated",
            "fee_variance_actual_minus_estimated",
            "actual_margin_pct_of_gross",
            "alert_margin_below_floor",
            "alert_fee_variance_high",
            "alert_any",
        ]
        alerts_df = alerts_df[[c for c in keep_cols if c in alerts_df.columns]]

    return {
        "by_sku": by_sku_df,
        "by_marketplace": by_marketplace_df,
        "alerts": alerts_df,
    }


def _summarize_fee_reconciliation(df: pd.DataFrame) -> dict[str, float | int]:
    if df.empty:
        return {
            "sales_count": 0,
            "total_actual_fee": 0.0,
            "total_estimated_fee": 0.0,
            "total_variance": 0.0,
            "estimate_coverage_count": 0,
        }
    return {
        "sales_count": int(len(df)),
        "total_actual_fee": float(df["actual_fee"].sum()),
        "total_estimated_fee": float(df["estimated_fee_scaled"].sum()),
        "total_variance": float(df["variance_actual_minus_estimate"].sum()),
        "estimate_coverage_count": int(df["fee_estimate_present"].astype(bool).sum()),
    }


def _fee_source_priority_counts(df: pd.DataFrame) -> dict[str, int]:
    if df.empty:
        return {
            "normalized_source_rows": 0,
            "notes_fallback_rows": 0,
            "sale_field_fallback_rows": 0,
        }
    count_map = {
        str(row.actual_fee_source or "").strip(): int(row.sales_count or 0)
        for row in df.itertuples(index=False)
    }
    return {
        "normalized_source_rows": int(
            count_map.get("normalized_order_finance_entries_marketplace_fee_sum", 0)
        ),
        "notes_fallback_rows": int(
            count_map.get("order_fee_breakdown_total_marketplace_fee", 0)
        ),
        "sale_field_fallback_rows": int(count_map.get("sale_fees_field", 0)),
    }


def _build_accounting_close_readiness_summary(
    *,
    inventory_df: pd.DataFrame,
    cogs_margin_df: pd.DataFrame,
    returns_df: pd.DataFrame,
    reconciliation_df: pd.DataFrame,
    shipping_economics_df: pd.DataFrame,
    ebay_fee_source_priority_df: pd.DataFrame,
    accounting_exceptions_df: pd.DataFrame,
    lot_allocation_source_summary_df: pd.DataFrame | None = None,
    cogs_source_summary_df: pd.DataFrame | None = None,
    qbo_adjustments_df: pd.DataFrame | None = None,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    inventory_value = (
        float(inventory_df.get("landed_inventory_value", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if inventory_df is not None and not inventory_df.empty
        else 0.0
    )
    sales_count = int(len(cogs_margin_df)) if cogs_margin_df is not None else 0
    gross_sales = (
        float(cogs_margin_df.get("gross_sales", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if cogs_margin_df is not None and not cogs_margin_df.empty
        else 0.0
    )
    net_before_cogs = (
        float(cogs_margin_df.get("net_before_cogs", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if cogs_margin_df is not None and not cogs_margin_df.empty
        else 0.0
    )
    fifo_cogs = (
        float(cogs_margin_df.get("fifo_cogs", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if cogs_margin_df is not None and not cogs_margin_df.empty
        else 0.0
    )
    fifo_margin = (
        float(cogs_margin_df.get("fifo_margin", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if cogs_margin_df is not None and not cogs_margin_df.empty
        else 0.0
    )
    negative_margin_rows = (
        int((cogs_margin_df.get("fifo_margin", pd.Series(dtype=float)).fillna(0).astype(float) < 0).sum())
        if cogs_margin_df is not None and not cogs_margin_df.empty
        else 0
    )
    returns_refund_total = 0.0
    if returns_df is not None and not returns_df.empty:
        returns_refund_total = (
            float(returns_df.get("refund_amount", pd.Series(dtype=float)).fillna(0).astype(float).sum())
            + float(returns_df.get("refund_fees", pd.Series(dtype=float)).fillna(0).astype(float).sum())
            + float(returns_df.get("refund_shipping", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        )
    returns_cogs_reversal_total = (
        float(qbo_adjustments_df.get("cogs_reversal_estimate", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if qbo_adjustments_df is not None and not qbo_adjustments_df.empty
        else 0.0
    )
    returns_estimated_profit_impact = -returns_refund_total + returns_cogs_reversal_total
    reconcile_flags = (
        int(reconciliation_df.get("reconcile_flag", pd.Series(dtype=bool)).fillna(False).astype(bool).sum())
        if reconciliation_df is not None and not reconciliation_df.empty
        else 0
    )
    shipping_rows = int(len(shipping_economics_df)) if shipping_economics_df is not None else 0
    shipping_label_covered = (
        int((shipping_economics_df.get("shipping_label_spend", pd.Series(dtype=float)).fillna(0).astype(float) > 0).sum())
        if shipping_economics_df is not None and not shipping_economics_df.empty
        else 0
    )
    shipping_charged_total = (
        float(shipping_economics_df.get("shipping_charged", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if shipping_economics_df is not None and not shipping_economics_df.empty
        else 0.0
    )
    shipping_label_spend_total = (
        float(shipping_economics_df.get("shipping_label_spend", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if shipping_economics_df is not None and not shipping_economics_df.empty
        else 0.0
    )
    shipping_delta_total = shipping_charged_total - shipping_label_spend_total
    shipping_label_coverage_pct = (
        round((shipping_label_covered / shipping_rows) * 100.0, 2) if shipping_rows > 0 else 0.0
    )
    fee_counts = _fee_source_priority_counts(ebay_fee_source_priority_df)
    fee_total = (
        float(
            pd.to_numeric(
                ebay_fee_source_priority_df.get("actual_fee_total", pd.Series(dtype=float)),
                errors="coerce",
            )
            .fillna(0.0)
            .sum()
        )
        if ebay_fee_source_priority_df is not None
        and not ebay_fee_source_priority_df.empty
        and "actual_fee_total" in ebay_fee_source_priority_df.columns
        else 0.0
    )
    p0_exceptions = (
        int((accounting_exceptions_df.get("severity", pd.Series(dtype=str)).astype(str) == "P0").sum())
        if accounting_exceptions_df is not None and not accounting_exceptions_df.empty
        else 0
    )
    p1_exceptions = (
        int((accounting_exceptions_df.get("severity", pd.Series(dtype=str)).astype(str) == "P1").sum())
        if accounting_exceptions_df is not None and not accounting_exceptions_df.empty
        else 0
    )
    total_exceptions = int(len(accounting_exceptions_df)) if accounting_exceptions_df is not None else 0
    lot_equal_fallback_assignments = 0
    lot_missing_cost_assignments = 0
    sold_equal_fallback_cogs = 0.0
    sold_missing_cost_cogs = 0.0
    if lot_allocation_source_summary_df is not None and not lot_allocation_source_summary_df.empty:
        source_df = lot_allocation_source_summary_df.copy()
        source_df["cost_source"] = source_df.get("cost_source", pd.Series(dtype=str)).fillna("").astype(str)
        source_df["assignment_count"] = pd.to_numeric(
            source_df.get("assignment_count", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0)
        lot_equal_fallback_assignments = int(
            source_df.loc[
                source_df["cost_source"] == "lot_equal_quantity_fallback",
                "assignment_count",
            ].sum()
        )
        lot_missing_cost_assignments = int(
            source_df.loc[
                source_df["cost_source"].isin(["missing_cost_basis", "unknown"]),
                "assignment_count",
            ].sum()
        )
    if cogs_source_summary_df is not None and not cogs_source_summary_df.empty:
        cogs_source_df = cogs_source_summary_df.copy()
        cogs_source_df["fifo_cost_source"] = (
            cogs_source_df.get("fifo_cost_source", pd.Series(dtype=str)).fillna("").astype(str)
        )
        cogs_source_df["fifo_cogs"] = pd.to_numeric(
            cogs_source_df.get("fifo_cogs", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        sold_equal_fallback_cogs = float(
            cogs_source_df.loc[
                cogs_source_df["fifo_cost_source"] == "lot_equal_quantity_fallback",
                "fifo_cogs",
            ].sum()
        )
        sold_missing_cost_cogs = float(
            cogs_source_df.loc[
                cogs_source_df["fifo_cost_source"].isin(["missing_cost_basis", "unknown"]),
                "fifo_cogs",
            ].sum()
        )

    blockers: list[str] = []
    warnings: list[str] = []
    if p0_exceptions > 0:
        blockers.append("P0 accounting exceptions")
    if reconcile_flags > 0:
        blockers.append("marketplace reconciliation flags")
    if negative_margin_rows > 0:
        warnings.append("negative FIFO margin rows")
    if p1_exceptions > 0:
        warnings.append("P1 accounting exceptions")
    if shipping_rows > 0 and shipping_label_coverage_pct < 100.0:
        warnings.append("incomplete shipping label spend coverage")
    if int(fee_counts.get("sale_field_fallback_rows") or 0) > 0:
        warnings.append("fee source fallback rows")
    if lot_equal_fallback_assignments > 0:
        warnings.append("lot equal quantity fallback assignments")
    if lot_missing_cost_assignments > 0:
        warnings.append("lot assignments missing cost basis")
    if sold_equal_fallback_cogs > 0:
        warnings.append("sold COGS uses equal quantity fallback")
    if sold_missing_cost_cogs > 0:
        warnings.append("sold COGS missing cost basis")

    profit_before_returns = round(fifo_margin, 2)
    estimated_profit_after_returns = round(fifo_margin + returns_estimated_profit_impact, 2)

    if blockers:
        readiness_status = "blocked"
    elif warnings:
        readiness_status = "review_needed"
    else:
        readiness_status = "close_ready"

    summary = {
        "readiness_status": readiness_status,
        "blocker_count": int(len(blockers)),
        "warning_count": int(len(warnings)),
        "blockers": ", ".join(blockers),
        "warnings": ", ".join(warnings),
        "inventory_value": round(inventory_value, 2),
        "sales_count": sales_count,
        "gross_sales": round(gross_sales, 2),
        "net_before_cogs": round(net_before_cogs, 2),
        "fifo_cogs": round(fifo_cogs, 2),
        "fifo_margin": round(fifo_margin, 2),
        "profit_before_returns": profit_before_returns,
        "returns_refund_total": round(returns_refund_total, 2),
        "returns_cogs_reversal_total": round(returns_cogs_reversal_total, 2),
        "returns_estimated_profit_impact": round(returns_estimated_profit_impact, 2),
        "net_after_returns_and_cogs": estimated_profit_after_returns,
        "estimated_profit_after_returns": estimated_profit_after_returns,
        "shipping_charged_total": round(shipping_charged_total, 2),
        "shipping_label_spend_total": round(shipping_label_spend_total, 2),
        "shipping_delta_total": round(shipping_delta_total, 2),
        "fee_total": round(fee_total, 2),
        "shipping_label_coverage_pct": shipping_label_coverage_pct,
        "normalized_fee_source_rows": int(fee_counts.get("normalized_source_rows") or 0),
        "sale_fee_field_fallback_rows": int(fee_counts.get("sale_field_fallback_rows") or 0),
        "p0_exceptions": p0_exceptions,
        "p1_exceptions": p1_exceptions,
        "total_exceptions": total_exceptions,
        "reconcile_flags": reconcile_flags,
        "negative_margin_rows": negative_margin_rows,
        "lot_equal_fallback_assignments": lot_equal_fallback_assignments,
        "lot_missing_cost_assignments": lot_missing_cost_assignments,
        "sold_equal_fallback_cogs": round(sold_equal_fallback_cogs, 2),
        "sold_missing_cost_cogs": round(sold_missing_cost_cogs, 2),
    }
    rows = [
        {"check": "P0 Exceptions", "status": "fail" if p0_exceptions else "pass", "value": p0_exceptions},
        {"check": "Reconciliation Flags", "status": "fail" if reconcile_flags else "pass", "value": reconcile_flags},
        {"check": "P1 Exceptions", "status": "warn" if p1_exceptions else "pass", "value": p1_exceptions},
        {
            "check": "Shipping Label Coverage",
            "status": "warn" if shipping_rows > 0 and shipping_label_coverage_pct < 100.0 else "pass",
            "value": shipping_label_coverage_pct,
        },
        {
            "check": "Fee Source Fallback Rows",
            "status": "warn" if int(fee_counts.get("sale_field_fallback_rows") or 0) else "pass",
            "value": int(fee_counts.get("sale_field_fallback_rows") or 0),
        },
        {
            "check": "Negative FIFO Margin Rows",
            "status": "warn" if negative_margin_rows else "pass",
            "value": negative_margin_rows,
        },
        {
            "check": "Lot Equal Fallback Assignments",
            "status": "warn" if lot_equal_fallback_assignments else "pass",
            "value": lot_equal_fallback_assignments,
        },
        {
            "check": "Lot Missing Cost Assignments",
            "status": "warn" if lot_missing_cost_assignments else "pass",
            "value": lot_missing_cost_assignments,
        },
        {
            "check": "Sold Equal Fallback COGS",
            "status": "warn" if sold_equal_fallback_cogs else "pass",
            "value": round(sold_equal_fallback_cogs, 2),
        },
        {
            "check": "Sold Missing Cost COGS",
            "status": "warn" if sold_missing_cost_cogs else "pass",
            "value": round(sold_missing_cost_cogs, 2),
        },
        {
            "check": "Return COGS Reversal",
            "status": "info" if returns_cogs_reversal_total > 0 else "pass",
            "value": round(returns_cogs_reversal_total, 2),
        },
    ]
    return summary, pd.DataFrame(rows)


def _build_lot_allocation_source_summary(lots_df: pd.DataFrame) -> pd.DataFrame:
    if lots_df is None or lots_df.empty or "cost_source" not in lots_df.columns:
        return pd.DataFrame()
    df = lots_df.copy()
    df["cost_source"] = df["cost_source"].fillna("unknown").astype(str).replace({"": "unknown"})
    df["quantity_acquired"] = pd.to_numeric(df.get("quantity_acquired", 0), errors="coerce").fillna(0)
    df["resolved_landed_total_cost"] = pd.to_numeric(
        df.get("resolved_landed_total_cost", 0),
        errors="coerce",
    ).fillna(0.0)
    grouped = (
        df.groupby(["cost_source"], dropna=False, as_index=False)
        .agg(
            assignment_count=("cost_source", "size"),
            quantity_acquired=("quantity_acquired", "sum"),
            resolved_landed_total_cost=("resolved_landed_total_cost", "sum"),
        )
        .sort_values(["resolved_landed_total_cost", "assignment_count"], ascending=[False, False])
    )
    total_cost = float(grouped["resolved_landed_total_cost"].sum()) if not grouped.empty else 0.0
    total_assignments = int(grouped["assignment_count"].sum()) if not grouped.empty else 0
    grouped["resolved_landed_total_cost"] = grouped["resolved_landed_total_cost"].round(2)
    grouped["assignment_share_pct"] = grouped["assignment_count"].apply(
        lambda value: round((float(value) / float(total_assignments)) * 100.0, 2)
        if total_assignments > 0
        else 0.0
    )
    grouped["cost_share_pct"] = grouped["resolved_landed_total_cost"].apply(
        lambda value: round((float(value) / float(total_cost)) * 100.0, 2)
        if total_cost > 0
        else 0.0
    )
    return grouped


def _build_cogs_source_summary(cogs_margin_df: pd.DataFrame) -> pd.DataFrame:
    if cogs_margin_df is None or cogs_margin_df.empty or "fifo_cost_source" not in cogs_margin_df.columns:
        return pd.DataFrame()
    df = cogs_margin_df.copy()
    df["fifo_cost_source"] = df["fifo_cost_source"].fillna("unknown").astype(str).replace({"": "unknown"})
    df["quantity"] = pd.to_numeric(df.get("quantity", 0), errors="coerce").fillna(0)
    df["gross_sales"] = pd.to_numeric(df.get("gross_sales", 0), errors="coerce").fillna(0.0)
    df["net_before_cogs"] = pd.to_numeric(df.get("net_before_cogs", 0), errors="coerce").fillna(0.0)
    df["fifo_cogs"] = pd.to_numeric(df.get("fifo_cogs", 0), errors="coerce").fillna(0.0)
    df["fifo_margin"] = pd.to_numeric(df.get("fifo_margin", 0), errors="coerce").fillna(0.0)
    if "listing_is_bundle" in df.columns:
        df["listing_is_bundle"] = df["listing_is_bundle"].fillna(False).astype(bool)
    else:
        df["listing_is_bundle"] = False
    if "listing_bundle_inventory_units_sold" in df.columns:
        bundle_units_series = df["listing_bundle_inventory_units_sold"]
    else:
        bundle_units_series = pd.Series([0] * len(df), index=df.index)
    df["listing_bundle_inventory_units_sold"] = pd.to_numeric(
        bundle_units_series,
        errors="coerce",
    ).fillna(0)
    grouped = (
        df.groupby(["fifo_cost_source"], dropna=False, as_index=False)
        .agg(
            sale_count=("fifo_cost_source", "size"),
            quantity=("quantity", "sum"),
            bundle_sale_count=("listing_is_bundle", "sum"),
            bundle_inventory_units_sold=("listing_bundle_inventory_units_sold", "sum"),
            gross_sales=("gross_sales", "sum"),
            net_before_cogs=("net_before_cogs", "sum"),
            fifo_cogs=("fifo_cogs", "sum"),
            fifo_margin=("fifo_margin", "sum"),
        )
        .sort_values(["fifo_cogs", "sale_count"], ascending=[False, False])
    )
    total_cogs = float(grouped["fifo_cogs"].sum()) if not grouped.empty else 0.0
    for column in ["quantity", "bundle_sale_count", "bundle_inventory_units_sold"]:
        grouped[column] = pd.to_numeric(grouped[column], errors="coerce").fillna(0).astype(int)
    for column in ["gross_sales", "net_before_cogs", "fifo_cogs", "fifo_margin"]:
        grouped[column] = grouped[column].round(2)
    grouped["cogs_share_pct"] = grouped["fifo_cogs"].apply(
        lambda value: round((float(value) / total_cogs) * 100.0, 2) if total_cogs > 0 else 0.0
    )
    return grouped


def _build_accounting_period_drift_checks(
    *,
    close_summary: dict[str, float | int | str],
    qbo_sales_df: pd.DataFrame,
    qbo_adjustments_df: pd.DataFrame,
    dashboard_live_metrics: dict[str, float | int | str] | None = None,
    slack_summary_metrics: dict[str, float | int | str] | None = None,
    ai_accounting_snapshot_metrics: dict[str, float | int | str] | None = None,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    def _count(df: pd.DataFrame) -> int:
        return 0 if df is None else int(len(df))

    def _profit_before_returns_metric(metrics: dict[str, float | int | str]) -> float:
        if metrics.get("profit_before_returns") is not None:
            return _safe_float(metrics.get("profit_before_returns"))
        return _safe_float(metrics.get("estimated_margin"))

    def _profit_before_returns_source(
        metrics: dict[str, float | int | str],
        observed_source: str,
    ) -> str:
        field_name = "profit_before_returns" if metrics.get("profit_before_returns") is not None else "estimated_margin"
        return f"{observed_source} {field_name}"

    def _close_profit_before_returns() -> float:
        return _accounting_close_profit_before_returns(close_summary)

    def _close_estimated_profit_after_returns() -> float:
        return _accounting_close_estimated_profit_after_returns(close_summary)

    qbo_profit_before_returns_column = (
        "profit_before_returns_estimate"
        if qbo_sales_df is not None and "profit_before_returns_estimate" in qbo_sales_df.columns
        else "gross_margin_estimate"
    )
    checks = [
        {
            "check": "sales_count_close_vs_qbo",
            "expected_source": "Accounting Close Readiness",
            "observed_source": "QuickBooks Sales Export",
            "expected": float(close_summary.get("sales_count") or 0),
            "observed": float(_count(qbo_sales_df)),
            "tolerance": 0.0,
        },
        {
            "check": "gross_sales_close_vs_qbo",
            "expected_source": "Accounting Close Readiness",
            "observed_source": "QuickBooks Sales Export.amount",
            "expected": _safe_float(close_summary.get("gross_sales")),
            "observed": _sum(qbo_sales_df, "amount"),
            "tolerance": tolerance,
        },
        {
            "check": "net_before_cogs_close_vs_qbo",
            "expected_source": "Accounting Close Readiness",
            "observed_source": "QuickBooks Sales Export.net_amount",
            "expected": _safe_float(close_summary.get("net_before_cogs")),
            "observed": _sum(qbo_sales_df, "net_amount"),
            "tolerance": tolerance,
        },
        {
            "check": "fifo_cogs_close_vs_qbo",
            "expected_source": "Accounting Close Readiness",
            "observed_source": "QuickBooks Sales Export.cogs_input_estimate",
            "expected": _safe_float(close_summary.get("fifo_cogs")),
            "observed": _sum(qbo_sales_df, "cogs_input_estimate"),
            "tolerance": tolerance,
        },
        {
            "check": "profit_before_returns_close_vs_qbo",
            "expected_source": "Accounting Close Readiness.profit_before_returns",
            "observed_source": f"QuickBooks Sales Export.{qbo_profit_before_returns_column}",
            "expected": _close_profit_before_returns(),
            "observed": _sum(qbo_sales_df, qbo_profit_before_returns_column),
            "tolerance": tolerance,
        },
        {
            "check": "qbo_sales_net_formula",
            "expected_source": "QuickBooks Sales Export amount + shipping_cost - fees - shipping_label_cost",
            "observed_source": "QuickBooks Sales Export.net_amount",
            "expected": _sum(qbo_sales_df, "amount")
            + _sum(qbo_sales_df, "shipping_cost")
            - _sum(qbo_sales_df, "fees")
            - _sum(qbo_sales_df, "shipping_label_cost"),
            "observed": _sum(qbo_sales_df, "net_amount"),
            "tolerance": tolerance,
        },
        {
            "check": "qbo_sales_profit_before_returns_formula",
            "expected_source": "QuickBooks Sales Export net_amount - cogs_input_estimate",
            "observed_source": f"QuickBooks Sales Export.{qbo_profit_before_returns_column}",
            "expected": _sum(qbo_sales_df, "net_amount") - _sum(qbo_sales_df, "cogs_input_estimate"),
            "observed": _sum(qbo_sales_df, qbo_profit_before_returns_column),
            "tolerance": tolerance,
        },
        {
            "check": "returns_refund_total_close_vs_qbo",
            "expected_source": "Accounting Close Readiness",
            "observed_source": "QuickBooks Refund/Adjustment Export refund components",
            "expected": _safe_float(close_summary.get("returns_refund_total")),
            "observed": _sum(qbo_adjustments_df, "refund_amount")
            + _sum(qbo_adjustments_df, "refund_fees")
            + _sum(qbo_adjustments_df, "refund_shipping"),
            "tolerance": tolerance,
        },
        {
            "check": "returns_cogs_reversal_close_vs_qbo",
            "expected_source": "Accounting Close Readiness",
            "observed_source": "QuickBooks Refund/Adjustment Export.cogs_reversal_estimate",
            "expected": _safe_float(close_summary.get("returns_cogs_reversal_total")),
            "observed": _sum(qbo_adjustments_df, "cogs_reversal_estimate"),
            "tolerance": tolerance,
        },
        {
            "check": "qbo_return_profit_impact_formula",
            "expected_source": "QuickBooks Refund/Adjustment Export -(refund_amount + refund_fees + refund_shipping) + cogs_reversal_estimate",
            "observed_source": "QuickBooks Refund/Adjustment Export.estimated_profit_impact",
            "expected": -(
                _sum(qbo_adjustments_df, "refund_amount")
                + _sum(qbo_adjustments_df, "refund_fees")
                + _sum(qbo_adjustments_df, "refund_shipping")
            )
            + _sum(qbo_adjustments_df, "cogs_reversal_estimate"),
            "observed": _sum(qbo_adjustments_df, "estimated_profit_impact"),
            "tolerance": tolerance,
        },
        {
            "check": "net_after_returns_and_cogs_close_vs_qbo",
            "expected_source": "Accounting Close Readiness.estimated_profit_after_returns",
            "observed_source": "QBO profit_before_returns_estimate + return estimated_profit_impact",
            "expected": _close_estimated_profit_after_returns(),
            "observed": _sum(qbo_sales_df, qbo_profit_before_returns_column)
            + _sum(qbo_adjustments_df, "estimated_profit_impact"),
            "tolerance": tolerance,
        },
    ]
    dashboard = dashboard_live_metrics or {}
    if dashboard:
        checks.extend(
            [
                {
                    "check": "sales_count_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": "Dashboard Live Metrics sales_30d_count",
                    "expected": float(close_summary.get("sales_count") or 0),
                    "observed": _safe_float(dashboard.get("sales_30d_count")),
                    "tolerance": 0.0,
                },
                {
                    "check": "gross_sales_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": "Dashboard Live Metrics sales_30d_gross",
                    "expected": _safe_float(close_summary.get("gross_sales")),
                    "observed": _safe_float(dashboard.get("sales_30d_gross")),
                    "tolerance": tolerance,
                },
                {
                    "check": "net_before_cogs_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": "Dashboard Live Metrics sales_30d_net",
                    "expected": _safe_float(close_summary.get("net_before_cogs")),
                    "observed": _safe_float(dashboard.get("sales_30d_net")),
                    "tolerance": tolerance,
                },
                {
                    "check": "fee_total_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.fee_total",
                    "observed_source": "Dashboard Live Metrics ebay_fees_30d_total",
                    "expected": _safe_float(close_summary.get("fee_total")),
                    "observed": _safe_float(dashboard.get("ebay_fees_30d_total")),
                    "tolerance": tolerance,
                },
                {
                    "check": "shipping_charged_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.shipping_charged_total",
                    "observed_source": "Dashboard Live Metrics sales_30d_shipping_charged",
                    "expected": _safe_float(close_summary.get("shipping_charged_total")),
                    "observed": _safe_float(dashboard.get("sales_30d_shipping_charged")),
                    "tolerance": tolerance,
                },
                {
                    "check": "shipping_label_spend_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.shipping_label_spend_total",
                    "observed_source": "Dashboard Live Metrics sales_30d_shipping_label_spend",
                    "expected": _safe_float(close_summary.get("shipping_label_spend_total")),
                    "observed": _safe_float(dashboard.get("sales_30d_shipping_label_spend")),
                    "tolerance": tolerance,
                },
                {
                    "check": "shipping_delta_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.shipping_delta_total",
                    "observed_source": "Dashboard Live Metrics sales_30d_shipping_delta",
                    "expected": _safe_float(close_summary.get("shipping_delta_total")),
                    "observed": _safe_float(dashboard.get("sales_30d_shipping_delta")),
                    "tolerance": tolerance,
                },
                {
                    "check": "fifo_cogs_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": "Dashboard Live Metrics sales_30d_est_cogs",
                    "expected": _safe_float(close_summary.get("fifo_cogs")),
                    "observed": _safe_float(dashboard.get("sales_30d_est_cogs")),
                    "tolerance": tolerance,
                },
                {
                    "check": "profit_before_returns_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.profit_before_returns",
                    "observed_source": "Dashboard Live Metrics sales_30d_profit_before_returns",
                    "expected": _close_profit_before_returns(),
                    "observed": (
                        _safe_float(dashboard.get("sales_30d_profit_before_returns"))
                        if dashboard.get("sales_30d_profit_before_returns") is not None
                        else _safe_float(dashboard.get("sales_30d_net"))
                        - _safe_float(dashboard.get("sales_30d_est_cogs"))
                    ),
                    "tolerance": tolerance,
                },
                {
                    "check": "estimated_profit_after_returns_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.estimated_profit_after_returns",
                    "observed_source": "Dashboard Live Metrics sales_30d_est_profit",
                    "expected": _close_estimated_profit_after_returns(),
                    "observed": _safe_float(dashboard.get("sales_30d_est_profit")),
                    "tolerance": tolerance,
                },
                {
                    "check": "returns_refund_total_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.returns_refund_total",
                    "observed_source": "Dashboard Live Metrics returns_30d_refund_total",
                    "expected": _safe_float(close_summary.get("returns_refund_total")),
                    "observed": _safe_float(dashboard.get("returns_30d_refund_total")),
                    "tolerance": tolerance,
                },
                {
                    "check": "returns_cogs_reversal_close_vs_dashboard_30d",
                    "expected_source": "Accounting Close Readiness.returns_cogs_reversal_total",
                    "observed_source": "Dashboard Live Metrics returns_30d_cogs_reversal",
                    "expected": _safe_float(close_summary.get("returns_cogs_reversal_total")),
                    "observed": _safe_float(dashboard.get("returns_30d_cogs_reversal")),
                    "tolerance": tolerance,
                },
                {
                    "check": "dashboard_30d_return_profit_impact_formula",
                    "expected_source": "Dashboard -returns_30d_refund_total + returns_30d_cogs_reversal",
                    "observed_source": "Dashboard Live Metrics returns_30d_profit_impact",
                    "expected": -_safe_float(dashboard.get("returns_30d_refund_total"))
                    + _safe_float(dashboard.get("returns_30d_cogs_reversal")),
                    "observed": _safe_float(dashboard.get("returns_30d_profit_impact")),
                    "tolerance": tolerance,
                },
                {
                    "check": "dashboard_30d_net_formula",
                    "expected_source": "Dashboard gross + shipping charged - fees - label spend",
                    "observed_source": "Dashboard Live Metrics sales_30d_net",
                    "expected": _safe_float(dashboard.get("sales_30d_gross"))
                    + _safe_float(dashboard.get("sales_30d_shipping_charged"))
                    - _safe_float(dashboard.get("ebay_fees_30d_total"))
                    - _safe_float(dashboard.get("sales_30d_shipping_label_spend")),
                    "observed": _safe_float(dashboard.get("sales_30d_net")),
                    "tolerance": tolerance,
                },
                {
                    "check": "dashboard_30d_shipping_delta_formula",
                    "expected_source": "Dashboard shipping charged - label spend",
                    "observed_source": "Dashboard Live Metrics sales_30d_shipping_delta",
                    "expected": _safe_float(dashboard.get("sales_30d_shipping_charged"))
                    - _safe_float(dashboard.get("sales_30d_shipping_label_spend")),
                    "observed": _safe_float(dashboard.get("sales_30d_shipping_delta")),
                    "tolerance": tolerance,
                },
                {
                    "check": "dashboard_30d_profit_formula",
                    "expected_source": "Dashboard profit before returns + return profit impact",
                    "observed_source": "Dashboard Live Metrics sales_30d_est_profit",
                    "expected": (
                        _safe_float(dashboard.get("sales_30d_profit_before_returns"))
                        if dashboard.get("sales_30d_profit_before_returns") is not None
                        else _safe_float(dashboard.get("sales_30d_net"))
                        - _safe_float(dashboard.get("sales_30d_est_cogs"))
                    )
                    + _safe_float(dashboard.get("returns_30d_profit_impact")),
                    "observed": _safe_float(dashboard.get("sales_30d_est_profit")),
                    "tolerance": tolerance,
                },
            ]
        )
    slack_metrics = slack_summary_metrics or {}
    if slack_metrics:
        slack_label = str(slack_metrics.get("window_label") or "summary").strip().lower().replace(" ", "_")
        observed_source = str(slack_metrics.get("observed_source") or "Slack business summary metrics")
        checks.extend(
            [
                {
                    "check": f"sales_count_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} sales_window_count",
                    "expected": float(close_summary.get("sales_count") or 0),
                    "observed": _safe_float(slack_metrics.get("sales_window_count")),
                    "tolerance": 0.0,
                },
                {
                    "check": f"gross_sales_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} gross_window",
                    "expected": _safe_float(close_summary.get("gross_sales")),
                    "observed": _safe_float(slack_metrics.get("gross_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"net_before_cogs_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} net_window",
                    "expected": _safe_float(close_summary.get("net_before_cogs")),
                    "observed": _safe_float(slack_metrics.get("net_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"fifo_cogs_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} cogs_window",
                    "expected": _safe_float(close_summary.get("fifo_cogs")),
                    "observed": _safe_float(slack_metrics.get("cogs_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"profit_before_returns_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness.profit_before_returns",
                    "observed_source": _profit_before_returns_source(slack_metrics, observed_source),
                    "expected": _close_profit_before_returns(),
                    "observed": _profit_before_returns_metric(slack_metrics),
                    "tolerance": tolerance,
                },
                {
                    "check": f"slack_{slack_label}_profit_before_returns_formula",
                    "expected_source": f"{observed_source} net_window - cogs_window",
                    "observed_source": _profit_before_returns_source(slack_metrics, observed_source),
                    "expected": _safe_float(slack_metrics.get("net_window"))
                    - _safe_float(slack_metrics.get("cogs_window")),
                    "observed": _profit_before_returns_metric(slack_metrics),
                    "tolerance": tolerance,
                },
                {
                    "check": f"returns_refund_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness returns_refund_total",
                    "observed_source": f"{observed_source} returns_refund_window",
                    "expected": _safe_float(close_summary.get("returns_refund_total")),
                    "observed": _safe_float(slack_metrics.get("returns_refund_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"returns_cogs_reversal_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness returns_cogs_reversal_total",
                    "observed_source": f"{observed_source} returns_cogs_reversal_window",
                    "expected": _safe_float(close_summary.get("returns_cogs_reversal_total")),
                    "observed": _safe_float(slack_metrics.get("returns_cogs_reversal_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"slack_{slack_label}_return_profit_impact_formula",
                    "expected_source": f"{observed_source} -returns_refund_window + returns_cogs_reversal_window",
                    "observed_source": f"{observed_source} returns_profit_impact_window",
                    "expected": -_safe_float(slack_metrics.get("returns_refund_window"))
                    + _safe_float(slack_metrics.get("returns_cogs_reversal_window")),
                    "observed": _safe_float(slack_metrics.get("returns_profit_impact_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"net_after_returns_and_cogs_close_vs_slack_{slack_label}",
                    "expected_source": "Accounting Close Readiness estimated_profit_after_returns",
                    "observed_source": f"{observed_source} estimated_profit_after_returns",
                    "expected": _close_estimated_profit_after_returns(),
                    "observed": _safe_float(slack_metrics.get("estimated_profit_after_returns")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"slack_{slack_label}_estimated_profit_after_returns_formula",
                    "expected_source": (
                        f"{observed_source} net_window - cogs_window + returns_profit_impact_window"
                    ),
                    "observed_source": f"{observed_source} estimated_profit_after_returns",
                    "expected": _safe_float(slack_metrics.get("net_window"))
                    - _safe_float(slack_metrics.get("cogs_window"))
                    + _safe_float(slack_metrics.get("returns_profit_impact_window")),
                    "observed": _safe_float(slack_metrics.get("estimated_profit_after_returns")),
                    "tolerance": tolerance,
                },
            ]
        )
    ai_metrics = ai_accounting_snapshot_metrics or {}
    if ai_metrics:
        ai_label = str(ai_metrics.get("window_label") or "30d").strip().lower().replace(" ", "_")
        observed_source = str(ai_metrics.get("observed_source") or "Ask/AI accounting snapshot metrics")
        checks.extend(
            [
                {
                    "check": f"sales_count_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} sales_window_count",
                    "expected": float(close_summary.get("sales_count") or 0),
                    "observed": _safe_float(ai_metrics.get("sales_window_count")),
                    "tolerance": 0.0,
                },
                {
                    "check": f"gross_sales_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} gross_window",
                    "expected": _safe_float(close_summary.get("gross_sales")),
                    "observed": _safe_float(ai_metrics.get("gross_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"net_before_cogs_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} net_window",
                    "expected": _safe_float(close_summary.get("net_before_cogs")),
                    "observed": _safe_float(ai_metrics.get("net_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"fifo_cogs_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness",
                    "observed_source": f"{observed_source} cogs_window",
                    "expected": _safe_float(close_summary.get("fifo_cogs")),
                    "observed": _safe_float(ai_metrics.get("cogs_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"profit_before_returns_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness.profit_before_returns",
                    "observed_source": _profit_before_returns_source(ai_metrics, observed_source),
                    "expected": _close_profit_before_returns(),
                    "observed": _profit_before_returns_metric(ai_metrics),
                    "tolerance": tolerance,
                },
                {
                    "check": f"ai_accounting_{ai_label}_profit_before_returns_formula",
                    "expected_source": f"{observed_source} net_window - cogs_window",
                    "observed_source": _profit_before_returns_source(ai_metrics, observed_source),
                    "expected": _safe_float(ai_metrics.get("net_window")) - _safe_float(ai_metrics.get("cogs_window")),
                    "observed": _profit_before_returns_metric(ai_metrics),
                    "tolerance": tolerance,
                },
                {
                    "check": f"returns_refund_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness returns_refund_total",
                    "observed_source": f"{observed_source} returns_refund_window",
                    "expected": _safe_float(close_summary.get("returns_refund_total")),
                    "observed": _safe_float(ai_metrics.get("returns_refund_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"returns_cogs_reversal_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness returns_cogs_reversal_total",
                    "observed_source": f"{observed_source} returns_cogs_reversal_window",
                    "expected": _safe_float(close_summary.get("returns_cogs_reversal_total")),
                    "observed": _safe_float(ai_metrics.get("returns_cogs_reversal_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"ai_accounting_{ai_label}_return_profit_impact_formula",
                    "expected_source": f"{observed_source} -returns_refund_window + returns_cogs_reversal_window",
                    "observed_source": f"{observed_source} returns_profit_impact_window",
                    "expected": -_safe_float(ai_metrics.get("returns_refund_window"))
                    + _safe_float(ai_metrics.get("returns_cogs_reversal_window")),
                    "observed": _safe_float(ai_metrics.get("returns_profit_impact_window")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"net_after_returns_and_cogs_close_vs_ai_accounting_{ai_label}",
                    "expected_source": "Accounting Close Readiness estimated_profit_after_returns",
                    "observed_source": f"{observed_source} estimated_profit_after_returns",
                    "expected": _close_estimated_profit_after_returns(),
                    "observed": _safe_float(ai_metrics.get("estimated_profit_after_returns")),
                    "tolerance": tolerance,
                },
                {
                    "check": f"ai_accounting_{ai_label}_estimated_profit_after_returns_formula",
                    "expected_source": (
                        f"{observed_source} net_window - cogs_window + returns_profit_impact_window"
                    ),
                    "observed_source": f"{observed_source} estimated_profit_after_returns",
                    "expected": _safe_float(ai_metrics.get("net_window"))
                    - _safe_float(ai_metrics.get("cogs_window"))
                    + _safe_float(ai_metrics.get("returns_profit_impact_window")),
                    "observed": _safe_float(ai_metrics.get("estimated_profit_after_returns")),
                    "tolerance": tolerance,
                },
            ]
        )
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _build_slack_summary_drift_metrics(
    cogs_margin_df: pd.DataFrame,
    *,
    window_label: str,
    returns_df: pd.DataFrame | None = None,
    qbo_adjustments_df: pd.DataFrame | None = None,
) -> dict[str, object]:
    returns_count = int(len(returns_df)) if returns_df is not None else 0
    returns_refund = (
        float(pd.to_numeric(returns_df.get("refund_amount", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())
        + float(pd.to_numeric(returns_df.get("refund_fees", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())
        + float(
            pd.to_numeric(returns_df.get("refund_shipping", pd.Series(dtype=float)), errors="coerce")
            .fillna(0.0)
            .sum()
        )
        if returns_df is not None and not returns_df.empty
        else 0.0
    )
    returns_cogs_reversal = (
        float(
            pd.to_numeric(
                qbo_adjustments_df.get("cogs_reversal_estimate", pd.Series(dtype=float)),
                errors="coerce",
            )
            .fillna(0.0)
            .sum()
        )
        if qbo_adjustments_df is not None and not qbo_adjustments_df.empty
        else 0.0
    )
    returns_profit_impact = -returns_refund + returns_cogs_reversal
    if cogs_margin_df is None or cogs_margin_df.empty:
        return {
            "window_label": window_label,
            "observed_source": f"Slack {window_label} business summary",
            "sales_window_count": 0,
            "gross_window": 0.0,
            "net_window": 0.0,
            "cogs_window": 0.0,
            "estimated_margin": 0.0,
            "profit_before_returns": 0.0,
            "returns_window_count": returns_count,
            "returns_refund_window": round(returns_refund, 2),
            "returns_cogs_reversal_window": round(returns_cogs_reversal, 2),
            "returns_profit_impact_window": round(returns_profit_impact, 2),
            "estimated_profit_after_returns": round(returns_profit_impact, 2),
        }
    gross = float(pd.to_numeric(cogs_margin_df.get("gross_sales", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())
    net = float(pd.to_numeric(cogs_margin_df.get("net_before_cogs", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())
    cogs = float(pd.to_numeric(cogs_margin_df.get("fifo_cogs", pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())
    profit_before_returns = net - cogs
    return {
        "window_label": window_label,
        "observed_source": f"Slack {window_label} business summary",
        "sales_window_count": int(len(cogs_margin_df)),
        "gross_window": round(gross, 2),
        "net_window": round(net, 2),
        "cogs_window": round(cogs, 2),
        "estimated_margin": round(profit_before_returns, 2),
        "profit_before_returns": round(profit_before_returns, 2),
        "returns_window_count": returns_count,
        "returns_refund_window": round(returns_refund, 2),
        "returns_cogs_reversal_window": round(returns_cogs_reversal, 2),
        "returns_profit_impact_window": round(returns_profit_impact, 2),
        "estimated_profit_after_returns": round(profit_before_returns + returns_profit_impact, 2),
    }


def _build_ai_accounting_snapshot_drift_metrics(
    cogs_margin_df: pd.DataFrame,
    *,
    window_label: str = "30d",
    returns_df: pd.DataFrame | None = None,
    qbo_adjustments_df: pd.DataFrame | None = None,
) -> dict[str, object]:
    metrics = _build_slack_summary_drift_metrics(
        cogs_margin_df,
        window_label=window_label,
        returns_df=returns_df,
        qbo_adjustments_df=qbo_adjustments_df,
    )
    metrics["observed_source"] = f"Ask/AI accounting snapshot {window_label}"
    return metrics


def _build_accounting_close_formula_checks(
    close_summary: dict[str, float | int | str],
    *,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    summary = close_summary or {}
    gross_sales = _safe_float(summary.get("gross_sales"))
    net_before_cogs = _safe_float(summary.get("net_before_cogs"))
    fifo_cogs = _safe_float(summary.get("fifo_cogs"))
    fifo_margin = _safe_float(summary.get("fifo_margin"))
    returns_refund_total = _safe_float(summary.get("returns_refund_total"))
    returns_cogs_reversal_total = _safe_float(summary.get("returns_cogs_reversal_total"))
    returns_estimated_profit_impact = _safe_float(summary.get("returns_estimated_profit_impact"))
    net_after_returns_and_cogs = _safe_float(summary.get("net_after_returns_and_cogs"))
    shipping_charged_total = _safe_float(summary.get("shipping_charged_total"))
    shipping_label_spend_total = _safe_float(summary.get("shipping_label_spend_total"))
    shipping_delta_total = _safe_float(summary.get("shipping_delta_total"))
    fee_total = _safe_float(summary.get("fee_total"))
    checks = [
        {
            "check": "net_before_cogs_component_formula",
            "formula": "gross_sales + shipping_charged_total - fee_total - shipping_label_spend_total",
            "expected": round(gross_sales + shipping_charged_total - fee_total - shipping_label_spend_total, 2),
            "observed": round(net_before_cogs, 2),
        },
        {
            "check": "net_before_cogs_minus_fifo_cogs_equals_fifo_margin",
            "formula": "net_before_cogs - fifo_cogs",
            "expected": round(net_before_cogs - fifo_cogs, 2),
            "observed": round(fifo_margin, 2),
        },
        {
            "check": "shipping_delta_total_formula",
            "formula": "shipping_charged_total - shipping_label_spend_total",
            "expected": round(shipping_charged_total - shipping_label_spend_total, 2),
            "observed": round(shipping_delta_total, 2),
        },
        {
            "check": "return_profit_impact_formula",
            "formula": "-returns_refund_total + returns_cogs_reversal_total",
            "expected": round(-returns_refund_total + returns_cogs_reversal_total, 2),
            "observed": round(returns_estimated_profit_impact, 2),
        },
        {
            "check": "net_after_returns_and_cogs_formula",
            "formula": "fifo_margin + returns_estimated_profit_impact",
            "expected": round(fifo_margin + returns_estimated_profit_impact, 2),
            "observed": round(net_after_returns_and_cogs, 2),
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= tolerance else "warn",
                "formula": str(row.get("formula") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(float(tolerance), 2),
            }
        )
    return pd.DataFrame(output)


def _accounting_close_profit_before_returns(close_summary: dict[str, float | int | str]) -> float:
    if close_summary.get("profit_before_returns") is not None:
        return _safe_float(close_summary.get("profit_before_returns"))
    return _safe_float(close_summary.get("fifo_margin"))


def _accounting_close_estimated_profit_after_returns(close_summary: dict[str, float | int | str]) -> float:
    if close_summary.get("estimated_profit_after_returns") is not None:
        return _safe_float(close_summary.get("estimated_profit_after_returns"))
    return _safe_float(close_summary.get("net_after_returns_and_cogs", close_summary.get("fifo_margin")))


def _apply_formula_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    formula_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    formula_warn_count = (
        int((formula_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if formula_df is not None and not formula_df.empty
        else 0
    )
    summary["formula_warn_count"] = formula_warn_count
    if formula_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "accounting formula warnings" not in blockers:
            blockers.append("accounting formula warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Accounting Formula Warnings",
                "status": "fail" if formula_warn_count > 0 else "pass",
                "value": formula_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_sales_component_checks(
    *,
    sales_df: pd.DataFrame,
    cogs_margin_df: pd.DataFrame,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    sales_count = 0 if sales_df is None else int(len(sales_df))
    margin_count = 0 if cogs_margin_df is None else int(len(cogs_margin_df))
    sales_gross = _sum(sales_df, "gross_sales")
    margin_gross = _sum(cogs_margin_df, "gross_sales")
    sales_component_net = (
        sales_gross
        + _sum(sales_df, "actual_shipping_charged")
        - _sum(sales_df, "actual_fee")
        - _sum(sales_df, "actual_shipping_label_cost")
    )
    sales_actual_net = _sum(sales_df, "actual_net_before_cogs")
    margin_net = _sum(cogs_margin_df, "net_before_cogs")
    checks = [
        {
            "check": "sales_detail_count_matches_cogs_margin",
            "expected_source": "COGS & Margin Detail",
            "observed_source": "Sales Detail",
            "expected": float(margin_count),
            "observed": float(sales_count),
            "tolerance": 0.0,
        },
        {
            "check": "sales_detail_gross_matches_cogs_margin",
            "expected_source": "COGS & Margin Detail.gross_sales",
            "observed_source": "Sales Detail.gross_sales",
            "expected": margin_gross,
            "observed": sales_gross,
            "tolerance": tolerance,
        },
        {
            "check": "sales_detail_component_net_formula",
            "expected_source": "gross + actual_shipping_charged - actual_fee - actual_shipping_label_cost",
            "observed_source": "Sales Detail.actual_net_before_cogs",
            "expected": sales_component_net,
            "observed": sales_actual_net,
            "tolerance": tolerance,
        },
        {
            "check": "sales_detail_net_matches_cogs_margin",
            "expected_source": "COGS & Margin Detail.net_before_cogs",
            "observed_source": "Sales Detail.actual_net_before_cogs",
            "expected": margin_net,
            "observed": sales_actual_net,
            "tolerance": tolerance,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_sales_component_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    component_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    component_warn_count = (
        int((component_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if component_df is not None and not component_df.empty
        else 0
    )
    summary["sales_component_warn_count"] = component_warn_count
    if component_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "sales component tie-out warnings" not in blockers:
            blockers.append("sales component tie-out warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Sales Component Tie-Out Warnings",
                "status": "fail" if component_warn_count > 0 else "pass",
                "value": component_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_return_tieout_checks(
    *,
    returns_df: pd.DataFrame,
    qbo_adjustments_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    returns_count = 0 if returns_df is None else int(len(returns_df))
    qbo_count = 0 if qbo_adjustments_df is None else int(len(qbo_adjustments_df))
    returns_refund_total = (
        _sum(returns_df, "refund_amount") + _sum(returns_df, "refund_fees") + _sum(returns_df, "refund_shipping")
    )
    qbo_refund_total = (
        _sum(qbo_adjustments_df, "refund_amount")
        + _sum(qbo_adjustments_df, "refund_fees")
        + _sum(qbo_adjustments_df, "refund_shipping")
    )
    qbo_cogs_reversal = _sum(qbo_adjustments_df, "cogs_reversal_estimate")
    qbo_profit_impact = _sum(qbo_adjustments_df, "estimated_profit_impact")
    qbo_profit_formula = -qbo_refund_total + qbo_cogs_reversal
    checks = [
        {
            "check": "returns_count_matches_qbo_adjustments",
            "expected_source": "Returns",
            "observed_source": "QuickBooks Refund/Adjustment Export",
            "expected": float(returns_count),
            "observed": float(qbo_count),
            "tolerance": 0.0,
        },
        {
            "check": "returns_refund_total_matches_qbo_adjustments",
            "expected_source": "Returns refund_amount + refund_fees + refund_shipping",
            "observed_source": "QuickBooks Refund/Adjustment Export refund components",
            "expected": returns_refund_total,
            "observed": qbo_refund_total,
            "tolerance": tolerance,
        },
        {
            "check": "return_cogs_reversal_matches_close_summary",
            "expected_source": "Accounting Close Readiness.returns_cogs_reversal_total",
            "observed_source": "QuickBooks Refund/Adjustment Export.cogs_reversal_estimate",
            "expected": _safe_float(close_summary.get("returns_cogs_reversal_total")),
            "observed": qbo_cogs_reversal,
            "tolerance": tolerance,
        },
        {
            "check": "qbo_return_profit_impact_formula",
            "expected_source": "-qbo_refund_total + qbo_cogs_reversal_estimate",
            "observed_source": "QuickBooks Refund/Adjustment Export.estimated_profit_impact",
            "expected": qbo_profit_formula,
            "observed": qbo_profit_impact,
            "tolerance": tolerance,
        },
        {
            "check": "return_profit_impact_matches_close_summary",
            "expected_source": "Accounting Close Readiness.returns_estimated_profit_impact",
            "observed_source": "QuickBooks Refund/Adjustment Export.estimated_profit_impact",
            "expected": _safe_float(close_summary.get("returns_estimated_profit_impact")),
            "observed": qbo_profit_impact,
            "tolerance": tolerance,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_return_tieout_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    return_tieout_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    return_warn_count = (
        int((return_tieout_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if return_tieout_df is not None and not return_tieout_df.empty
        else 0
    )
    summary["return_tieout_warn_count"] = return_warn_count
    if return_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "return tie-out warnings" not in blockers:
            blockers.append("return tie-out warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Return Tie-Out Warnings",
                "status": "fail" if return_warn_count > 0 else "pass",
                "value": return_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_inventory_valuation_checks(
    *,
    inventory_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    tolerance: float = 0.01,
) -> pd.DataFrame:
    if inventory_df is None or inventory_df.empty:
        formula_value = 0.0
        snapshot_value = 0.0
        stocked_missing_cost = 0
        stocked_rows = 0
    else:
        df = inventory_df.copy()
        df["qty_on_hand"] = pd.to_numeric(df.get("qty_on_hand", pd.Series(dtype=float)), errors="coerce").fillna(0.0)
        df["landed_unit_cost"] = pd.to_numeric(
            df.get("landed_unit_cost", pd.Series(dtype=float)),
            errors="coerce",
        )
        df["landed_inventory_value"] = pd.to_numeric(
            df.get("landed_inventory_value", pd.Series(dtype=float)),
            errors="coerce",
        )
        stocked = df["qty_on_hand"] > 0
        stocked_rows = int(stocked.sum())
        stocked_missing_cost = int((stocked & df["landed_unit_cost"].isna()).sum())
        formula_value = float((df["qty_on_hand"].fillna(0.0) * df["landed_unit_cost"].fillna(0.0)).sum())
        snapshot_value = float(df["landed_inventory_value"].fillna(0.0).sum())
    checks = [
        {
            "check": "stocked_inventory_rows_have_landed_cost",
            "status": "pass" if stocked_missing_cost == 0 else "warn",
            "expected_source": "Inventory Snapshot stocked rows",
            "observed_source": "Inventory Snapshot landed_unit_cost",
            "expected": 0.0,
            "observed": float(stocked_missing_cost),
            "delta_observed_minus_expected": float(stocked_missing_cost),
            "tolerance": 0.0,
            "details": f"{stocked_rows} stocked inventory row(s) evaluated.",
        },
        {
            "check": "inventory_snapshot_value_formula",
            "expected_source": "sum(qty_on_hand * landed_unit_cost)",
            "observed_source": "Inventory Snapshot.landed_inventory_value",
            "expected": formula_value,
            "observed": snapshot_value,
            "tolerance": tolerance,
        },
        {
            "check": "close_inventory_value_matches_inventory_snapshot",
            "expected_source": "Accounting Close Readiness.inventory_value",
            "observed_source": "Inventory Snapshot.landed_inventory_value",
            "expected": _safe_float(close_summary.get("inventory_value")),
            "observed": snapshot_value,
            "tolerance": tolerance,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = _safe_float(row.get("delta_observed_minus_expected", observed - expected))
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": str(row.get("status") or ("pass" if abs(delta) <= row_tolerance else "warn")),
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
                "details": str(row.get("details") or ""),
            }
        )
    return pd.DataFrame(output)


def _apply_inventory_valuation_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    valuation_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    valuation_warn_count = (
        int((valuation_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if valuation_df is not None and not valuation_df.empty
        else 0
    )
    summary["inventory_valuation_warn_count"] = valuation_warn_count
    if valuation_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "inventory valuation warnings" not in blockers:
            blockers.append("inventory valuation warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Inventory Valuation Warnings",
                "status": "fail" if valuation_warn_count > 0 else "pass",
                "value": valuation_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_fee_evidence_checks(
    *,
    sales_df: pd.DataFrame,
    fee_reconciliation_df: pd.DataFrame,
    fee_source_priority_df: pd.DataFrame,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    sales_count = 0 if sales_df is None else int(len(sales_df))
    fee_row_count = 0 if fee_reconciliation_df is None else int(len(fee_reconciliation_df))
    source_row_count = 0
    sale_field_fallback_rows = 0
    source_actual_fee_total = 0.0
    if fee_source_priority_df is not None and not fee_source_priority_df.empty:
        source_df = fee_source_priority_df.copy()
        source_df["actual_fee_source"] = source_df.get("actual_fee_source", pd.Series(dtype=str)).fillna("").astype(str)
        source_df["sales_count"] = pd.to_numeric(
            source_df.get("sales_count", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        source_df["actual_fee_total"] = pd.to_numeric(
            source_df.get("actual_fee_total", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        source_row_count = int(source_df["sales_count"].sum())
        sale_field_fallback_rows = int(
            source_df.loc[source_df["actual_fee_source"] == "sale_fees_field", "sales_count"].sum()
        )
        source_actual_fee_total = float(source_df["actual_fee_total"].sum())
    sales_actual_fee_total = _sum(sales_df, "actual_fee")
    reconciliation_actual_fee_total = _sum(fee_reconciliation_df, "actual_fee")
    checks = [
        {
            "check": "fee_reconciliation_rows_cover_sales_detail",
            "expected_source": "Sales Detail",
            "observed_source": "eBay Fee Reconciliation",
            "expected": float(sales_count),
            "observed": float(fee_row_count),
            "tolerance": 0.0,
        },
        {
            "check": "fee_reconciliation_total_matches_sales_detail",
            "expected_source": "Sales Detail.actual_fee",
            "observed_source": "eBay Fee Reconciliation.actual_fee",
            "expected": sales_actual_fee_total,
            "observed": reconciliation_actual_fee_total,
            "tolerance": tolerance,
        },
        {
            "check": "fee_source_priority_rows_match_reconciliation",
            "expected_source": "eBay Fee Reconciliation",
            "observed_source": "eBay Fee Source Priority.sales_count",
            "expected": float(fee_row_count),
            "observed": float(source_row_count),
            "tolerance": 0.0,
        },
        {
            "check": "fee_source_priority_total_matches_reconciliation",
            "expected_source": "eBay Fee Reconciliation.actual_fee",
            "observed_source": "eBay Fee Source Priority.actual_fee_total",
            "expected": reconciliation_actual_fee_total,
            "observed": source_actual_fee_total,
            "tolerance": tolerance,
        },
        {
            "check": "sale_fee_field_fallback_rows",
            "expected_source": "Fee source priority",
            "observed_source": "sale_fees_field fallback rows",
            "expected": 0.0,
            "observed": float(sale_field_fallback_rows),
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_fee_evidence_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    fee_evidence_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    fee_warn_count = (
        int((fee_evidence_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if fee_evidence_df is not None and not fee_evidence_df.empty
        else 0
    )
    summary["fee_evidence_warn_count"] = fee_warn_count
    if fee_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "fee evidence warnings" not in blockers:
            blockers.append("fee evidence warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Fee Evidence Warnings",
                "status": "fail" if fee_warn_count > 0 else "pass",
                "value": fee_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_shipping_evidence_checks(
    *,
    sales_df: pd.DataFrame,
    shipping_economics_df: pd.DataFrame,
    shipping_econ_summary_df: pd.DataFrame,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    sales_shipping_charged = _sum(sales_df, "actual_shipping_charged")
    sales_label_spend = _sum(sales_df, "actual_shipping_label_cost")
    detail_shipping_charged = _sum(shipping_economics_df, "shipping_charged_to_buyer")
    detail_label_spend = _sum(shipping_economics_df, "shipping_label_spend")
    detail_delta = _sum(shipping_economics_df, "shipping_delta_charged_minus_spend")
    summary_shipping_charged = _sum(shipping_econ_summary_df, "total_shipping_charged")
    summary_label_spend = _sum(shipping_econ_summary_df, "total_label_spend")
    summary_delta = _sum(shipping_econ_summary_df, "shipping_delta_charged_minus_spend")
    missing_label_rows = 0
    if shipping_economics_df is not None and not shipping_economics_df.empty:
        detail_df = shipping_economics_df.copy()
        detail_df["shipping_charged_to_buyer"] = pd.to_numeric(
            detail_df.get("shipping_charged_to_buyer", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        detail_df["shipping_label_spend"] = pd.to_numeric(
            detail_df.get("shipping_label_spend", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        missing_label_rows = int(
            ((detail_df["shipping_charged_to_buyer"] > 0) & (detail_df["shipping_label_spend"] <= 0)).sum()
        )
    checks = [
        {
            "check": "shipping_charged_sales_detail_matches_shipping_economics",
            "expected_source": "Sales Detail.actual_shipping_charged",
            "observed_source": "Shipping Economics.shipping_charged_to_buyer",
            "expected": sales_shipping_charged,
            "observed": detail_shipping_charged,
            "tolerance": tolerance,
        },
        {
            "check": "label_spend_sales_detail_matches_shipping_economics",
            "expected_source": "Sales Detail.actual_shipping_label_cost",
            "observed_source": "Shipping Economics.shipping_label_spend",
            "expected": sales_label_spend,
            "observed": detail_label_spend,
            "tolerance": tolerance,
        },
        {
            "check": "shipping_economics_delta_formula",
            "expected_source": "shipping_charged_to_buyer - shipping_label_spend",
            "observed_source": "Shipping Economics.shipping_delta_charged_minus_spend",
            "expected": detail_shipping_charged - detail_label_spend,
            "observed": detail_delta,
            "tolerance": tolerance,
        },
        {
            "check": "shipping_summary_charged_matches_detail",
            "expected_source": "Shipping Economics Detail.shipping_charged_to_buyer",
            "observed_source": "Shipping Economics Summary.total_shipping_charged",
            "expected": detail_shipping_charged,
            "observed": summary_shipping_charged,
            "tolerance": tolerance,
        },
        {
            "check": "shipping_summary_label_spend_matches_detail",
            "expected_source": "Shipping Economics Detail.shipping_label_spend",
            "observed_source": "Shipping Economics Summary.total_label_spend",
            "expected": detail_label_spend,
            "observed": summary_label_spend,
            "tolerance": tolerance,
        },
        {
            "check": "shipping_summary_delta_matches_detail",
            "expected_source": "Shipping Economics Detail shipping delta",
            "observed_source": "Shipping Economics Summary.shipping_delta_charged_minus_spend",
            "expected": detail_shipping_charged - detail_label_spend,
            "observed": summary_delta,
            "tolerance": tolerance,
        },
        {
            "check": "paid_shipping_rows_missing_label_spend",
            "expected_source": "Shipping Economics",
            "observed_source": "paid shipping rows with zero label spend",
            "expected": 0.0,
            "observed": float(missing_label_rows),
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_shipping_evidence_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    shipping_evidence_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    shipping_warn_count = (
        int((shipping_evidence_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if shipping_evidence_df is not None and not shipping_evidence_df.empty
        else 0
    )
    summary["shipping_evidence_warn_count"] = shipping_warn_count
    if shipping_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "shipping evidence warnings" not in blockers:
            blockers.append("shipping evidence warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Shipping Evidence Warnings",
                "status": "fail" if shipping_warn_count > 0 else "pass",
                "value": shipping_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_reconciliation_tieout_checks(
    *,
    sales_df: pd.DataFrame,
    returns_df: pd.DataFrame,
    reconciliation_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    sales_count = 0 if sales_df is None else int(len(sales_df))
    returns_count = 0 if returns_df is None else int(len(returns_df))
    reconcile_sales_count = _sum(reconciliation_df, "sales_count")
    reconcile_returns_count = _sum(reconciliation_df, "returns_count")
    sales_gross = _sum(sales_df, "gross_sales")
    sales_net = _sum(sales_df, "actual_net_before_cogs")
    reconcile_gross = _sum(reconciliation_df, "sales_gross")
    reconcile_net = _sum(reconciliation_df, "sales_net_before_returns")
    returns_refund_total = _sum(returns_df, "refund_amount") + _sum(returns_df, "refund_fees") + _sum(
        returns_df, "refund_shipping"
    )
    reconcile_returns_total = _sum(reconciliation_df, "returns_refund_total")
    reconcile_net_after_returns = _sum(reconciliation_df, "net_after_returns")
    reconcile_flag_count = (
        int(reconciliation_df.get("reconcile_flag", pd.Series(dtype=bool)).fillna(False).astype(bool).sum())
        if reconciliation_df is not None and not reconciliation_df.empty
        else 0
    )
    checks = [
        {
            "check": "reconciliation_sales_count_matches_sales_detail",
            "expected_source": "Sales Detail",
            "observed_source": "Reconciliation by Marketplace.sales_count",
            "expected": float(sales_count),
            "observed": reconcile_sales_count,
            "tolerance": 0.0,
        },
        {
            "check": "reconciliation_sales_gross_matches_sales_detail",
            "expected_source": "Sales Detail.gross_sales",
            "observed_source": "Reconciliation by Marketplace.sales_gross",
            "expected": sales_gross,
            "observed": reconcile_gross,
            "tolerance": tolerance,
        },
        {
            "check": "reconciliation_net_before_returns_matches_sales_detail",
            "expected_source": "Sales Detail.actual_net_before_cogs",
            "observed_source": "Reconciliation by Marketplace.sales_net_before_returns",
            "expected": sales_net,
            "observed": reconcile_net,
            "tolerance": tolerance,
        },
        {
            "check": "reconciliation_returns_count_matches_returns",
            "expected_source": "Returns",
            "observed_source": "Reconciliation by Marketplace.returns_count",
            "expected": float(returns_count),
            "observed": reconcile_returns_count,
            "tolerance": 0.0,
        },
        {
            "check": "reconciliation_returns_total_matches_returns",
            "expected_source": "Returns refund components",
            "observed_source": "Reconciliation by Marketplace.returns_refund_total",
            "expected": returns_refund_total,
            "observed": reconcile_returns_total,
            "tolerance": tolerance,
        },
        {
            "check": "reconciliation_net_after_returns_formula",
            "expected_source": "sales_net_before_returns - returns_refund_total",
            "observed_source": "Reconciliation by Marketplace.net_after_returns",
            "expected": reconcile_net - reconcile_returns_total,
            "observed": reconcile_net_after_returns,
            "tolerance": tolerance,
        },
        {
            "check": "reconciliation_flags_match_close_summary",
            "expected_source": "Accounting Close Readiness.reconcile_flags",
            "observed_source": "Reconciliation by Marketplace.reconcile_flag",
            "expected": _safe_float(close_summary.get("reconcile_flags")),
            "observed": float(reconcile_flag_count),
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_reconciliation_tieout_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    reconciliation_tieout_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    reconciliation_warn_count = (
        int((reconciliation_tieout_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if reconciliation_tieout_df is not None and not reconciliation_tieout_df.empty
        else 0
    )
    summary["reconciliation_tieout_warn_count"] = reconciliation_warn_count
    if reconciliation_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "reconciliation tie-out warnings" not in blockers:
            blockers.append("reconciliation tie-out warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Reconciliation Tie-Out Warnings",
                "status": "fail" if reconciliation_warn_count > 0 else "pass",
                "value": reconciliation_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_cogs_source_checks(
    *,
    cogs_margin_df: pd.DataFrame,
    cogs_source_summary_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    sale_fifo_cogs_evidence_df: pd.DataFrame | None = None,
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    margin_sale_count = 0 if cogs_margin_df is None else int(len(cogs_margin_df))
    source_sale_count = int(_sum(cogs_source_summary_df, "sale_count"))
    margin_quantity = _sum(cogs_margin_df, "quantity")
    source_quantity = _sum(cogs_source_summary_df, "quantity")
    margin_fifo_cogs = _sum(cogs_margin_df, "fifo_cogs")
    source_fifo_cogs = _sum(cogs_source_summary_df, "fifo_cogs")
    margin_fifo_margin = _sum(cogs_margin_df, "fifo_margin")
    source_fifo_margin = _sum(cogs_source_summary_df, "fifo_margin")
    expected_evidence_row_count = int(_sum(cogs_margin_df, "fifo_cogs_evidence_rows"))
    evidence_row_count = 0 if sale_fifo_cogs_evidence_df is None else int(len(sale_fifo_cogs_evidence_df))
    evidence_sale_count = 0
    evidence_total_cost = _sum(sale_fifo_cogs_evidence_df, "total_cost")
    if (
        sale_fifo_cogs_evidence_df is not None
        and not sale_fifo_cogs_evidence_df.empty
        and "sale_id" in sale_fifo_cogs_evidence_df.columns
    ):
        evidence_sale_count = int(sale_fifo_cogs_evidence_df["sale_id"].dropna().nunique())
    equal_fallback_cogs = 0.0
    missing_cost_cogs = 0.0
    if cogs_source_summary_df is not None and not cogs_source_summary_df.empty:
        source_df = cogs_source_summary_df.copy()
        source_df["fifo_cost_source"] = source_df.get("fifo_cost_source", pd.Series(dtype=str)).fillna("").astype(str)
        source_df["fifo_cogs"] = pd.to_numeric(
            source_df.get("fifo_cogs", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        equal_fallback_cogs = float(
            source_df.loc[source_df["fifo_cost_source"] == "lot_equal_quantity_fallback", "fifo_cogs"].sum()
        )
        missing_cost_cogs = float(
            source_df.loc[source_df["fifo_cost_source"].isin(["missing_cost_basis", "unknown"]), "fifo_cogs"].sum()
        )
    checks = [
        {
            "check": "cogs_source_sale_count_matches_margin_detail",
            "expected_source": "COGS & Margin Detail",
            "observed_source": "Sold COGS Source Summary.sale_count",
            "expected": float(margin_sale_count),
            "observed": float(source_sale_count),
            "tolerance": 0.0,
        },
        {
            "check": "cogs_source_quantity_matches_margin_detail",
            "expected_source": "COGS & Margin Detail.quantity",
            "observed_source": "Sold COGS Source Summary.quantity",
            "expected": margin_quantity,
            "observed": source_quantity,
            "tolerance": tolerance,
        },
        {
            "check": "cogs_source_fifo_cogs_matches_margin_detail",
            "expected_source": "COGS & Margin Detail.fifo_cogs",
            "observed_source": "Sold COGS Source Summary.fifo_cogs",
            "expected": margin_fifo_cogs,
            "observed": source_fifo_cogs,
            "tolerance": tolerance,
        },
        {
            "check": "cogs_source_fifo_cogs_matches_close_summary",
            "expected_source": "Accounting Close Readiness.fifo_cogs",
            "observed_source": "Sold COGS Source Summary.fifo_cogs",
            "expected": _safe_float(close_summary.get("fifo_cogs")),
            "observed": source_fifo_cogs,
            "tolerance": tolerance,
        },
        {
            "check": "cogs_source_fifo_margin_matches_margin_detail",
            "expected_source": "COGS & Margin Detail.fifo_margin",
            "observed_source": "Sold COGS Source Summary.fifo_margin",
            "expected": margin_fifo_margin,
            "observed": source_fifo_margin,
            "tolerance": tolerance,
        },
        {
            "check": "fifo_cogs_evidence_total_matches_margin_detail",
            "expected_source": "COGS & Margin Detail.fifo_cogs",
            "observed_source": "Sale FIFO COGS Evidence.total_cost",
            "expected": margin_fifo_cogs,
            "observed": evidence_total_cost,
            "tolerance": tolerance,
        },
        {
            "check": "fifo_cogs_evidence_sale_count_matches_margin_detail",
            "expected_source": "COGS & Margin Detail sale rows",
            "observed_source": "Sale FIFO COGS Evidence distinct sale_id",
            "expected": float(margin_sale_count),
            "observed": float(evidence_sale_count),
            "tolerance": 0.0,
        },
        {
            "check": "fifo_cogs_evidence_row_count_matches_margin_detail",
            "expected_source": "COGS & Margin Detail.fifo_cogs_evidence_rows",
            "observed_source": "Sale FIFO COGS Evidence row count",
            "expected": float(expected_evidence_row_count),
            "observed": float(evidence_row_count),
            "tolerance": 0.0,
        },
        {
            "check": "sold_equal_fallback_cogs_matches_close_summary",
            "expected_source": "Accounting Close Readiness.sold_equal_fallback_cogs",
            "observed_source": "Sold COGS Source Summary lot_equal_quantity_fallback",
            "expected": _safe_float(close_summary.get("sold_equal_fallback_cogs")),
            "observed": equal_fallback_cogs,
            "tolerance": tolerance,
        },
        {
            "check": "sold_missing_cost_cogs_matches_close_summary",
            "expected_source": "Accounting Close Readiness.sold_missing_cost_cogs",
            "observed_source": "Sold COGS Source Summary missing/unknown",
            "expected": _safe_float(close_summary.get("sold_missing_cost_cogs")),
            "observed": missing_cost_cogs,
            "tolerance": tolerance,
        },
        {
            "check": "sold_equal_fallback_cogs_present",
            "expected_source": "Sold COGS Source Summary",
            "observed_source": "lot_equal_quantity_fallback COGS",
            "expected": 0.0,
            "observed": equal_fallback_cogs,
            "tolerance": 0.0,
        },
        {
            "check": "sold_missing_cost_cogs_present",
            "expected_source": "Sold COGS Source Summary",
            "observed_source": "missing/unknown COGS",
            "expected": 0.0,
            "observed": missing_cost_cogs,
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_cogs_source_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    cogs_source_checks_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    cogs_source_warn_count = (
        int((cogs_source_checks_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if cogs_source_checks_df is not None and not cogs_source_checks_df.empty
        else 0
    )
    summary["cogs_source_warn_count"] = cogs_source_warn_count
    if cogs_source_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "COGS source warnings" not in blockers:
            blockers.append("COGS source warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "COGS Source Warnings",
                "status": "fail" if cogs_source_warn_count > 0 else "pass",
                "value": cogs_source_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_lot_allocation_checks(
    *,
    lots_df: pd.DataFrame,
    lot_allocation_source_summary_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    tolerance: float = 0.01,
) -> pd.DataFrame:
    def _sum(df: pd.DataFrame, column: str) -> float:
        if df is None or df.empty or column not in df.columns:
            return 0.0
        return float(pd.to_numeric(df.get(column, pd.Series(dtype=float)), errors="coerce").fillna(0.0).sum())

    lot_assignment_count = 0 if lots_df is None else int(len(lots_df))
    summary_assignment_count = int(_sum(lot_allocation_source_summary_df, "assignment_count"))
    lot_quantity = _sum(lots_df, "quantity_acquired")
    summary_quantity = _sum(lot_allocation_source_summary_df, "quantity_acquired")
    lot_resolved_cost = _sum(lots_df, "resolved_landed_total_cost")
    summary_resolved_cost = _sum(lot_allocation_source_summary_df, "resolved_landed_total_cost")
    equal_fallback_assignments = 0
    missing_cost_assignments = 0
    equal_fallback_cost = 0.0
    missing_cost = 0.0
    if lot_allocation_source_summary_df is not None and not lot_allocation_source_summary_df.empty:
        source_df = lot_allocation_source_summary_df.copy()
        source_df["cost_source"] = source_df.get("cost_source", pd.Series(dtype=str)).fillna("").astype(str)
        source_df["assignment_count"] = pd.to_numeric(
            source_df.get("assignment_count", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        source_df["resolved_landed_total_cost"] = pd.to_numeric(
            source_df.get("resolved_landed_total_cost", pd.Series(dtype=float)),
            errors="coerce",
        ).fillna(0.0)
        equal_rows = source_df["cost_source"] == "lot_equal_quantity_fallback"
        missing_rows = source_df["cost_source"].isin(["missing_cost_basis", "unknown"])
        equal_fallback_assignments = int(source_df.loc[equal_rows, "assignment_count"].sum())
        missing_cost_assignments = int(source_df.loc[missing_rows, "assignment_count"].sum())
        equal_fallback_cost = float(source_df.loc[equal_rows, "resolved_landed_total_cost"].sum())
        missing_cost = float(source_df.loc[missing_rows, "resolved_landed_total_cost"].sum())
    checks = [
        {
            "check": "lot_allocation_assignment_count_matches_detail",
            "expected_source": "Lot Assignment",
            "observed_source": "Lot Allocation Source Summary.assignment_count",
            "expected": float(lot_assignment_count),
            "observed": float(summary_assignment_count),
            "tolerance": 0.0,
        },
        {
            "check": "lot_allocation_quantity_matches_detail",
            "expected_source": "Lot Assignment.quantity_acquired",
            "observed_source": "Lot Allocation Source Summary.quantity_acquired",
            "expected": lot_quantity,
            "observed": summary_quantity,
            "tolerance": tolerance,
        },
        {
            "check": "lot_allocation_resolved_cost_matches_detail",
            "expected_source": "Lot Assignment.resolved_landed_total_cost",
            "observed_source": "Lot Allocation Source Summary.resolved_landed_total_cost",
            "expected": lot_resolved_cost,
            "observed": summary_resolved_cost,
            "tolerance": tolerance,
        },
        {
            "check": "lot_equal_fallback_assignments_matches_close_summary",
            "expected_source": "Accounting Close Readiness.lot_equal_fallback_assignments",
            "observed_source": "Lot Allocation Source Summary lot_equal_quantity_fallback assignments",
            "expected": _safe_float(close_summary.get("lot_equal_fallback_assignments")),
            "observed": float(equal_fallback_assignments),
            "tolerance": 0.0,
        },
        {
            "check": "lot_missing_cost_assignments_matches_close_summary",
            "expected_source": "Accounting Close Readiness.lot_missing_cost_assignments",
            "observed_source": "Lot Allocation Source Summary missing/unknown assignments",
            "expected": _safe_float(close_summary.get("lot_missing_cost_assignments")),
            "observed": float(missing_cost_assignments),
            "tolerance": 0.0,
        },
        {
            "check": "lot_equal_fallback_assignments_present",
            "expected_source": "Lot Allocation Source Summary",
            "observed_source": "lot_equal_quantity_fallback assignment count",
            "expected": 0.0,
            "observed": float(equal_fallback_assignments),
            "tolerance": 0.0,
        },
        {
            "check": "lot_missing_cost_assignments_present",
            "expected_source": "Lot Allocation Source Summary",
            "observed_source": "missing/unknown assignment count",
            "expected": 0.0,
            "observed": float(missing_cost_assignments),
            "tolerance": 0.0,
        },
        {
            "check": "lot_equal_fallback_cost_present",
            "expected_source": "Lot Allocation Source Summary",
            "observed_source": "lot_equal_quantity_fallback resolved cost",
            "expected": 0.0,
            "observed": equal_fallback_cost,
            "tolerance": 0.0,
        },
        {
            "check": "lot_missing_cost_basis_cost_present",
            "expected_source": "Lot Allocation Source Summary",
            "observed_source": "missing/unknown resolved cost",
            "expected": 0.0,
            "observed": missing_cost,
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_lot_allocation_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    lot_allocation_checks_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    lot_allocation_warn_count = (
        int((lot_allocation_checks_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if lot_allocation_checks_df is not None and not lot_allocation_checks_df.empty
        else 0
    )
    summary["lot_allocation_warn_count"] = lot_allocation_warn_count
    if lot_allocation_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "lot allocation warnings" not in blockers:
            blockers.append("lot allocation warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Lot Allocation Warnings",
                "status": "fail" if lot_allocation_warn_count > 0 else "pass",
                "value": lot_allocation_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_exception_queue_checks(
    *,
    accounting_exceptions_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    tolerance: float = 0.01,
) -> pd.DataFrame:
    exceptions = (
        accounting_exceptions_df.copy()
        if accounting_exceptions_df is not None and not accounting_exceptions_df.empty
        else pd.DataFrame()
    )
    total_count = int(len(exceptions))
    severity = (
        exceptions.get("severity", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.upper()
        if not exceptions.empty
        else pd.Series(dtype=str)
    )
    p0_count = int((severity == "P0").sum())
    p1_count = int((severity == "P1").sum())
    allowed_severities = {"P0", "P1", "P2", "P3", "WARN", "WARNING", "INFO"}
    missing_or_unknown_severity_count = int((~severity.isin(allowed_severities)).sum()) if not severity.empty else 0
    exception_type = (
        exceptions.get("exception_type", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        if not exceptions.empty
        else pd.Series(dtype=str)
    )
    missing_type_count = int((exception_type == "").sum()) if not exception_type.empty else 0
    checks = [
        {
            "check": "total_exception_count_matches_close_summary",
            "expected_source": "Accounting Close Readiness.total_exceptions",
            "observed_source": "Accounting Exception Queue row count",
            "expected": _safe_float(close_summary.get("total_exceptions")),
            "observed": float(total_count),
            "tolerance": 0.0,
        },
        {
            "check": "p0_exception_count_matches_close_summary",
            "expected_source": "Accounting Close Readiness.p0_exceptions",
            "observed_source": "Accounting Exception Queue severity=P0",
            "expected": _safe_float(close_summary.get("p0_exceptions")),
            "observed": float(p0_count),
            "tolerance": 0.0,
        },
        {
            "check": "p1_exception_count_matches_close_summary",
            "expected_source": "Accounting Close Readiness.p1_exceptions",
            "observed_source": "Accounting Exception Queue severity=P1",
            "expected": _safe_float(close_summary.get("p1_exceptions")),
            "observed": float(p1_count),
            "tolerance": 0.0,
        },
        {
            "check": "p0_exceptions_present",
            "expected_source": "Close readiness blocking policy",
            "observed_source": "Accounting Exception Queue severity=P0",
            "expected": 0.0,
            "observed": float(p0_count),
            "tolerance": 0.0,
        },
        {
            "check": "exception_rows_have_severity",
            "expected_source": "Accounting Exception Queue schema",
            "observed_source": "blank/unknown severity rows",
            "expected": 0.0,
            "observed": float(missing_or_unknown_severity_count),
            "tolerance": 0.0,
        },
        {
            "check": "exception_rows_have_type",
            "expected_source": "Accounting Exception Queue schema",
            "observed_source": "blank exception_type rows",
            "expected": 0.0,
            "observed": float(missing_type_count),
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_exception_queue_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    exception_queue_checks_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    exception_queue_warn_count = (
        int((exception_queue_checks_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if exception_queue_checks_df is not None and not exception_queue_checks_df.empty
        else 0
    )
    summary["exception_queue_warn_count"] = exception_queue_warn_count
    if exception_queue_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "exception queue warnings" not in blockers:
            blockers.append("exception queue warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Exception Queue Warnings",
                "status": "fail" if exception_queue_warn_count > 0 else "pass",
                "value": exception_queue_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_margin_anomaly_checks(
    *,
    cogs_margin_df: pd.DataFrame,
    accounting_exceptions_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    tolerance: float = 0.01,
) -> pd.DataFrame:
    margin = (
        cogs_margin_df.copy()
        if cogs_margin_df is not None and not cogs_margin_df.empty
        else pd.DataFrame()
    )
    if not margin.empty:
        fifo_margin = pd.to_numeric(margin.get("fifo_margin", pd.Series(dtype=float)), errors="coerce").fillna(0.0)
    else:
        fifo_margin = pd.Series(dtype=float)
    negative_margin_rows = int((fifo_margin < 0).sum()) if not fifo_margin.empty else 0
    nonpositive_margin_rows = int((fifo_margin <= 0).sum()) if not fifo_margin.empty else 0

    exceptions = (
        accounting_exceptions_df.copy()
        if accounting_exceptions_df is not None and not accounting_exceptions_df.empty
        else pd.DataFrame()
    )
    exception_type = (
        exceptions.get("exception_type", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
        if not exceptions.empty
        else pd.Series(dtype=str)
    )
    nonpositive_exception_count = int((exception_type == "nonpositive_margin").sum()) if not exception_type.empty else 0

    checks = [
        {
            "check": "negative_fifo_margin_rows_match_close_summary",
            "expected_source": "Accounting Close Readiness.negative_margin_rows",
            "observed_source": "COGS & Margin Detail.fifo_margin < 0",
            "expected": _safe_float(close_summary.get("negative_margin_rows")),
            "observed": float(negative_margin_rows),
            "tolerance": 0.0,
        },
        {
            "check": "nonpositive_fifo_margin_rows_have_exception",
            "expected_source": "COGS & Margin Detail.fifo_margin <= 0",
            "observed_source": "Accounting Exception Queue.exception_type=nonpositive_margin",
            "expected": float(nonpositive_margin_rows),
            "observed": float(nonpositive_exception_count),
            "tolerance": 0.0,
        },
        {
            "check": "negative_fifo_margin_rows_present",
            "expected_source": "Close readiness margin policy",
            "observed_source": "COGS & Margin Detail.fifo_margin < 0",
            "expected": 0.0,
            "observed": float(negative_margin_rows),
            "tolerance": 0.0,
        },
        {
            "check": "nonpositive_fifo_margin_rows_present",
            "expected_source": "Close readiness margin policy",
            "observed_source": "COGS & Margin Detail.fifo_margin <= 0",
            "expected": 0.0,
            "observed": float(nonpositive_margin_rows),
            "tolerance": 0.0,
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        row_tolerance = _safe_float(row.get("tolerance"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= row_tolerance else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": round(row_tolerance, 2),
            }
        )
    return pd.DataFrame(output)


def _apply_margin_anomaly_checks_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    margin_anomaly_checks_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    margin_anomaly_warn_count = (
        int((margin_anomaly_checks_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if margin_anomaly_checks_df is not None and not margin_anomaly_checks_df.empty
        else 0
    )
    summary["margin_anomaly_warn_count"] = margin_anomaly_warn_count
    if margin_anomaly_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "margin anomaly warnings" not in blockers:
            blockers.append("margin anomaly warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Margin Anomaly Warnings",
                "status": "fail" if margin_anomaly_warn_count > 0 else "pass",
                "value": margin_anomaly_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_close_consistency_checks(
    *,
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
) -> pd.DataFrame:
    summary = dict(close_summary or {})
    checks_df = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    blockers = [
        part.strip()
        for part in str(summary.get("blockers") or "").split(",")
        if part.strip()
    ]
    warnings = [
        part.strip()
        for part in str(summary.get("warnings") or "").split(",")
        if part.strip()
    ]
    status_series = (
        checks_df.get("status", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.lower()
        if not checks_df.empty
        else pd.Series(dtype=str)
    )
    fail_count = int((status_series == "fail").sum()) if not status_series.empty else 0
    warn_count = int((status_series == "warn").sum()) if not status_series.empty else 0
    readiness_status = str(summary.get("readiness_status") or "").strip().lower()
    close_ready_risk_count = int(_safe_float(summary.get("blocker_count")) + warn_count)
    checks = [
        {
            "check": "blocker_count_matches_blocker_list",
            "expected_source": "Accounting Close Readiness.blocker_count",
            "observed_source": "Accounting Close Readiness.blockers list length",
            "expected": _safe_float(summary.get("blocker_count")),
            "observed": float(len(blockers)),
        },
        {
            "check": "warning_count_matches_warning_list",
            "expected_source": "Accounting Close Readiness.warning_count",
            "observed_source": "Accounting Close Readiness.warnings list length",
            "expected": _safe_float(summary.get("warning_count")),
            "observed": float(len(warnings)),
        },
        {
            "check": "failed_close_checks_have_blocked_status",
            "expected_source": "Close readiness status policy",
            "observed_source": "fail rows while readiness_status is not blocked",
            "expected": 0.0,
            "observed": float(fail_count if fail_count > 0 and readiness_status != "blocked" else 0),
        },
        {
            "check": "warning_close_checks_prevent_close_ready",
            "expected_source": "Close readiness status policy",
            "observed_source": "warn rows while readiness_status is close_ready",
            "expected": 0.0,
            "observed": float(warn_count if warn_count > 0 and readiness_status == "close_ready" else 0),
        },
        {
            "check": "close_ready_has_no_blockers_or_warnings",
            "expected_source": "Close readiness status policy",
            "observed_source": "blockers plus warn check rows while readiness_status is close_ready",
            "expected": 0.0,
            "observed": float(close_ready_risk_count if readiness_status == "close_ready" else 0),
        },
    ]
    output: list[dict[str, object]] = []
    for row in checks:
        expected = _safe_float(row.get("expected"))
        observed = _safe_float(row.get("observed"))
        delta = observed - expected
        output.append(
            {
                "check": str(row.get("check") or ""),
                "status": "pass" if abs(delta) <= 0.0 else "warn",
                "expected_source": str(row.get("expected_source") or ""),
                "observed_source": str(row.get("observed_source") or ""),
                "expected": round(expected, 2),
                "observed": round(observed, 2),
                "delta_observed_minus_expected": round(delta, 2),
                "tolerance": 0.0,
            }
        )
    return pd.DataFrame(output)


def _build_accounting_close_packet_completeness_checks(
    *,
    report_frames: dict[str, pd.DataFrame],
    close_summary: dict[str, float | int | str],
) -> pd.DataFrame:
    summary = close_summary or {}
    sales_count = int(_safe_float(summary.get("sales_count")))
    has_return_activity = (
        abs(_safe_float(summary.get("returns_refund_total"))) > 0
        or abs(_safe_float(summary.get("returns_cogs_reversal_total"))) > 0
        or abs(_safe_float(summary.get("returns_estimated_profit_impact"))) > 0
    )
    required_specs = [
        ("accounting_close_readiness_checks", 1, "always"),
        ("accounting_close_formula_checks", 1, "always"),
        ("accounting_sales_component_checks", 1, "always"),
        ("accounting_return_tieout_checks", 1, "always"),
        ("accounting_inventory_valuation_checks", 1, "always"),
        ("accounting_fee_evidence_checks", 1, "always"),
        ("accounting_shipping_evidence_checks", 1, "always"),
        ("accounting_reconciliation_tieout_checks", 1, "always"),
        ("accounting_cogs_source_checks", 1, "always"),
        ("accounting_lot_allocation_checks", 1, "always"),
        ("accounting_exception_queue_checks", 1, "always"),
        ("accounting_margin_anomaly_checks", 1, "always"),
        ("accounting_close_consistency_checks", 1, "always"),
        ("accounting_period_drift_checks", 1, "always"),
        ("inventory_snapshot", 1, "always"),
        ("sales_detail", 1 if sales_count > 0 else 0, "when sales_count > 0"),
        ("cogs_margin_detail", 1 if sales_count > 0 else 0, "when sales_count > 0"),
        ("sale_fifo_cogs_evidence", 1 if sales_count > 0 else 0, "when sales_count > 0"),
        ("qbo_sales_export", 1 if sales_count > 0 else 0, "when sales_count > 0"),
        ("qbo_adjustments_export", 1 if has_return_activity else 0, "when return activity exists"),
    ]
    rows: list[dict[str, object]] = []
    for prefix, expected_min_rows, required_when in required_specs:
        df = report_frames.get(prefix)
        present = df is not None
        row_count = int(len(df)) if present and df is not None else 0
        expected_min_rows = int(expected_min_rows)
        missing_required = expected_min_rows > 0 and (not present or row_count < expected_min_rows)
        rows.append(
            {
                "artifact": f"{prefix}.csv",
                "status": "warn" if missing_required else "pass",
                "required_when": required_when,
                "expected_min_rows": expected_min_rows,
                "observed_rows": row_count,
                "present_in_report_list": bool(present),
                "details": (
                    "Required close-packet evidence is missing or empty."
                    if missing_required
                    else "Required close-packet evidence is present."
                ),
            }
        )
    return pd.DataFrame(rows)


def _accounting_close_packet_prefixes() -> set[str]:
    return {
        "accounting_close_readiness_checks",
        "accounting_close_signoffs",
        "accounting_close_signoff_review",
        "accounting_close_formula_checks",
        "accounting_sales_component_checks",
        "accounting_return_tieout_checks",
        "accounting_inventory_valuation_checks",
        "accounting_fee_evidence_checks",
        "accounting_shipping_evidence_checks",
        "accounting_reconciliation_tieout_checks",
        "accounting_cogs_source_checks",
        "accounting_lot_allocation_checks",
        "accounting_exception_queue_checks",
        "accounting_margin_anomaly_checks",
        "accounting_close_consistency_checks",
        "accounting_close_packet_completeness_checks",
        "accounting_close_packet_manifest_checks",
        "accounting_close_packet_hash_checks",
        "accounting_close_packet_evidence_hash",
        "accounting_period_drift_checks",
        "ai_review_outcomes",
        "accounting_exception_queue",
        "sales_detail",
        "cogs_margin_detail",
        "sale_fifo_cogs_evidence",
        "margin_by_sku",
        "reconciliation_marketplace",
        "shipping_economics_detail",
        "shipping_economics_summary",
        "ebay_fee_estimate_vs_actual",
        "ebay_fee_source_priority",
        "tax_summary_estimated",
        "tax_by_marketplace_estimated",
        "tax_detail_estimated",
        "tax_exceptions_advisor_review",
        "tax_reporting_signoffs",
        "tax_reporting_signoff_review",
        "lot_assignment",
        "lot_allocation_source_summary",
        "cogs_source_summary",
        "inventory_snapshot",
        "returns",
        "qbo_sales_export",
        "qbo_adjustments_export",
    }


def _accounting_close_packet_evidence_hash_prefixes() -> set[str]:
    excluded = {
        "accounting_close_signoffs",
        "accounting_close_signoff_review",
        "accounting_close_packet_completeness_checks",
        "accounting_close_packet_manifest_checks",
        "accounting_close_packet_hash_checks",
        "accounting_close_packet_evidence_hash",
        "tax_reporting_signoff_review",
        "ai_review_outcomes",
    }
    return _accounting_close_packet_prefixes() - excluded


def _build_accounting_close_packet_evidence_hash_rows(
    *,
    evidence_hash: str,
    from_date,
    to_date,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "artifact": "accounting_close_packet",
                "hash_key": "accounting_close_packet_evidence_hash_sha256",
                "sha256": str(evidence_hash or "").strip(),
                "from_date": str(from_date),
                "to_date": str(to_date),
                "hash_scope": "selected close CSV payloads, date range, and close summary; excludes sign-off/review/hash evidence tables",
            }
        ]
    )


def _accounting_close_packet_evidence_hash_from_frames(
    *,
    report_frames: dict[str, pd.DataFrame],
    close_summary: dict[str, float | int | str],
    from_date,
    to_date,
) -> str:
    report_csv_by_prefix: dict[str, str] = {}
    for prefix in sorted(_accounting_close_packet_evidence_hash_prefixes()):
        if prefix not in report_frames:
            continue
        export_df = report_frames.get(prefix) if report_frames.get(prefix) is not None else pd.DataFrame()
        report_csv_by_prefix[str(prefix)] = export_df.to_csv(index=False)
    hash_payload = {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "close_summary": close_summary or {},
        "reports": report_csv_by_prefix,
    }
    return hashlib.sha256(json.dumps(hash_payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def _build_accounting_close_packet_manifest_checks(
    *,
    report_frames: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    selected_prefixes = sorted(_accounting_close_packet_prefixes())
    rows: list[dict[str, object]] = []
    self_generated_prefixes = {"accounting_close_packet_manifest_checks", "accounting_close_packet_hash_checks"}
    for prefix in selected_prefixes:
        df = report_frames.get(prefix)
        self_generated = prefix in self_generated_prefixes and df is None
        present = df is not None or self_generated
        row_count = len(selected_prefixes) if self_generated else (int(len(df)) if df is not None else 0)
        rows.append(
            {
                "artifact": f"{prefix}.csv",
                "status": "pass" if present else "warn",
                "manifest_key": f"row_count_{prefix}",
                "manifest_value": row_count,
                "observed_rows": row_count,
                "present_in_report_list": bool(present),
                "details": (
                    "Manifest row count will match the exported dataframe row count."
                    if present
                    else "Close packet prefix is selected for export but missing from the report list."
                ),
            }
        )
    return pd.DataFrame(rows)


def _dataframe_csv_sha256(df: pd.DataFrame | None) -> str:
    export_df = df if df is not None else pd.DataFrame()
    return hashlib.sha256(export_df.to_csv(index=False).encode("utf-8")).hexdigest()


def _stable_json_sha256(payload: object) -> str:
    return hashlib.sha256(json.dumps(payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def _build_accounting_close_packet_hash_checks(
    *,
    report_frames: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    selected_prefixes = sorted(_accounting_close_packet_prefixes())
    rows: list[dict[str, object]] = []
    self_generated_prefixes = {"accounting_close_packet_hash_checks"}
    for prefix in selected_prefixes:
        df = report_frames.get(prefix)
        self_generated = prefix in self_generated_prefixes and df is None
        present = df is not None or self_generated
        rows.append(
            {
                "artifact": f"{prefix}.csv",
                "status": "pass" if present else "warn",
                "manifest_hash_key": f"sha256_{prefix}",
                "sha256": "self-generated-in-packet" if self_generated else (_dataframe_csv_sha256(df) if present else ""),
                "present_in_report_list": bool(present),
                "details": (
                    "CSV hash is available for packet integrity review."
                    if present
                    else "Close packet prefix is selected for export but missing from the report list."
                ),
            }
        )
    return pd.DataFrame(rows)


def _apply_period_drift_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    drift_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    drift_warn_count = (
        int((drift_df.get("status", pd.Series(dtype=str)).astype(str) == "warn").sum())
        if drift_df is not None and not drift_df.empty
        else 0
    )
    summary["period_drift_warn_count"] = drift_warn_count
    if drift_warn_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "period drift warnings" not in blockers:
            blockers.append("period drift warnings")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "Period Drift Warnings",
                "status": "fail" if drift_warn_count > 0 else "pass",
                "value": drift_warn_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _apply_ai_review_outcomes_to_close_readiness(
    close_summary: dict[str, float | int | str],
    close_checks_df: pd.DataFrame,
    ai_review_outcomes_df: pd.DataFrame,
) -> tuple[dict[str, float | int | str], pd.DataFrame]:
    summary = dict(close_summary or {})
    checks = close_checks_df.copy() if close_checks_df is not None and not close_checks_df.empty else pd.DataFrame()
    latest_by_type: dict[str, dict[str, object]] = {}
    if ai_review_outcomes_df is not None and not ai_review_outcomes_df.empty:
        sort_col = "recorded_at_utc" if "recorded_at_utc" in ai_review_outcomes_df.columns else None
        rows_df = (
            ai_review_outcomes_df.sort_values(sort_col, ascending=False)
            if sort_col
            else ai_review_outcomes_df
        )
        for row in rows_df.to_dict("records"):
            review_type = str(row.get("review_type") or row.get("intent") or "unknown").strip()
            if review_type and review_type not in latest_by_type:
                latest_by_type[review_type] = row
    followup_rows = [
        row
        for row in latest_by_type.values()
        if str(row.get("outcome") or "").strip().lower() in {"edited", "rejected"}
    ]
    followup_count = int(len(followup_rows))
    summary["ai_review_followup_count"] = followup_count
    if followup_count > 0:
        blockers = [
            part.strip()
            for part in str(summary.get("blockers") or "").split(",")
            if part.strip()
        ]
        if "AI review outcome follow-up" not in blockers:
            blockers.append("AI review outcome follow-up")
        summary["blockers"] = ", ".join(blockers)
        summary["blocker_count"] = int(len(blockers))
        summary["readiness_status"] = "blocked"
    check_row = pd.DataFrame(
        [
            {
                "check": "AI Review Outcome Follow-Up",
                "status": "fail" if followup_count > 0 else "pass",
                "value": followup_count,
            }
        ]
    )
    if checks.empty:
        return summary, check_row
    return summary, pd.concat([checks, check_row], ignore_index=True)


def _build_accounting_close_signoff_review(
    *,
    signoff_df: pd.DataFrame,
    close_summary: dict[str, float | int | str],
    from_date,
    to_date,
    current_packet_hash: str = "",
) -> pd.DataFrame:
    expected_periods = {
        str(from_date or "").strip(),
        str(to_date or "").strip(),
        f"{from_date}..{to_date}",
        f"{from_date} to {to_date}",
    }
    try:
        if getattr(from_date, "year", None) == getattr(to_date, "year", None) and getattr(
            from_date, "month", None
        ) == getattr(to_date, "month", None):
            expected_periods.add(f"{int(from_date.year):04d}-{int(from_date.month):02d}")
    except Exception:
        pass
    expected_periods = {value for value in expected_periods if value}

    def _col(name: str) -> pd.Series:
        if signoff_df is None or signoff_df.empty or name not in signoff_df.columns:
            return pd.Series(dtype=str)
        return signoff_df[name].fillna("").astype(str)

    matching_df = pd.DataFrame()
    if signoff_df is not None and not signoff_df.empty:
        period_series = _col("close_period").str.strip()
        type_series = _col("signoff_type").str.strip().str.lower()
        matching_df = signoff_df.loc[
            period_series.isin(expected_periods)
            & type_series.isin({"monthly_close_review", "accounting_close_review", "close_review"})
        ].copy()
    approved_df = pd.DataFrame()
    if not matching_df.empty and "status" in matching_df.columns:
        approved_df = matching_df.loc[matching_df["status"].fillna("").astype(str).str.lower() == "approved"].copy()
        if not approved_df.empty:
            approved_df["_signoff_sort_date"] = pd.to_datetime(
                approved_df.get("signoff_date", pd.Series(dtype=str)),
                errors="coerce",
                utc=True,
            )
            approved_df["_recorded_sort_date"] = pd.to_datetime(
                approved_df.get("recorded_at_utc", pd.Series(dtype=str)),
                errors="coerce",
                utc=True,
            )
            approved_df = approved_df.sort_values(
                ["_signoff_sort_date", "_recorded_sort_date"],
                ascending=[False, False],
                na_position="last",
                kind="mergesort",
            ).drop(columns=["_signoff_sort_date", "_recorded_sort_date"])

    current_status = str(close_summary.get("readiness_status") or "").strip().lower()
    current_exceptions = int(_safe_float(close_summary.get("total_exceptions")))
    current_blockers = int(_safe_float(close_summary.get("blocker_count")))
    current_drift_warnings = int(_safe_float(close_summary.get("period_drift_warn_count")))
    current_ai_followups = int(_safe_float(close_summary.get("ai_review_followup_count")))
    rows: list[dict[str, object]] = [
        {
            "check": "Close Sign-Off Evidence Present",
            "status": "pass" if not approved_df.empty else ("warn" if current_status == "close_ready" else "info"),
            "expected": "approved monthly close sign-off for selected period",
            "observed": "approved" if not approved_df.empty else "missing",
            "details": (
                "Approved close sign-off evidence found for the selected period."
                if not approved_df.empty
                else "No approved monthly close sign-off evidence found for the selected period."
            ),
        }
    ]
    if approved_df.empty:
        return pd.DataFrame(rows)

    latest = approved_df.iloc[0].to_dict()
    signoff_readiness = str(latest.get("close_readiness_status") or "").strip().lower()
    signoff_exceptions = int(_safe_float(latest.get("exception_count")))
    signoff_blockers = int(_safe_float(latest.get("unresolved_blocker_count")))
    signoff_drift_warnings = int(_safe_float(latest.get("period_drift_warn_count")))
    signoff_ai_followups = int(_safe_float(latest.get("ai_review_followup_count")))
    signoff_packet_ref = str(latest.get("accounting_packet_ref") or "").strip()
    signoff_packet_hash = str(latest.get("accounting_packet_hash") or "").strip().lower()
    current_packet_hash = str(current_packet_hash or "").strip().lower()
    signoff_evidence_link = str(latest.get("evidence_link") or "").strip()
    signoff_owner = str(latest.get("owner") or "").strip()
    signoff_date = str(latest.get("signoff_date") or "").strip()

    def _coerce_review_date(value):
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if all(hasattr(value, attr) for attr in ("year", "month", "day")):
            try:
                return datetime(int(value.year), int(value.month), int(value.day)).date()
            except Exception:
                return None
        raw = str(value).strip()
        if not raw:
            return None
        try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
        except Exception:
            try:
                return datetime.fromisoformat(raw[:10]).date()
            except Exception:
                return None

    signoff_review_date = _coerce_review_date(signoff_date)
    close_end_date = _coerce_review_date(to_date)
    today = utc_today()
    signoff_date_valid = bool(
        signoff_review_date
        and (close_end_date is None or signoff_review_date >= close_end_date)
        and signoff_review_date <= today
    )
    if not signoff_date:
        signoff_date_valid_observed = "missing"
    elif signoff_review_date is None:
        signoff_date_valid_observed = f"{signoff_date} (unparseable)"
    elif close_end_date is not None and signoff_review_date < close_end_date:
        signoff_date_valid_observed = f"{signoff_review_date.isoformat()} before {close_end_date.isoformat()}"
    elif signoff_review_date > today:
        signoff_date_valid_observed = f"{signoff_review_date.isoformat()} after {today.isoformat()}"
    else:
        signoff_date_valid_observed = signoff_review_date.isoformat()

    rows.extend(
        [
            {
                "check": "Approved Sign-Off Readiness Match",
                "status": "pass" if signoff_readiness == current_status else "warn",
                "expected": current_status,
                "observed": signoff_readiness or "missing",
                "details": "Latest approved sign-off readiness status should match the recalculated close readiness.",
            },
            {
                "check": "Approved Sign-Off Blocker Count",
                "status": "pass" if signoff_blockers == current_blockers else "warn",
                "expected": current_blockers,
                "observed": signoff_blockers,
                "details": "Latest approved sign-off unresolved blocker count should match the recalculated close blockers.",
            },
            {
                "check": "Approved Sign-Off Exception Count",
                "status": "pass" if signoff_exceptions == current_exceptions else "warn",
                "expected": current_exceptions,
                "observed": signoff_exceptions,
                "details": "Latest approved sign-off exception count should match the recalculated close exception count.",
            },
            {
                "check": "Approved Sign-Off Drift Warning Count",
                "status": "pass" if signoff_drift_warnings == current_drift_warnings else "warn",
                "expected": current_drift_warnings,
                "observed": signoff_drift_warnings,
                "details": "Latest approved sign-off drift warning count should match recalculated period drift warnings.",
            },
            {
                "check": "Approved Sign-Off AI Review Follow-Up Count",
                "status": "pass" if signoff_ai_followups == current_ai_followups else "warn",
                "expected": current_ai_followups,
                "observed": signoff_ai_followups,
                "details": "Latest approved sign-off AI review follow-up count should match recalculated Copilot/AI Accountant outcome blockers.",
            },
            {
                "check": "Approved Sign-Off Owner Present",
                "status": "pass" if signoff_owner else "warn",
                "expected": "review owner",
                "observed": signoff_owner or "missing",
                "details": "Approved monthly close sign-off should identify the reviewer or owner.",
            },
            {
                "check": "Approved Sign-Off Date Present",
                "status": "pass" if signoff_date else "warn",
                "expected": "sign-off date",
                "observed": signoff_date or "missing",
                "details": "Approved monthly close sign-off should include the approval date.",
            },
            {
                "check": "Approved Sign-Off Date Validity",
                "status": "pass" if signoff_date_valid else ("info" if not signoff_date else "warn"),
                "expected": "parseable date from period end through today",
                "observed": signoff_date_valid_observed,
                "details": "Approved monthly close sign-off date should be parseable, on or after the close period end, and not future-dated.",
            },
            {
                "check": "Approved Sign-Off Is Close Ready",
                "status": "pass" if current_status == "close_ready" and current_blockers == 0 else "warn",
                "expected": "close_ready with 0 blockers",
                "observed": f"{current_status or 'missing'} with {current_blockers} blocker(s)",
                "details": "Approved monthly close sign-off should not be treated as current if the recalculated close is blocked.",
            },
            {
                "check": "Approved Sign-Off Packet Evidence",
                "status": "pass" if signoff_packet_ref or signoff_evidence_link else "warn",
                "expected": "accounting packet reference or evidence link",
                "observed": signoff_packet_ref or signoff_evidence_link or "missing",
                "details": "Approved monthly close sign-off should reference the close packet or external evidence reviewed.",
            },
            {
                "check": "Approved Sign-Off Packet Hash",
                "status": (
                    "pass"
                    if signoff_packet_hash and current_packet_hash and signoff_packet_hash == current_packet_hash
                    else ("info" if not current_packet_hash else "warn")
                ),
                "expected": current_packet_hash or "current packet hash unavailable",
                "observed": signoff_packet_hash or "missing",
                "details": (
                    "Approved monthly close sign-off packet hash should match the recalculated close-packet evidence hash."
                ),
            },
        ]
    )
    return pd.DataFrame(rows)


def _build_accounting_close_export_packet(
    *,
    reports: list[tuple[str, pd.DataFrame, str]],
    close_summary: dict[str, float | int | str],
    from_date,
    to_date,
) -> bytes:
    close_report_prefixes = _accounting_close_packet_prefixes()
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    manifest_rows = [
        {"key": "generated_at", "value": generated_at},
        {"key": "from_date", "value": str(from_date)},
        {"key": "to_date", "value": str(to_date)},
    ]
    for key, value in sorted((close_summary or {}).items()):
        manifest_rows.append({"key": str(key), "value": str(value)})

    selected_reports = [
        (label, df, prefix)
        for label, df, prefix in reports
        if str(prefix or "").strip() in close_report_prefixes
    ]
    evidence_hash = _accounting_close_packet_evidence_hash_from_frames(
        report_frames={
            str(prefix): (df if df is not None else pd.DataFrame())
            for _label, df, prefix in selected_reports
        },
        close_summary=close_summary,
        from_date=from_date,
        to_date=to_date,
    )
    manifest_rows.append({"key": "accounting_close_packet_evidence_hash_sha256", "value": evidence_hash})

    for label, df, prefix in selected_reports:
        manifest_rows.append(
            {
                "key": f"row_count_{prefix}",
                "value": str(0 if df is None else int(len(df))),
            }
        )
        manifest_rows.append(
            {
                "key": f"sha256_{prefix}",
                "value": _dataframe_csv_sha256(df),
            }
        )

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as packet:
        manifest_df = pd.DataFrame(manifest_rows)
        packet.writestr("manifest.csv", manifest_df.to_csv(index=False))
        packet.writestr(
            "README.txt",
            (
                "GoldenStackers accounting close export packet.\n"
                "Tax outputs are estimates for operational planning. Validate local/state tax treatment with a tax advisor.\n"
                "Profit convention before returns: gross + shipping charged - fees - label spend - COGS.\n"
                "Estimated profit after returns: profit before returns - return refunds + return COGS reversal.\n"
                "QuickBooks Sales Export uses profit_before_returns_estimate for before-return profit; "
                "gross_margin_estimate is retained as a legacy compatibility alias.\n"
            ),
        )
        for label, df, prefix in selected_reports:
            export_df = df if df is not None else pd.DataFrame()
            packet.writestr(f"{prefix}.csv", export_df.to_csv(index=False))
    buffer.seek(0)
    return buffer.getvalue()


def _report_context_caption(file_prefix: str) -> str:
    prefix = str(file_prefix or "").strip()
    captions = {
        "qbo_sales_export": (
            "`profit_before_returns_estimate` is the preferred before-return profit field; "
            "`gross_margin_estimate` remains as a legacy compatibility alias."
        ),
        "qbo_adjustments_export": (
            "`estimated_profit_impact` is `-(refund_amount + refund_fees + refund_shipping) "
            "+ cogs_reversal_estimate`."
        ),
        "cogs_margin_detail": (
            "`fifo_margin` is before-return profit by sale. Use close-readiness `Est. Profit After Returns` "
            "for final return-adjusted profit."
        ),
        "sale_fifo_cogs_evidence": (
            "One row per FIFO COGS allocation consumed by a sale; ties sale margin back to product, lot, "
            "assignment, quantity, unit cost, total cost, and cost source."
        ),
        "accounting_period_drift_checks": (
            "Drift checks tie out before-return profit and estimated profit after returns separately."
        ),
    }
    return captions.get(prefix, "")


def _build_tax_review_export_packet(
    *,
    reports: list[tuple[str, pd.DataFrame, str]],
    from_date,
    to_date,
    tax_jurisdiction: str,
    tax_rate_percent: float,
    shipping_taxable: bool,
    marketplace_scope: str,
    facilitator_channels: set[str] | None,
    tax_exempt_categories: set[str] | None,
    tax_profile: dict[str, object] | None = None,
    extra_artifacts: list[tuple[str, bytes]] | None = None,
) -> bytes:
    extra_artifact_hashes = {
        str(name or "").strip(): hashlib.sha256(payload or b"").hexdigest()
        for name, payload in (extra_artifacts or [])
        if str(name or "").strip()
    }
    evidence_hash = _tax_review_packet_evidence_hash_from_reports(
        reports=reports,
        from_date=from_date,
        to_date=to_date,
        tax_jurisdiction=tax_jurisdiction,
        tax_rate_percent=tax_rate_percent,
        shipping_taxable=shipping_taxable,
        marketplace_scope=marketplace_scope,
        facilitator_channels=facilitator_channels,
        tax_exempt_categories=tax_exempt_categories,
        tax_profile=tax_profile,
        extra_artifact_hashes=extra_artifact_hashes,
    )
    tax_report_prefixes = {
        "tax_summary_estimated",
        "tax_by_marketplace_estimated",
        "tax_detail_estimated",
        "tax_exceptions_advisor_review",
        "tax_reporting_signoffs",
        "tax_reporting_signoff_review",
    }
    selected_reports = [
        (label, df, prefix)
        for label, df, prefix in reports
        if str(prefix or "").strip() in tax_report_prefixes
    ]
    report_csv_by_prefix: dict[str, str] = {}
    for _label, df, prefix in selected_reports:
        export_df = df if df is not None else pd.DataFrame()
        report_csv_by_prefix[str(prefix)] = export_df.to_csv(index=False)
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    manifest_rows = [
        {"key": "generated_at", "value": generated_at},
        {"key": "tax_packet_evidence_hash_sha256", "value": evidence_hash},
        {"key": "from_date", "value": str(from_date)},
        {"key": "to_date", "value": str(to_date)},
        {"key": "tax_jurisdiction", "value": str(tax_jurisdiction or "")},
        {"key": "tax_rate_percent", "value": str(float(tax_rate_percent or 0.0))},
        {"key": "shipping_taxable", "value": str(bool(shipping_taxable)).lower()},
        {"key": "marketplace_scope", "value": str(marketplace_scope or "all")},
        {"key": "facilitator_channels", "value": ",".join(sorted(facilitator_channels or set()))},
        {"key": "tax_exempt_categories", "value": ",".join(sorted(tax_exempt_categories or set()))},
    ]
    if tax_profile:
        for key in [
            "profile_key",
            "profile_name",
            "effective_from",
            "effective_to",
            "human_validation_status",
            "advisor_evidence_link",
        ]:
            manifest_rows.append({"key": f"tax_profile_{key}", "value": str(tax_profile.get(key) or "")})
    for label, df, prefix in selected_reports:
        manifest_rows.append(
            {
                "key": f"row_count_{prefix}",
                "value": str(0 if df is None else int(len(df))),
            }
        )
    for name, payload in extra_artifacts or []:
        safe_name = str(name or "").strip()
        if not safe_name:
            continue
        manifest_rows.append({"key": f"artifact_{safe_name}", "value": str(len(payload or b""))})
        manifest_rows.append(
            {
                "key": f"sha256_{safe_name}",
                "value": hashlib.sha256(payload or b"").hexdigest(),
            }
        )

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as packet:
        packet.writestr("manifest.csv", pd.DataFrame(manifest_rows).to_csv(index=False))
        packet.writestr(
            "README.txt",
            (
                "GoldenStackers tax review export packet.\n"
                "Tax outputs are estimates for operational planning and advisor review.\n"
                "Validate local/state treatment, marketplace facilitator rules, shipping taxability, "
                "and bullion/coin exemptions with your tax advisor before filing or remittance decisions.\n"
            ),
        )
        for label, df, prefix in selected_reports:
            packet.writestr(f"{prefix}.csv", report_csv_by_prefix.get(str(prefix), ""))
        for name, payload in extra_artifacts or []:
            safe_name = str(name or "").strip()
            if safe_name:
                packet.writestr(safe_name, payload or b"")
    buffer.seek(0)
    return buffer.getvalue()


def _tax_review_packet_evidence_hash_from_reports(
    *,
    reports: list[tuple[str, pd.DataFrame, str]],
    from_date,
    to_date,
    tax_jurisdiction: str,
    tax_rate_percent: float,
    shipping_taxable: bool,
    marketplace_scope: str,
    facilitator_channels: set[str] | None,
    tax_exempt_categories: set[str] | None,
    tax_profile: dict[str, object] | None = None,
    extra_artifact_hashes: dict[str, str] | None = None,
) -> str:
    tax_report_prefixes = {
        "tax_summary_estimated",
        "tax_by_marketplace_estimated",
        "tax_detail_estimated",
        "tax_exceptions_advisor_review",
        "tax_reporting_signoffs",
    }
    report_csv_by_prefix: dict[str, str] = {}
    for _label, df, prefix in reports:
        if str(prefix or "").strip() not in tax_report_prefixes:
            continue
        export_df = df if df is not None else pd.DataFrame()
        report_csv_by_prefix[str(prefix)] = export_df.to_csv(index=False)
    hash_payload = {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "tax_jurisdiction": str(tax_jurisdiction or ""),
        "tax_rate_percent": float(tax_rate_percent or 0.0),
        "shipping_taxable": bool(shipping_taxable),
        "marketplace_scope": str(marketplace_scope or "all"),
        "facilitator_channels": sorted(facilitator_channels or set()),
        "tax_exempt_categories": sorted(tax_exempt_categories or set()),
        "tax_profile": tax_profile or {},
        "extra_artifact_hashes": extra_artifact_hashes or {},
        "reports": report_csv_by_prefix,
    }
    return hashlib.sha256(json.dumps(hash_payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def _top_n_records(
    df: pd.DataFrame,
    *,
    sort_by: str,
    n: int = 10,
    ascending: bool = False,
    columns: list[str] | None = None,
) -> list[dict]:
    if df.empty or sort_by not in df.columns:
        return []
    slice_df = df
    if columns:
        keep = [c for c in columns if c in df.columns]
        if keep:
            slice_df = df[keep]
    top_df = (
        slice_df.nsmallest(int(n), sort_by)
        if ascending
        else slice_df.nlargest(int(n), sort_by)
    )
    return top_df.to_dict("records")


def _top_n_by_abs_records(
    df: pd.DataFrame,
    *,
    value_col: str,
    n: int = 10,
    drop_cols: list[str] | None = None,
) -> list[dict]:
    if df.empty or value_col not in df.columns:
        return []
    top_idx = df[value_col].abs().nlargest(int(n)).index
    top_df = df.loc[top_idx]
    if drop_cols:
        drop_existing = [c for c in drop_cols if c in top_df.columns]
        if drop_existing:
            top_df = top_df.drop(columns=drop_existing)
    return top_df.to_dict("records")


def _default_tax_marketplace_scope(
    sales_marketplace_options: list[str],
    facilitator_channels: set[str],
) -> list[str]:
    options = [str(v or "").strip().lower() for v in (sales_marketplace_options or []) if str(v or "").strip()]
    unique_options: list[str] = []
    seen: set[str] = set()
    for value in options:
        if value in seen:
            continue
        seen.add(value)
        unique_options.append(value)
    filtered = [m for m in unique_options if m not in set(facilitator_channels or set())]
    return filtered


def _bounded_dataframe(
    df: pd.DataFrame,
    *,
    render_full_tables: bool,
    preview_row_limit: int,
) -> tuple[pd.DataFrame, bool]:
    if render_full_tables:
        return df, False
    limit = max(1, int(preview_row_limit or 1))
    return df.head(limit), int(len(df)) > limit


def _extract_order_fee_breakdown_from_notes(notes: str | None) -> dict:
    raw = str(notes or "").strip()
    if not raw:
        return {}
    marker = "fee_breakdown_json="
    idx = raw.find(marker)
    if idx < 0:
        return {}
    json_raw = raw[idx + len(marker):].strip()
    if "; " in json_raw:
        json_raw = json_raw.split("; ", 1)[0].strip()
    if not json_raw:
        return {}
    try:
        payload = json.loads(json_raw)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _parse_marketplace_payload_json(raw_payload: str | None) -> dict:
    raw = str(raw_payload or "").strip()
    if not raw:
        return {}
    try:
        payload = json.loads(raw)
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _parse_ai_json_sections(raw_payload: str | None, section_keys: list[str]) -> dict[str, list[str]]:
    raw = str(raw_payload or "").strip()
    if not raw:
        return {}
    if raw.startswith("```"):
        lines = raw.splitlines()
        if lines and lines[0].strip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        raw = "\n".join(lines).strip()
    if not raw.startswith("{"):
        start_idx = raw.find("{")
        end_idx = raw.rfind("}")
        if start_idx >= 0 and end_idx > start_idx:
            raw = raw[start_idx : end_idx + 1].strip()
    try:
        payload = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(payload, dict):
        return {}
    sections: dict[str, list[str]] = {}
    for key in section_keys:
        value = payload.get(key)
        if isinstance(value, list):
            items = [str(item).strip() for item in value if str(item).strip()]
        elif value is None:
            items = []
        else:
            text = str(value).strip()
            items = [text] if text else []
        if items:
            sections[key] = items
    return sections


def _render_ai_json_sections(raw_payload: str | None, section_labels: list[tuple[str, str]]) -> bool:
    sections = _parse_ai_json_sections(raw_payload, [key for key, _label in section_labels])
    if not sections:
        return False
    for key, label in section_labels:
        items = sections.get(key) or []
        if not items:
            continue
        st.markdown(f"**{label}**")
        for item in items:
            st.markdown(f"- {item}")
    return True


def _log_reports_ai_outcome(
    repo,
    *,
    actor: str,
    review_type: str,
    outcome: str,
    answer_text: str,
    review_metadata: dict[str, object] | None,
) -> None:
    if not hasattr(repo, "log_ai_chat_interaction"):
        return
    clean_review_type = str(review_type or "reports_ai_review").strip() or "reports_ai_review"
    clean_outcome = str(outcome or "").strip().lower()
    if clean_outcome not in {"accepted", "edited", "rejected"}:
        clean_outcome = "reviewed"
    metadata = dict(review_metadata or {})
    metadata.update(
        {
            "event_type": f"{clean_review_type}_outcome",
            "review_type": clean_review_type,
            "outcome": clean_outcome,
            "answer_hash_sha256": _stable_json_sha256(str(answer_text or "")),
            "requires_human_approval_for_writes": True,
        }
    )
    repo.log_ai_chat_interaction(
        actor=actor,
        prompt=f"{clean_review_type} outcome: {clean_outcome}",
        intent=f"{clean_review_type}_outcome",
        allowed_domains=["accounting", "reports", "sales", "orders", "inventory", "tax"],
        citations=[],
        answer_preview=str(answer_text or "").strip(),
        denied=False,
        elapsed_ms=0,
        metadata=metadata,
    )


def _parse_iso_datetime(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        return None


def _sale_from_row(row: dict):
    product_id = row.get("product_id")
    product_obj = None
    if product_id is not None:
        product_obj = SimpleNamespace(
            id=int(product_id),
            sku=str(row.get("sku") or "").strip() or None,
            title=str(row.get("product_title") or "").strip() or None,
            acquisition_cost=row.get("product_acquisition_cost"),
            product_cost=row.get("product_cost"),
        )
    return SimpleNamespace(
        id=int(row.get("sale_id") or 0),
        sold_at=_parse_iso_datetime(row.get("sold_at")),
        marketplace=str(row.get("marketplace") or "").strip(),
        order_id=row.get("order_id"),
        product=product_obj,
        listing_id=row.get("listing_id"),
        external_order_id=str(row.get("external_order_id") or "").strip(),
        quantity_sold=int(row.get("quantity_sold") or 0),
        sold_price=_safe_float(row.get("sold_price")),
        fees=_safe_float(row.get("fees")),
        shipping_cost=_safe_float(row.get("shipping_cost")),
        shipping_provider=str(row.get("shipping_provider") or "").strip(),
        shipping_service=str(row.get("shipping_service") or "").strip(),
        shipping_package_type=str(row.get("shipping_package_type") or "").strip(),
        tracking_number=str(row.get("tracking_number") or "").strip(),
        tracking_status=str(row.get("tracking_status") or "").strip(),
        shipping_exception_code=str(row.get("shipping_exception_code") or "").strip(),
        shipping_exception_action=str(row.get("shipping_exception_action") or "").strip(),
        shipping_exception_notes=str(row.get("shipping_exception_notes") or "").strip(),
        shipping_exception_resolved_at=_parse_iso_datetime(row.get("shipping_exception_resolved_at")),
        shipping_exception_resolved_by=str(row.get("shipping_exception_resolved_by") or "").strip(),
        shipment_exported_at=_parse_iso_datetime(row.get("shipment_exported_at")),
        shipped_at=_parse_iso_datetime(row.get("shipped_at")),
        delivered_at=_parse_iso_datetime(row.get("delivered_at")),
        shipping_label_cost=row.get("shipping_label_cost"),
    )


def _order_from_row(row: dict):
    return SimpleNamespace(
        id=int(row.get("order_id") or 0),
        sold_at=_parse_iso_datetime(row.get("sold_at")),
        marketplace=str(row.get("marketplace") or "").strip(),
        external_order_id=str(row.get("external_order_id") or "").strip(),
        order_status=str(row.get("status") or "").strip(),
        subtotal_amount=_safe_float(row.get("subtotal_amount")),
        fees=_safe_float(row.get("fees")),
        shipping_cost=_safe_float(row.get("shipping_cost")),
        shipping_label_cost=row.get("shipping_label_cost"),
        total_amount=_safe_float(row.get("total_amount")),
        notes=str(row.get("notes") or "").strip(),
        items=[None] * int(row.get("item_count") or 0),
    )


def _product_from_row(row: dict):
    return SimpleNamespace(
        id=int(row.get("product_id") or 0),
        sku=str(row.get("sku") or "").strip(),
        title=str(row.get("title") or "").strip(),
        description=str(row.get("description") or "").strip(),
        category=str(row.get("category") or "").strip(),
        metal_type=str(row.get("metal_type") or "").strip(),
        current_quantity=int(row.get("current_quantity") or 0),
        acquisition_cost=row.get("acquisition_cost"),
        product_cost=row.get("product_cost"),
        acquired_at=_parse_iso_datetime(row.get("acquired_at")),
        acquisition_tax_paid=row.get("acquisition_tax_paid"),
        acquisition_shipping_paid=row.get("acquisition_shipping_paid"),
        acquisition_handling_paid=row.get("acquisition_handling_paid"),
        weight_oz=row.get("weight_oz"),
        package_weight_oz=row.get("package_weight_oz"),
        package_length_in=row.get("package_length_in"),
        package_width_in=row.get("package_width_in"),
        package_height_in=row.get("package_height_in"),
    )


def _listing_from_row(row: dict):
    product_obj = None
    product_id = row.get("product_id")
    sku = str(row.get("sku") or "").strip()
    if product_id is not None or sku:
        product_obj = SimpleNamespace(
            id=int(product_id or 0) if product_id is not None else None,
            sku=sku or None,
        )
    return SimpleNamespace(
        id=int(row.get("listing_id") or 0),
        listed_at=_parse_iso_datetime(row.get("listed_at")),
        marketplace=str(row.get("marketplace") or "").strip(),
        product=product_obj,
        listing_title=str(row.get("listing_title") or "").strip(),
        listing_status=str(row.get("listing_status") or "").strip(),
        marketplace_url=str(row.get("marketplace_url") or "").strip(),
        marketplace_details=str(row.get("marketplace_details") or "").strip(),
        quantity_listed=int(row.get("quantity_listed") or 0),
        listing_price=_safe_float(row.get("listing_price")),
        external_listing_id=str(row.get("external_listing_id") or "").strip(),
    )


def _build_ebay_marketplace_fee_rows(repo: InventoryRepository, orders) -> list[dict]:
    rows: list[dict] = []
    order_lookup: dict[int, dict] = {}
    for order in orders:
        if str(getattr(order, "marketplace", "") or "").strip().lower() != "ebay":
            continue
        order_lookup[int(getattr(order, "id", 0) or 0)] = {
            "external_order_id": str(getattr(order, "external_order_id", "") or "").strip(),
            "sold_at": iso_or_none(getattr(order, "sold_at", None)),
        }

    order_ids = [oid for oid in order_lookup.keys() if oid > 0]
    if order_ids:
        normalized_entries = repo.db.scalars(
            select(OrderFinanceEntry).where(
                OrderFinanceEntry.order_id.in_(order_ids),
                OrderFinanceEntry.entry_kind == "marketplace_fee",
            )
        ).all()
        for entry in normalized_entries:
            order_meta = order_lookup.get(int(getattr(entry, "order_id", 0) or 0), {})
            rows.append(
                {
                    "order_id": int(getattr(entry, "order_id", 0) or 0),
                    "sold_at": order_meta.get("sold_at") or "",
                    "external_order_id": order_meta.get("external_order_id") or "",
                    "line_item_id": str(getattr(entry, "line_item_id", "") or "").strip(),
                    "sku": str(getattr(entry, "sku", "") or "").strip(),
                    "product_title": "",
                    "legacy_item_id": str(getattr(entry, "legacy_item_id", "") or "").strip(),
                    "fee_type": str(getattr(entry, "fee_type", "") or "").strip(),
                    "fee_amount": _safe_float(getattr(entry, "amount", 0)),
                    "fee_currency": str(getattr(entry, "currency", "") or "").strip(),
                    "fee_memo": str(getattr(entry, "memo", "") or "").strip(),
                    "transaction_id": str(getattr(entry, "transaction_id", "") or "").strip(),
                    "transaction_date": iso_or_none(getattr(entry, "transaction_date", None)) or "",
                    "transaction_type": str(getattr(entry, "transaction_type", "") or "").strip(),
                    "transaction_status": str(getattr(entry, "transaction_status", "") or "").strip(),
                    "source": str(getattr(entry, "source", "") or "").strip() or "normalized_order_finance_entries",
                }
            )
        if rows:
            return rows

    for order in orders:
        if str(getattr(order, "marketplace", "") or "").strip().lower() != "ebay":
            continue
        order_id = int(getattr(order, "id", 0) or 0)
        external_order_id = str(getattr(order, "external_order_id", "") or "").strip()
        sold_at = iso_or_none(getattr(order, "sold_at", None))
        payload = _parse_marketplace_payload_json(getattr(order, "marketplace_payload_json", None))
        if not payload:
            continue

        order_line_lookup: dict[str, dict] = {}
        for line in payload.get("lineItems") or []:
            if not isinstance(line, dict):
                continue
            line_id = str(line.get("lineItemId") or "").strip()
            if not line_id:
                continue
            order_line_lookup[line_id] = {
                "sku": str(line.get("sku") or "").strip(),
                "title": str(line.get("title") or "").strip(),
                "legacy_item_id": str(line.get("legacyItemId") or "").strip(),
            }

        for line in payload.get("lineItems") or []:
            if not isinstance(line, dict):
                continue
            line_id = str(line.get("lineItemId") or "").strip()
            fee_list = line.get("marketplaceFees") or []
            if not isinstance(fee_list, list):
                continue
            for fee in fee_list:
                if not isinstance(fee, dict):
                    continue
                amount = fee.get("amount") or {}
                rows.append(
                    {
                        "order_id": order_id,
                        "sold_at": sold_at,
                        "external_order_id": external_order_id,
                        "line_item_id": line_id,
                        "sku": str(line.get("sku") or "").strip(),
                        "product_title": str(line.get("title") or "").strip(),
                        "legacy_item_id": str(line.get("legacyItemId") or "").strip(),
                        "fee_type": str(fee.get("feeType") or "").strip(),
                        "fee_amount": _safe_float(amount.get("value")),
                        "fee_currency": str(amount.get("currency") or "").strip(),
                        "fee_memo": str(fee.get("feeMemo") or "").strip(),
                        "transaction_id": "",
                        "transaction_date": "",
                        "transaction_type": "",
                        "transaction_status": "",
                        "source": "order_payload_lineItems",
                    }
                )

        for tx in payload.get("_finance_transactions") or []:
            if not isinstance(tx, dict):
                continue
            tx_id = str(tx.get("transactionId") or "").strip()
            tx_date = str(tx.get("transactionDate") or "").strip()
            tx_type = str(tx.get("transactionType") or "").strip()
            tx_status = str(tx.get("transactionStatus") or "").strip()
            tx_fee_memo = str(tx.get("transactionMemo") or "").strip()
            tx_line_items = tx.get("orderLineItems") or []
            if not isinstance(tx_line_items, list):
                continue
            for tx_line in tx_line_items:
                if not isinstance(tx_line, dict):
                    continue
                line_id = str(tx_line.get("lineItemId") or "").strip()
                line_meta = order_line_lookup.get(line_id) or {}
                fee_list = tx_line.get("marketplaceFees") or []
                if not isinstance(fee_list, list):
                    continue
                for fee in fee_list:
                    if not isinstance(fee, dict):
                        continue
                    amount = fee.get("amount") or {}
                    rows.append(
                        {
                            "order_id": order_id,
                            "sold_at": sold_at,
                            "external_order_id": external_order_id,
                            "line_item_id": line_id,
                            "sku": str(line_meta.get("sku") or "").strip(),
                            "product_title": str(line_meta.get("title") or "").strip(),
                            "legacy_item_id": str(line_meta.get("legacy_item_id") or "").strip(),
                            "fee_type": str(fee.get("feeType") or "").strip(),
                            "fee_amount": _safe_float(amount.get("value")),
                            "fee_currency": str(amount.get("currency") or "").strip(),
                            "fee_memo": str(fee.get("feeMemo") or tx_fee_memo).strip(),
                            "transaction_id": tx_id,
                            "transaction_date": tx_date,
                            "transaction_type": tx_type,
                            "transaction_status": tx_status,
                            "source": "finance_transactions_orderLineItems",
                        }
                    )
    return rows


@st.cache_data(show_spinner=False, max_entries=64)
def _build_fee_source_priority_summary(reconciliation_df: pd.DataFrame) -> pd.DataFrame:
    if reconciliation_df is None or reconciliation_df.empty or "actual_fee_source" not in reconciliation_df.columns:
        return pd.DataFrame()
    source_priority = {
        "normalized_order_finance_entries_marketplace_fee_sum": 1,
        "order_fee_breakdown_total_marketplace_fee": 2,
        "sale_fees_field": 3,
    }
    grouped = (
        reconciliation_df.groupby(["actual_fee_source"], dropna=False, as_index=False)
        .agg(
            sales_count=("sale_id", "count"),
            actual_fee_total=("actual_fee", "sum"),
        )
    )
    total_rows = float(len(reconciliation_df))
    grouped["coverage_pct"] = grouped["sales_count"].map(
        lambda n: round((float(n or 0) / total_rows) * 100.0, 2) if total_rows > 0 else 0.0
    )
    grouped["priority_rank"] = grouped["actual_fee_source"].map(
        lambda s: source_priority.get(str(s or "").strip().lower(), 99)
    )
    grouped = grouped.sort_values(["priority_rank", "sales_count"], ascending=[True, False]).reset_index(drop=True)
    return grouped


@st.cache_data(show_spinner=False, max_entries=64)
def _build_fee_source_priority_trend(reconciliation_df: pd.DataFrame) -> pd.DataFrame:
    if (
        reconciliation_df is None
        or reconciliation_df.empty
        or "actual_fee_source" not in reconciliation_df.columns
        or "sold_at" not in reconciliation_df.columns
    ):
        return pd.DataFrame()
    base = reconciliation_df.copy()
    base["sold_at_dt"] = pd.to_datetime(base["sold_at"], errors="coerce")
    base = base[base["sold_at_dt"].notna()].copy()
    if base.empty:
        return pd.DataFrame()

    source_priority = {
        "normalized_order_finance_entries_marketplace_fee_sum": 1,
        "order_fee_breakdown_total_marketplace_fee": 2,
        "sale_fees_field": 3,
    }
    base["source_rank"] = base["actual_fee_source"].map(
        lambda s: source_priority.get(str(s or "").strip().lower(), 99)
    )

    daily = (
        base.assign(
            bucket_granularity="daily",
            bucket_date=base["sold_at_dt"].dt.strftime("%Y-%m-%d"),
        )
        .groupby(
            ["bucket_granularity", "bucket_date", "actual_fee_source", "source_rank"],
            dropna=False,
            as_index=False,
        )
        .agg(
            sales_count=("sale_id", "count"),
            actual_fee_total=("actual_fee", "sum"),
        )
    )
    weekly = (
        base.assign(
            bucket_granularity="weekly",
            bucket_date=base["sold_at_dt"].dt.to_period("W-MON").map(lambda p: str(p.start_time.date())),
        )
        .groupby(
            ["bucket_granularity", "bucket_date", "actual_fee_source", "source_rank"],
            dropna=False,
            as_index=False,
        )
        .agg(
            sales_count=("sale_id", "count"),
            actual_fee_total=("actual_fee", "sum"),
        )
    )
    out = pd.concat([daily, weekly], ignore_index=True)
    out = out.sort_values(
        ["bucket_granularity", "bucket_date", "source_rank", "sales_count"],
        ascending=[True, True, True, False],
    ).reset_index(drop=True)
    return out


@st.cache_data(show_spinner=False, max_entries=64)
def _build_normalized_source_weekly_coverage(trend_df: pd.DataFrame) -> pd.DataFrame:
    if (
        trend_df is None
        or trend_df.empty
        or "bucket_granularity" not in trend_df.columns
        or "bucket_date" not in trend_df.columns
        or "actual_fee_source" not in trend_df.columns
        or "sales_count" not in trend_df.columns
    ):
        return pd.DataFrame()
    weekly = trend_df[trend_df["bucket_granularity"] == "weekly"].copy()
    if weekly.empty:
        return pd.DataFrame()
    total_by_week = (
        weekly.groupby(["bucket_date"], dropna=False, as_index=False)
        .agg(total_sales_count=("sales_count", "sum"))
    )
    normalized_by_week = (
        weekly[weekly["actual_fee_source"] == "normalized_order_finance_entries_marketplace_fee_sum"]
        .groupby(["bucket_date"], dropna=False, as_index=False)
        .agg(normalized_sales_count=("sales_count", "sum"))
    )
    merged = total_by_week.merge(normalized_by_week, on="bucket_date", how="left")
    merged["normalized_sales_count"] = merged["normalized_sales_count"].fillna(0).astype(float)
    merged["total_sales_count"] = merged["total_sales_count"].fillna(0).astype(float)
    merged["normalized_coverage_pct"] = merged.apply(
        lambda r: round((float(r["normalized_sales_count"]) / float(r["total_sales_count"]) * 100.0), 2)
        if float(r["total_sales_count"] or 0) > 0
        else 0.0,
        axis=1,
    )
    return merged.sort_values(["bucket_date"], ascending=[True]).reset_index(drop=True)


@st.cache_data(show_spinner=False, max_entries=64)
def _build_weekly_fee_source_count_chart_data(trend_df: pd.DataFrame) -> pd.DataFrame:
    if (
        trend_df is None
        or trend_df.empty
        or "bucket_granularity" not in trend_df.columns
        or "bucket_date" not in trend_df.columns
        or "actual_fee_source" not in trend_df.columns
        or "sales_count" not in trend_df.columns
    ):
        return pd.DataFrame()
    weekly = trend_df[trend_df["bucket_granularity"] == "weekly"].copy()
    if weekly.empty:
        return pd.DataFrame()
    source_order = [
        "normalized_order_finance_entries_marketplace_fee_sum",
        "order_fee_breakdown_total_marketplace_fee",
        "sale_fees_field",
    ]
    pivot = (
        weekly.pivot_table(
            index="bucket_date",
            columns="actual_fee_source",
            values="sales_count",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
        .rename_axis(None, axis=1)
    )
    for col in source_order:
        if col not in pivot.columns:
            pivot[col] = 0
    keep_cols = ["bucket_date"] + source_order
    pivot = pivot[keep_cols].copy()
    pivot = pivot.rename(
        columns={
            "bucket_date": "week_start",
            "normalized_order_finance_entries_marketplace_fee_sum": "normalized_source",
            "order_fee_breakdown_total_marketplace_fee": "notes_fallback",
            "sale_fees_field": "sale_field_fallback",
        }
    )
    return pivot.sort_values(["week_start"], ascending=[True]).reset_index(drop=True)


@st.cache_data(show_spinner=False, max_entries=128)
def _filter_tax_drilldown_rows(
    tax_detail_df: pd.DataFrame,
    *,
    marketplace: str,
    taxability: str,
) -> pd.DataFrame:
    if tax_detail_df is None or tax_detail_df.empty:
        return pd.DataFrame()
    filtered = tax_detail_df.copy()
    marketplace_norm = str(marketplace or "all").strip().lower()
    taxability_norm = str(taxability or "all").strip().lower()
    if marketplace_norm != "all":
        filtered = filtered[
            filtered["marketplace"].astype(str).str.strip().str.lower() == marketplace_norm
        ]
    if taxability_norm == "taxable_only":
        filtered = filtered[filtered["taxable_subtotal"].astype(float) > 0.0]
    elif taxability_norm == "exempt_only":
        filtered = filtered[filtered["is_tax_exempt_category"].astype(bool)]
    return filtered


@st.cache_data(show_spinner=False, max_entries=128)
def _build_tax_drilldown_sale_option_rows(filtered_tax_detail: pd.DataFrame) -> list[dict]:
    if filtered_tax_detail is None or filtered_tax_detail.empty:
        return []
    rows: list[dict] = []
    for row in filtered_tax_detail.itertuples(index=False):
        sale_id = int(getattr(row, "sale_id", 0) or 0)
        if sale_id <= 0:
            continue
        label = (
            f"sale#{sale_id} | {str(getattr(row, 'marketplace', '') or '').strip()} | "
            f"{str(getattr(row, 'sku', '') or '').strip()} | "
            f"tax={float(getattr(row, 'estimated_tax_collected', 0.0) or 0.0):,.2f}"
        )
        rows.append({"label": label, "sale_id": sale_id})
    return rows


@st.cache_data(show_spinner=False, max_entries=128)
def _tax_drilldown_kpis(filtered_tax_detail: pd.DataFrame) -> dict[str, float | int]:
    if filtered_tax_detail is None or filtered_tax_detail.empty:
        return {
            "rows": 0,
            "taxable_subtotal": 0.0,
            "estimated_tax": 0.0,
        }
    return {
        "rows": int(len(filtered_tax_detail)),
        "taxable_subtotal": float(filtered_tax_detail["taxable_subtotal"].sum()),
        "estimated_tax": float(filtered_tax_detail["estimated_tax_collected"].sum()),
    }


@st.cache_data(show_spinner=False, max_entries=8)
def _load_colorado_suts_jurisdiction_options(
    template_path: str = str(COLORADO_SUTS_TEMPLATE_PATH),
) -> list[dict[str, object]]:
    path = Path(template_path)
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=False)
    if "Upload Data" not in wb.sheetnames:
        return []
    ws = wb["Upload Data"]
    rows: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        account_type = str(ws.cell(row_idx, 2).value or "").strip()
        jurisdiction_code = str(ws.cell(row_idx, 3).value or "").strip()
        jurisdiction_name = str(ws.cell(row_idx, 4).value or "").strip()
        if not jurisdiction_code or not jurisdiction_name:
            continue
        rows.append(
            {
                "row": row_idx,
                "account_type": account_type,
                "jurisdiction_code": jurisdiction_code,
                "jurisdiction_name": jurisdiction_name,
                "label": f"{jurisdiction_code} | {jurisdiction_name} | {account_type}",
            }
        )
    return rows


def _format_suts_gross_text(value: object) -> str:
    amount = round(_safe_float(value), 2)
    if abs(amount) < 0.005:
        return "0"
    return f"{amount:.2f}"


def _month_bounds_from_yyyy_mm(month_value: str) -> tuple[datetime, datetime]:
    raw = str(month_value or "").strip()
    try:
        month_start = datetime.strptime(raw, "%Y-%m")
    except Exception as exc:
        raise ValueError("Month must be in YYYY-MM format.") from exc
    if month_start.month == 12:
        next_month = datetime(month_start.year + 1, 1, 1)
    else:
        next_month = datetime(month_start.year, month_start.month + 1, 1)
    return month_start, next_month - timedelta(microseconds=1)


def _filter_tax_detail_for_month(tax_detail_df: pd.DataFrame, *, month_value: str) -> pd.DataFrame:
    if tax_detail_df is None or tax_detail_df.empty:
        return pd.DataFrame()
    month_start, month_end = _month_bounds_from_yyyy_mm(month_value)
    if "sold_at" not in tax_detail_df.columns:
        return tax_detail_df.copy()
    filtered = tax_detail_df.copy()
    sold_at = pd.to_datetime(filtered["sold_at"], errors="coerce")
    filtered = filtered[(sold_at >= month_start) & (sold_at <= month_end)].copy()
    return filtered


def _build_colorado_suts_scope_summary_rows(
    reportable_tax_detail_df: pd.DataFrame,
    excluded_facilitator_tax_detail_df: pd.DataFrame,
    *,
    selected_marketplaces: set[str] | None,
    facilitator_channels: set[str] | None,
) -> list[dict[str, object]]:
    selected = {str(v or "").strip().lower() for v in (selected_marketplaces or set()) if str(v or "").strip()}
    facilitators = {str(v or "").strip().lower() for v in (facilitator_channels or set()) if str(v or "").strip()}

    def _summary_row(
        df: pd.DataFrame,
        *,
        scope: str,
        marketplace_scope: set[str],
        suts_treatment: str,
        note: str,
    ) -> dict[str, object]:
        if df is None or df.empty:
            sales_count = 0
            gross_sales = 0.0
            marketplaces_seen: list[str] = []
        else:
            sales_count = int(len(df))
            gross_sales = round(float(df.get("gross_sales", pd.Series(dtype=float)).fillna(0).astype(float).sum()), 2)
            marketplaces_seen = sorted(
                {
                    str(v or "").strip().lower()
                    for v in df.get("marketplace", pd.Series(dtype=str)).dropna().tolist()
                    if str(v or "").strip()
                }
            )
        return {
            "scope": scope,
            "marketplace_scope": ", ".join(sorted(marketplace_scope)) if marketplace_scope else "(none)",
            "marketplaces_seen": ", ".join(marketplaces_seen) if marketplaces_seen else "(none)",
            "sales_count": sales_count,
            "gross_sales": gross_sales,
            "suts_treatment": suts_treatment,
            "note": note,
        }

    excluded_facilitators = facilitators - selected
    rows = [
        _summary_row(
            reportable_tax_detail_df,
            scope="Reportable SUTS upload gross",
            marketplace_scope=selected,
            suts_treatment="included",
            note="Direct/local marketplace scope selected above; this is what the SUTS gross-sales rows use.",
        )
    ]
    if excluded_facilitators or (excluded_facilitator_tax_detail_df is not None and not excluded_facilitator_tax_detail_df.empty):
        rows.append(
            _summary_row(
                excluded_facilitator_tax_detail_df,
                scope="Marketplace facilitator gross",
                marketplace_scope=excluded_facilitators,
                suts_treatment="excluded by default",
                note="eBay/facilitator sales are kept out of normal SUTS remittance unless advisor-confirmed.",
            )
        )
    if selected.intersection(facilitators):
        rows.append(
            _summary_row(
                pd.DataFrame(),
                scope="Advisor override",
                marketplace_scope=selected.intersection(facilitators),
                suts_treatment="facilitator selected",
                note="A marketplace facilitator channel is selected; only submit that scope with advisor-confirmed reporting instructions.",
            )
        )
    return rows


def _suts_jurisdiction_key(jurisdiction_code: object, account_type: object) -> str:
    return (
        f"{str(jurisdiction_code or '').strip()}|"
        f"{str(account_type or '').strip().upper()}"
    )


def _build_colorado_suts_upload_workbook(
    tax_detail_df: pd.DataFrame,
    *,
    account_number: str = COLORADO_SUTS_ACCOUNT_NUMBER,
    gross_jurisdiction_code: str = "",
    gross_jurisdiction_key: str = "",
    gross_jurisdiction_codes: list[str] | None = None,
    gross_jurisdiction_keys: list[str] | None = None,
    zero_filing_jurisdiction_codes: list[str] | None = None,
    zero_filing_jurisdiction_keys: list[str] | None = None,
    account_number_by_jurisdiction_code: dict[str, str] | None = None,
    account_number_by_jurisdiction_key: dict[str, str] | None = None,
    allow_blank_account_jurisdiction_keys: set[str] | None = None,
    custom_jurisdictions: list[dict[str, object]] | None = None,
    remove_unselected_template_rows: bool = True,
    template_path: str = str(COLORADO_SUTS_TEMPLATE_PATH),
) -> tuple[bytes, pd.DataFrame]:
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Colorado SUTS template not found: {path}")
    wb = load_workbook(path, data_only=False)
    if "Upload Data" not in wb.sheetnames:
        raise ValueError("Colorado SUTS template must contain an `Upload Data` sheet.")
    ws = wb["Upload Data"]
    ws["A1"].value = None

    account_digits = "".join(ch for ch in str(account_number or "").strip() if ch.isdigit())
    if not account_digits:
        raise ValueError("Colorado SUTS account number is required.")
    gross_code = str(gross_jurisdiction_code or "").strip()
    gross_key = str(gross_jurisdiction_key or "").strip()
    gross_codes = {str(code).strip() for code in (gross_jurisdiction_codes or []) if str(code).strip()}
    gross_keys = {str(key).strip() for key in (gross_jurisdiction_keys or []) if str(key).strip()}
    if gross_code:
        gross_codes.add(gross_code)
    if gross_key:
        gross_keys.add(gross_key)
    zero_codes = {str(code).strip() for code in (zero_filing_jurisdiction_codes or []) if str(code).strip()}
    zero_keys = {str(key).strip() for key in (zero_filing_jurisdiction_keys or []) if str(key).strip()}
    zero_codes -= gross_codes
    zero_keys -= gross_keys
    account_override_digits_by_code = {
        str(code or "").strip(): "".join(ch for ch in str(value or "").strip() if ch.isdigit())
        for code, value in (account_number_by_jurisdiction_code or {}).items()
        if str(code or "").strip()
    }
    account_override_digits_by_key = {
        str(key or "").strip(): "".join(ch for ch in str(value or "").strip() if ch.isdigit())
        for key, value in (account_number_by_jurisdiction_key or {}).items()
        if str(key or "").strip()
    }
    allow_blank_keys = {str(key or "").strip() for key in (allow_blank_account_jurisdiction_keys or set()) if str(key or "").strip()}

    gross_sales_total = 0.0
    if tax_detail_df is not None and not tax_detail_df.empty and "gross_sales" in tax_detail_df.columns:
        gross_sales_total = float(tax_detail_df["gross_sales"].fillna(0).astype(float).sum())

    selected_codes = {*gross_codes, *zero_codes}
    selected_codes.discard("")
    selected_keys = {*gross_keys, *zero_keys}
    selected_keys.discard("")
    existing_keys = {
        _suts_jurisdiction_key(ws.cell(row_idx, 3).value, ws.cell(row_idx, 2).value)
        for row_idx in range(4, ws.max_row + 1)
    }
    for custom in custom_jurisdictions or []:
        code = str(custom.get("jurisdiction_code") or "").strip()
        account_type = str(custom.get("account_type") or "LOCAL").strip().upper() or "LOCAL"
        key = _suts_jurisdiction_key(code, account_type)
        if not code or key in existing_keys or (code not in selected_codes and key not in selected_keys):
            continue
        row_idx = ws.max_row + 1
        ws.cell(row_idx, 2).value = account_type
        ws.cell(row_idx, 3).value = int(code) if code.isdigit() else code
        ws.cell(row_idx, 4).value = str(custom.get("jurisdiction_name") or "Custom Jurisdiction").strip()
        existing_keys.add(key)

    summary_rows: list[dict[str, object]] = []
    for row_idx in range(4, ws.max_row + 1):
        jurisdiction_code = str(ws.cell(row_idx, 3).value or "").strip()
        account_type = str(ws.cell(row_idx, 2).value or "").strip().upper()
        jurisdiction_key = _suts_jurisdiction_key(jurisdiction_code, account_type)
        if (
            not jurisdiction_code
            or (jurisdiction_code not in selected_codes and jurisdiction_key not in selected_keys)
        ):
            continue
        if jurisdiction_key in account_override_digits_by_key:
            row_account_digits = account_override_digits_by_key.get(jurisdiction_key) or ""
        elif jurisdiction_code in account_override_digits_by_code:
            row_account_digits = account_override_digits_by_code.get(jurisdiction_code) or ""
        elif jurisdiction_key in allow_blank_keys:
            row_account_digits = ""
        else:
            row_account_digits = account_digits
        if not row_account_digits and jurisdiction_key not in allow_blank_keys:
            raise ValueError(f"Account number is required for Colorado SUTS jurisdiction {jurisdiction_code}.")
        account_cell = ws.cell(row_idx, 1)
        if row_account_digits:
            account_cell.value = int(row_account_digits)
            account_cell.number_format = "0" * max(1, len(row_account_digits))
        else:
            account_cell.value = None
            account_cell.number_format = "General"

        is_gross_row = (
            jurisdiction_key in gross_keys
            if gross_keys
            else jurisdiction_code in gross_codes
        )
        gross_text = _format_suts_gross_text(gross_sales_total if is_gross_row else 0.0)
        gross_cell = ws.cell(row_idx, 5)
        gross_cell.value = gross_text
        gross_cell.number_format = "@"
        summary_rows.append(
            {
                "account_number": row_account_digits,
                "account_type": account_type,
                "jurisdiction_code": jurisdiction_code,
                "jurisdiction_key": jurisdiction_key,
                "jurisdiction_name": str(ws.cell(row_idx, 4).value or "").strip(),
                "gross_amount": gross_text,
                "filing_type": "gross_sales" if is_gross_row else "zero_filing",
                "worksheet_row": row_idx,
            }
        )

    if (selected_codes or selected_keys) and not summary_rows:
        raise ValueError("No selected Colorado SUTS jurisdiction codes were found in the template.")
    if remove_unselected_template_rows:
        selected_row_indexes = {
            int(row["worksheet_row"])
            for row in summary_rows
            if int(row.get("worksheet_row") or 0) >= 4
        }
        for row_idx in range(ws.max_row, 3, -1):
            if row_idx not in selected_row_indexes:
                ws.delete_rows(row_idx)

    output = BytesIO()
    wb.save(output)
    return output.getvalue(), pd.DataFrame(summary_rows)


def _colorado_suts_summary_warnings(summary_df: pd.DataFrame) -> list[str]:
    if summary_df is None or summary_df.empty:
        return ["No Colorado SUTS rows are selected for export."]
    warnings: list[str] = []
    rows = summary_df.to_dict(orient="records")
    gross_rows = [row for row in rows if str(row.get("filing_type") or "") == "gross_sales"]
    zero_rows = [row for row in rows if str(row.get("filing_type") or "") == "zero_filing"]
    if not gross_rows and not zero_rows:
        warnings.append("Selected SUTS rows do not contain a gross-sales row or zero-filing row.")
    for row in rows:
        account_type = str(row.get("account_type") or "").strip().upper()
        jurisdiction_code = str(row.get("jurisdiction_code") or "").strip()
        account_number = str(row.get("account_number") or "").strip()
        filing_type = str(row.get("filing_type") or "").strip()
        if account_type == "STATE" and jurisdiction_code == "110042" and not account_number:
            warnings.append("Golden STATE row is missing SUTS account 970074130001.")
        if filing_type == "gross_sales" and not account_number:
            warnings.append(
                f"{str(row.get('jurisdiction_name') or jurisdiction_code).strip()} "
                f"{account_type} gross-sales row is missing an account number."
            )
        if (
            account_type == "LOCAL"
            and jurisdiction_code == "110042"
            and filing_type == "gross_sales"
            and not account_number
        ):
            warnings.append(
                "Golden LOCAL gross-sales row has a blank account number. SUTS accepted blank account only "
                "for zero filing in observed testing; confirm whether nonzero self-collected Golden filing "
                "requires a local account before submitting."
            )
    duplicate_codes = {
        str(code)
        for code, count in summary_df["jurisdiction_code"].astype(str).value_counts().items()
        if int(count) > 1
    } if "jurisdiction_code" in summary_df.columns else set()
    for code in sorted(duplicate_codes):
        account_types = sorted(
            {
                str(row.get("account_type") or "").strip().upper()
                for row in rows
                if str(row.get("jurisdiction_code") or "").strip() == code
            }
        )
        if len(account_types) != len(
            [
                row
                for row in rows
                if str(row.get("jurisdiction_code") or "").strip() == code
            ]
        ):
            warnings.append(f"Jurisdiction {code} has duplicate rows with the same account type.")
    return warnings


def _build_tax_exception_rows(
    tax_detail_df: pd.DataFrame,
    *,
    tax_jurisdiction: str,
    tax_rate_percent: float,
    shipping_taxable: bool,
    facilitator_channels: set[str] | None,
    tax_exempt_categories: set[str] | None,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    jurisdiction = str(tax_jurisdiction or "").strip()
    facilitator_set = {
        str(v).strip().lower()
        for v in (facilitator_channels or set())
        if str(v).strip()
    }
    exempt_set = {
        str(v).strip().lower()
        for v in (tax_exempt_categories or set())
        if str(v).strip()
    }
    tax_rate = _safe_float(tax_rate_percent)

    def _append(
        *,
        exception_type: str,
        severity: str,
        message: str,
        sale_id: int | None = None,
        marketplace: str = "",
        category: str = "",
        evidence: str = "",
    ) -> None:
        rows.append(
            {
                "severity": severity,
                "exception_type": exception_type,
                "sale_id": int(sale_id or 0) if sale_id else None,
                "marketplace": marketplace,
                "category": category,
                "message": message,
                "evidence": evidence,
            }
        )

    if not jurisdiction:
        _append(
            exception_type="missing_tax_jurisdiction",
            severity="P1",
            message="Tax report has no jurisdiction configured for the selected scope.",
            evidence="tax_jurisdiction is blank",
        )
    if tax_rate <= 0 and tax_detail_df is not None and not tax_detail_df.empty:
        taxable_total = float(tax_detail_df.get("taxable_subtotal", pd.Series(dtype=float)).fillna(0).astype(float).sum())
        if taxable_total > 0:
            _append(
                exception_type="missing_or_zero_tax_rate",
                severity="P1",
                message="Taxable sales exist but the configured tax rate is zero or missing.",
                evidence=f"taxable_subtotal={taxable_total:.2f}; tax_rate_percent={tax_rate:.4f}",
            )
    if tax_detail_df is None or tax_detail_df.empty:
        return rows

    for raw in tax_detail_df.to_dict(orient="records"):
        sale_id = int(raw.get("sale_id") or 0)
        marketplace = str(raw.get("marketplace") or "").strip().lower()
        category = str(raw.get("category") or "").strip().lower()
        gross_sales = _safe_float(raw.get("gross_sales"))
        shipping_cost = _safe_float(raw.get("shipping_cost"))
        taxable_subtotal = _safe_float(raw.get("taxable_subtotal"))
        taxable_shipping = _safe_float(raw.get("taxable_shipping_subtotal"))
        estimated_tax = _safe_float(raw.get("estimated_tax_collected"))
        is_exempt = bool(raw.get("is_tax_exempt_category"))

        if marketplace in facilitator_set and taxable_subtotal > 0:
            _append(
                exception_type="facilitator_channel_in_tax_scope",
                severity="P2",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Marketplace facilitator channel is included in local tax liability scope.",
                evidence=f"taxable_subtotal={taxable_subtotal:.2f}; estimated_tax={estimated_tax:.2f}",
            )
        if gross_sales > 0 and not category:
            _append(
                exception_type="missing_tax_category",
                severity="P2",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Sale has no product category for taxable/exempt classification.",
                evidence=f"gross_sales={gross_sales:.2f}",
            )
        if is_exempt:
            _append(
                exception_type="exempt_category_review_needed",
                severity="P3",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Sale uses configured tax-exempt category treatment; confirm exemption basis with advisor.",
                evidence=f"category={category}; configured_exempt_categories={','.join(sorted(exempt_set))}",
            )
        if is_exempt and bool(shipping_taxable) and shipping_cost > 0 and taxable_shipping > 0:
            _append(
                exception_type="exempt_item_taxable_shipping_review_needed",
                severity="P2",
                sale_id=sale_id,
                marketplace=marketplace,
                category=category,
                message="Exempt item has taxable shipping under current report settings.",
                evidence=f"shipping_cost={shipping_cost:.2f}; taxable_shipping={taxable_shipping:.2f}",
            )
    return rows


@st.cache_data(show_spinner=False, max_entries=128)
def _build_documents_handoff_sale_option_rows(sales_df: pd.DataFrame) -> list[dict]:
    if sales_df is None or sales_df.empty:
        return []
    rows: list[dict] = []
    for row in sales_df.itertuples(index=False):
        sale_id = int(getattr(row, "sale_id", 0) or 0)
        if sale_id <= 0:
            continue
        label = (
            f"sale#{sale_id} | {str(getattr(row, 'sold_at', '') or '')} | "
            f"{str(getattr(row, 'marketplace', '') or '').strip()} | "
            f"{str(getattr(row, 'sku', '') or '').strip()} | "
            f"gross=${float(getattr(row, 'gross_sales', 0.0) or 0.0):,.2f}"
        )
        rows.append({"label": label, "source_id": sale_id})
    return rows


@st.cache_data(show_spinner=False, max_entries=128)
def _build_documents_handoff_order_option_rows(orders_df: pd.DataFrame) -> list[dict]:
    if orders_df is None or orders_df.empty:
        return []
    rows: list[dict] = []
    for row in orders_df.itertuples(index=False):
        order_id = int(getattr(row, "order_id", 0) or 0)
        if order_id <= 0:
            continue
        label = (
            f"order#{order_id} | {str(getattr(row, 'sold_at', '') or '')} | "
            f"{str(getattr(row, 'marketplace', '') or '').strip()} | "
            f"ext={str(getattr(row, 'external_order_id', '') or '').strip()} | "
            f"total=${float(getattr(row, 'total_amount', 0.0) or 0.0):,.2f}"
        )
        rows.append({"label": label, "source_id": order_id})
    return rows


def _tax_report_presets(
    *,
    default_jurisdiction: str,
    default_tax_rate_percent: float,
    default_shipping_taxable: bool,
) -> dict[str, dict]:
    return {
        "Golden Local Retail": {
            "jurisdiction": default_jurisdiction or "Golden, Colorado",
            "tax_rate_percent": float(default_tax_rate_percent),
            "shipping_taxable": bool(default_shipping_taxable),
            "marketplace_mode": "local_only",
        },
        "Marketplace Shipped": {
            "jurisdiction": default_jurisdiction or "Golden, Colorado",
            "tax_rate_percent": float(default_tax_rate_percent),
            "shipping_taxable": False,
            "marketplace_mode": "all",
        },
        "Bullion Exempt Focus": {
            "jurisdiction": default_jurisdiction or "Golden, Colorado",
            "tax_rate_percent": float(default_tax_rate_percent),
            "shipping_taxable": False,
            "marketplace_mode": "all",
        },
    }


def _build_fifo_unit_cost_map(
    all_sales,
    all_assignments,
    default_unit_cost_by_product: dict[int, float],
) -> dict[int, float]:
    fifo_unit_cost_by_sale, _remaining = _build_fifo_cost_maps(
        all_sales,
        all_assignments,
        default_unit_cost_by_product,
    )
    return fifo_unit_cost_by_sale


def _build_fifo_remaining_unit_cost_map(
    all_sales,
    all_assignments,
    default_unit_cost_by_product: dict[int, float],
) -> dict[int, float]:
    _sale_costs, remaining_unit_cost_by_product = _build_fifo_cost_maps(
        all_sales,
        all_assignments,
        default_unit_cost_by_product,
    )
    return remaining_unit_cost_by_product


def _build_fifo_cost_maps(
    all_sales,
    all_assignments,
    default_unit_cost_by_product: dict[int, float],
) -> tuple[dict[int, float], dict[int, float]]:
    lots_by_product: dict[int, list[dict]] = defaultdict(list)
    lot_fallback_unit_costs, assignment_fallback_unit_costs = _lot_fallback_unit_cost_maps(all_assignments)
    for a in sorted(all_assignments, key=lambda x: (x.acquired_at or datetime.min, x.id)):
        product_id = _safe_product_id(a)
        if product_id is None:
            continue
        unit_cost = _landed_unit_cost_from_assignment(
            a,
            lot_fallback_unit_costs=lot_fallback_unit_costs,
            assignment_fallback_unit_costs=assignment_fallback_unit_costs,
        )
        lots_by_product[int(product_id)].append(
            {
                "remaining_qty": max(0, int(a.quantity_acquired or 0)),
                "unit_cost": unit_cost,
                "acquired_at": a.acquired_at or datetime.min,
                "assignment_id": int(getattr(a, "id", 0) or 0),
            }
        )

    lot_sources: dict[int, list[dict]] = {
        product_id: sorted(
            lots,
            key=lambda row: (row.get("acquired_at") or datetime.min, int(row.get("assignment_id") or 0)),
        )
        for product_id, lots in lots_by_product.items()
    }
    source_index_by_product = {product_id: 0 for product_id in lot_sources}
    queues: dict[int, deque] = {
        product_id: deque() for product_id in set(lot_sources) | set(default_unit_cost_by_product)
    }

    def _queue_available_lots(product_id: int, cutoff: datetime) -> None:
        lots = lot_sources.get(product_id) or []
        idx = int(source_index_by_product.get(product_id, 0))
        queue = queues.setdefault(product_id, deque())
        while idx < len(lots):
            acquired_at = lots[idx].get("acquired_at") or datetime.min
            if acquired_at > cutoff:
                break
            queue.append(dict(lots[idx]))
            idx += 1
        source_index_by_product[product_id] = idx

    fifo_unit_cost_by_sale: dict[int, float] = {}
    sales_sorted = sorted(all_sales, key=lambda s: (s.sold_at or datetime.min, s.id))
    for sale in sales_sorted:
        product_id = _safe_product_id(sale)
        qty = max(1, int(sale.quantity_sold or 1))
        if product_id is None:
            fifo_unit_cost_by_sale[sale.id] = 0.0
            continue

        _queue_available_lots(product_id, sale.sold_at or datetime.min)
        queue = queues.setdefault(product_id, deque())
        default_cost = max(0.0, _safe_float(default_unit_cost_by_product.get(product_id)))

        qty_remaining = qty
        total_cost = 0.0
        while qty_remaining > 0:
            if queue and int(queue[0]["remaining_qty"]) > 0:
                use_qty = min(qty_remaining, int(queue[0]["remaining_qty"]))
                total_cost += float(use_qty) * _safe_float(queue[0]["unit_cost"])
                queue[0]["remaining_qty"] = int(queue[0]["remaining_qty"]) - use_qty
                qty_remaining -= use_qty
                if int(queue[0]["remaining_qty"]) <= 0:
                    queue.popleft()
            else:
                total_cost += float(qty_remaining) * default_cost
                qty_remaining = 0

        fifo_unit_cost_by_sale[sale.id] = (total_cost / float(qty)) if qty > 0 else 0.0

    max_sale_dt = max(
        [s.sold_at for s in sales_sorted if getattr(s, "sold_at", None) is not None],
        default=datetime.max,
    )
    for product_id in lot_sources:
        _queue_available_lots(product_id, max_sale_dt)

    remaining_unit_cost_by_product: dict[int, float] = {}
    for product_id, queue in queues.items():
        remaining_qty = 0.0
        remaining_cost = 0.0
        for row in list(queue):
            qty = float(max(0, int(row.get("remaining_qty") or 0)))
            if qty <= 0:
                continue
            remaining_qty += qty
            remaining_cost += qty * _safe_float(row.get("unit_cost"))
        if remaining_qty > 0:
            remaining_unit_cost_by_product[int(product_id)] = remaining_cost / remaining_qty
    for product_id, default_cost in default_unit_cost_by_product.items():
        remaining_unit_cost_by_product.setdefault(product_id, max(0.0, _safe_float(default_cost)))
    return fifo_unit_cost_by_sale, remaining_unit_cost_by_product


def _build_lot_weighted_unit_cost_map(
    all_assignments,
    default_unit_cost_by_product: dict[int, float],
) -> dict[int, float]:
    totals: dict[int, dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "cost": 0.0})
    lot_fallback_unit_costs, assignment_fallback_unit_costs = _lot_fallback_unit_cost_maps(all_assignments)
    for a in all_assignments:
        pid = _safe_product_id(a)
        if pid is None:
            continue
        qty = float(max(0, int(a.quantity_acquired or 0)))
        if qty <= 0:
            continue
        unit_cost = _landed_unit_cost_from_assignment(
            a,
            lot_fallback_unit_costs=lot_fallback_unit_costs,
            assignment_fallback_unit_costs=assignment_fallback_unit_costs,
        )
        totals[pid]["qty"] += qty
        totals[pid]["cost"] += unit_cost * qty

    result: dict[int, float] = {}
    for pid, agg in totals.items():
        if agg["qty"] > 0:
            result[pid] = agg["cost"] / agg["qty"]
    for pid, default_cost in default_unit_cost_by_product.items():
        result.setdefault(pid, max(0.0, _safe_float(default_cost)))
    return result


def _build_inventory_cycle_rows(
    products,
    movements,
    sales,
    actual_econ_by_sale_id: dict[int, dict[str, float | str]] | None = None,
) -> list[dict]:
    actual_by_sale = actual_econ_by_sale_id or {}
    product_by_id = {int(p.id): p for p in products if p is not None and p.id is not None}
    movements_by_product: dict[int, list] = defaultdict(list)
    for m in movements:
        product_id = _safe_product_id(m)
        if product_id is None:
            continue
        movements_by_product[int(product_id)].append(m)
    sales_by_product: dict[int, list] = defaultdict(list)
    for s in sales:
        product_id = _safe_product_id(s)
        if product_id is None:
            continue
        sales_by_product[int(product_id)].append(s)

    rows: list[dict] = []
    for product_id, product_movements in movements_by_product.items():
        product = product_by_id.get(product_id)
        product_sales = sorted(
            sales_by_product.get(product_id, []),
            key=lambda x: (x.sold_at or datetime.min, x.id),
        )
        sales_idx = 0
        sorted_movements = sorted(product_movements, key=lambda x: (x.occurred_at or datetime.min, x.id))
        current_cycle: dict | None = None
        cycle_number = 0

        def _add_sale_to_cycle(cycle: dict, sale) -> None:
            sale_id = int(getattr(sale, "id", 0) or 0)
            actual = actual_by_sale.get(sale_id) or {}
            fee = (
                _safe_float(actual.get("allocated_fee_actual"))
                if actual
                else _safe_float(sale.fees)
            )
            shipping_charged = (
                _safe_float(actual.get("allocated_shipping_charged"))
                if actual
                else _safe_float(sale.shipping_cost)
            )
            shipping_label = (
                _safe_float(actual.get("allocated_shipping_actual"))
                if actual
                else _safe_float(getattr(sale, "shipping_label_cost", None))
            )
            net_sales = (
                _safe_float(actual.get("net_before_cogs_actual"))
                if actual
                else _sale_net_before_cogs_from_fields(sale)
            )
            cycle["sale_count"] += 1
            cycle["qty_sold_sales"] += int(sale.quantity_sold or 0)
            cycle["gross_sales"] += _safe_float(sale.sold_price)
            cycle["fees"] += fee
            cycle["shipping_cost"] += shipping_charged
            cycle["shipping_label_cost"] += shipping_label
            cycle["net_sales"] += net_sales

        for mv in sorted_movements:
            before_qty = int(mv.quantity_before or 0)
            after_qty = int(mv.quantity_after or 0)
            qty_delta = int(mv.quantity_delta or 0)
            started_new_cycle = current_cycle is None and after_qty > 0
            if started_new_cycle:
                cycle_number += 1
                current_cycle = {
                    "product_id": product_id,
                    "sku": product.sku if product else None,
                    "product_title": product.title if product else None,
                    "cycle_number": cycle_number,
                    "cycle_id": f"{product.sku or product_id}-C{cycle_number}",
                    "cycle_start": mv.occurred_at,
                    "cycle_end": None,
                    "cycle_status": "open",
                    "start_qty_before": before_qty,
                    "end_qty_after": after_qty,
                    "qty_in": 0,
                    "qty_out_movements": 0,
                    "acquisition_cost_known": 0.0,
                    "movement_count": 0,
                    "sale_count": 0,
                    "qty_sold_sales": 0,
                    "gross_sales": 0.0,
                    "fees": 0.0,
                    "shipping_cost": 0.0,
                    "shipping_label_cost": 0.0,
                    "net_sales": 0.0,
                }
            if current_cycle is None:
                continue

            current_cycle["movement_count"] += 1
            current_cycle["end_qty_after"] = after_qty
            if qty_delta > 0:
                current_cycle["qty_in"] += qty_delta
                if mv.unit_cost is not None:
                    current_cycle["acquisition_cost_known"] += _safe_float(mv.unit_cost) * float(qty_delta)
            elif qty_delta < 0:
                current_cycle["qty_out_movements"] += abs(qty_delta)

            cycle_start = current_cycle["cycle_start"] or datetime.min
            cycle_end_candidate = mv.occurred_at or datetime.min
            while sales_idx < len(product_sales):
                sale = product_sales[sales_idx]
                sold_at = sale.sold_at or datetime.min
                if sold_at < cycle_start:
                    sales_idx += 1
                    continue
                if sold_at > cycle_end_candidate:
                    break
                _add_sale_to_cycle(current_cycle, sale)
                sales_idx += 1

            if after_qty <= 0:
                current_cycle["cycle_end"] = mv.occurred_at
                current_cycle["cycle_status"] = "closed"
                known_cost = _safe_float(current_cycle["acquisition_cost_known"])
                current_cycle["estimated_margin_vs_known_cost"] = (
                    _safe_float(current_cycle["net_sales"]) - known_cost
                )
                rows.append(current_cycle)
                current_cycle = None

        if current_cycle is not None:
            while sales_idx < len(product_sales):
                sale = product_sales[sales_idx]
                sold_at = sale.sold_at or datetime.min
                if sold_at < (current_cycle["cycle_start"] or datetime.min):
                    sales_idx += 1
                    continue
                _add_sale_to_cycle(current_cycle, sale)
                sales_idx += 1
            current_cycle["cycle_status"] = "open"
            known_cost = _safe_float(current_cycle["acquisition_cost_known"])
            current_cycle["estimated_margin_vs_known_cost"] = (
                _safe_float(current_cycle["net_sales"]) - known_cost
            )
            rows.append(current_cycle)

    output = []
    for row in rows:
        output.append(
            {
                "product_id": row["product_id"],
                "sku": row["sku"],
                "product_title": row["product_title"],
                "cycle_number": row["cycle_number"],
                "cycle_id": row["cycle_id"],
                "cycle_status": row["cycle_status"],
                "cycle_start": iso_or_none(row["cycle_start"]),
                "cycle_end": iso_or_none(row["cycle_end"]),
                "start_qty_before": int(row["start_qty_before"]),
                "end_qty_after": int(row["end_qty_after"]),
                "qty_in": int(row["qty_in"]),
                "qty_out_movements": int(row["qty_out_movements"]),
                "qty_sold_sales": int(row["qty_sold_sales"]),
                "movement_count": int(row["movement_count"]),
                "sale_count": int(row["sale_count"]),
                "acquisition_cost_known": round(_safe_float(row["acquisition_cost_known"]), 2),
                "gross_sales": round(_safe_float(row["gross_sales"]), 2),
                "fees": round(_safe_float(row["fees"]), 2),
                "shipping_cost": round(_safe_float(row["shipping_cost"]), 2),
                "shipping_label_cost": round(_safe_float(row.get("shipping_label_cost")), 2),
                "net_sales": round(_safe_float(row["net_sales"]), 2),
                "estimated_margin_vs_known_cost": round(
                    _safe_float(row["estimated_margin_vs_known_cost"]),
                    2,
                ),
            }
        )
    return sorted(
        output,
        key=lambda x: (x.get("sku") or "", x.get("cycle_number") or 0),
    )


def _build_inventory_cycle_summary_rows(cycle_rows: list[dict]) -> list[dict]:
    if not cycle_rows:
        return []

    grouped: dict[tuple[int, str, str], dict] = {}
    for row in cycle_rows:
        product_id = int(row.get("product_id") or 0)
        sku = str(row.get("sku") or "").strip()
        product_title = str(row.get("product_title") or "").strip()
        key = (product_id, sku, product_title)
        if key not in grouped:
            grouped[key] = {
                "product_id": product_id,
                "sku": sku,
                "product_title": product_title,
                "cycle_count": 0,
                "open_cycle_count": 0,
                "closed_cycle_count": 0,
                "qty_in_total": 0,
                "qty_out_total": 0,
                "qty_sold_total": 0,
                "sale_count_total": 0,
                "net_sales_total": 0.0,
                "acquisition_cost_known_total": 0.0,
                "estimated_margin_vs_known_cost_total": 0.0,
                "avg_closed_cycle_days": 0.0,
                "last_cycle_status": "",
                "last_cycle_end": "",
                "_closed_durations": [],
                "_last_cycle_number": 0,
            }
        agg = grouped[key]
        agg["cycle_count"] += 1
        status = str(row.get("cycle_status") or "").strip().lower()
        if status == "closed":
            agg["closed_cycle_count"] += 1
        else:
            agg["open_cycle_count"] += 1

        agg["qty_in_total"] += int(row.get("qty_in") or 0)
        agg["qty_out_total"] += int(row.get("qty_out_movements") or 0)
        agg["qty_sold_total"] += int(row.get("qty_sold_sales") or 0)
        agg["sale_count_total"] += int(row.get("sale_count") or 0)
        agg["net_sales_total"] += _safe_float(row.get("net_sales"))
        agg["acquisition_cost_known_total"] += _safe_float(row.get("acquisition_cost_known"))
        agg["estimated_margin_vs_known_cost_total"] += _safe_float(row.get("estimated_margin_vs_known_cost"))

        cycle_num = int(row.get("cycle_number") or 0)
        if cycle_num >= int(agg.get("_last_cycle_number") or 0):
            agg["_last_cycle_number"] = cycle_num
            agg["last_cycle_status"] = str(row.get("cycle_status") or "").strip()
            agg["last_cycle_end"] = str(row.get("cycle_end") or "").strip()

        if status == "closed":
            start_raw = str(row.get("cycle_start") or "").strip()
            end_raw = str(row.get("cycle_end") or "").strip()
            if start_raw and end_raw:
                try:
                    start_dt = datetime.fromisoformat(start_raw)
                    end_dt = datetime.fromisoformat(end_raw)
                    duration = (end_dt - start_dt).total_seconds() / 86400.0
                    if duration >= 0:
                        agg["_closed_durations"].append(float(duration))
                except Exception:
                    pass

    output: list[dict] = []
    for agg in grouped.values():
        durations = agg.pop("_closed_durations", [])
        agg.pop("_last_cycle_number", None)
        if durations:
            agg["avg_closed_cycle_days"] = round(float(sum(durations) / len(durations)), 2)
        else:
            agg["avg_closed_cycle_days"] = 0.0
        agg["net_sales_total"] = round(float(agg["net_sales_total"]), 2)
        agg["acquisition_cost_known_total"] = round(float(agg["acquisition_cost_known_total"]), 2)
        agg["estimated_margin_vs_known_cost_total"] = round(float(agg["estimated_margin_vs_known_cost_total"]), 2)
        output.append(agg)
    return sorted(
        output,
        key=lambda r: (
            -int(r.get("closed_cycle_count") or 0),
            str(r.get("sku") or ""),
        ),
    )


def _build_rebuy_cost_trend_rows(
    products,
    assignments,
    movements,
) -> list[dict]:
    product_by_id = {int(p.id): p for p in products if p is not None and p.id is not None}
    assignment_keys = set()
    acquisition_events: dict[int, list[dict]] = defaultdict(list)

    for a in assignments:
        pid = _safe_product_id(a)
        if pid is None:
            continue
        qty = max(0, int(a.quantity_acquired or 0))
        unit_cost = _safe_float(a.unit_cost)
        if qty <= 0 or unit_cost <= 0:
            continue
        ts = a.acquired_at or datetime.min
        key = (pid, ts, qty, round(unit_cost, 6))
        assignment_keys.add(key)
        acquisition_events[pid].append(
            {
                "occurred_at": ts,
                "event_type": "lot_assignment",
                "qty_in": qty,
                "unit_cost": unit_cost,
                "source_ref": f"assignment:{a.id}",
            }
        )

    for m in movements:
        pid = _safe_product_id(m)
        if pid is None:
            continue
        mv_type = (m.movement_type or "").strip().lower()
        if mv_type not in {"initial_stock", "repurchase_in"}:
            continue
        qty = max(0, int(m.quantity_delta or 0))
        unit_cost = _safe_float(m.unit_cost)
        if qty <= 0 or unit_cost <= 0:
            continue
        ts = m.occurred_at or datetime.min
        key = (pid, ts, qty, round(unit_cost, 6))
        if key in assignment_keys:
            continue
        acquisition_events[pid].append(
            {
                "occurred_at": ts,
                "event_type": mv_type,
                "qty_in": qty,
                "unit_cost": unit_cost,
                "source_ref": f"movement:{m.id}",
            }
        )

    rows: list[dict] = []
    for pid, events in acquisition_events.items():
        product = product_by_id.get(pid)
        cumulative_qty = 0.0
        cumulative_cost = 0.0
        for idx, event in enumerate(
            sorted(events, key=lambda x: (x["occurred_at"], x["event_type"], x["source_ref"])),
            start=1,
        ):
            qty = float(event["qty_in"])
            unit_cost = _safe_float(event["unit_cost"])
            cumulative_qty += qty
            cumulative_cost += qty * unit_cost
            weighted_unit_cost = (cumulative_cost / cumulative_qty) if cumulative_qty > 0 else 0.0
            rows.append(
                {
                    "product_id": pid,
                    "sku": product.sku if product else None,
                    "product_title": product.title if product else None,
                    "event_index": idx,
                    "as_of": iso_or_none(event["occurred_at"]),
                    "event_type": event["event_type"],
                    "qty_in": int(qty),
                    "unit_cost": round(unit_cost, 4),
                    "acquisition_value": round(qty * unit_cost, 2),
                    "cumulative_qty_acquired": round(cumulative_qty, 2),
                    "cumulative_acquisition_cost": round(cumulative_cost, 2),
                    "weighted_unit_cost": round(weighted_unit_cost, 4),
                    "source_ref": event["source_ref"],
                }
            )

    return sorted(rows, key=lambda x: (x.get("sku") or "", x.get("event_index") or 0))


def _build_listing_review_activity_rows(
    listings,
    *,
    start_dt: datetime,
    end_dt: datetime,
) -> list[dict]:
    rows: list[dict] = []
    for listing in listings:
        marketplace = (listing.marketplace or "").strip().lower()
        sku = listing.product.sku if listing.product else None
        title = listing.listing_title
        payload_raw = (listing.marketplace_details or "").strip()
        if not payload_raw:
            continue
        try:
            payload = json.loads(payload_raw)
        except Exception:
            continue
        history = payload.get("review_history")
        if not isinstance(history, list):
            continue
        for event in history:
            if not isinstance(event, dict):
                continue
            reviewed_at_raw = str(event.get("reviewed_at") or "").strip()
            if not reviewed_at_raw:
                continue
            reviewed_at_dt: datetime | None = None
            try:
                reviewed_at_dt = datetime.fromisoformat(reviewed_at_raw.replace("Z", "+00:00")).replace(tzinfo=None)
            except Exception:
                reviewed_at_dt = None
            if reviewed_at_dt is None:
                continue
            if not (start_dt <= reviewed_at_dt <= end_dt):
                continue
            rows.append(
                {
                    "listing_id": listing.id,
                    "marketplace": marketplace,
                    "sku": sku,
                    "listing_title": title,
                    "review_decision": str(event.get("decision") or "").strip().lower(),
                    "reviewed_by": str(event.get("actor") or "").strip(),
                    "reviewed_at": reviewed_at_dt.isoformat(),
                    "review_date": reviewed_at_dt.date().isoformat(),
                    "review_notes": str(event.get("notes") or "").strip(),
                }
            )
    return sorted(rows, key=lambda x: (x.get("reviewed_at") or "", x.get("listing_id") or 0), reverse=True)


def _build_listing_format_outcome_rows(
    listings,
    *,
    start_dt: datetime,
    end_dt: datetime,
) -> list[dict]:
    rows: list[dict] = []
    for listing in listings:
        listed_at = listing.listed_at
        if listed_at is not None and not (start_dt <= listed_at <= end_dt):
            continue
        meta = {}
        details_raw = str(listing.marketplace_details or "").strip()
        if details_raw:
            try:
                parsed = json.loads(details_raw)
                if isinstance(parsed, dict):
                    publish_meta = parsed.get("ebay_publish")
                    if isinstance(publish_meta, dict):
                        meta = publish_meta
            except Exception:
                meta = {}
        intent_format = str(
            meta.get("format") or meta.get("format_type") or "FIXED_PRICE"
        ).strip().upper()
        if intent_format not in {"FIXED_PRICE", "AUCTION"}:
            intent_format = "FIXED_PRICE"
        intent_duration = str(meta.get("listing_duration") or "").strip().upper()
        publish_history = meta.get("history") if isinstance(meta.get("history"), list) else []
        publish_attempt_count = len(publish_history)
        publish_success_count = len(
            [h for h in publish_history if str((h or {}).get("status") or "").strip().lower() in {"published", "success"}]
        )
        publish_error_events = [
            h for h in publish_history if str((h or {}).get("status") or "").strip().lower() in {"error", "failed"}
        ]
        publish_error_count = len(publish_error_events)
        last_error = ""
        if publish_error_events:
            last_error = str((publish_error_events[-1] or {}).get("error") or "").strip()
        published_at = str(meta.get("published_at") or "").strip()
        external_listing_id = str(listing.external_listing_id or "").strip()
        listing_state = str(listing.listing_status or "").strip().lower()
        if external_listing_id and listing_state in {"active", "ended", "sold"}:
            publish_outcome = "published"
        elif publish_error_count > 0:
            publish_outcome = "publish_error"
        elif publish_attempt_count > 0:
            publish_outcome = "attempted_no_publish"
        else:
            publish_outcome = "not_attempted"
        rows.append(
            {
                "listing_id": int(listing.id),
                "listed_at": iso_or_none(listed_at),
                "marketplace": str(listing.marketplace or "").strip().lower(),
                "sku": listing.product.sku if listing.product else None,
                "listing_title": str(listing.listing_title or "").strip(),
                "review_status": str(listing.review_status or "").strip().lower(),
                "listing_status": listing_state,
                "intent_format": intent_format,
                "intent_duration": intent_duration,
                "intent_best_offer_enabled": bool(meta.get("best_offer_enabled")),
                "intent_auction_start_price": _safe_float(meta.get("auction_start_price")),
                "intent_auction_reserve_price": _safe_float(meta.get("auction_reserve_price")),
                "intent_auction_buy_now_price": _safe_float(meta.get("auction_buy_now_price")),
                "publish_attempt_count": int(publish_attempt_count),
                "publish_success_count": int(publish_success_count),
                "publish_error_count": int(publish_error_count),
                "publish_outcome": publish_outcome,
                "published_at": published_at or None,
                "published_listing_id": external_listing_id or None,
                "last_publish_error": last_error or None,
            }
        )
    return sorted(
        rows,
        key=lambda x: (str(x.get("listed_at") or ""), int(x.get("listing_id") or 0)),
        reverse=True,
    )


def render_reports(repo: InventoryRepository) -> None:
    user = current_user()
    st.subheader("Reports")
    st.caption("Operational reports and export files for bookkeeping and QuickBooks workflows.")
    render_help_panel(
        section_title="Reports",
        goal="Generate operational and accounting exports with a selectable date range.",
        steps=[
            "Set report start/end dates to scope records for inventory, listings, and sales.",
            "Review report tables in-app before exporting to CSV or XLSX.",
            "Use QuickBooks export files as staging inputs for accounting sync workflows.",
            "Re-run exports after data corrections to keep downstream books accurate.",
        ],
        roadmap_phase="v0.3 Channel Sync + Accounting Readiness",
    )

    c1, c2 = st.columns(2)
    with c1:
        from_date = st.date_input("From Date", value=utc_today().replace(day=1))
    with c2:
        to_date = st.date_input("To Date", value=utc_today())

    start_dt = datetime.combine(from_date, datetime.min.time())
    end_dt = datetime.combine(to_date, datetime.max.time())
    load_extended_analytics = st.checkbox(
        "Load Extended Analytics (slower; includes fee reconciliation, lifecycle trends, and deep margin rollups)",
        value=False,
        key="reports_load_extended_analytics",
    )
    load_inventory_cycle_analytics = st.checkbox(
        "Load Inventory Cycle + Rebuy Analytics (slower)",
        value=False,
        key="reports_load_inventory_cycle_analytics",
    )
    load_shipping_tax_analytics = st.checkbox(
        "Load Shipping + Tax/SUTS Analytics (slower)",
        value=True,
        key="reports_load_shipping_tax_analytics",
        help="Shows shipping economics, tax drilldown, and the Colorado SUTS XLSX upload generator.",
    )
    render_full_tables = st.checkbox(
        "Render full report tables (slower)",
        value=False,
        key="reports_render_full_tables",
    )
    preview_row_limit = int(
        st.number_input(
            "Report preview row limit",
            min_value=25,
            max_value=5000,
            value=250,
            step=25,
            key="reports_preview_row_limit",
            help="Used when full table rendering is disabled.",
        )
    )

    def _render_df_with_preview(df: pd.DataFrame, *, hide_index: bool = False) -> None:
        bounded_df, truncated = _bounded_dataframe(
            df,
            render_full_tables=render_full_tables,
            preview_row_limit=preview_row_limit,
        )
        if truncated:
            st.caption(
                f"Showing preview rows only (`{int(len(bounded_df))}` of `{int(len(df))}`). "
                "Enable `Render full report tables` for complete in-app rendering."
            )
        st.dataframe(bounded_df, use_container_width=True, hide_index=hide_index)

    def _rollback_report_session() -> None:
        db = getattr(repo, "db", None)
        rollback = getattr(db, "rollback", None)
        if not callable(rollback):
            return
        try:
            rollback()
        except Exception:
            pass

    def _load_rollup_rows(method_name: str, *, enabled: bool, **kwargs) -> tuple[list[dict], bool]:
        if not enabled:
            return [], False
        method = getattr(repo, method_name, None)
        if method is None:
            return [], False
        try:
            rows = method(**kwargs) or []
            return list(rows), True
        except Exception:
            _rollback_report_session()
            return [], False

    all_products: list | None = None
    all_listings: list | None = None
    supports_products_rollup = hasattr(repo, "report_products_rows")
    supports_listings_rollup = hasattr(repo, "report_listings_rows")
    supports_sales_rollup = hasattr(repo, "report_sales_rows")
    all_sales: list | None = None
    all_orders: list | None = None
    supports_movement_rollup = hasattr(repo, "report_inventory_movement_rows")
    supports_cycle_rollup = hasattr(repo, "report_inventory_cycle_rows")
    supports_rebuy_rollup = hasattr(repo, "report_rebuy_cost_trend_rows")
    supports_orders_rollup = hasattr(repo, "report_orders_rows")
    supports_order_items_rollup = hasattr(repo, "report_order_items_rows")
    supports_returns_rollup = hasattr(repo, "report_returns_rows")
    supports_lot_assignment_rollup = hasattr(repo, "report_lot_assignment_rows")
    supports_cost_maps_rollup = hasattr(repo, "report_sale_unit_cost_maps")
    supports_accounting_exception_rollup = hasattr(repo, "report_accounting_exception_rows")
    supports_listing_review_rollup = hasattr(repo, "report_listing_review_activity_rows")
    supports_listing_format_rollup = hasattr(repo, "report_listing_format_outcome_rows")

    product_rows, product_rows_loaded = _load_rollup_rows(
        "report_products_rows",
        enabled=supports_products_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )

    listing_rows, listing_rows_loaded = _load_rollup_rows(
        "report_listings_rows",
        enabled=supports_listings_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    order_rows, order_rows_loaded = _load_rollup_rows(
        "report_orders_rows",
        enabled=supports_orders_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    sales_rows, sales_rows_loaded = _load_rollup_rows(
        "report_sales_rows",
        enabled=supports_sales_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    all_returns: list | None = None
    all_order_items: list | None = None
    all_movements: list | None = None
    all_assignments: list | None = None

    def _get_all_products() -> list:
        nonlocal all_products
        if all_products is None:
            all_products = repo.list_products()
        return all_products

    def _get_all_listings() -> list:
        nonlocal all_listings
        if all_listings is None:
            all_listings = repo.list_listings()
        return all_listings

    def _get_all_sales() -> list:
        nonlocal all_sales
        if all_sales is None:
            all_sales = repo.list_sales()
        return all_sales

    def _get_all_orders() -> list:
        nonlocal all_orders
        if all_orders is None:
            all_orders = repo.list_orders()
        return all_orders

    def _get_all_order_items() -> list:
        nonlocal all_order_items
        if all_order_items is None:
            all_order_items = repo.list_order_items()
        return all_order_items

    def _get_all_returns() -> list:
        nonlocal all_returns
        if all_returns is None:
            all_returns = repo.list_returns()
        return all_returns

    def _get_all_movements() -> list:
        nonlocal all_movements
        if all_movements is None:
            all_movements = repo.list_inventory_movements(limit=20000)
        return all_movements

    def _get_all_assignments() -> list:
        nonlocal all_assignments
        if all_assignments is None:
            all_assignments = repo.list_product_lot_assignments()
        return all_assignments

    if product_rows_loaded:
        products = [_product_from_row(row) for row in product_rows]
    else:
        products = [
            p
            for p in _get_all_products()
            if p.acquired_at is None or start_dt <= p.acquired_at <= end_dt
        ]
    if listing_rows_loaded:
        listings = [_listing_from_row(row) for row in listing_rows]
    else:
        listings = [
            l
            for l in _get_all_listings()
            if l.listed_at is None or start_dt <= l.listed_at <= end_dt
        ]
    if sales_rows_loaded:
        sales = [_sale_from_row(row) for row in sales_rows]
    else:
        sales = [
            s
            for s in _get_all_sales()
            if s.sold_at is not None and start_dt <= s.sold_at <= end_dt
        ]
    if order_rows_loaded:
        orders = [_order_from_row(row) for row in order_rows]
    else:
        orders = [
            o
            for o in _get_all_orders()
            if o.sold_at is not None and start_dt <= o.sold_at <= end_dt
        ]
    order_items = []
    returns = []
    assignments = []
    movement_rows, movement_rows_loaded = _load_rollup_rows(
        "report_inventory_movement_rows",
        enabled=supports_movement_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    movements = []
    if not movement_rows_loaded:
        movements = [
            m
            for m in _get_all_movements()
            if m.occurred_at is None or start_dt <= m.occurred_at <= end_dt
        ]
    return_rows, return_rows_loaded = _load_rollup_rows(
        "report_returns_rows",
        enabled=supports_returns_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    if not return_rows_loaded:
        returns = [
            r
            for r in _get_all_returns()
            if r.returned_at is not None and start_dt <= r.returned_at <= end_dt
        ]
    assignment_rows, assignment_rows_loaded = _load_rollup_rows(
        "report_lot_assignment_rows",
        enabled=supports_lot_assignment_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    if not assignment_rows_loaded:
        assignments = [
            a
            for a in _get_all_assignments()
            if a.acquired_at is None or start_dt <= a.acquired_at <= end_dt
        ]
    default_unit_cost_by_product = {int(p.id): _landed_unit_cost_from_product(p) for p in products}
    fifo_unit_cost_by_sale: dict[int, float] = {}
    fifo_unit_cost_source_by_sale: dict[int, str] = {}
    fifo_cogs_evidence_by_sale: dict[int, list[dict[str, object]]] = {}
    lot_weighted_unit_cost_by_product: dict[int, float] = {}
    lot_weighted_unit_cost_source_by_product: dict[int, str] = {}
    fifo_remaining_unit_cost_by_product: dict[int, float] = {}
    if supports_cost_maps_rollup:
        try:
            maps_payload = repo.report_sale_unit_cost_maps(
                end_dt=end_dt,
                default_unit_cost_by_product=default_unit_cost_by_product,
            )
            fifo_unit_cost_by_sale = {
                int(k): _safe_float(v)
                for k, v in dict(maps_payload.get("fifo_unit_cost_by_sale") or {}).items()
            }
            fifo_unit_cost_source_by_sale = {
                int(k): str(v or "").strip() or "unknown"
                for k, v in dict(maps_payload.get("fifo_unit_cost_source_by_sale") or {}).items()
            }
            fifo_cogs_evidence_by_sale = {
                int(k): [dict(item) for item in list(v or []) if isinstance(item, dict)]
                for k, v in dict(maps_payload.get("fifo_cogs_evidence_by_sale") or {}).items()
            }
            lot_weighted_unit_cost_by_product = {
                int(k): _safe_float(v)
                for k, v in dict(maps_payload.get("lot_weighted_unit_cost_by_product") or {}).items()
            }
            lot_weighted_unit_cost_source_by_product = {
                int(k): str(v or "").strip() or "unknown"
                for k, v in dict(maps_payload.get("lot_weighted_unit_cost_source_by_product") or {}).items()
            }
            fifo_remaining_unit_cost_by_product = {
                int(k): _safe_float(v)
                for k, v in dict(maps_payload.get("fifo_remaining_unit_cost_by_product") or {}).items()
            }
        except Exception:
            _rollback_report_session()
            fifo_unit_cost_by_sale = {}
            fifo_unit_cost_source_by_sale = {}
            fifo_cogs_evidence_by_sale = {}
            lot_weighted_unit_cost_by_product = {}
            lot_weighted_unit_cost_source_by_product = {}
            fifo_remaining_unit_cost_by_product = {}
    if not fifo_unit_cost_by_sale and not lot_weighted_unit_cost_by_product:
        all_sales = _get_all_sales()
        all_assignments = _get_all_assignments()
        fifo_unit_cost_by_sale = _build_fifo_unit_cost_map(
            all_sales=all_sales,
            all_assignments=all_assignments,
            default_unit_cost_by_product=default_unit_cost_by_product,
        )
        lot_weighted_unit_cost_by_product = _build_lot_weighted_unit_cost_map(
            all_assignments=all_assignments,
            default_unit_cost_by_product=default_unit_cost_by_product,
        )
        fifo_unit_cost_source_by_sale = {
            int(sale_id): "legacy_reports_fifo_cost_map"
            for sale_id, unit_cost in fifo_unit_cost_by_sale.items()
            if _safe_float(unit_cost) > 0
        }
        fifo_cogs_evidence_by_sale = {}
        lot_weighted_unit_cost_source_by_product = {
            int(product_id): "legacy_reports_lot_weighted_cost_map"
            for product_id, unit_cost in lot_weighted_unit_cost_by_product.items()
            if _safe_float(unit_cost) > 0
        }
        fifo_remaining_unit_cost_by_product = _build_fifo_remaining_unit_cost_map(
            all_sales=all_sales,
            all_assignments=all_assignments,
            default_unit_cost_by_product=default_unit_cost_by_product,
        )

    st.markdown("### Tax Reporting Scope")
    tax_default_jurisdiction = get_runtime_str(repo, "invoicing_tax_jurisdiction", "Golden, Colorado")
    tax_default_rate_raw = get_runtime_str(repo, "invoicing_tax_rate_percent_default", "7.50")
    try:
        tax_default_rate = float(tax_default_rate_raw)
    except Exception:
        tax_default_rate = 0.0
    tax_shipping_taxable_default = get_runtime_bool(
        repo,
        "invoicing_tax_shipping_taxable_default",
        False,
    )
    tax_exempt_categories_default_csv = get_runtime_str(repo, "invoicing_tax_exempt_categories_csv", "bullion,coins")
    facilitator_channels_default_csv = get_runtime_str(repo, "marketplace_facilitator_channels_csv", "ebay")
    if "reports_tax_exempt_categories_csv" not in st.session_state:
        st.session_state["reports_tax_exempt_categories_csv"] = str(tax_exempt_categories_default_csv or "bullion,coins")
    if "reports_tax_facilitator_channels_csv" not in st.session_state:
        st.session_state["reports_tax_facilitator_channels_csv"] = str(facilitator_channels_default_csv or "ebay")
    tax_exempt_categories = _parse_csv_set(str(st.session_state.get("reports_tax_exempt_categories_csv") or ""))
    sales_marketplace_options = sorted(
        {
            str((s.marketplace or "")).strip().lower()
            for s in sales
            if str((s.marketplace or "")).strip()
        }
    )
    facilitator_channels = _parse_csv_set(str(st.session_state.get("reports_tax_facilitator_channels_csv") or ""))
    default_tax_marketplaces = _default_tax_marketplace_scope(
        sales_marketplace_options=sales_marketplace_options,
        facilitator_channels=facilitator_channels,
    )
    if "reports_tax_jurisdiction" not in st.session_state:
        st.session_state["reports_tax_jurisdiction"] = str(tax_default_jurisdiction or "Golden, Colorado")
    if "reports_tax_rate_percent" not in st.session_state:
        st.session_state["reports_tax_rate_percent"] = float(max(0.0, tax_default_rate))
    if "reports_tax_shipping_taxable" not in st.session_state:
        st.session_state["reports_tax_shipping_taxable"] = bool(tax_shipping_taxable_default)
    if "reports_tax_marketplaces" not in st.session_state:
        st.session_state["reports_tax_marketplaces"] = list(default_tax_marketplaces)
    tr1, tr2, tr3 = st.columns(3)
    with tr1:
        tax_jurisdiction = st.text_input(
            "Tax Jurisdiction Context",
            key="reports_tax_jurisdiction",
        ).strip()
    with tr2:
        tax_rate_percent = st.number_input(
            "Estimated Tax Rate (%)",
            min_value=0.0,
            step=0.01,
            key="reports_tax_rate_percent",
        )
    with tr3:
        tax_shipping_taxable = st.checkbox(
            "Treat shipping as taxable",
            key="reports_tax_shipping_taxable",
        )
    if facilitator_channels:
        st.caption(
            "Marketplace facilitator channels are excluded by default in local tax scope: "
            + ", ".join(sorted(facilitator_channels))
            + "."
        )
    tax_profile_rows = _latest_tax_profile_rows(repo)
    tax_signoff_df = pd.DataFrame(_latest_tax_signoff_rows(repo))
    accounting_close_signoff_df = pd.DataFrame(_latest_accounting_close_signoff_rows(repo))
    ai_review_outcomes_df = pd.DataFrame(_latest_ai_review_outcome_rows(repo))
    selected_tax_profile_context: dict[str, object] = {}
    if tax_profile_rows:
        profile_options = {
            f"{row.get('profile_name') or row.get('profile_key')} [{row.get('profile_key')}]": row
            for row in tax_profile_rows
        }
        sp1, sp2 = st.columns([2, 1])
        with sp1:
            selected_tax_profile_label = st.selectbox(
                "Saved Tax Profile",
                options=list(profile_options.keys()),
                key="reports_saved_tax_profile",
            )
        with sp2:
            if st.button("Apply Saved Tax Profile", key="reports_apply_saved_tax_profile_btn"):
                profile = profile_options.get(selected_tax_profile_label) or {}
                st.session_state["reports_tax_jurisdiction"] = str(
                    profile.get("jurisdiction") or tax_default_jurisdiction or "Golden, Colorado"
                )
                st.session_state["reports_tax_rate_percent"] = float(max(0.0, _safe_float(profile.get("tax_rate_percent"))))
                st.session_state["reports_tax_shipping_taxable"] = bool(profile.get("shipping_taxable"))
                st.session_state["reports_tax_facilitator_channels_csv"] = str(profile.get("facilitator_channels") or "")
                st.session_state["reports_tax_exempt_categories_csv"] = str(profile.get("tax_exempt_categories") or "")
                st.session_state["reports_tax_marketplaces"] = _default_tax_marketplace_scope(
                    sales_marketplace_options=sales_marketplace_options,
                    facilitator_channels=_parse_csv_set(str(profile.get("facilitator_channels") or "")),
                )
                st.success(f"Applied saved tax profile `{profile.get('profile_key')}`.")
                st.rerun()
        selected_profile = profile_options.get(str(st.session_state.get("reports_saved_tax_profile") or "")) or {}
        selected_tax_profile_context = dict(selected_profile)
        validation_status = str(selected_profile.get("human_validation_status") or "").strip()
        if validation_status and validation_status != "advisor_validated":
            st.warning(
                "Selected tax profile is not advisor-validated "
                f"(`{validation_status}`). Treat outputs as review-only estimates."
            )
        elif selected_profile:
            st.caption("Selected saved tax profile is advisor-validated.")
    preset_map = _tax_report_presets(
        default_jurisdiction=tax_default_jurisdiction,
        default_tax_rate_percent=float(max(0.0, tax_default_rate)),
        default_shipping_taxable=bool(tax_shipping_taxable_default),
    )
    tp1, tp2 = st.columns([2, 1])
    with tp1:
        tax_preset_name = st.selectbox(
            "Tax Report Preset",
            options=list(preset_map.keys()),
            key="reports_tax_preset_name",
        )
    with tp2:
        if st.button("Apply Tax Report Preset", key="reports_apply_tax_preset_btn"):
            preset = preset_map.get(tax_preset_name) or {}
            st.session_state["reports_tax_jurisdiction"] = str(
                preset.get("jurisdiction") or tax_default_jurisdiction or "Golden, Colorado"
            )
            st.session_state["reports_tax_rate_percent"] = float(
                max(0.0, float(preset.get("tax_rate_percent") or 0.0))
            )
            st.session_state["reports_tax_shipping_taxable"] = bool(preset.get("shipping_taxable", False))
            marketplace_mode = str(preset.get("marketplace_mode") or "all").strip().lower()
            if marketplace_mode == "local_only":
                local_candidates = [m for m in sales_marketplace_options if m in {"local", "in_person", "pos"}]
                st.session_state["reports_tax_marketplaces"] = local_candidates
            else:
                st.session_state["reports_tax_marketplaces"] = list(sales_marketplace_options)
            st.success(f"Applied tax report preset `{tax_preset_name}`.")
            st.rerun()
    # Keep state value normalized to available options before rendering keyed widget.
    current_tax_marketplaces = st.session_state.get("reports_tax_marketplaces")
    if current_tax_marketplaces is None:
        current_tax_marketplaces = list(default_tax_marketplaces)
    st.session_state["reports_tax_marketplaces"] = [
        m for m in current_tax_marketplaces if m in sales_marketplace_options
    ]
    selected_tax_marketplaces = st.multiselect(
        "Tax Marketplace Filter",
        options=sales_marketplace_options,
        key="reports_tax_marketplaces",
        help=(
            "Choose marketplaces you are responsible for reporting/remitting. "
            "Leave eBay/facilitator channels unselected for normal SUTS filing unless your advisor tells you otherwise."
        ),
    )
    selected_tax_marketplace_set = {str(v).strip().lower() for v in selected_tax_marketplaces if str(v).strip()}
    tax_query_marketplace_set = (
        set(selected_tax_marketplace_set)
        if selected_tax_marketplace_set
        else {"__no_tax_marketplace_selected__"}
    )
    tax_marketplace_scope_label = (
        ",".join(sorted(selected_tax_marketplace_set)) if selected_tax_marketplace_set else "none"
    )
    if sales_marketplace_options and not selected_tax_marketplace_set:
        st.warning(
            "No non-facilitator tax marketplaces are selected. Tax/SUTS outputs for this scope will be zero "
            "unless you intentionally select a marketplace for advisor-reviewed reporting."
        )
    selected_facilitator_marketplaces = selected_tax_marketplace_set.intersection(facilitator_channels)
    if selected_facilitator_marketplaces:
        st.warning(
            "Marketplace facilitator channel(s) selected for tax/SUTS scope: "
            + ", ".join(sorted(selected_facilitator_marketplaces))
            + ". For normal SUTS remittance, keep eBay/facilitator channels unselected because the facilitator "
            "generally collects/remits those taxes. Select only with advisor-confirmed reporting instructions."
        )
    else:
        st.caption(
            "For SUTS filing, keep eBay/facilitator channels unselected unless your advisor says to report "
            "marketplace-facilitator gross sales. Direct/local channels are the normal SUTS remittance scope."
        )
    st.caption(
        "Tax-exempt categories (runtime): "
        + (", ".join(sorted(tax_exempt_categories)) if tax_exempt_categories else "(none)")
    )
    st.info(
        "Tax outputs in this report are estimates for operational planning. "
        "Validate local/state tax treatment (including bullion/coin exemptions) with your tax advisor."
    )

    sales_row_by_id = {
        int(row.get("sale_id") or 0): row
        for row in (sales_rows if sales_rows_loaded else [])
        if int(row.get("sale_id") or 0) > 0
    }
    sales_detail_actual_by_sale_id: dict[int, dict[str, object]] = {}
    if hasattr(repo, "report_sales_actual_econ_rows"):
        try:
            sales_detail_actual_by_sale_id = {
                int(row.get("sale_id") or 0): row
                for row in repo.report_sales_actual_econ_rows(start_dt=start_dt, end_dt=end_dt)
                if int(row.get("sale_id") or 0) > 0
            }
        except Exception:
            _rollback_report_session()
            sales_detail_actual_by_sale_id = {}

    sales_df = pd.DataFrame(
        [
            {
                "sale_id": s.id,
                "sold_at": iso_or_none(s.sold_at),
                "marketplace": s.marketplace,
                "order_id": s.order_id,
                "sku": s.product.sku if s.product else None,
                "product_title": s.product.title if s.product else None,
                "listing_id": s.listing_id,
                "external_order_id": s.external_order_id,
                "qty": s.quantity_sold,
                "gross_sales": float(s.sold_price),
                "fees": float(s.fees),
                "shipping_cost": float(s.shipping_cost),
                "shipping_provider": s.shipping_provider,
                "shipping_service": s.shipping_service,
                "shipping_package_type": s.shipping_package_type,
                "tracking_number": s.tracking_number,
                "tracking_status": s.tracking_status,
                "shipping_exception_code": s.shipping_exception_code,
                "shipping_exception_action": s.shipping_exception_action,
                "shipping_exception_notes": s.shipping_exception_notes,
                "shipping_exception_resolved_at": iso_or_none(s.shipping_exception_resolved_at),
                "shipping_exception_resolved_by": s.shipping_exception_resolved_by,
                "shipment_exported_at": iso_or_none(s.shipment_exported_at),
                "shipped_at": iso_or_none(s.shipped_at),
                "delivered_at": iso_or_none(s.delivered_at),
                "shipping_label_cost": _safe_float(getattr(s, "shipping_label_cost", None)),
                "field_net_before_cogs": _safe_float(
                    sales_row_by_id.get(int(s.id), {}).get("field_net_before_cogs")
                )
                if int(s.id) in sales_row_by_id
                else _sale_net_before_cogs_from_fields(s),
                "actual_fee": _safe_float(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_fee",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get("allocated_fee_actual", s.fees),
                    )
                ),
                "actual_shipping_charged": _safe_float(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_shipping_charged",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get(
                            "allocated_shipping_charged",
                            s.shipping_cost,
                        ),
                    )
                ),
                "actual_shipping_label_cost": _safe_float(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_shipping_label_cost",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get(
                            "allocated_shipping_actual",
                            getattr(s, "shipping_label_cost", None),
                        ),
                    )
                ),
                "actual_net_before_cogs": _safe_float(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_net_before_cogs",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get(
                            "net_before_cogs_actual",
                            _sale_net_before_cogs_from_fields(s),
                        ),
                    )
                ),
                "actual_fee_source": str(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_fee_source",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get(
                            "actual_fee_source",
                            "sale_fees_field",
                        ),
                    )
                    or ""
                ),
                "actual_shipping_source": str(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_shipping_source",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get(
                            "actual_shipping_source",
                            "sale_shipping_label_field",
                        ),
                    )
                    or ""
                ),
                "net_sales": _safe_float(
                    sales_row_by_id.get(int(s.id), {}).get(
                        "actual_net_before_cogs",
                        sales_detail_actual_by_sale_id.get(int(s.id), {}).get(
                            "net_before_cogs_actual",
                            _sale_net_before_cogs_from_fields(s),
                        ),
                    )
                ),
            }
            for s in sales
        ]
    )
    shipping_economics_df = pd.DataFrame()
    shipping_econ_summary_df = pd.DataFrame()
    tax_detail_rows = []
    tax_detail_df = pd.DataFrame()
    tax_summary_df = pd.DataFrame()
    tax_by_marketplace_df = pd.DataFrame()
    tax_exceptions_df = pd.DataFrame()
    if load_shipping_tax_analytics:
        shipping_marketplaces = {"ebay", "facebook", "craigslist", "local", "in_person", "pos"}
        shipping_loaded_from_repo = False
        if hasattr(repo, "report_shipping_economics_rows"):
            try:
                shipping_rows = repo.report_shipping_economics_rows(
                    start_dt=start_dt,
                    end_dt=end_dt,
                    marketplaces=shipping_marketplaces,
                )
                shipping_economics_df = pd.DataFrame(shipping_rows)
                if not shipping_economics_df.empty:
                    shipping_economics_df["sold_at"] = shipping_economics_df["sold_at"].apply(iso_or_none)
                    shipping_economics_df["shipping_label_purchased_at"] = shipping_economics_df[
                        "shipping_label_purchased_at"
                    ].apply(iso_or_none)
                summary_rows = repo.report_shipping_economics_summary(
                    start_dt=start_dt,
                    end_dt=end_dt,
                    marketplaces=shipping_marketplaces,
                )
                shipping_econ_summary_df = pd.DataFrame(summary_rows)
                shipping_loaded_from_repo = True
            except Exception:
                _rollback_report_session()
        if not shipping_loaded_from_repo:
            shipping_economics_df = pd.DataFrame(
                [
                    {
                        "sale_id": int(s.id),
                        "sold_at": iso_or_none(s.sold_at),
                        "marketplace": str(s.marketplace or "").strip().lower(),
                        "external_order_id": str(s.external_order_id or "").strip(),
                        "order_id": int(s.order_id) if s.order_id is not None else None,
                        "sku": s.product.sku if s.product else None,
                        "product_title": s.product.title if s.product else None,
                        "qty": int(s.quantity_sold or 0),
                        "shipping_charged_to_buyer": round(_safe_float(s.shipping_cost), 2),
                        "shipping_label_spend": round(_safe_float(getattr(s, "shipping_label_cost", None)), 2),
                        "shipping_delta_charged_minus_spend": round(
                            _safe_float(s.shipping_cost) - _safe_float(getattr(s, "shipping_label_cost", None)),
                            2,
                        ),
                        "shipping_label_currency": str(getattr(s, "shipping_label_currency", "") or "").strip(),
                        "shipping_label_id": str(getattr(s, "shipping_label_id", "") or "").strip(),
                        "shipping_label_purchased_at": iso_or_none(getattr(s, "shipping_label_purchased_at", None)),
                        "shipping_provider": str(s.shipping_provider or "").strip(),
                        "shipping_service": str(s.shipping_service or "").strip(),
                        "tracking_number": str(s.tracking_number or "").strip(),
                    }
                    for s in sales
                    if str(s.marketplace or "").strip().lower() in shipping_marketplaces
                ]
            )
            shipping_econ_summary_df = pd.DataFrame()
            if not shipping_economics_df.empty:
                shipping_econ_summary_df = (
                    shipping_economics_df.groupby(["marketplace"], dropna=False, as_index=False)
                    .agg(
                        sales_count=("sale_id", "count"),
                        total_shipping_charged=("shipping_charged_to_buyer", "sum"),
                        total_label_spend=("shipping_label_spend", "sum"),
                    )
                    .sort_values(["sales_count"], ascending=[False])
                )
                shipping_econ_summary_df["shipping_delta_charged_minus_spend"] = (
                    shipping_econ_summary_df["total_shipping_charged"] - shipping_econ_summary_df["total_label_spend"]
                )
                # Compute coverage once per marketplace and merge, avoiding per-row
                # DataFrame scans over the full shipping dataset.
                coverage_counts = (
                    shipping_economics_df.assign(
                        _has_label_spend=shipping_economics_df["shipping_label_spend"] > 0
                    )
                    .groupby(["marketplace"], dropna=False, as_index=False)
                    .agg(label_spend_covered_count=("_has_label_spend", "sum"))
                )
                shipping_econ_summary_df = shipping_econ_summary_df.merge(
                    coverage_counts,
                    on="marketplace",
                    how="left",
                )
                shipping_econ_summary_df["label_spend_covered_count"] = (
                    shipping_econ_summary_df["label_spend_covered_count"].fillna(0).astype(int)
                )
                shipping_econ_summary_df["label_spend_coverage_percent"] = (
                    (
                        shipping_econ_summary_df["label_spend_covered_count"]
                        / shipping_econ_summary_df["sales_count"].replace(0, pd.NA)
                    )
                    * 100.0
                ).fillna(0.0)

        tax_loaded_from_repo = False
        if hasattr(repo, "report_tax_estimate_detail_rows"):
            try:
                tax_detail_rows = repo.report_tax_estimate_detail_rows(
                    start_dt=start_dt,
                    end_dt=end_dt,
                    tax_rate_percent=float(tax_rate_percent),
                    shipping_taxable=bool(tax_shipping_taxable),
                    tax_exempt_categories=tax_exempt_categories,
                    marketplaces=tax_query_marketplace_set,
                )
                for row in tax_detail_rows:
                    row["sold_at"] = iso_or_none(row.get("sold_at"))
                    row["tax_jurisdiction"] = tax_jurisdiction or tax_default_jurisdiction
                    row["estimated_tax_rate_percent"] = float(tax_rate_percent)
                tax_loaded_from_repo = True
            except Exception:
                _rollback_report_session()
                tax_detail_rows = []
        if not tax_loaded_from_repo:
            for s in sales:
                marketplace = str(s.marketplace or "").strip().lower()
                if marketplace not in tax_query_marketplace_set:
                    continue
                sold_price = _safe_float(s.sold_price)
                shipping_cost = _safe_float(s.shipping_cost)
                category = str((s.product.category if s.product else "") or "").strip().lower()
                is_exempt = bool(category and category in tax_exempt_categories)
                taxable_item_subtotal = 0.0 if is_exempt else sold_price
                taxable_shipping = shipping_cost if tax_shipping_taxable else 0.0
                taxable_subtotal = max(0.0, taxable_item_subtotal + taxable_shipping)
                estimated_tax = round(taxable_subtotal * (float(tax_rate_percent) / 100.0), 2)
                tax_detail_rows.append(
                    {
                        "sale_id": s.id,
                        "sold_at": iso_or_none(s.sold_at),
                        "marketplace": marketplace,
                        "sku": s.product.sku if s.product else None,
                        "product_title": s.product.title if s.product else None,
                        "category": category,
                        "gross_sales": sold_price,
                        "shipping_cost": shipping_cost,
                        "is_tax_exempt_category": bool(is_exempt),
                        "taxable_item_subtotal": round(taxable_item_subtotal, 2),
                        "taxable_shipping_subtotal": round(taxable_shipping, 2),
                        "taxable_subtotal": round(taxable_subtotal, 2),
                        "estimated_tax_collected": estimated_tax,
                        "tax_jurisdiction": tax_jurisdiction or tax_default_jurisdiction,
                        "estimated_tax_rate_percent": float(tax_rate_percent),
                    }
                )
        tax_detail_df = pd.DataFrame(tax_detail_rows)
        tax_summary_rows = []
        if not tax_detail_df.empty:
            taxable_subtotal_sum = float(tax_detail_df["taxable_subtotal"].sum())
            gross_sales_sum = float(tax_detail_df["gross_sales"].sum())
            exempt_subtotal_sum = max(0.0, gross_sales_sum - float(tax_detail_df["taxable_item_subtotal"].sum()))
            tax_summary_rows.append(
                {
                    "jurisdiction": tax_jurisdiction or tax_default_jurisdiction,
                    "tax_rate_percent": float(tax_rate_percent),
                    "shipping_taxable": bool(tax_shipping_taxable),
                    "marketplace_scope": tax_marketplace_scope_label,
                    "sales_count": int(len(tax_detail_df)),
                    "gross_sales_subtotal": round(gross_sales_sum, 2),
                    "taxable_subtotal": round(taxable_subtotal_sum, 2),
                    "exempt_subtotal": round(exempt_subtotal_sum, 2),
                    "estimated_tax_collected": round(float(tax_detail_df["estimated_tax_collected"].sum()), 2),
                }
            )
        tax_summary_df = pd.DataFrame(tax_summary_rows)
        tax_by_marketplace_df = pd.DataFrame()
        if not tax_detail_df.empty:
            tax_by_marketplace_df = (
                tax_detail_df.groupby(["marketplace"], as_index=False)
                .agg(
                    sales_count=("sale_id", "count"),
                    gross_sales_subtotal=("gross_sales", "sum"),
                    taxable_subtotal=("taxable_subtotal", "sum"),
                    estimated_tax_collected=("estimated_tax_collected", "sum"),
                )
                .sort_values(["estimated_tax_collected"], ascending=[False])
            )
        tax_exceptions_df = pd.DataFrame(
            _build_tax_exception_rows(
                tax_detail_df,
                tax_jurisdiction=tax_jurisdiction or tax_default_jurisdiction,
                tax_rate_percent=float(tax_rate_percent),
                shipping_taxable=bool(tax_shipping_taxable),
                facilitator_channels=facilitator_channels,
                tax_exempt_categories=tax_exempt_categories,
            )
        )

    def _inventory_report_landed_unit_cost(product) -> float | None:
        product_landed = _landed_unit_cost_from_product(product)
        has_product_cost = (
            getattr(product, "acquisition_cost", None) is not None
            or getattr(product, "acquisition_tax_paid", None) is not None
            or getattr(product, "acquisition_shipping_paid", None) is not None
            or getattr(product, "acquisition_handling_paid", None) is not None
        )
        product_id = getattr(product, "id", None)
        if product_id is not None and int(product_id) in fifo_remaining_unit_cost_by_product:
            return _safe_float(fifo_remaining_unit_cost_by_product.get(int(product_id)))
        if product_id is not None and int(product_id) in lot_weighted_unit_cost_by_product:
            return _safe_float(lot_weighted_unit_cost_by_product.get(int(product_id)))
        if has_product_cost and product_landed > 0:
            return product_landed
        return product_landed if has_product_cost else None

    inventory_df = pd.DataFrame(
        [
            {
                "product_id": p.id,
                "sku": p.sku,
                "title": p.title,
                "category": p.category,
                "metal_type": p.metal_type,
                "acquired_at": iso_or_none(p.acquired_at),
                "unit_cost": float(p.acquisition_cost) if p.acquisition_cost is not None else None,
                "product_cost": float(getattr(p, "product_cost", 0)) if getattr(p, "product_cost", None) is not None else None,
                "unit_tax_paid": float(getattr(p, "acquisition_tax_paid", None)) if getattr(p, "acquisition_tax_paid", None) is not None else None,
                "unit_shipping_paid": float(getattr(p, "acquisition_shipping_paid", None)) if getattr(p, "acquisition_shipping_paid", None) is not None else None,
                "unit_handling_paid": float(getattr(p, "acquisition_handling_paid", None)) if getattr(p, "acquisition_handling_paid", None) is not None else None,
                "landed_unit_cost": (
                    _inventory_report_landed_unit_cost(p)
                ),
                "item_weight_oz": float(p.weight_oz) if p.weight_oz is not None else None,
                "package_weight_oz": float(p.package_weight_oz) if p.package_weight_oz is not None else None,
                "package_length_in": float(p.package_length_in) if p.package_length_in is not None else None,
                "package_width_in": float(p.package_width_in) if p.package_width_in is not None else None,
                "package_height_in": float(p.package_height_in) if p.package_height_in is not None else None,
                "qty_on_hand": p.current_quantity,
                "inventory_value": (
                    float(p.current_quantity or 0)
                    * float(_inventory_report_landed_unit_cost(p) or 0)
                    if _inventory_report_landed_unit_cost(p) is not None
                    else None
                ),
                "landed_inventory_value": (
                    (
                        float(p.current_quantity or 0)
                        * float(_inventory_report_landed_unit_cost(p) or 0)
                    )
                    if _inventory_report_landed_unit_cost(p) is not None
                    else None
                ),
            }
            for p in products
        ]
    )

    listings_df = pd.DataFrame(
        [
            {
                "listing_id": l.id,
                "listed_at": iso_or_none(l.listed_at),
                "marketplace": l.marketplace,
                "sku": l.product.sku if l.product else None,
                "title": l.listing_title,
                "status": l.listing_status,
                "marketplace_url": l.marketplace_url,
                "marketplace_details": l.marketplace_details,
                "qty_listed": l.quantity_listed,
                "price": float(l.listing_price),
                "external_listing_id": l.external_listing_id,
            }
            for l in listings
        ]
    )

    if assignment_rows_loaded:
        lots_df = pd.DataFrame(assignment_rows)
    else:
        lots_df = pd.DataFrame(
            [
                {
                    "assignment_id": a.id,
                    "lot_code": a.lot.lot_code if a.lot else None,
                    "source_name": a.lot.source.name if a.lot and a.lot.source else None,
                    "source_type": a.lot.source.source_type if a.lot and a.lot.source else None,
                    "vendor": a.lot.vendor if a.lot else None,
                    "purchase_date": iso_or_none(a.lot.purchase_date) if a.lot else None,
                    "sku": a.product.sku if a.product else None,
                    "product_title": a.product.title if a.product else None,
                    "quantity_acquired": a.quantity_acquired,
                    "unit_cost": float(a.unit_cost) if a.unit_cost is not None else None,
                    "allocated_cost": float(a.allocated_cost) if a.allocated_cost is not None else None,
                    "acquired_at": iso_or_none(a.acquired_at),
                }
                for a in assignments
            ]
        )
    lot_allocation_source_summary_df = _build_lot_allocation_source_summary(lots_df)

    if return_rows_loaded:
        qbo_adjustments_df = pd.DataFrame(
            _build_qbo_adjustment_export_rows(
                return_rows,
                fifo_unit_cost_by_sale,
                fifo_unit_cost_source_by_sale=fifo_unit_cost_source_by_sale,
            )
        )
    else:
        qbo_adjustments_df = pd.DataFrame(
            _build_qbo_adjustment_export_rows(
                returns,
                fifo_unit_cost_by_sale,
                fifo_unit_cost_source_by_sale=fifo_unit_cost_source_by_sale,
            )
        )

    if order_rows_loaded:
        orders_df = pd.DataFrame(order_rows)
    else:
        orders_df = pd.DataFrame(
            [
                {
                    "order_id": o.id,
                    "sold_at": iso_or_none(o.sold_at),
                    "marketplace": o.marketplace,
                    "external_order_id": o.external_order_id,
                    "status": o.order_status,
                    "subtotal_amount": float(o.subtotal_amount),
                    "fees": float(o.fees),
                    "field_fees": float(o.fees),
                    "shipping_cost": float(o.shipping_cost),
                    "shipping_label_cost": _safe_float(getattr(o, "shipping_label_cost", None)),
                    "field_shipping_label_cost": _safe_float(getattr(o, "shipping_label_cost", None)),
                    "shipping_label_currency": str(getattr(o, "shipping_label_currency", "") or "").strip(),
                    "shipping_delta_charged_minus_actual": round(
                        _safe_float(o.shipping_cost) - _safe_float(getattr(o, "shipping_label_cost", None)),
                        2,
                    ),
                    "actual_fee": _safe_float(getattr(o, "fees", None)),
                    "actual_shipping_label_cost": _safe_float(getattr(o, "shipping_label_cost", None)),
                    "actual_shipping_delta_charged_minus_label": round(
                        _safe_float(o.shipping_cost) - _safe_float(getattr(o, "shipping_label_cost", None)),
                        2,
                    ),
                    "actual_net_before_cogs": round(
                        _safe_float(o.subtotal_amount)
                        + _safe_float(o.shipping_cost)
                        - _safe_float(o.fees)
                        - _safe_float(getattr(o, "shipping_label_cost", None)),
                        2,
                    ),
                    "actual_fee_source": "order_fees_field",
                    "actual_shipping_source": "order_shipping_label_field",
                    "total_amount": float(o.total_amount),
                    "item_count": len(o.items),
                    "notes": o.notes,
                }
                for o in orders
            ]
        )

    order_item_rows, order_item_rows_loaded = _load_rollup_rows(
        "report_order_items_rows",
        enabled=supports_order_items_rollup,
        start_dt=start_dt,
        end_dt=end_dt,
    )
    if not order_item_rows_loaded:
        order_items = [
            oi
            for oi in _get_all_order_items()
            if oi.order is not None and oi.order.sold_at is not None and start_dt <= oi.order.sold_at <= end_dt
        ]
    if order_item_rows_loaded:
        order_items_df = pd.DataFrame(order_item_rows)
    else:
        order_items_df = pd.DataFrame(
            [
                {
                    "order_item_id": oi.id,
                    "order_id": oi.order_id,
                    "sold_at": iso_or_none(oi.order.sold_at) if oi.order else None,
                    "marketplace": oi.order.marketplace if oi.order else None,
                    "external_order_id": oi.order.external_order_id if oi.order else None,
                    "product_id": oi.product_id,
                    "listing_id": oi.listing_id,
                    "sku": oi.product.sku if oi.product else None,
                    "product_title": oi.product.title if oi.product else None,
                    "quantity": oi.quantity,
                    "unit_price": float(oi.unit_price),
                    "line_total": float(oi.line_total),
                    "line_fees": float(oi.line_fees),
                    "line_shipping": float(oi.line_shipping),
                    "notes": oi.notes,
                }
                for oi in order_items
            ]
        )
    ebay_order_fee_breakdown_df = pd.DataFrame(
        [
            {
                "order_id": o.id,
                "sold_at": iso_or_none(o.sold_at),
                "external_order_id": o.external_order_id,
                "fee_price_subtotal": _safe_float(breakdown.get("price_subtotal")),
                "fee_delivery_cost": _safe_float(breakdown.get("delivery_cost")),
                "fee_delivery_shipping_cost": _safe_float(breakdown.get("delivery_shipping_cost")),
                "fee_total_marketplace_fee": _safe_float(breakdown.get("total_marketplace_fee")),
                "fee_total_tax": _safe_float(breakdown.get("total_tax")),
                "fee_sales_tax": _safe_float(breakdown.get("sales_tax")),
                "fee_order_total": _safe_float(breakdown.get("order_total")),
                "order_fees_field": _safe_float(o.fees),
                "delta_order_fees_vs_breakdown_total_marketplace_fee": (
                    _safe_float(o.fees) - _safe_float(breakdown.get("total_marketplace_fee"))
                ),
            }
            for o in orders
            for breakdown in [_extract_order_fee_breakdown_from_notes(o.notes)]
            if (o.marketplace or "").strip().lower() == "ebay" and breakdown
        ]
    )
    ebay_marketplace_fee_rows: list[dict] = []
    if hasattr(repo, "report_ebay_marketplace_fee_rows"):
        try:
            ebay_marketplace_fee_rows = repo.report_ebay_marketplace_fee_rows(
                start_dt=start_dt,
                end_dt=end_dt,
            )
        except Exception:
            _rollback_report_session()
            ebay_marketplace_fee_rows = []
    if not ebay_marketplace_fee_rows:
        ebay_marketplace_fee_rows = _build_ebay_marketplace_fee_rows(repo, orders)
    ebay_marketplace_fee_detail_df = pd.DataFrame(ebay_marketplace_fee_rows)
    ebay_marketplace_fee_summary_df = pd.DataFrame()
    ebay_marketplace_fee_by_sku_df = pd.DataFrame()
    ebay_marketplace_fee_by_category_df = pd.DataFrame()
    if not ebay_marketplace_fee_detail_df.empty:
        category_by_sku: dict[str, str] = {}
        for p in products:
            sku_key = str(getattr(p, "sku", "") or "").strip()
            if not sku_key:
                continue
            if sku_key in category_by_sku:
                continue
            category_by_sku[sku_key] = str(getattr(p, "category", "") or "").strip()
        ebay_marketplace_fee_detail_df["product_category"] = ebay_marketplace_fee_detail_df["sku"].map(
            lambda sku: category_by_sku.get(str(sku or "").strip(), "")
        )
        ebay_marketplace_fee_summary_df = (
            ebay_marketplace_fee_detail_df.groupby(
                ["fee_type", "fee_currency", "source"],
                dropna=False,
                as_index=False,
            )
            .agg(
                fee_row_count=("fee_type", "count"),
                order_count=("external_order_id", "nunique"),
                fee_amount_total=("fee_amount", "sum"),
                first_seen=("transaction_date", "min"),
                last_seen=("transaction_date", "max"),
            )
            .sort_values(["fee_amount_total"], ascending=[False])
        )
        ebay_marketplace_fee_by_sku_df = (
            ebay_marketplace_fee_detail_df.groupby(
                ["sku", "product_title", "product_category", "fee_type", "fee_currency"],
                dropna=False,
                as_index=False,
            )
            .agg(
                fee_row_count=("fee_type", "count"),
                order_count=("external_order_id", "nunique"),
                fee_amount_total=("fee_amount", "sum"),
            )
            .sort_values(["fee_amount_total"], ascending=[False])
        )
        ebay_marketplace_fee_by_category_df = (
            ebay_marketplace_fee_detail_df.groupby(
                ["product_category", "fee_type", "fee_currency"],
                dropna=False,
                as_index=False,
            )
            .agg(
                fee_row_count=("fee_type", "count"),
                order_count=("external_order_id", "nunique"),
                sku_count=("sku", "nunique"),
                fee_amount_total=("fee_amount", "sum"),
            )
            .sort_values(["fee_amount_total"], ascending=[False])
        )

    if return_rows_loaded:
        returns_df = pd.DataFrame(return_rows)
    else:
        returns_df = pd.DataFrame(
            [
                {
                    "return_id": r.id,
                    "returned_at": iso_or_none(r.returned_at),
                    "processed_at": iso_or_none(r.processed_at),
                    "marketplace": r.marketplace,
                    "external_return_id": r.external_return_id,
                    "sale_id": r.sale_id,
                    "order_id": r.order_id,
                    "product_id": r.product_id,
                    "sku": (
                        r.product.sku
                        if r.product
                        else (
                            r.sale.listing.product.sku
                            if r.sale and r.sale.listing and r.sale.listing.product
                            else None
                        )
                    ),
                    "product_title": (
                        r.product.title
                        if r.product
                        else (
                            r.sale.listing.product.title
                            if r.sale and r.sale.listing and r.sale.listing.product
                            else None
                        )
                    ),
                    "status": r.return_status,
                    "reason": r.reason,
                    "disposition": r.disposition,
                    "quantity": r.quantity,
                    "refund_amount": float(r.refund_amount),
                    "refund_fees": float(r.refund_fees),
                    "refund_shipping": float(r.refund_shipping),
                    "restocked": r.restocked,
                    "notes": r.notes,
                    "source_order": r.sale.external_order_id if r.sale and r.sale.external_order_id else "",
                }
                for r in returns
            ]
        )

    if movement_rows_loaded:
        movements_df = pd.DataFrame(movement_rows)
    else:
        movements_df = pd.DataFrame(
            [
                {
                    "movement_id": m.id,
                    "occurred_at": iso_or_none(m.occurred_at),
                    "product_id": m.product_id,
                    "sku": m.product.sku if m.product else None,
                    "product_title": m.product.title if m.product else None,
                    "movement_type": m.movement_type,
                    "quantity_delta": m.quantity_delta,
                    "quantity_before": m.quantity_before,
                    "quantity_after": m.quantity_after,
                    "unit_cost": float(m.unit_cost) if m.unit_cost is not None else None,
                    "reference_type": m.reference_type,
                    "reference_id": m.reference_id,
                    "notes": m.notes,
                }
                for m in movements
            ]
        )

    marketplace_rows = []
    if hasattr(repo, "report_marketplace_reconciliation_rows"):
        try:
            marketplace_rows = repo.report_marketplace_reconciliation_rows(
                start_dt=start_dt,
                end_dt=end_dt,
            )
        except Exception:
            _rollback_report_session()
            marketplace_rows = []
    if not marketplace_rows:
        fallback_actual_by_sale_id: dict[int, dict[str, float | str]] = {}
        if hasattr(repo, "report_sales_actual_econ_rows"):
            try:
                fallback_actual_by_sale_id = {
                    int(row.get("sale_id") or 0): row
                    for row in repo.report_sales_actual_econ_rows(start_dt=start_dt, end_dt=end_dt)
                    if int(row.get("sale_id") or 0) > 0
                }
            except Exception:
                _rollback_report_session()
                fallback_actual_by_sale_id = {}
        marketplace_rows = _build_marketplace_reconciliation_fallback_rows(
            sales,
            orders,
            returns_df,
            actual_econ_by_sale_id=fallback_actual_by_sale_id,
        )
    reconciliation_df = pd.DataFrame(marketplace_rows)

    validation_rows = []
    for s in sales:
        reasons = []
        if (s.marketplace or "").strip() and not (s.external_order_id or "").strip():
            reasons.append("missing_external_order_id")
        if s.order_id is None:
            reasons.append("missing_order_link")
        if (
            _safe_float(s.sold_price)
            + _safe_float(s.shipping_cost)
            - _safe_float(s.fees)
            - _safe_float(getattr(s, "shipping_label_cost", None))
        ) < 0:
            reasons.append("negative_net_sale")
        if reasons:
            validation_rows.append(
                {
                    "entity_type": "sale",
                    "entity_id": s.id,
                    "marketplace": s.marketplace,
                    "reference": s.external_order_id or "",
                    "issues": ",".join(reasons),
                    "sold_at": iso_or_none(s.sold_at),
                }
            )
    if not returns_df.empty:
        for row in returns_df.to_dict("records"):
            reasons = []
            if row.get("sale_id") is None:
                reasons.append("return_missing_sale_link")
            if (
                _safe_float(row.get("refund_amount"))
                + _safe_float(row.get("refund_fees"))
                + _safe_float(row.get("refund_shipping"))
            ) <= 0:
                reasons.append("return_non_positive_refund_total")
            if reasons:
                validation_rows.append(
                    {
                        "entity_type": "return",
                        "entity_id": int(row.get("return_id") or 0),
                        "marketplace": str(row.get("marketplace") or "").strip(),
                        "reference": str(row.get("external_return_id") or "").strip(),
                        "issues": ",".join(reasons),
                        "sold_at": str(row.get("returned_at") or "").strip(),
                    }
                )
    accounting_validation_df = pd.DataFrame(validation_rows)

    actual_econ_by_sale_id: dict[int, dict[str, float]] = {}
    actual_econ_rows: list[dict] = []
    economics_intel_df = pd.DataFrame()
    economics_intel_by_sku_df = pd.DataFrame()
    economics_intel_by_marketplace_df = pd.DataFrame()
    economics_intel_alerts_df = pd.DataFrame()
    if hasattr(repo, "report_sales_actual_econ_rows"):
        try:
            actual_econ_rows = repo.report_sales_actual_econ_rows(
                start_dt=start_dt,
                end_dt=end_dt,
            )
        except Exception:
            _rollback_report_session()
            actual_econ_rows = []
    if not actual_econ_rows:
        sales_by_order_id: dict[int, list] = defaultdict(list)
        for s in sales:
            if s.order_id is not None:
                sales_by_order_id[int(s.order_id)].append(s)
        order_by_id = {int(o.id): o for o in orders if o.id is not None}
        for s in sales:
            sold_price = _safe_float(s.sold_price)
            charged_shipping_sale = _safe_float(s.shipping_cost)
            sale_fee_fallback = _safe_float(s.fees)
            sale_label_fallback = _safe_float(getattr(s, "shipping_label_cost", None))

            order_fee_total = sale_fee_fallback
            order_shipping_charged_total = charged_shipping_sale
            order_shipping_actual_total = sale_label_fallback
            weight = 1.0

            if s.order_id is not None and int(s.order_id) in order_by_id:
                order = order_by_id[int(s.order_id)]
                siblings = sales_by_order_id.get(int(s.order_id), [])
                sibling_gross_total = sum(_safe_float(x.sold_price) for x in siblings)
                sibling_fee_total = sum(_safe_float(x.fees) for x in siblings)
                sibling_shipping_charged_total = sum(_safe_float(x.shipping_cost) for x in siblings)
                sibling_shipping_label_total = sum(
                    _safe_float(getattr(x, "shipping_label_cost", None)) for x in siblings
                )
                if sibling_gross_total > 0:
                    weight = sold_price / sibling_gross_total
                elif len(siblings) > 0:
                    weight = 1.0 / float(len(siblings))
                order_fee_total = _safe_float(order.fees) or sibling_fee_total
                order_shipping_charged_total = _safe_float(order.shipping_cost) or sibling_shipping_charged_total
                order_shipping_actual_total = (
                    _safe_float(getattr(order, "shipping_label_cost", None)) or sibling_shipping_label_total
                )

            allocated_fee_actual = order_fee_total * weight
            allocated_shipping_charged = order_shipping_charged_total * weight
            allocated_shipping_actual = order_shipping_actual_total * weight
            net_before_cogs_actual = (
                sold_price + allocated_shipping_charged - allocated_fee_actual - allocated_shipping_actual
            )

            actual_econ_rows.append(
                {
                    "sale_id": int(s.id),
                    "order_id": int(s.order_id) if s.order_id is not None else None,
                    "marketplace": str(s.marketplace or "").strip().lower(),
                    "external_order_id": str(s.external_order_id or "").strip(),
                    "sku": (
                        s.product.sku
                        if s.product
                        else (
                            s.listing.product.sku
                            if s.listing and s.listing.product
                            else None
                        )
                    ),
                    "product_title": (
                        s.product.title
                        if s.product
                        else (
                            s.listing.product.title
                            if s.listing and s.listing.product
                            else None
                        )
                    ),
                    "qty": int(s.quantity_sold or 0),
                    "sold_price": round(sold_price, 2),
                    "allocation_weight": round(weight, 6),
                    "order_fee_total_actual": round(order_fee_total, 2),
                    "order_shipping_charged_total": round(order_shipping_charged_total, 2),
                    "order_shipping_actual_total": round(order_shipping_actual_total, 2),
                    "allocated_fee_actual": round(allocated_fee_actual, 2),
                    "allocated_shipping_charged": round(allocated_shipping_charged, 2),
                    "allocated_shipping_actual": round(allocated_shipping_actual, 2),
                    "shipping_delta_charged_minus_actual": round(
                        allocated_shipping_charged - allocated_shipping_actual,
                        2,
                    ),
                    "net_before_cogs_actual": round(net_before_cogs_actual, 2),
                }
            )
    if hasattr(repo, "report_economics_intelligence_fact_rows"):
        try:
            economics_intel_df = pd.DataFrame(
                repo.report_economics_intelligence_fact_rows(
                    start_dt=start_dt,
                    end_dt=end_dt,
                    marketplaces={str(m or "").strip().lower() for m in sales_df.get("marketplace", []) if str(m or "").strip()},
                )
            )
            if not economics_intel_df.empty:
                economics_intel_df["sold_at"] = economics_intel_df["sold_at"].apply(iso_or_none)
        except Exception:
            _rollback_report_session()
            economics_intel_df = pd.DataFrame()
    for row in actual_econ_rows:
        sid = int(_safe_float(row.get("sale_id") or 0))
        if sid <= 0:
            continue
        actual_econ_by_sale_id[sid] = {
            "allocated_fee_actual": _safe_float(row.get("allocated_fee_actual")),
            "allocated_shipping_charged": _safe_float(row.get("allocated_shipping_charged")),
            "allocated_shipping_actual": _safe_float(row.get("allocated_shipping_actual")),
            "net_before_cogs_actual": _safe_float(row.get("net_before_cogs_actual")),
            "actual_fee_source": str(row.get("actual_fee_source") or "").strip(),
            "actual_shipping_source": str(row.get("actual_shipping_source") or "").strip(),
        }
    order_actual_econ_df = pd.DataFrame(actual_econ_rows)
    qbo_sales_df = pd.DataFrame(
        _build_qbo_sales_export_rows(
            sales,
            fifo_unit_cost_by_sale,
            fifo_unit_cost_source_by_sale=fifo_unit_cost_source_by_sale,
            actual_econ_by_sale_id=actual_econ_by_sale_id,
        )
    )

    cogs_margin_rows: list[dict[str, object]] = []
    sale_fifo_cogs_evidence_rows: list[dict[str, object]] = []
    for s in sales:
        bundle_summary = _sale_listing_bundle_summary(s)
        sale_product_id = getattr(s, "product_id", None)
        if sale_product_id is None:
            sale_product = getattr(s, "product", None)
            sale_product_id = getattr(sale_product, "id", None)
        lot_unit_cost = _safe_float(
            lot_weighted_unit_cost_by_product.get(int(sale_product_id))
            if sale_product_id is not None
            else 0.0
        )
        qty_sold = int(getattr(s, "quantity_sold", 0) or 0)
        fifo_cogs = _safe_float(fifo_unit_cost_by_sale.get(s.id)) * qty_sold
        sale_cogs_evidence = list(fifo_cogs_evidence_by_sale.get(int(s.id)) or [])
        lot_cogs = lot_unit_cost * qty_sold
        fifo_cost_source = str(fifo_unit_cost_source_by_sale.get(int(s.id)) or "unknown").strip() or "unknown"
        lot_cost_source = (
            str(lot_weighted_unit_cost_source_by_product.get(int(sale_product_id)) or "unknown").strip()
            if sale_product_id is not None
            else "unknown"
        ) or "unknown"
        field_net_before_cogs = (
            _safe_float(s.sold_price)
            + _safe_float(s.shipping_cost)
            - _safe_float(s.fees)
            - _safe_float(getattr(s, "shipping_label_cost", None))
        )
        actual_net_before_cogs = _safe_float(
            actual_econ_by_sale_id.get(int(s.id), {}).get("net_before_cogs_actual")
        )
        net_before_cogs = actual_net_before_cogs if actual_econ_by_sale_id.get(int(s.id)) else field_net_before_cogs
        cogs_margin_rows.append(
            {
                "sale_id": s.id,
                "sold_at": iso_or_none(s.sold_at),
                "marketplace": s.marketplace,
                "sku": s.product.sku if s.product else None,
                "product_title": s.product.title if s.product else None,
                "quantity": qty_sold,
                **bundle_summary,
                "gross_sales": _safe_float(s.sold_price),
                "fees": _safe_float(s.fees),
                "shipping_cost": _safe_float(s.shipping_cost),
                "field_net_before_cogs": field_net_before_cogs,
                "net_before_cogs": net_before_cogs,
                "actual_fee_alloc": _safe_float(actual_econ_by_sale_id.get(int(s.id), {}).get("allocated_fee_actual")),
                "actual_shipping_alloc": _safe_float(
                    actual_econ_by_sale_id.get(int(s.id), {}).get("allocated_shipping_actual")
                ),
                "actual_net_before_cogs": actual_net_before_cogs,
                "fifo_unit_cost": _safe_float(fifo_unit_cost_by_sale.get(s.id)),
                "fifo_cost_source": fifo_cost_source,
                "fifo_cogs_evidence_rows": int(len(sale_cogs_evidence)),
                "fifo_cogs": fifo_cogs,
                "fifo_margin": net_before_cogs - fifo_cogs,
                "lot_unit_cost": lot_unit_cost,
                "lot_cost_source": lot_cost_source,
                "lot_cogs": lot_cogs,
                "lot_margin": net_before_cogs - lot_cogs,
                "fifo_margin_actual": (
                    _safe_float(actual_econ_by_sale_id.get(int(s.id), {}).get("net_before_cogs_actual"))
                    - fifo_cogs
                ),
                "lot_margin_actual": (
                    _safe_float(actual_econ_by_sale_id.get(int(s.id), {}).get("net_before_cogs_actual"))
                    - lot_cogs
                ),
                "margin_method_delta": (net_before_cogs - fifo_cogs) - (net_before_cogs - lot_cogs),
            }
        )
        for evidence_idx, evidence in enumerate(sale_cogs_evidence, start=1):
            sale_fifo_cogs_evidence_rows.append(
                {
                    "sale_id": s.id,
                    "sold_at": iso_or_none(s.sold_at),
                    "marketplace": s.marketplace,
                    "external_order_id": getattr(s, "external_order_id", None),
                    "sku": s.product.sku if s.product else None,
                    "product_title": s.product.title if s.product else None,
                    "sale_quantity": qty_sold,
                    "allocation_index": evidence_idx,
                    "evidence_product_id": evidence.get("product_id"),
                    "lot_id": evidence.get("lot_id"),
                    "assignment_id": evidence.get("assignment_id"),
                    "quantity": int(evidence.get("quantity") or 0),
                    "unit_cost": _safe_float(evidence.get("unit_cost")),
                    "total_cost": _safe_float(evidence.get("total_cost")),
                    "cost_source": str(evidence.get("cost_source") or "unknown").strip() or "unknown",
                }
            )

    cogs_margin_df = pd.DataFrame(cogs_margin_rows)
    sale_fifo_cogs_evidence_df = pd.DataFrame(sale_fifo_cogs_evidence_rows)
    cogs_source_summary_df = _build_cogs_source_summary(cogs_margin_df)

    margin_by_sku_df = (
        cogs_margin_df.groupby(["sku", "product_title", "marketplace"], dropna=False, as_index=False)
        .agg(
            quantity=("quantity", "sum"),
            gross_sales=("gross_sales", "sum"),
            fees=("fees", "sum"),
            shipping_cost=("shipping_cost", "sum"),
            actual_fee_alloc=("actual_fee_alloc", "sum"),
            actual_shipping_alloc=("actual_shipping_alloc", "sum"),
            actual_net_before_cogs=("actual_net_before_cogs", "sum"),
            fifo_cogs=("fifo_cogs", "sum"),
            lot_cogs=("lot_cogs", "sum"),
            fifo_margin=("fifo_margin", "sum"),
            lot_margin=("lot_margin", "sum"),
            fifo_margin_actual=("fifo_margin_actual", "sum"),
            lot_margin_actual=("lot_margin_actual", "sum"),
        )
        if not cogs_margin_df.empty
        else pd.DataFrame()
    )
    if not margin_by_sku_df.empty:
        margin_by_sku_df = _add_margin_pct_columns(margin_by_sku_df)

    margin_by_channel_df = (
        cogs_margin_df.groupby(["marketplace"], dropna=False, as_index=False)
        .agg(
            quantity=("quantity", "sum"),
            gross_sales=("gross_sales", "sum"),
            fees=("fees", "sum"),
            shipping_cost=("shipping_cost", "sum"),
            actual_fee_alloc=("actual_fee_alloc", "sum"),
            actual_shipping_alloc=("actual_shipping_alloc", "sum"),
            actual_net_before_cogs=("actual_net_before_cogs", "sum"),
            fifo_cogs=("fifo_cogs", "sum"),
            lot_cogs=("lot_cogs", "sum"),
            fifo_margin=("fifo_margin", "sum"),
            lot_margin=("lot_margin", "sum"),
            fifo_margin_actual=("fifo_margin_actual", "sum"),
            lot_margin_actual=("lot_margin_actual", "sum"),
        )
        if not cogs_margin_df.empty
        else pd.DataFrame()
    )
    if not margin_by_channel_df.empty:
        margin_by_channel_df = _add_margin_pct_columns(margin_by_channel_df)

    margin_by_period_df = pd.DataFrame()
    if not cogs_margin_df.empty:
        period_df = cogs_margin_df.copy()
        period_df["period_month"] = period_df["sold_at"].fillna("").astype(str).str.slice(0, 7)
        margin_by_period_df = (
            period_df.groupby(["period_month", "marketplace"], dropna=False, as_index=False)
            .agg(
                quantity=("quantity", "sum"),
                gross_sales=("gross_sales", "sum"),
                fees=("fees", "sum"),
                shipping_cost=("shipping_cost", "sum"),
                actual_fee_alloc=("actual_fee_alloc", "sum"),
                actual_shipping_alloc=("actual_shipping_alloc", "sum"),
                actual_net_before_cogs=("actual_net_before_cogs", "sum"),
                fifo_cogs=("fifo_cogs", "sum"),
                lot_cogs=("lot_cogs", "sum"),
                fifo_margin=("fifo_margin", "sum"),
                lot_margin=("lot_margin", "sum"),
                fifo_margin_actual=("fifo_margin_actual", "sum"),
                lot_margin_actual=("lot_margin_actual", "sum"),
            )
        )
        margin_by_period_df = _add_margin_pct_columns(margin_by_period_df)

    inventory_cycles_df = pd.DataFrame()
    rebuy_cost_trend_df = pd.DataFrame()
    review_activity_df = pd.DataFrame()
    review_summary_df = pd.DataFrame()
    listing_format_outcome_df = pd.DataFrame()
    inventory_cycle_summary_df = pd.DataFrame()
    ebay_fee_reconciliation_df = pd.DataFrame()
    ebay_fee_reconciliation_by_marketplace_df = pd.DataFrame()
    ebay_fee_actual_source_df = pd.DataFrame()
    ebay_fee_source_priority_df = pd.DataFrame()
    ebay_fee_source_priority_trend_df = pd.DataFrame()
    accounting_exceptions_df = pd.DataFrame()
    fee_reconciliation_summary: dict[str, float | int] = {}
    fee_source_counts: dict[str, int] = {}
    if load_inventory_cycle_analytics:
        try:
            inventory_cycle_rows = (
                repo.report_inventory_cycle_rows(end_dt=end_dt)
                if supports_cycle_rollup
                else _build_inventory_cycle_rows(
                    products=_get_all_products(),
                    movements=_get_all_movements(),
                    sales=_get_all_sales(),
                    actual_econ_by_sale_id=actual_econ_by_sale_id,
                )
            )
        except Exception:
            _rollback_report_session()
            inventory_cycle_rows = []
        inventory_cycles_df = pd.DataFrame(inventory_cycle_rows)
        inventory_cycle_summary_df = pd.DataFrame(
            _build_inventory_cycle_summary_rows(
                inventory_cycles_df.to_dict("records")
                if not inventory_cycles_df.empty
                else []
            )
        )
        try:
            rebuy_cost_trend_rows = (
                repo.report_rebuy_cost_trend_rows(end_dt=end_dt)
                if supports_rebuy_rollup
                else _build_rebuy_cost_trend_rows(
                    products=_get_all_products(),
                    assignments=_get_all_assignments(),
                    movements=_get_all_movements(),
                )
            )
        except Exception:
            _rollback_report_session()
            rebuy_cost_trend_rows = []
        rebuy_cost_trend_df = pd.DataFrame(rebuy_cost_trend_rows)
    if load_extended_analytics:
        try:
            review_activity_rows = (
                repo.report_listing_review_activity_rows(
                    start_dt=start_dt,
                    end_dt=end_dt,
                )
                if hasattr(repo, "report_listing_review_activity_rows")
                else _build_listing_review_activity_rows(
                    listings=_get_all_listings(),
                    start_dt=start_dt,
                    end_dt=end_dt,
                )
            )
        except Exception:
            _rollback_report_session()
            review_activity_rows = []
        review_activity_df = pd.DataFrame(review_activity_rows)
        if not review_activity_df.empty:
            review_summary_df = (
                review_activity_df.groupby(
                    ["reviewed_by", "marketplace", "review_decision"],
                    dropna=False,
                    as_index=False,
                )
                .size()
                .rename(columns={"size": "review_events"})
                .sort_values(["review_events"], ascending=[False])
            )
        try:
            listing_format_outcome_rows = (
                repo.report_listing_format_outcome_rows(
                    start_dt=start_dt,
                    end_dt=end_dt,
                )
                if hasattr(repo, "report_listing_format_outcome_rows")
                else _build_listing_format_outcome_rows(
                    listings=_get_all_listings(),
                    start_dt=start_dt,
                    end_dt=end_dt,
                )
            )
        except Exception:
            _rollback_report_session()
            listing_format_outcome_rows = []
        listing_format_outcome_df = pd.DataFrame(listing_format_outcome_rows)
        if hasattr(repo, "report_ebay_fee_reconciliation_rows"):
            try:
                ebay_fee_reconciliation_df = pd.DataFrame(
                    repo.report_ebay_fee_reconciliation_rows(
                        start_dt=start_dt,
                        end_dt=end_dt,
                    )
                )
            except Exception:
                _rollback_report_session()
                ebay_fee_reconciliation_df = pd.DataFrame()
        else:
            ebay_fee_reconciliation_df = pd.DataFrame(build_ebay_fee_reconciliation_rows(sales))
        if not ebay_fee_reconciliation_df.empty:
            fee_reconciliation_summary = _summarize_fee_reconciliation(ebay_fee_reconciliation_df)
            ebay_fee_reconciliation_by_marketplace_df = (
                ebay_fee_reconciliation_df.groupby(["fee_estimate_present"], dropna=False, as_index=False)
                .agg(
                    sales_count=("sale_id", "count"),
                    gross_sales=("sale_gross", "sum"),
                    actual_fee=("actual_fee", "sum"),
                    estimated_fee_scaled=("estimated_fee_scaled", "sum"),
                    variance_actual_minus_estimate=("variance_actual_minus_estimate", "sum"),
                )
            )
            ebay_fee_reconciliation_by_marketplace_df["variance_pct_of_estimate"] = ebay_fee_reconciliation_by_marketplace_df.apply(
                lambda row: (
                    (float(row["variance_actual_minus_estimate"]) / float(row["estimated_fee_scaled"]) * 100.0)
                    if float(row["estimated_fee_scaled"] or 0.0) > 0
                    else 0.0
                ),
                axis=1,
            )
            ebay_fee_reconciliation_by_marketplace_df["estimate_bucket"] = ebay_fee_reconciliation_by_marketplace_df[
                "fee_estimate_present"
            ].map({True: "estimate_present", False: "estimate_missing"})
            ebay_fee_reconciliation_by_marketplace_df = ebay_fee_reconciliation_by_marketplace_df[
                [
                    "estimate_bucket",
                    "sales_count",
                    "gross_sales",
                    "actual_fee",
                    "estimated_fee_scaled",
                    "variance_actual_minus_estimate",
                    "variance_pct_of_estimate",
                ]
            ]
            ebay_fee_actual_source_df = (
                ebay_fee_reconciliation_df.groupby(["actual_fee_source"], dropna=False, as_index=False)
                .agg(
                    sales_count=("sale_id", "count"),
                    actual_fee=("actual_fee", "sum"),
                    estimated_fee_scaled=("estimated_fee_scaled", "sum"),
                    variance_actual_minus_estimate=("variance_actual_minus_estimate", "sum"),
                )
                .sort_values(["sales_count"], ascending=[False])
            )
            ebay_fee_source_priority_df = _build_fee_source_priority_summary(ebay_fee_reconciliation_df)
            fee_source_counts = _fee_source_priority_counts(ebay_fee_source_priority_df)
            ebay_fee_source_priority_trend_df = _build_fee_source_priority_trend(ebay_fee_reconciliation_df)
        accounting_exception_rows, _ = _load_rollup_rows(
            "report_accounting_exception_rows",
            enabled=supports_accounting_exception_rollup,
            start_dt=start_dt,
            end_dt=end_dt,
        )
        accounting_exceptions_df = pd.DataFrame(accounting_exception_rows)

    report_scalar_cache = {
        "shipping_rows": int(len(shipping_economics_df)),
        "shipping_total_charged": (
            float(shipping_economics_df["shipping_charged_to_buyer"].sum())
            if not shipping_economics_df.empty
            else 0.0
        ),
        "shipping_total_label_spend": (
            float(shipping_economics_df["shipping_label_spend"].sum())
            if not shipping_economics_df.empty
            else 0.0
        ),
        "shipping_label_covered_count": (
            int((shipping_economics_df["shipping_label_spend"] > 0).sum())
            if not shipping_economics_df.empty
            else 0
        ),
        "cogs_gross_sales_total": (
            float(cogs_margin_df["gross_sales"].sum()) if not cogs_margin_df.empty else 0.0
        ),
        "cogs_fifo_margin_total": (
            float(cogs_margin_df["fifo_margin"].sum()) if not cogs_margin_df.empty else 0.0
        ),
        "cogs_lot_margin_total": (
            float(cogs_margin_df["lot_margin"].sum()) if not cogs_margin_df.empty else 0.0
        ),
        "cogs_negative_fifo_rows": (
            int((cogs_margin_df["fifo_margin"] < 0).sum()) if not cogs_margin_df.empty else 0
        ),
        "reconcile_flags": (
            int(reconciliation_df["reconcile_flag"].sum()) if not reconciliation_df.empty else 0
        ),
        "table_row_counts": {
            "sales": int(len(sales_df)),
            "tax_summary": int(len(tax_summary_df)),
            "tax_by_marketplace": int(len(tax_by_marketplace_df)),
            "tax_detail": int(len(tax_detail_df)),
            "tax_exceptions": int(len(tax_exceptions_df)),
            "inventory": int(len(inventory_df)),
            "listings": int(len(listings_df)),
            "orders": int(len(orders_df)),
            "order_items": int(len(order_items_df)),
            "returns": int(len(returns_df)),
            "movements": int(len(movements_df)),
            "reconciliation": int(len(reconciliation_df)),
            "accounting_validation": int(len(accounting_validation_df)),
            "accounting_exceptions": int(len(accounting_exceptions_df)),
        },
    }
    accounting_close_summary, accounting_close_checks_df = _build_accounting_close_readiness_summary(
        inventory_df=inventory_df,
        cogs_margin_df=cogs_margin_df,
        returns_df=returns_df,
        reconciliation_df=reconciliation_df,
        shipping_economics_df=shipping_economics_df,
        ebay_fee_source_priority_df=ebay_fee_source_priority_df,
        accounting_exceptions_df=accounting_exceptions_df,
        lot_allocation_source_summary_df=lot_allocation_source_summary_df,
        cogs_source_summary_df=cogs_source_summary_df,
        qbo_adjustments_df=qbo_adjustments_df,
    )
    accounting_close_formula_df = _build_accounting_close_formula_checks(accounting_close_summary)
    accounting_close_summary, accounting_close_checks_df = _apply_formula_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_close_formula_df,
    )
    accounting_sales_component_df = _build_accounting_sales_component_checks(
        sales_df=sales_df,
        cogs_margin_df=cogs_margin_df,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_sales_component_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_sales_component_df,
    )
    accounting_return_tieout_df = _build_accounting_return_tieout_checks(
        returns_df=returns_df,
        qbo_adjustments_df=qbo_adjustments_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_return_tieout_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_return_tieout_df,
    )
    accounting_inventory_valuation_df = _build_accounting_inventory_valuation_checks(
        inventory_df=inventory_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_inventory_valuation_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_inventory_valuation_df,
    )
    accounting_fee_evidence_df = _build_accounting_fee_evidence_checks(
        sales_df=sales_df,
        fee_reconciliation_df=ebay_fee_reconciliation_df,
        fee_source_priority_df=ebay_fee_source_priority_df,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_fee_evidence_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_fee_evidence_df,
    )
    accounting_shipping_evidence_df = _build_accounting_shipping_evidence_checks(
        sales_df=sales_df,
        shipping_economics_df=shipping_economics_df,
        shipping_econ_summary_df=shipping_econ_summary_df,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_shipping_evidence_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_shipping_evidence_df,
    )
    accounting_reconciliation_tieout_df = _build_accounting_reconciliation_tieout_checks(
        sales_df=sales_df,
        returns_df=returns_df,
        reconciliation_df=reconciliation_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_reconciliation_tieout_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_reconciliation_tieout_df,
    )
    accounting_cogs_source_df = _build_accounting_cogs_source_checks(
        cogs_margin_df=cogs_margin_df,
        cogs_source_summary_df=cogs_source_summary_df,
        sale_fifo_cogs_evidence_df=sale_fifo_cogs_evidence_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_cogs_source_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_cogs_source_df,
    )
    accounting_lot_allocation_df = _build_accounting_lot_allocation_checks(
        lots_df=lots_df,
        lot_allocation_source_summary_df=lot_allocation_source_summary_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_lot_allocation_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_lot_allocation_df,
    )
    accounting_exception_queue_checks_df = _build_accounting_exception_queue_checks(
        accounting_exceptions_df=accounting_exceptions_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_exception_queue_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_exception_queue_checks_df,
    )
    accounting_margin_anomaly_checks_df = _build_accounting_margin_anomaly_checks(
        cogs_margin_df=cogs_margin_df,
        accounting_exceptions_df=accounting_exceptions_df,
        close_summary=accounting_close_summary,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_margin_anomaly_checks_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_margin_anomaly_checks_df,
    )
    dashboard_live_drift_metrics: dict[str, object] = {}
    dashboard_compatible_start = (end_dt - timedelta(days=30)).date()
    if from_date == dashboard_compatible_start and hasattr(repo, "dashboard_live_metrics"):
        try:
            dashboard_live_drift_metrics = dict(
                repo.dashboard_live_metrics(now=end_dt, include_fee_type_breakdown=False) or {}
            )
        except TypeError:
            try:
                dashboard_live_drift_metrics = dict(repo.dashboard_live_metrics(now=end_dt) or {})
            except Exception:
                _rollback_report_session()
                dashboard_live_drift_metrics = {}
        except Exception:
            _rollback_report_session()
            dashboard_live_drift_metrics = {}
    slack_summary_drift_metrics: dict[str, object] = {}
    if from_date == (end_dt - timedelta(days=1)).date():
        slack_summary_drift_metrics = _build_slack_summary_drift_metrics(
            cogs_margin_df,
            window_label="daily",
            returns_df=returns_df,
            qbo_adjustments_df=qbo_adjustments_df,
        )
    elif from_date == (end_dt - timedelta(days=7)).date():
        slack_summary_drift_metrics = _build_slack_summary_drift_metrics(
            cogs_margin_df,
            window_label="weekly",
            returns_df=returns_df,
            qbo_adjustments_df=qbo_adjustments_df,
        )
    ai_accounting_drift_metrics: dict[str, object] = {}
    if from_date == dashboard_compatible_start:
        ai_accounting_drift_metrics = _build_ai_accounting_snapshot_drift_metrics(
            cogs_margin_df,
            window_label="30d",
            returns_df=returns_df,
            qbo_adjustments_df=qbo_adjustments_df,
        )
    accounting_period_drift_df = _build_accounting_period_drift_checks(
        close_summary=accounting_close_summary,
        qbo_sales_df=qbo_sales_df,
        qbo_adjustments_df=qbo_adjustments_df,
        dashboard_live_metrics=dashboard_live_drift_metrics,
        slack_summary_metrics=slack_summary_drift_metrics,
        ai_accounting_snapshot_metrics=ai_accounting_drift_metrics,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_period_drift_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        accounting_period_drift_df,
    )
    accounting_close_summary, accounting_close_checks_df = _apply_ai_review_outcomes_to_close_readiness(
        accounting_close_summary,
        accounting_close_checks_df,
        ai_review_outcomes_df,
    )
    accounting_close_consistency_checks_df = _build_accounting_close_consistency_checks(
        close_summary=accounting_close_summary,
        close_checks_df=accounting_close_checks_df,
    )
    accounting_close_report_frames = {
        "accounting_close_readiness_checks": accounting_close_checks_df,
        "accounting_close_formula_checks": accounting_close_formula_df,
        "accounting_sales_component_checks": accounting_sales_component_df,
        "accounting_return_tieout_checks": accounting_return_tieout_df,
        "accounting_inventory_valuation_checks": accounting_inventory_valuation_df,
        "accounting_fee_evidence_checks": accounting_fee_evidence_df,
        "accounting_shipping_evidence_checks": accounting_shipping_evidence_df,
        "accounting_reconciliation_tieout_checks": accounting_reconciliation_tieout_df,
        "accounting_cogs_source_checks": accounting_cogs_source_df,
        "accounting_lot_allocation_checks": accounting_lot_allocation_df,
        "accounting_exception_queue_checks": accounting_exception_queue_checks_df,
        "accounting_margin_anomaly_checks": accounting_margin_anomaly_checks_df,
        "accounting_close_consistency_checks": accounting_close_consistency_checks_df,
        "accounting_period_drift_checks": accounting_period_drift_df,
        "inventory_snapshot": inventory_df,
        "sales_detail": sales_df,
        "cogs_margin_detail": cogs_margin_df,
        "sale_fifo_cogs_evidence": sale_fifo_cogs_evidence_df,
        "qbo_sales_export": qbo_sales_df,
    }
    accounting_close_packet_completeness_df = _build_accounting_close_packet_completeness_checks(
        report_frames=accounting_close_report_frames,
        close_summary=accounting_close_summary,
    )
    accounting_close_report_frames["accounting_close_packet_completeness_checks"] = (
        accounting_close_packet_completeness_df
    )
    accounting_close_report_frames.update(
        {
            "accounting_period_drift_checks": accounting_period_drift_df,
            "accounting_close_signoffs": accounting_close_signoff_df,
            "accounting_exception_queue": accounting_exceptions_df,
            "margin_by_sku": margin_by_sku_df,
            "reconciliation_marketplace": reconciliation_df,
            "shipping_economics_detail": shipping_economics_df,
            "shipping_economics_summary": shipping_econ_summary_df,
            "ebay_fee_estimate_vs_actual": ebay_fee_reconciliation_df,
            "ebay_fee_source_priority": ebay_fee_source_priority_df,
            "tax_summary_estimated": tax_summary_df,
            "tax_by_marketplace_estimated": tax_by_marketplace_df,
            "tax_detail_estimated": tax_detail_df,
            "tax_exceptions_advisor_review": tax_exceptions_df,
            "tax_reporting_signoffs": tax_signoff_df,
            "ai_review_outcomes": ai_review_outcomes_df,
            "lot_assignment": lots_df,
            "lot_allocation_source_summary": lot_allocation_source_summary_df,
            "cogs_source_summary": cogs_source_summary_df,
            "returns": returns_df,
            "qbo_adjustments_export": qbo_adjustments_df,
        }
    )
    current_accounting_close_packet_hash = _accounting_close_packet_evidence_hash_from_frames(
        report_frames=accounting_close_report_frames,
        close_summary=accounting_close_summary,
        from_date=from_date,
        to_date=to_date,
    )
    accounting_close_packet_evidence_hash_df = _build_accounting_close_packet_evidence_hash_rows(
        evidence_hash=current_accounting_close_packet_hash,
        from_date=from_date,
        to_date=to_date,
    )
    accounting_close_report_frames["accounting_close_packet_evidence_hash"] = accounting_close_packet_evidence_hash_df
    accounting_close_signoff_review_df = _build_accounting_close_signoff_review(
        signoff_df=accounting_close_signoff_df,
        close_summary=accounting_close_summary,
        from_date=from_date,
        to_date=to_date,
        current_packet_hash=current_accounting_close_packet_hash,
    )
    accounting_close_report_frames["accounting_close_signoff_review"] = accounting_close_signoff_review_df
    accounting_close_packet_manifest_df = _build_accounting_close_packet_manifest_checks(
        report_frames=accounting_close_report_frames,
    )
    accounting_close_report_frames["accounting_close_packet_manifest_checks"] = accounting_close_packet_manifest_df
    accounting_close_packet_hash_df = _build_accounting_close_packet_hash_checks(
        report_frames=accounting_close_report_frames,
    )
    accounting_close_report_frames["accounting_close_packet_hash_checks"] = accounting_close_packet_hash_df

    st.markdown("### Accounting Review / Close Readiness")
    close_status = str(accounting_close_summary.get("readiness_status") or "review_needed")
    if close_status == "close_ready":
        st.success("Close readiness: no blocking accounting checks found for the selected range.")
    elif close_status == "blocked":
        st.error("Close readiness: blocked by accounting exceptions or reconciliation flags.")
    else:
        st.warning("Close readiness: review needed before treating this range as close-ready.")
    cr1, cr2, cr3, cr4 = st.columns(4)
    cr1.metric("Inventory Value", f"${float(accounting_close_summary.get('inventory_value') or 0.0):,.2f}")
    cr2.metric("Sales Net Before COGS", f"${float(accounting_close_summary.get('net_before_cogs') or 0.0):,.2f}")
    cr3.metric("FIFO COGS", f"${float(accounting_close_summary.get('fifo_cogs') or 0.0):,.2f}")
    cr4.metric(
        "Est. Profit After Returns",
        f"${_accounting_close_estimated_profit_after_returns(accounting_close_summary):,.2f}",
    )
    cr5, cr6, cr7, cr8 = st.columns(4)
    cr5.metric(
        "Profit Before Returns",
        f"${_accounting_close_profit_before_returns(accounting_close_summary):,.2f}",
    )
    cr6.metric("Return Refunds", f"${float(accounting_close_summary.get('returns_refund_total') or 0.0):,.2f}")
    cr7.metric(
        "Return COGS Reversal",
        f"${float(accounting_close_summary.get('returns_cogs_reversal_total') or 0.0):,.2f}",
    )
    cr8.metric(
        "Return Profit Impact",
        f"${float(accounting_close_summary.get('returns_estimated_profit_impact') or 0.0):,.2f}",
    )
    cr9, cr10, cr11, cr12 = st.columns(4)
    cr9.metric("P0 Exceptions", f"{int(accounting_close_summary.get('p0_exceptions') or 0)}")
    cr10.metric("P1 Exceptions", f"{int(accounting_close_summary.get('p1_exceptions') or 0)}")
    cr11.metric("Reconcile Flags", f"{int(accounting_close_summary.get('reconcile_flags') or 0)}")
    cr12.metric(
        "Label Coverage",
        f"{float(accounting_close_summary.get('shipping_label_coverage_pct') or 0.0):.1f}%",
    )
    if str(accounting_close_summary.get("blockers") or "").strip():
        st.caption("Blockers: " + str(accounting_close_summary.get("blockers")))
    if str(accounting_close_summary.get("warnings") or "").strip():
        st.caption("Warnings: " + str(accounting_close_summary.get("warnings")))
    st.caption(f"Accounting close packet evidence hash: `{current_accounting_close_packet_hash}`")
    returns_cogs_reversal_total = float(accounting_close_summary.get("returns_cogs_reversal_total") or 0.0)
    if returns_cogs_reversal_total > 0:
        st.caption(
            "Returns include estimated COGS reversal for restocked items: "
            f"${returns_cogs_reversal_total:,.2f}; "
            "net after returns and COGS includes this reversal."
        )
    close_profit_summary = {
        "gross_sales_total": float(accounting_close_summary.get("gross_sales") or 0.0),
        "net_before_cogs_total": float(accounting_close_summary.get("net_before_cogs") or 0.0),
        "fifo_cogs_total": float(accounting_close_summary.get("fifo_cogs") or 0.0),
        "profit_before_returns": _accounting_close_profit_before_returns(accounting_close_summary),
        "returns_refund_total": float(accounting_close_summary.get("returns_refund_total") or 0.0),
        "returns_cogs_reversal_total": float(accounting_close_summary.get("returns_cogs_reversal_total") or 0.0),
        "returns_profit_impact": float(accounting_close_summary.get("returns_estimated_profit_impact") or 0.0),
        "estimated_profit_after_returns": _accounting_close_estimated_profit_after_returns(accounting_close_summary),
        "profit_formula": (
            "estimated_profit_after_returns = profit_before_returns - returns_refund_total "
            "+ returns_cogs_reversal_total"
        ),
    }
    with st.expander("Accounting Field Semantics", expanded=False):
        st.caption(
            "Canonical cost-basis rules used by dashboard, Reports, close packet exports, and accounting review."
        )
        _render_df_with_preview(pd.DataFrame(ACCOUNTING_FIELD_SEMANTICS_ROWS), hide_index=True)
    with st.expander("Close Readiness Checks", expanded=False):
        _render_df_with_preview(accounting_close_checks_df, hide_index=True)
    with st.expander("Close Consistency Checks", expanded=False):
        if accounting_close_consistency_checks_df.empty:
            st.caption("No close consistency checks are available for the selected window.")
        else:
            close_consistency_warn_count = int(
                (accounting_close_consistency_checks_df["status"].astype(str) == "warn").sum()
            )
            if close_consistency_warn_count:
                st.warning(f"{close_consistency_warn_count} close consistency check(s) need review.")
            else:
                st.caption("Close readiness status, blockers, warnings, and check rows are internally consistent.")
            _render_df_with_preview(accounting_close_consistency_checks_df, hide_index=True)
    with st.expander("Close Packet Completeness Checks", expanded=False):
        if accounting_close_packet_completeness_df.empty:
            st.caption("No close packet completeness checks are available for the selected window.")
        else:
            packet_completeness_warn_count = int(
                (accounting_close_packet_completeness_df["status"].astype(str) == "warn").sum()
            )
            if packet_completeness_warn_count:
                st.warning(f"{packet_completeness_warn_count} close packet artifact(s) need review.")
            else:
                st.caption("Required close-packet evidence tables are present for the selected window.")
            _render_df_with_preview(accounting_close_packet_completeness_df, hide_index=True)
    with st.expander("Close Packet Manifest Checks", expanded=False):
        if accounting_close_packet_manifest_df.empty:
            st.caption("No close packet manifest checks are available for the selected window.")
        else:
            packet_manifest_warn_count = int(
                (accounting_close_packet_manifest_df["status"].astype(str) == "warn").sum()
            )
            if packet_manifest_warn_count:
                st.warning(f"{packet_manifest_warn_count} close packet manifest artifact(s) need review.")
            else:
                st.caption("Close packet manifest row counts match the selected export dataframes.")
            _render_df_with_preview(accounting_close_packet_manifest_df, hide_index=True)
    with st.expander("Close Packet Hash Checks", expanded=False):
        if accounting_close_packet_hash_df.empty:
            st.caption("No close packet hash checks are available for the selected window.")
        else:
            packet_hash_warn_count = int((accounting_close_packet_hash_df["status"].astype(str) == "warn").sum())
            if packet_hash_warn_count:
                st.warning(f"{packet_hash_warn_count} close packet hash artifact(s) need review.")
            else:
                st.caption("Close packet CSV hashes are available for selected export dataframes.")
            _render_df_with_preview(accounting_close_packet_hash_df, hide_index=True)
    with st.expander("Close Packet Evidence Hash", expanded=False):
        _render_df_with_preview(accounting_close_packet_evidence_hash_df, hide_index=True)
    with st.expander("Accounting Formula Checks", expanded=False):
        if accounting_close_formula_df.empty:
            st.caption("No accounting formula checks are available for the selected window.")
        else:
            formula_warn_count = int((accounting_close_formula_df["status"].astype(str) == "warn").sum())
            if formula_warn_count:
                st.warning(f"{formula_warn_count} accounting formula check(s) need review.")
            else:
                st.caption("Core close arithmetic ties out for the selected window.")
            _render_df_with_preview(accounting_close_formula_df, hide_index=True)
    with st.expander("Sales Component Tie-Out Checks", expanded=False):
        if accounting_sales_component_df.empty:
            st.caption("No sales component tie-out checks are available for the selected window.")
        else:
            component_warn_count = int((accounting_sales_component_df["status"].astype(str) == "warn").sum())
            if component_warn_count:
                st.warning(f"{component_warn_count} sales component tie-out check(s) need review.")
            else:
                st.caption("Sales Detail components tie to COGS & Margin close totals for the selected window.")
            _render_df_with_preview(accounting_sales_component_df, hide_index=True)
    with st.expander("Return Tie-Out Checks", expanded=False):
        if accounting_return_tieout_df.empty:
            st.caption("No return tie-out checks are available for the selected window.")
        else:
            return_warn_count = int((accounting_return_tieout_df["status"].astype(str) == "warn").sum())
            if return_warn_count:
                st.warning(f"{return_warn_count} return tie-out check(s) need review.")
            else:
                st.caption("Return/refund totals tie to QBO adjustment staging and close return impact.")
            _render_df_with_preview(accounting_return_tieout_df, hide_index=True)
    with st.expander("Inventory Valuation Checks", expanded=False):
        if accounting_inventory_valuation_df.empty:
            st.caption("No inventory valuation checks are available for the selected window.")
        else:
            valuation_warn_count = int((accounting_inventory_valuation_df["status"].astype(str) == "warn").sum())
            if valuation_warn_count:
                st.warning(f"{valuation_warn_count} inventory valuation check(s) need review.")
            else:
                st.caption("Inventory Snapshot valuation ties to close readiness for stocked items.")
            _render_df_with_preview(accounting_inventory_valuation_df, hide_index=True)
    with st.expander("Fee Evidence Checks", expanded=False):
        if accounting_fee_evidence_df.empty:
            st.caption("No fee evidence checks are available for the selected window.")
        else:
            fee_warn_count = int((accounting_fee_evidence_df["status"].astype(str) == "warn").sum())
            if fee_warn_count:
                st.warning(f"{fee_warn_count} fee evidence check(s) need review.")
            else:
                st.caption("Fee reconciliation and source-priority evidence tie to Sales Detail.")
            _render_df_with_preview(accounting_fee_evidence_df, hide_index=True)
    with st.expander("Shipping Evidence Checks", expanded=False):
        if accounting_shipping_evidence_df.empty:
            st.caption("No shipping evidence checks are available for the selected window.")
        else:
            shipping_warn_count = int((accounting_shipping_evidence_df["status"].astype(str) == "warn").sum())
            if shipping_warn_count:
                st.warning(f"{shipping_warn_count} shipping evidence check(s) need review.")
            else:
                st.caption("Shipping charged, label spend, and shipping delta tie across Sales Detail and Shipping Economics.")
            _render_df_with_preview(accounting_shipping_evidence_df, hide_index=True)
    with st.expander("Reconciliation Tie-Out Checks", expanded=False):
        if accounting_reconciliation_tieout_df.empty:
            st.caption("No reconciliation tie-out checks are available for the selected window.")
        else:
            reconciliation_warn_count = int(
                (accounting_reconciliation_tieout_df["status"].astype(str) == "warn").sum()
            )
            if reconciliation_warn_count:
                st.warning(f"{reconciliation_warn_count} reconciliation tie-out check(s) need review.")
            else:
                st.caption("Marketplace reconciliation totals tie to Sales Detail, Returns, and close flags.")
            _render_df_with_preview(accounting_reconciliation_tieout_df, hide_index=True)
    with st.expander("COGS Source Checks", expanded=False):
        if accounting_cogs_source_df.empty:
            st.caption("No COGS source checks are available for the selected window.")
        else:
            cogs_source_warn_count = int((accounting_cogs_source_df["status"].astype(str) == "warn").sum())
            if cogs_source_warn_count:
                st.warning(f"{cogs_source_warn_count} COGS source check(s) need review.")
            else:
                st.caption("Sold COGS source summary ties to COGS & Margin and close readiness.")
            _render_df_with_preview(accounting_cogs_source_df, hide_index=True)
    with st.expander("Lot Allocation Checks", expanded=False):
        if accounting_lot_allocation_df.empty:
            st.caption("No lot allocation checks are available for the selected window.")
        else:
            lot_allocation_warn_count = int((accounting_lot_allocation_df["status"].astype(str) == "warn").sum())
            if lot_allocation_warn_count:
                st.warning(f"{lot_allocation_warn_count} lot allocation check(s) need review.")
            else:
                st.caption("Lot Allocation Source Summary ties to Lot Assignment detail and close readiness.")
            _render_df_with_preview(accounting_lot_allocation_df, hide_index=True)
    with st.expander("Exception Queue Checks", expanded=False):
        if accounting_exception_queue_checks_df.empty:
            st.caption("No exception queue checks are available for the selected window.")
        else:
            exception_queue_warn_count = int(
                (accounting_exception_queue_checks_df["status"].astype(str) == "warn").sum()
            )
            if exception_queue_warn_count:
                st.warning(f"{exception_queue_warn_count} exception queue check(s) need review.")
            else:
                st.caption("Exception queue severity counts tie to close readiness.")
            _render_df_with_preview(accounting_exception_queue_checks_df, hide_index=True)
    with st.expander("Margin Anomaly Checks", expanded=False):
        if accounting_margin_anomaly_checks_df.empty:
            st.caption("No margin anomaly checks are available for the selected window.")
        else:
            margin_anomaly_warn_count = int(
                (accounting_margin_anomaly_checks_df["status"].astype(str) == "warn").sum()
            )
            if margin_anomaly_warn_count:
                st.warning(f"{margin_anomaly_warn_count} margin anomaly check(s) need review.")
            else:
                st.caption("COGS & Margin anomalies tie to close readiness and exception queue evidence.")
            _render_df_with_preview(accounting_margin_anomaly_checks_df, hide_index=True)
    with st.expander("Close Period Drift Checks", expanded=False):
        if accounting_period_drift_df.empty:
            st.caption("No close-period drift checks are available for the selected window.")
        else:
            drift_warn_count = int((accounting_period_drift_df["status"].astype(str) == "warn").sum())
            if drift_warn_count:
                st.warning(f"{drift_warn_count} close-period drift check(s) need review.")
            else:
                st.caption("Dashboard/report export close-period totals are aligned for available checks.")
            if dashboard_live_drift_metrics:
                st.caption("Includes Dashboard Live Metrics 30-day comparisons for this selected window.")
            else:
                st.caption(
                    "Dashboard Live Metrics comparisons are included only when the selected range matches "
                    "the dashboard 30-day window ending on `To Date`."
                )
            if slack_summary_drift_metrics:
                st.caption("Includes Slack-style daily/weekly business summary comparisons for this selected window.")
            else:
                st.caption(
                    "Slack summary comparisons are included only when the selected range matches the daily or weekly "
                    "business summary window ending on `To Date`."
                )
            _render_df_with_preview(accounting_period_drift_df, hide_index=True)
    with st.expander("Close Sign-Off Evidence Review", expanded=False):
        _render_df_with_preview(accounting_close_signoff_review_df, hide_index=True)
    with st.expander("AI Review Outcome Evidence", expanded=False):
        if ai_review_outcomes_df.empty:
            st.caption("No Reports Copilot or AI Accountant outcome audit events have been recorded yet.")
        else:
            _render_df_with_preview(ai_review_outcomes_df, hide_index=True)
    with st.expander("Record Accounting Close Sign-Off", expanded=False):
        current_close_packet_ref = f"accounting_close_packet_{from_date}_{to_date}.zip"
        st.caption(
            "Record sign-off only after downloading and reviewing the Accounting Close Packet for this exact date range."
        )
        st.code(current_accounting_close_packet_hash or "packet hash unavailable", language="text")
        st.caption(
            "Captured with this sign-off: "
            f"readiness={accounting_close_summary.get('readiness_status') or 'unknown'}, "
            f"blockers={int(_safe_float(accounting_close_summary.get('blocker_count')))}, "
            f"drift warnings={int(_safe_float(accounting_close_summary.get('period_drift_warn_count')))}, "
            f"AI review follow-ups={int(_safe_float(accounting_close_summary.get('ai_review_followup_count')))}."
        )
        can_record_close_signoff = has_permission(user.role, "manage_settings")
        signoff_form_context = (
            st.form("reports_accounting_close_signoff_form")
            if hasattr(st, "form")
            else st.expander("Sign-Off Fields", expanded=True)
        )
        with signoff_form_context:
            cs1, cs2 = st.columns(2)
            with cs1:
                close_signoff_period = st.text_input(
                    "Close Period",
                    value=_default_accounting_close_period(from_date, to_date),
                    key="reports_accounting_close_signoff_period",
                )
                close_signoff_owner = st.text_input(
                    "Owner",
                    value=str(user.username or ""),
                    key="reports_accounting_close_signoff_owner",
                )
                close_signoff_date = st.date_input(
                    "Sign-Off Date",
                    value=utc_today(),
                    key="reports_accounting_close_signoff_date",
                )
            with cs2:
                close_signoff_status = st.selectbox(
                    "Status",
                    options=["approved", "blocked", "needs_followup"],
                    index=0 if close_status == "close_ready" else 2,
                    key="reports_accounting_close_signoff_status",
                )
                close_signoff_packet_ref = st.text_input(
                    "Accounting Packet Ref",
                    value=current_close_packet_ref,
                    key="reports_accounting_close_signoff_packet_ref",
                )
                close_signoff_evidence_link = st.text_input(
                    "Evidence Link",
                    placeholder="ticket/runbook/shared artifact URL",
                    key="reports_accounting_close_signoff_evidence_link",
                )
            if hasattr(st, "text_area"):
                close_signoff_notes = st.text_area(
                    "Notes",
                    placeholder="Reviewer notes, unresolved follow-up, or approval context.",
                    key="reports_accounting_close_signoff_notes",
                )
            else:
                close_signoff_notes = st.text_input(
                    "Notes",
                    value="",
                    key="reports_accounting_close_signoff_notes",
                )
            if hasattr(st, "form_submit_button"):
                save_close_signoff = st.form_submit_button(
                    "Record Accounting Close Sign-Off",
                    disabled=not can_record_close_signoff,
                )
            else:
                save_close_signoff = st.button(
                    "Record Accounting Close Sign-Off",
                    key="reports_accounting_close_signoff_submit_btn",
                    disabled=not can_record_close_signoff,
                )
        if not can_record_close_signoff:
            st.caption("You need `manage_settings` permission to record accounting close sign-off evidence.")
        if save_close_signoff:
            try:
                payload = _build_accounting_close_signoff_payload(
                    target_env=settings.app_env,
                    close_period=close_signoff_period,
                    status=close_signoff_status,
                    owner=close_signoff_owner,
                    signoff_date=close_signoff_date,
                    close_summary=accounting_close_summary,
                    accounting_packet_ref=close_signoff_packet_ref,
                    accounting_packet_hash=current_accounting_close_packet_hash,
                    evidence_link=close_signoff_evidence_link,
                    notes=close_signoff_notes,
                )
                repo.record_audit_event(
                    entity_type="accounting_close_signoff",
                    entity_id=None,
                    action="signoff",
                    actor=user.username,
                    changes=payload,
                )
                st.success("Accounting close sign-off recorded.")
                st.rerun()
            except Exception as exc:
                st.error(f"Unable to record accounting close sign-off: {exc}")
    with st.expander("Lot Allocation Source Summary", expanded=False):
        if lot_allocation_source_summary_df.empty:
            st.caption("No lot allocation source summary is available for the selected window.")
        else:
            explicit_sources = {
                "assignment_unit_landed_cost",
                "assignment_allocated_landed_cost",
                "lot_allocation_weight",
            }
            fallback_sources = {
                "lot_expected_quantity_fallback",
                "lot_equal_quantity_fallback",
                "missing_cost_basis",
                "unknown",
            }
            explicit_total = float(
                lot_allocation_source_summary_df[
                    lot_allocation_source_summary_df["cost_source"].isin(explicit_sources)
                ]["resolved_landed_total_cost"].sum()
            )
            fallback_total = float(
                lot_allocation_source_summary_df[
                    lot_allocation_source_summary_df["cost_source"].isin(fallback_sources)
                ]["resolved_landed_total_cost"].sum()
            )
            as1, as2, as3 = st.columns(3)
            as1.metric("Explicit/Weighted Basis", f"${explicit_total:,.2f}")
            as2.metric("Fallback Basis", f"${fallback_total:,.2f}")
            as3.metric("Allocation Sources", f"{len(lot_allocation_source_summary_df)}")
            _render_df_with_preview(lot_allocation_source_summary_df, hide_index=True)
    with st.expander("Sold COGS Source Summary", expanded=False):
        if cogs_source_summary_df.empty:
            st.caption("No sold COGS source summary is available for the selected window.")
        else:
            fallback_sources = {
                "lot_equal_quantity_fallback",
                "missing_cost_basis",
                "unknown",
                "mixed_fifo_cost",
                "legacy_reports_fifo_cost_map",
            }
            fallback_cogs = float(
                cogs_source_summary_df[
                    cogs_source_summary_df["fifo_cost_source"].isin(fallback_sources)
                ]["fifo_cogs"].sum()
            )
            cs1, cs2, cs3, cs4 = st.columns(4)
            cs1.metric("Sold FIFO COGS", f"${float(cogs_source_summary_df['fifo_cogs'].sum()):,.2f}")
            cs2.metric("Fallback/Review COGS", f"${fallback_cogs:,.2f}")
            cs3.metric("COGS Sources", f"{len(cogs_source_summary_df)}")
            cs4.metric(
                "Bundle Units Sold",
                f"{int(cogs_source_summary_df.get('bundle_inventory_units_sold', pd.Series(dtype=int)).sum())}",
            )
            _render_df_with_preview(cogs_source_summary_df, hide_index=True)

    st.markdown("### eBay Fee Reconciliation")
    if not load_extended_analytics:
        st.info("Enable `Load Extended Analytics` to run fee reconciliation and calibration.")
    elif ebay_fee_reconciliation_df.empty:
        st.info("No eBay sales in the selected date range for fee reconciliation.")
    else:
        er1, er2, er3, er4 = st.columns(4)
        sales_count = int(fee_reconciliation_summary.get("sales_count") or len(ebay_fee_reconciliation_df))
        total_actual_fee = float(fee_reconciliation_summary.get("total_actual_fee") or 0.0)
        total_estimated_fee = float(fee_reconciliation_summary.get("total_estimated_fee") or 0.0)
        total_variance = float(fee_reconciliation_summary.get("total_variance") or 0.0)
        coverage_count = int(fee_reconciliation_summary.get("estimate_coverage_count") or 0)
        er1.metric("eBay Sales", f"{sales_count}")
        er2.metric("Actual Fees", f"${total_actual_fee:,.2f}")
        er3.metric("Estimated Fees", f"${total_estimated_fee:,.2f}")
        er4.metric("Variance", f"${total_variance:,.2f}")
        st.caption(
            f"Estimate coverage: {coverage_count}/{sales_count} sales have listing-time fee estimates."
        )
        if not ebay_fee_source_priority_df.empty:
            st.markdown("#### Actual Fee Source Priority")
            p1, p2, p3 = st.columns(3)
            p1.metric(
                "Normalized Source Rows",
                f"{int(fee_source_counts.get('normalized_source_rows') or 0)}",
            )
            p2.metric(
                "Notes Fallback Rows",
                f"{int(fee_source_counts.get('notes_fallback_rows') or 0)}",
            )
            p3.metric(
                "Sale Field Fallback Rows",
                f"{int(fee_source_counts.get('sale_field_fallback_rows') or 0)}",
            )
            with st.expander("Actual Fee Source Priority Breakdown", expanded=False):
                _render_df_with_preview(ebay_fee_source_priority_df, hide_index=True)
        if not ebay_fee_source_priority_trend_df.empty:
            weekly_coverage_df = _build_normalized_source_weekly_coverage(ebay_fee_source_priority_trend_df)
            if not weekly_coverage_df.empty:
                st.markdown("#### Normalized Source Coverage Trend (Weekly)")
                chart_df = weekly_coverage_df[["bucket_date", "normalized_coverage_pct"]].copy()
                chart_df = chart_df.rename(columns={"bucket_date": "week_start"})
                chart_df["week_start"] = pd.to_datetime(chart_df["week_start"], errors="coerce")
                chart_df = chart_df.set_index("week_start").sort_index()
                st.line_chart(chart_df, y="normalized_coverage_pct", use_container_width=True)
                with st.expander("Normalized Source Coverage Data (Weekly)", expanded=False):
                    _render_df_with_preview(weekly_coverage_df, hide_index=True)
                stacked_weekly_df = _build_weekly_fee_source_count_chart_data(ebay_fee_source_priority_trend_df)
                if not stacked_weekly_df.empty:
                    st.markdown("#### Fee Source Counts by Week")
                    stack_chart_df = stacked_weekly_df.copy()
                    stack_chart_df["week_start"] = pd.to_datetime(stack_chart_df["week_start"], errors="coerce")
                    stack_chart_df = stack_chart_df.set_index("week_start").sort_index()
                    st.bar_chart(
                        stack_chart_df[["normalized_source", "notes_fallback", "sale_field_fallback"]],
                        use_container_width=True,
                    )
                    with st.expander("Fee Source Counts by Week Data", expanded=False):
                        _render_df_with_preview(stacked_weekly_df, hide_index=True)
        if not ebay_fee_actual_source_df.empty:
            with st.expander("Actual Fee Source Breakdown", expanded=False):
                _render_df_with_preview(ebay_fee_actual_source_df, hide_index=True)
        current_final_value_rate = float(
            get_runtime_float(repo, "ebay_fee_estimate_final_value_rate_percent", 13.25)
        )
        calibration = build_final_value_rate_calibration(
            ebay_fee_reconciliation_df.to_dict(orient="records"),
            current_final_value_rate_percent=current_final_value_rate,
        )
        st.markdown("#### Fee Calibration Assist")
        cc1, cc2, cc3, cc4 = st.columns(4)
        cc1.metric("Calibration Samples", f"{int(calibration.get('sample_count') or 0)}")
        cc2.metric("Current Final Value %", f"{current_final_value_rate:.3f}%")
        cc3.metric(
            "Suggested Final Value %",
            f"{float(calibration.get('suggested_final_value_rate_percent') or current_final_value_rate):.3f}%",
        )
        cc4.metric(
            "Suggested Delta",
            f"{float(calibration.get('delta_percent') or 0.0):+.3f}%",
        )
        st.caption(
            "Suggestion is based on implied final-value rates from recent eBay sales with estimate metadata. "
            "Use as calibration guidance, not an automatic accounting source-of-truth."
        )
        can_apply_calibration = has_permission(user.role, "manage_settings")
        apply_col1, apply_col2 = st.columns([1, 3])
        with apply_col1:
            if st.button(
                "Apply Suggested Final Value %",
                key="reports_apply_fee_calibration_btn",
                disabled=not can_apply_calibration
                or int(calibration.get("sample_count") or 0) <= 0,
            ):
                try:
                    suggested_value = float(
                        calibration.get("suggested_final_value_rate_percent")
                        or current_final_value_rate
                    )
                    repo.upsert_runtime_setting(
                        environment=settings.app_env,
                        key="ebay_fee_estimate_final_value_rate_percent",
                        value=f"{suggested_value:.4f}",
                        value_type="float",
                        description="Calibrated from Reports fee reconciliation assist.",
                        is_active=True,
                        actor=user.username,
                    )
                    st.success(
                        f"Updated `ebay_fee_estimate_final_value_rate_percent` to {suggested_value:.4f}%."
                    )
                    st.rerun()
                except Exception as exc:
                    st.error(f"Unable to apply fee calibration setting: {exc}")
        with apply_col2:
            if not can_apply_calibration:
                    st.caption("You need `manage_settings` permission to apply runtime calibration values.")

    st.markdown("### Accounting Exception Queue")
    if not load_extended_analytics:
        st.info("Enable `Load Extended Analytics` to run accounting exception checks.")
    elif accounting_exceptions_df.empty:
        st.success("No accounting exceptions found for the selected date range.")
    else:
        severity_counts = (
            accounting_exceptions_df.groupby(["severity"], dropna=False, as_index=False)
            .size()
            .rename(columns={"size": "count"})
            .sort_values(["severity"], ascending=[True])
        )
        exception_counts = (
            accounting_exceptions_df.groupby(["exception_type"], dropna=False, as_index=False)
            .size()
            .rename(columns={"size": "count"})
            .sort_values(["count"], ascending=[False])
        )
        aq1, aq2, aq3 = st.columns(3)
        aq1.metric("Total Exceptions", f"{len(accounting_exceptions_df)}")
        aq2.metric("P0 Exceptions", f"{int((accounting_exceptions_df['severity'] == 'P0').sum())}")
        aq3.metric("P1 Exceptions", f"{int((accounting_exceptions_df['severity'] == 'P1').sum())}")
        with st.expander("Accounting Exception Rows", expanded=True):
            _render_df_with_preview(accounting_exceptions_df, hide_index=True)
        with st.expander("Accounting Exception Summary", expanded=False):
            _render_df_with_preview(severity_counts, hide_index=True)
            _render_df_with_preview(exception_counts, hide_index=True)

    st.markdown("### Shipping Economics (Charged vs Label Spend)")
    if not load_shipping_tax_analytics:
        st.caption(
            "Shipping economics is deferred. Enable `Load Shipping + Tax Analytics (slower)` to compute this section."
        )
    elif shipping_economics_df.empty:
        st.info("No sales in selected range with shipping economics data.")
    else:
        se1, se2, se3, se4 = st.columns(4)
        total_shipping_charged = float(report_scalar_cache.get("shipping_total_charged") or 0.0)
        total_label_spend = float(report_scalar_cache.get("shipping_total_label_spend") or 0.0)
        total_shipping_delta = total_shipping_charged - total_label_spend
        label_spend_covered_count = int(report_scalar_cache.get("shipping_label_covered_count") or 0)
        shipping_rows_count = int(report_scalar_cache.get("shipping_rows") or 0)
        se1.metric("Sales Rows", f"{shipping_rows_count}")
        se2.metric("Shipping Charged", f"${total_shipping_charged:,.2f}")
        se3.metric("Label Spend", f"${total_label_spend:,.2f}")
        se4.metric("Delta (Charged-Spend)", f"${total_shipping_delta:,.2f}")
        st.caption(
            f"Label-spend coverage: {label_spend_covered_count}/{shipping_rows_count} sales rows."
        )
        if not shipping_econ_summary_df.empty:
            _render_df_with_preview(shipping_econ_summary_df, hide_index=True)

    st.markdown("### Economics Intelligence Drilldowns + Alerts")
    if economics_intel_df.empty:
        st.info("No estimate-vs-actual economics rows in selected date range.")
    else:
        ed1, ed2, ed3 = st.columns(3)
        with ed1:
            economics_min_margin_alert_pct = float(
                st.number_input(
                    "Min Actual Margin % Alert",
                    min_value=-100.0,
                    max_value=100.0,
                    value=5.0,
                    step=0.5,
                    key="reports_econ_min_margin_alert_pct",
                )
            )
        with ed2:
            economics_max_fee_variance_alert_usd = float(
                st.number_input(
                    "Max Avg Fee Variance Alert ($)",
                    min_value=0.0,
                    value=3.0,
                    step=0.25,
                    key="reports_econ_max_fee_variance_alert_usd",
                )
            )
        with ed3:
            economics_min_group_sales_for_alert = int(
                st.number_input(
                    "Min Sales per Group for Alert",
                    min_value=1,
                    value=3,
                    step=1,
                    key="reports_econ_min_group_sales_for_alert",
                )
            )
        economics_drilldowns = _build_economics_intelligence_drilldowns(
            economics_intel_df,
            min_margin_alert_pct=float(economics_min_margin_alert_pct),
            max_fee_variance_alert_usd=float(economics_max_fee_variance_alert_usd),
            min_group_sales_for_alert=int(economics_min_group_sales_for_alert),
        )
        economics_intel_by_sku_df = economics_drilldowns.get("by_sku", pd.DataFrame())
        economics_intel_by_marketplace_df = economics_drilldowns.get("by_marketplace", pd.DataFrame())
        economics_intel_alerts_df = economics_drilldowns.get("alerts", pd.DataFrame())

        total_groups = int(len(economics_intel_by_sku_df))
        sku_alert_groups = (
            int(economics_intel_by_sku_df["alert_any"].astype(bool).sum())
            if (not economics_intel_by_sku_df.empty and "alert_any" in economics_intel_by_sku_df.columns)
            else 0
        )
        channel_alert_groups = (
            int(economics_intel_by_marketplace_df["alert_any"].astype(bool).sum())
            if (not economics_intel_by_marketplace_df.empty and "alert_any" in economics_intel_by_marketplace_df.columns)
            else 0
        )
        alert_rows = int(len(economics_intel_alerts_df))
        ea1, ea2, ea3, ea4 = st.columns(4)
        ea1.metric("SKU Groups", total_groups)
        ea2.metric("SKU Alert Groups", sku_alert_groups)
        ea3.metric("Marketplace Alert Groups", channel_alert_groups)
        ea4.metric("Alert Sale Rows", alert_rows)

        with st.expander("Economics by SKU", expanded=False):
            _render_df_with_preview(economics_intel_by_sku_df, hide_index=True)
        with st.expander("Economics by Marketplace", expanded=False):
            _render_df_with_preview(economics_intel_by_marketplace_df, hide_index=True)
        with st.expander("Economics Alert Rows", expanded=False):
            _render_df_with_preview(economics_intel_alerts_df, hide_index=True)

    st.markdown("### Purchase Document -> Lot Apply Audit")
    load_purchase_doc_apply_audit = st.checkbox(
        "Load Purchase-Document Lot-Apply Audit (slower)",
        value=False,
        key="reports_load_purchase_doc_lot_apply_audit",
    )
    if not load_purchase_doc_apply_audit:
        st.caption(
            "Audit table is deferred. Enable `Load Purchase-Document Lot-Apply Audit (slower)` to query events."
        )
    else:
        purchase_doc_audit_limit = int(
            st.number_input(
                "Purchase-Document Audit Lookback Rows",
                min_value=100,
                max_value=5000,
                value=1000,
                step=100,
                key="reports_purchase_doc_apply_audit_limit",
            )
        )
        audit_rows = repo.list_audit_logs(limit=purchase_doc_audit_limit)
        event_rows: list[dict] = []
        for row in audit_rows:
            if str(getattr(row, "entity_type", "") or "").strip().lower() != "purchase_document":
                continue
            action = str(getattr(row, "action", "") or "").strip().lower()
            if action not in {"auto_apply_extracted_fields_to_lot", "manual_apply_extracted_fields_to_lot"}:
                continue
            created_at = getattr(row, "created_at", None)
            if created_at is None:
                continue
            if created_at < start_dt or created_at > end_dt:
                continue
            payload: dict = {}
            try:
                parsed = json.loads(str(getattr(row, "changes_json", "") or "{}"))
                if isinstance(parsed, dict):
                    payload = parsed
            except Exception:
                payload = {}
            applied_fields_raw = payload.get("applied_fields")
            if isinstance(applied_fields_raw, list):
                applied_fields = [str(v).strip() for v in applied_fields_raw if str(v).strip()]
            else:
                applied_fields = []
            event_rows.append(
                {
                    "created_at": created_at.isoformat(),
                    "actor": str(getattr(row, "actor", "") or "").strip(),
                    "action": action,
                    "mode": str(payload.get("mode") or "").strip().lower(),
                    "workflow": str(payload.get("workflow") or "").strip(),
                    "purchase_document_id": int(getattr(row, "entity_id", 0) or 0),
                    "lot_id": int(payload.get("lot_id") or 0) if payload.get("lot_id") is not None else 0,
                    "applied_field_count": int(len(applied_fields)),
                    "applied_fields": ", ".join(applied_fields),
                }
            )
        if not event_rows:
            st.info("No purchase-document lot-apply audit events found in selected date range.")
        else:
            events_df = pd.DataFrame(event_rows).sort_values("created_at", ascending=False)
            auto_count = int(
                (events_df["action"].astype(str) == "auto_apply_extracted_fields_to_lot").sum()
            )
            manual_count = int(
                (events_df["action"].astype(str) == "manual_apply_extracted_fields_to_lot").sum()
            )
            pd1, pd2, pd3, pd4 = st.columns(4)
            pd1.metric("Events", int(len(events_df)))
            pd2.metric("Auto", auto_count)
            pd3.metric("Manual", manual_count)
            pd4.metric("Distinct Lots", int(events_df["lot_id"].replace(0, pd.NA).dropna().nunique()))
            _render_df_with_preview(events_df, hide_index=True)
            st.download_button(
                "Download Purchase-Document Lot-Apply Audit CSV",
                data=events_df.to_csv(index=False).encode("utf-8"),
                file_name=f"purchase_document_lot_apply_audit_{settings.app_env}.csv",
                mime="text/csv",
                key="reports_purchase_doc_lot_apply_audit_download",
            )

    st.markdown("### Tax Drilldown")
    colorado_suts_packet_artifacts: list[tuple[str, bytes]] = []
    if not load_shipping_tax_analytics:
        st.caption(
            "Tax drilldown is deferred. Enable `Load Shipping + Tax Analytics (slower)` to compute this section."
        )
    elif tax_detail_df.empty:
        st.info("No tax detail rows in selected date range/scope.")
    else:
        if not tax_exceptions_df.empty:
            with st.expander("Tax Exceptions / Advisor Review", expanded=False):
                st.caption(
                    "Tax exception rows are operational review aids. Validate jurisdiction, marketplace "
                    "facilitator, shipping-taxability, and bullion/coin exemption treatment with your tax advisor."
                )
                _render_df_with_preview(tax_exceptions_df, hide_index=True)
        drill_marketplace_options = ["all"] + sorted(
            {
                str(v).strip().lower()
                for v in tax_detail_df["marketplace"].dropna().unique().tolist()
                if str(v).strip()
            }
        )
        td1, td2 = st.columns(2)
        with td1:
            drill_marketplace = st.selectbox(
                "Drilldown Marketplace",
                options=drill_marketplace_options,
                index=0,
                key="reports_tax_drill_marketplace",
            )
        with td2:
            drill_taxability = st.selectbox(
                "Drilldown Segment",
                options=["all", "taxable_only", "exempt_only"],
                index=0,
                key="reports_tax_drill_taxability",
            )
        filtered_tax_detail = _filter_tax_drilldown_rows(
            tax_detail_df,
            marketplace=drill_marketplace,
            taxability=drill_taxability,
        )
        tax_drill_kpis = _tax_drilldown_kpis(filtered_tax_detail)
        dt1, dt2, dt3 = st.columns(3)
        dt1.metric("Rows", int(tax_drill_kpis.get("rows") or 0))
        dt2.metric(
            "Taxable Subtotal",
            f"${float(tax_drill_kpis.get('taxable_subtotal') or 0.0):,.2f}",
        )
        dt3.metric(
            "Estimated Tax",
            f"${float(tax_drill_kpis.get('estimated_tax') or 0.0):,.2f}",
        )
        sale_option_rows = _build_tax_drilldown_sale_option_rows(filtered_tax_detail)
        sale_option_map = {str(row.get("label") or ""): int(row.get("sale_id") or 0) for row in sale_option_rows}
        if sale_option_map:
            hx1, hx2 = st.columns([3, 1])
            with hx1:
                selected_sale_label = st.selectbox(
                    "Create Invoice From Sale",
                    options=list(sale_option_map.keys()),
                    key="reports_tax_drill_sale_pick",
                )
            with hx2:
                if st.button("Open in Documents", key="reports_tax_drill_to_documents_btn"):
                    handoff_to_documents_draft(
                        source_type="Sale",
                        source_id=int(sale_option_map[selected_sale_label]),
                        doc_type="invoice",
                        handoff_from="reports_tax_drilldown",
                        tax_jurisdiction=str(tax_jurisdiction or "").strip(),
                        tax_rate_percent=float(tax_rate_percent or 0.0),
                        tax_shipping_taxable=bool(tax_shipping_taxable),
                        repo=repo,
                        actor=user.username,
                    )
        _render_df_with_preview(filtered_tax_detail)
        dx1, dx2 = st.columns(2)
        with dx1:
            st.download_button(
                label="Download Tax Drilldown CSV",
                data=filtered_tax_detail.to_csv(index=False).encode("utf-8"),
                file_name=f"tax_drilldown_{from_date}_{to_date}.csv",
                mime="text/csv",
                disabled=filtered_tax_detail.empty,
            )
        with dx2:
            st.download_button(
                label="Download Tax Drilldown XLSX",
                data=dataframe_to_xlsx_bytes(filtered_tax_detail, sheet_name="tax_drilldown"),
                file_name=f"tax_drilldown_{from_date}_{to_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                disabled=filtered_tax_detail.empty,
            )
        st.markdown("#### Colorado SUTS Upload")
        st.caption(
            "Creates a Colorado SUTS workbook from the official template. Gross sales are written as text, "
            "account numbers are numeric, A1 is blank, and deduction/exemption columns are left blank for "
            "advisor review because local deduction codes vary by jurisdiction. Marketplace-facilitator "
            "channels such as eBay are excluded by default in Tax Reporting Scope; include only channels your "
            "tax advisor says you must report."
        )
        st.info(
            "SUTS gross sales use the Tax Marketplace Filter above. For normal filing, leave eBay/facilitator "
            "channels unselected; selecting eBay will add those marketplace sales to this SUTS upload."
        )
        suts_options = _load_colorado_suts_jurisdiction_options()
        if not suts_options:
            st.warning("Colorado SUTS template could not be loaded from app assets.")
        else:
            suts_month_value = st.text_input(
                "SUTS Filing Month (YYYY-MM)",
                value=from_date.strftime("%Y-%m"),
                key="reports_colorado_suts_filing_month",
                help="Generates the SUTS workbook for the selected filing month using the current Tax Reporting Scope.",
            )
            suts_month_valid = True
            try:
                suts_month_start, suts_month_end = _month_bounds_from_yyyy_mm(suts_month_value)
                suts_tax_detail_df = _filter_tax_detail_for_month(tax_detail_df, month_value=suts_month_value)
            except Exception as exc:
                suts_month_valid = False
                suts_month_start, suts_month_end = start_dt, end_dt
                suts_tax_detail_df = pd.DataFrame()
                st.warning(f"Colorado SUTS filing month is invalid: {exc}")
            suts_excluded_facilitator_df = pd.DataFrame()
            if suts_month_valid and hasattr(repo, "report_tax_estimate_detail_rows"):
                try:
                    suts_month_rows = repo.report_tax_estimate_detail_rows(
                        start_dt=suts_month_start,
                        end_dt=suts_month_end,
                        tax_rate_percent=float(tax_rate_percent),
                        shipping_taxable=bool(tax_shipping_taxable),
                        tax_exempt_categories=tax_exempt_categories,
                        marketplaces=tax_query_marketplace_set,
                    )
                    suts_tax_detail_df = pd.DataFrame(suts_month_rows or [])
                except Exception:
                    _rollback_report_session()
                excluded_facilitator_scope = set(facilitator_channels) - set(selected_tax_marketplace_set)
                if excluded_facilitator_scope:
                    try:
                        excluded_facilitator_rows = repo.report_tax_estimate_detail_rows(
                            start_dt=suts_month_start,
                            end_dt=suts_month_end,
                            tax_rate_percent=float(tax_rate_percent),
                            shipping_taxable=bool(tax_shipping_taxable),
                            tax_exempt_categories=tax_exempt_categories,
                            marketplaces=excluded_facilitator_scope,
                        )
                        suts_excluded_facilitator_df = pd.DataFrame(excluded_facilitator_rows or [])
                    except Exception:
                        _rollback_report_session()
            suts_scope_summary_df = pd.DataFrame(
                _build_colorado_suts_scope_summary_rows(
                    suts_tax_detail_df,
                    suts_excluded_facilitator_df,
                    selected_marketplaces=selected_tax_marketplace_set,
                    facilitator_channels=facilitator_channels,
                )
            )
            if not suts_scope_summary_df.empty:
                st.markdown("##### SUTS Scope Check")
                _render_df_with_preview(suts_scope_summary_df, hide_index=True)
            suts_labels = [str(row["label"]) for row in suts_options]
            suts_code_by_label = {
                str(row["label"]): str(row["jurisdiction_code"])
                for row in suts_options
            }
            suts_key_by_label = {
                str(row["label"]): _suts_jurisdiction_key(row["jurisdiction_code"], row.get("account_type"))
                for row in suts_options
            }
            suts_custom_jurisdictions: list[dict[str, object]] = []
            suts_account_override_by_key: dict[str, str] = {}
            suts_allow_blank_account_keys: set[str] = set()
            include_golden_custom = st.checkbox(
                "Include Golden SUTS rows (11-0042)",
                value=True,
                key="reports_colorado_suts_include_golden_custom",
                help="The provided SUTS template does not include Golden; this appends the Golden State and Local rows shown in SUTS.",
            )
            if include_golden_custom:
                golden_state_key = _suts_jurisdiction_key("110042", "STATE")
                golden_local_key = _suts_jurisdiction_key("110042", "LOCAL")
                golden_state_label = "110042 | GOLDEN | STATE (custom)"
                golden_local_label = "110042 | GOLDEN | LOCAL (custom)"
                suts_labels.extend([golden_state_label, golden_local_label])
                suts_code_by_label[golden_state_label] = "110042"
                suts_code_by_label[golden_local_label] = "110042"
                suts_key_by_label[golden_state_label] = golden_state_key
                suts_key_by_label[golden_local_label] = golden_local_key
                golden_state_account_number = st.text_input(
                    "Golden State Account Number",
                    value=COLORADO_SUTS_GOLDEN_STATE_ACCOUNT_NUMBER,
                    key="reports_colorado_suts_golden_state_account_number",
                    help=(
                        "Use the Local Account # shown by SUTS for the selected Golden location. "
                        "Your SUTS State location shows 970074130001 for Golden 110042."
                    ),
                )
                if "".join(ch for ch in str(golden_state_account_number or "") if ch.isdigit()):
                    suts_account_override_by_key[golden_state_key] = golden_state_account_number
                golden_local_account_number = st.text_input(
                    "Golden Local/Self-Collected Account Number",
                    value="",
                    key="reports_colorado_suts_golden_local_account_number",
                    help=(
                        "Your SUTS Local/Self-Collected Golden row shows no local account number. "
                        "Leave blank for zero filings if SUTS accepts it, or enter the local account number if one is assigned."
                    ),
                )
                if "".join(ch for ch in str(golden_local_account_number or "") if ch.isdigit()):
                    suts_account_override_by_key[golden_local_key] = golden_local_account_number
                else:
                    suts_allow_blank_account_keys.add(golden_local_key)
                suts_custom_jurisdictions.extend(
                    [
                        {"account_type": "STATE", "jurisdiction_code": "110042", "jurisdiction_name": "GOLDEN"},
                        {"account_type": "LOCAL", "jurisdiction_code": "110042", "jurisdiction_name": "GOLDEN"},
                    ]
                )
            suts_default_index = suts_labels.index(golden_state_label) if include_golden_custom and golden_state_label in suts_labels else 0
            s1, s2 = st.columns([1, 2])
            with s1:
                suts_account_number = st.text_input(
                    "Colorado SUTS Account Number",
                    value=COLORADO_SUTS_ACCOUNT_NUMBER,
                    key="reports_colorado_suts_account_number",
                )
            with s2:
                default_gross_labels = [suts_labels[suts_default_index]] if suts_labels else []
                if include_golden_custom and not suts_tax_detail_df.empty:
                    default_gross_labels = [
                        label
                        for label in [golden_state_label, golden_local_label]
                        if label in suts_labels
                    ]
                suts_gross_labels = st.multiselect(
                    "Gross Sales Jurisdiction Rows",
                    options=suts_labels,
                    default=[] if suts_tax_detail_df.empty else default_gross_labels,
                    key="reports_colorado_suts_gross_jurisdictions",
                    help="Choose all SUTS jurisdiction rows that should receive the selected report scope's gross sales.",
                )
            suts_zero_options = (
                suts_labels
                if suts_tax_detail_df.empty
                else [label for label in suts_labels if label not in set(suts_gross_labels)]
            )
            suts_zero_labels = st.multiselect(
                "Zero Filing Jurisdiction Rows",
                options=suts_zero_options,
                default=default_gross_labels if suts_tax_detail_df.empty else [],
                key="reports_colorado_suts_zero_jurisdictions",
                help="Use only for jurisdictions where you have a filing obligation but zero gross sales.",
            )
            try:
                if not suts_month_valid:
                    raise ValueError("Enter a valid SUTS filing month in YYYY-MM format.")
                gross_codes = [
                    suts_code_by_label[label]
                    for label in suts_gross_labels
                    if label in suts_code_by_label
                ]
                gross_keys = [
                    suts_key_by_label[label]
                    for label in suts_gross_labels
                    if label in suts_key_by_label
                ]
                zero_codes = [
                    suts_code_by_label[label]
                    for label in suts_zero_labels
                    if label in suts_code_by_label
                ]
                zero_keys = [
                    suts_key_by_label[label]
                    for label in suts_zero_labels
                    if label in suts_key_by_label
                ]
                golden_state_key = _suts_jurisdiction_key("110042", "STATE")
                if (
                    (golden_state_key in gross_keys or golden_state_key in zero_keys)
                    and golden_state_key not in suts_account_override_by_key
                ):
                    raise ValueError(
                        "Golden STATE (110042) requires the account number shown in SUTS before exporting."
                    )
                suts_bytes, suts_summary_df = _build_colorado_suts_upload_workbook(
                    suts_tax_detail_df,
                    account_number=suts_account_number,
                    gross_jurisdiction_codes=gross_codes,
                    gross_jurisdiction_keys=gross_keys,
                    zero_filing_jurisdiction_codes=zero_codes,
                    zero_filing_jurisdiction_keys=zero_keys,
                    account_number_by_jurisdiction_key=suts_account_override_by_key,
                    allow_blank_account_jurisdiction_keys=suts_allow_blank_account_keys,
                    custom_jurisdictions=suts_custom_jurisdictions,
                )
                if not suts_summary_df.empty:
                    _render_df_with_preview(suts_summary_df, hide_index=True)
                for warning in _colorado_suts_summary_warnings(suts_summary_df):
                    st.warning(warning)
                colorado_suts_filename = f"colorado_suts_upload_{suts_month_value}.xlsx"
                colorado_suts_packet_artifacts = [(colorado_suts_filename, suts_bytes)]
                st.download_button(
                    label="Download Colorado SUTS Upload XLSX",
                    data=suts_bytes,
                    file_name=colorado_suts_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="reports_colorado_suts_upload_xlsx",
                )
            except Exception as exc:
                st.warning(f"Colorado SUTS upload workbook could not be generated: {exc}")
    if load_shipping_tax_analytics and tax_detail_df.empty:
        st.markdown("#### Colorado SUTS Upload")
        st.caption(
            "Creates a Colorado SUTS workbook from the official template. This remains available when "
            "the selected report scope has no tax rows so you can generate zero-filing uploads for "
            "jurisdictions where you still have a filing obligation."
        )
        st.info(
            "SUTS gross sales use the Tax Marketplace Filter above. For normal filing, leave eBay/facilitator "
            "channels unselected; selecting eBay will add those marketplace sales to this SUTS upload."
        )
        suts_options = _load_colorado_suts_jurisdiction_options()
        if not suts_options:
            st.warning("Colorado SUTS template could not be loaded from app assets.")
        else:
            suts_month_value = st.text_input(
                "SUTS Filing Month (YYYY-MM)",
                value=from_date.strftime("%Y-%m"),
                key="reports_colorado_suts_empty_filing_month",
                help="Generates the SUTS workbook for the selected filing month using the current Tax Reporting Scope.",
            )
            suts_month_valid = True
            try:
                suts_month_start, suts_month_end = _month_bounds_from_yyyy_mm(suts_month_value)
                suts_tax_detail_df = _filter_tax_detail_for_month(tax_detail_df, month_value=suts_month_value)
            except Exception as exc:
                suts_month_valid = False
                suts_month_start, suts_month_end = start_dt, end_dt
                suts_tax_detail_df = pd.DataFrame()
                st.warning(f"Colorado SUTS filing month is invalid: {exc}")
            suts_excluded_facilitator_df = pd.DataFrame()
            if suts_month_valid and hasattr(repo, "report_tax_estimate_detail_rows"):
                try:
                    suts_month_rows = repo.report_tax_estimate_detail_rows(
                        start_dt=suts_month_start,
                        end_dt=suts_month_end,
                        tax_rate_percent=float(tax_rate_percent),
                        shipping_taxable=bool(tax_shipping_taxable),
                        tax_exempt_categories=tax_exempt_categories,
                        marketplaces=tax_query_marketplace_set,
                    )
                    suts_tax_detail_df = pd.DataFrame(suts_month_rows or [])
                except Exception:
                    _rollback_report_session()
                excluded_facilitator_scope = set(facilitator_channels) - set(selected_tax_marketplace_set)
                if excluded_facilitator_scope:
                    try:
                        excluded_facilitator_rows = repo.report_tax_estimate_detail_rows(
                            start_dt=suts_month_start,
                            end_dt=suts_month_end,
                            tax_rate_percent=float(tax_rate_percent),
                            shipping_taxable=bool(tax_shipping_taxable),
                            tax_exempt_categories=tax_exempt_categories,
                            marketplaces=excluded_facilitator_scope,
                        )
                        suts_excluded_facilitator_df = pd.DataFrame(excluded_facilitator_rows or [])
                    except Exception:
                        _rollback_report_session()
            suts_scope_summary_df = pd.DataFrame(
                _build_colorado_suts_scope_summary_rows(
                    suts_tax_detail_df,
                    suts_excluded_facilitator_df,
                    selected_marketplaces=selected_tax_marketplace_set,
                    facilitator_channels=facilitator_channels,
                )
            )
            if not suts_scope_summary_df.empty:
                st.markdown("##### SUTS Scope Check")
                _render_df_with_preview(suts_scope_summary_df, hide_index=True)
            suts_labels = [str(row["label"]) for row in suts_options]
            suts_code_by_label = {
                str(row["label"]): str(row["jurisdiction_code"])
                for row in suts_options
            }
            suts_key_by_label = {
                str(row["label"]): _suts_jurisdiction_key(row["jurisdiction_code"], row.get("account_type"))
                for row in suts_options
            }
            suts_custom_jurisdictions: list[dict[str, object]] = []
            suts_account_override_by_key: dict[str, str] = {}
            suts_allow_blank_account_keys: set[str] = set()
            include_golden_custom = st.checkbox(
                "Include Golden SUTS rows (11-0042)",
                value=True,
                key="reports_colorado_suts_empty_include_golden_custom",
                help="The provided SUTS template does not include Golden; this appends the Golden State and Local rows shown in SUTS.",
            )
            if include_golden_custom:
                golden_state_key = _suts_jurisdiction_key("110042", "STATE")
                golden_local_key = _suts_jurisdiction_key("110042", "LOCAL")
                golden_state_label = "110042 | GOLDEN | STATE (custom)"
                golden_local_label = "110042 | GOLDEN | LOCAL (custom)"
                suts_labels.extend([golden_state_label, golden_local_label])
                suts_code_by_label[golden_state_label] = "110042"
                suts_code_by_label[golden_local_label] = "110042"
                suts_key_by_label[golden_state_label] = golden_state_key
                suts_key_by_label[golden_local_label] = golden_local_key
                golden_state_account_number = st.text_input(
                    "Golden State Account Number",
                    value=COLORADO_SUTS_GOLDEN_STATE_ACCOUNT_NUMBER,
                    key="reports_colorado_suts_empty_golden_state_account_number",
                    help=(
                        "Use the Local Account # shown by SUTS for the selected Golden location. "
                        "Your SUTS State location shows 970074130001 for Golden 110042."
                    ),
                )
                if "".join(ch for ch in str(golden_state_account_number or "") if ch.isdigit()):
                    suts_account_override_by_key[golden_state_key] = golden_state_account_number
                golden_local_account_number = st.text_input(
                    "Golden Local/Self-Collected Account Number",
                    value="",
                    key="reports_colorado_suts_empty_golden_local_account_number",
                    help=(
                        "Your SUTS Local/Self-Collected Golden row shows no local account number. "
                        "Leave blank for zero filings if SUTS accepts it, or enter the local account number if one is assigned."
                    ),
                )
                if "".join(ch for ch in str(golden_local_account_number or "") if ch.isdigit()):
                    suts_account_override_by_key[golden_local_key] = golden_local_account_number
                else:
                    suts_allow_blank_account_keys.add(golden_local_key)
                suts_custom_jurisdictions.extend(
                    [
                        {"account_type": "STATE", "jurisdiction_code": "110042", "jurisdiction_name": "GOLDEN"},
                        {"account_type": "LOCAL", "jurisdiction_code": "110042", "jurisdiction_name": "GOLDEN"},
                    ]
                )
            suts_default_index = suts_labels.index(golden_state_label) if include_golden_custom and golden_state_label in suts_labels else 0
            s1, s2 = st.columns([1, 2])
            with s1:
                suts_account_number = st.text_input(
                    "Colorado SUTS Account Number",
                    value=COLORADO_SUTS_ACCOUNT_NUMBER,
                    key="reports_colorado_suts_empty_account_number",
                )
            with s2:
                default_gross_labels = [suts_labels[suts_default_index]] if suts_labels else []
                if include_golden_custom and not suts_tax_detail_df.empty:
                    default_gross_labels = [
                        label
                        for label in [golden_state_label, golden_local_label]
                        if label in suts_labels
                    ]
                suts_gross_labels = st.multiselect(
                    "Gross Sales Jurisdiction Rows",
                    options=suts_labels,
                    default=[] if suts_tax_detail_df.empty else default_gross_labels,
                    key="reports_colorado_suts_empty_gross_jurisdictions",
                    help="Choose all SUTS jurisdiction rows that should receive the selected report scope's gross sales.",
                )
            suts_zero_options = (
                suts_labels
                if suts_tax_detail_df.empty
                else [label for label in suts_labels if label not in set(suts_gross_labels)]
            )
            suts_zero_labels = st.multiselect(
                "Zero Filing Jurisdiction Rows",
                options=suts_zero_options,
                default=default_gross_labels if suts_tax_detail_df.empty else [],
                key="reports_colorado_suts_empty_zero_jurisdictions",
                help="Use only for jurisdictions where you have a filing obligation but zero gross sales.",
            )
            try:
                if not suts_month_valid:
                    raise ValueError("Enter a valid SUTS filing month in YYYY-MM format.")
                gross_codes = [
                    suts_code_by_label[label]
                    for label in suts_gross_labels
                    if label in suts_code_by_label
                ]
                gross_keys = [
                    suts_key_by_label[label]
                    for label in suts_gross_labels
                    if label in suts_key_by_label
                ]
                zero_codes = [
                    suts_code_by_label[label]
                    for label in suts_zero_labels
                    if label in suts_code_by_label
                ]
                zero_keys = [
                    suts_key_by_label[label]
                    for label in suts_zero_labels
                    if label in suts_key_by_label
                ]
                golden_state_key = _suts_jurisdiction_key("110042", "STATE")
                if (
                    (golden_state_key in gross_keys or golden_state_key in zero_keys)
                    and golden_state_key not in suts_account_override_by_key
                ):
                    raise ValueError(
                        "Golden STATE (110042) requires the account number shown in SUTS before exporting."
                    )
                suts_bytes, suts_summary_df = _build_colorado_suts_upload_workbook(
                    suts_tax_detail_df,
                    account_number=suts_account_number,
                    gross_jurisdiction_codes=gross_codes,
                    gross_jurisdiction_keys=gross_keys,
                    zero_filing_jurisdiction_codes=zero_codes,
                    zero_filing_jurisdiction_keys=zero_keys,
                    account_number_by_jurisdiction_key=suts_account_override_by_key,
                    allow_blank_account_jurisdiction_keys=suts_allow_blank_account_keys,
                    custom_jurisdictions=suts_custom_jurisdictions,
                )
                if not suts_summary_df.empty:
                    _render_df_with_preview(suts_summary_df, hide_index=True)
                for warning in _colorado_suts_summary_warnings(suts_summary_df):
                    st.warning(warning)
                colorado_suts_filename = f"colorado_suts_upload_{suts_month_value}.xlsx"
                colorado_suts_packet_artifacts = [(colorado_suts_filename, suts_bytes)]
                st.download_button(
                    label="Download Colorado SUTS Upload XLSX",
                    data=suts_bytes,
                    file_name=colorado_suts_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="reports_colorado_suts_empty_upload_xlsx",
                )
            except Exception as exc:
                st.warning(f"Colorado SUTS upload workbook could not be generated: {exc}")

    st.markdown("### Document Draft Handoff")
    st.caption("Open Documents with prefilled source context from either a sale or an order.")
    h1, h2, h3 = st.columns([1, 2, 1])
    with h1:
        handoff_source_type = st.selectbox(
            "Source Type",
            options=["Sale", "Order"],
            index=0,
            key="reports_documents_handoff_source_type",
        )
    with h2:
        handoff_doc_type = st.selectbox(
            "Document Type",
            options=["invoice", "receipt"],
            index=0,
            key="reports_documents_handoff_doc_type",
        )
    source_option_map: dict[str, int] = {}
    if handoff_source_type == "Sale":
        source_option_rows = _build_documents_handoff_sale_option_rows(sales_df)
        source_option_map = {
            str(row.get("label") or ""): int(row.get("source_id") or 0)
            for row in source_option_rows
        }
    else:
        source_option_rows = _build_documents_handoff_order_option_rows(orders_df)
        source_option_map = {
            str(row.get("label") or ""): int(row.get("source_id") or 0)
            for row in source_option_rows
        }
    if source_option_map:
        selected_source_label = st.selectbox(
            "Select Source",
            options=list(source_option_map.keys()),
            key="reports_documents_handoff_source_pick",
        )
        with h3:
            if st.button(
                "Open in Documents",
                key="reports_documents_handoff_open_btn",
            ):
                source_id = source_option_map.get(str(selected_source_label or ""))
                if source_id:
                    handoff_to_documents_draft(
                        source_type=handoff_source_type,
                        source_id=int(source_id),
                        doc_type=handoff_doc_type,
                        handoff_from="reports_documents_handoff",
                        tax_jurisdiction=str(tax_jurisdiction or "").strip(),
                        tax_rate_percent=float(tax_rate_percent or 0.0),
                        tax_shipping_taxable=bool(tax_shipping_taxable),
                        repo=repo,
                        actor=user.username,
                    )
    else:
        st.info(f"No {handoff_source_type.lower()} records in selected date range.")

    st.markdown("### Rebuy Cost Trend (Weighted/Lot)")
    if not load_inventory_cycle_analytics:
        st.caption("Enable `Load Inventory Cycle + Rebuy Analytics` to run cycle and rebuy trend reports.")
    elif rebuy_cost_trend_df.empty:
        st.info("No acquisition events with unit cost found for trend analysis.")
    else:
        sku_options = sorted({str(v) for v in rebuy_cost_trend_df["sku"].dropna().unique() if str(v).strip()})
        selected_sku = st.selectbox(
            "Trend SKU",
            options=sku_options,
            index=0,
            key="reports_rebuy_cost_trend_sku",
        )
        sku_rows = rebuy_cost_trend_df[rebuy_cost_trend_df["sku"] == selected_sku].copy()
        if sku_rows.empty:
            st.info("No rows for selected SKU.")
        else:
            chart_df = sku_rows[["as_of", "weighted_unit_cost", "unit_cost"]].copy()
            chart_df = chart_df.rename(
                columns={"weighted_unit_cost": "weighted_unit_cost_running", "unit_cost": "event_unit_cost"}
            )
            chart_df = chart_df.set_index("as_of")
            st.line_chart(chart_df, use_container_width=True)
            _render_df_with_preview(sku_rows)

    tax_evidence_reports = [
        ("Tax Summary (Estimated)", tax_summary_df, "tax_summary_estimated"),
        ("Tax by Marketplace (Estimated)", tax_by_marketplace_df, "tax_by_marketplace_estimated"),
        ("Tax Detail (Estimated)", tax_detail_df, "tax_detail_estimated"),
        ("Tax Exceptions / Advisor Review", tax_exceptions_df, "tax_exceptions_advisor_review"),
        ("Tax Reporting Sign-Off Evidence", tax_signoff_df, "tax_reporting_signoffs"),
    ]
    current_tax_packet_hash = _tax_review_packet_evidence_hash_from_reports(
        reports=tax_evidence_reports,
        from_date=from_date,
        to_date=to_date,
        tax_jurisdiction=tax_jurisdiction or tax_default_jurisdiction,
        tax_rate_percent=float(tax_rate_percent),
        shipping_taxable=bool(tax_shipping_taxable),
        marketplace_scope=tax_marketplace_scope_label,
        facilitator_channels=facilitator_channels,
        tax_exempt_categories=tax_exempt_categories,
        tax_profile=selected_tax_profile_context,
        extra_artifact_hashes={
            str(name or "").strip(): hashlib.sha256(payload or b"").hexdigest()
            for name, payload in colorado_suts_packet_artifacts
            if str(name or "").strip()
        },
    )
    tax_reporting_signoff_review_df = _build_tax_reporting_signoff_review(
        signoff_df=tax_signoff_df,
        tax_period=_default_accounting_close_period(from_date, to_date),
        jurisdiction=tax_jurisdiction or tax_default_jurisdiction,
        profile_key=str(selected_tax_profile_context.get("profile_key") or ""),
        tax_exception_count=int(len(tax_exceptions_df)),
        current_packet_hash=current_tax_packet_hash,
        to_date=to_date,
    )

    st.markdown("### Reports Copilot")
    st.caption("AI narrative summary for margin anomalies, reconciliation risk, and export recommendations.")
    if not load_extended_analytics:
        st.caption("Enable `Load Extended Analytics` to run Reports Copilot on full fee/margin context.")
    if st.button(
        "Analyze Report Snapshot",
        key="reports_copilot_analyze_btn",
        disabled=not load_extended_analytics,
    ):
        if not ensure_permission(user, "ai_comp_use", "Use Reports Copilot"):
            st.stop()
        try:
            context = {
                "date_range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
                "table_row_counts": dict(report_scalar_cache.get("table_row_counts") or {}),
                "margin_snapshot": {
                    "gross_sales_total": float(report_scalar_cache.get("cogs_gross_sales_total") or 0.0),
                    "fifo_margin_before_returns_total": float(
                        report_scalar_cache.get("cogs_fifo_margin_total") or 0.0
                    ),
                    "lot_margin_total": float(report_scalar_cache.get("cogs_lot_margin_total") or 0.0),
                    "negative_fifo_margin_rows": int(report_scalar_cache.get("cogs_negative_fifo_rows") or 0),
                },
                "profit_after_returns_summary": close_profit_summary,
                "reconciliation_flags": int(report_scalar_cache.get("reconcile_flags") or 0),
                "accounting_close_readiness": dict(accounting_close_summary),
                "accounting_close_packet_evidence_hash": current_accounting_close_packet_hash,
                "accounting_close_formula_checks": (
                    accounting_close_formula_df.to_dict("records")
                    if not accounting_close_formula_df.empty
                    else []
                ),
                "accounting_sales_component_checks": (
                    accounting_sales_component_df.to_dict("records")
                    if not accounting_sales_component_df.empty
                    else []
                ),
                "accounting_return_tieout_checks": (
                    accounting_return_tieout_df.to_dict("records")
                    if not accounting_return_tieout_df.empty
                    else []
                ),
                "accounting_inventory_valuation_checks": (
                    accounting_inventory_valuation_df.to_dict("records")
                    if not accounting_inventory_valuation_df.empty
                    else []
                ),
                "accounting_fee_evidence_checks": (
                    accounting_fee_evidence_df.to_dict("records")
                    if not accounting_fee_evidence_df.empty
                    else []
                ),
                "accounting_shipping_evidence_checks": (
                    accounting_shipping_evidence_df.to_dict("records")
                    if not accounting_shipping_evidence_df.empty
                    else []
                ),
                "accounting_reconciliation_tieout_checks": (
                    accounting_reconciliation_tieout_df.to_dict("records")
                    if not accounting_reconciliation_tieout_df.empty
                    else []
                ),
                "accounting_cogs_source_checks": (
                    accounting_cogs_source_df.to_dict("records")
                    if not accounting_cogs_source_df.empty
                    else []
                ),
                "sale_fifo_cogs_evidence_rows": (
                    sale_fifo_cogs_evidence_df.head(50).to_dict("records")
                    if not sale_fifo_cogs_evidence_df.empty
                    else []
                ),
                "accounting_lot_allocation_checks": (
                    accounting_lot_allocation_df.to_dict("records")
                    if not accounting_lot_allocation_df.empty
                    else []
                ),
                "accounting_exception_queue_checks": (
                    accounting_exception_queue_checks_df.to_dict("records")
                    if not accounting_exception_queue_checks_df.empty
                    else []
                ),
                "accounting_margin_anomaly_checks": (
                    accounting_margin_anomaly_checks_df.to_dict("records")
                    if not accounting_margin_anomaly_checks_df.empty
                    else []
                ),
                "accounting_close_consistency_checks": (
                    accounting_close_consistency_checks_df.to_dict("records")
                    if not accounting_close_consistency_checks_df.empty
                    else []
                ),
                "accounting_close_packet_completeness_checks": (
                    accounting_close_packet_completeness_df.to_dict("records")
                    if not accounting_close_packet_completeness_df.empty
                    else []
                ),
                "accounting_close_packet_manifest_checks": (
                    accounting_close_packet_manifest_df.to_dict("records")
                    if not accounting_close_packet_manifest_df.empty
                    else []
                ),
                "accounting_close_packet_hash_checks": (
                    accounting_close_packet_hash_df.to_dict("records")
                    if not accounting_close_packet_hash_df.empty
                    else []
                ),
                "accounting_close_packet_evidence_hash_rows": (
                    accounting_close_packet_evidence_hash_df.to_dict("records")
                    if not accounting_close_packet_evidence_hash_df.empty
                    else []
                ),
                "accounting_period_drift_checks": (
                    accounting_period_drift_df.to_dict("records")
                    if not accounting_period_drift_df.empty
                    else []
                ),
                "tax_review_summary": {
                    "jurisdiction": str(tax_jurisdiction or tax_default_jurisdiction or ""),
                    "tax_rate_percent": float(tax_rate_percent or 0.0),
                    "shipping_taxable": bool(tax_shipping_taxable),
                    "tax_packet_evidence_hash": current_tax_packet_hash,
                    "marketplace_scope": tax_marketplace_scope_label,
                    "facilitator_channels": sorted(facilitator_channels),
                    "tax_exempt_categories": sorted(tax_exempt_categories),
                    "tax_exception_count": int(len(tax_exceptions_df)),
                },
                "tax_profile_evidence": selected_tax_profile_context,
                "tax_reporting_signoff_rows": (
                    tax_signoff_df.head(20).to_dict("records")
                    if not tax_signoff_df.empty
                    else []
                ),
                "tax_reporting_signoff_review": (
                    tax_reporting_signoff_review_df.to_dict("records")
                    if not tax_reporting_signoff_review_df.empty
                    else []
                ),
                "top_negative_fifo_margin_rows": _top_n_records(
                    cogs_margin_df,
                    sort_by="fifo_margin",
                    ascending=True,
                    n=10,
                ),
                "accounting_exception_rows": (
                    accounting_exceptions_df.head(20).to_dict("records")
                    if not accounting_exceptions_df.empty
                    else []
                ),
                "top_margin_by_sku_rows": _top_n_records(
                    margin_by_sku_df,
                    sort_by="fifo_margin",
                    ascending=False,
                    n=10,
                ),
                "marketplace_reconciliation_rows": _top_n_by_abs_records(
                    reconciliation_df,
                    value_col="delta_order_total_vs_sales_gross",
                    n=10,
                ),
                "validation_issue_rows": (
                    accounting_validation_df.head(20).to_dict("records")
                    if not accounting_validation_df.empty
                    else []
                ),
            }
            copilot_system_message = get_runtime_str(
                repo,
                "comp_llm_system_message",
                "You are an accounting and operations reporting copilot.",
            ).strip()
            copilot_instruction = (
                "Return ONLY JSON with keys: `executive_summary`, `margin_anomalies`, "
                "`reconciliation_findings`, `tax_review_findings`, `recommended_exports`, `next_actions`. "
                "Each key must be an array of concise bullet strings. When discussing profit, use "
                "`profit_after_returns_summary.estimated_profit_after_returns` as the final estimated profit "
                "and label `margin_snapshot.fifo_margin_before_returns_total` as before-return margin."
            )
            copilot_prompt = "Reports narrative summary and export recommendations"
            copilot_started = time.perf_counter()
            copilot_prompt_hash = _stable_json_sha256(
                {
                    "query": copilot_prompt,
                    "system_message": copilot_system_message,
                    "instruction": copilot_instruction,
                }
            )
            copilot_data_scope = {
                "date_range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
                "context_keys": sorted(context.keys()),
                "context_hash_sha256": _stable_json_sha256(context),
                "tax_packet_evidence_hash": current_tax_packet_hash,
                "accounting_close_packet_evidence_hash": current_accounting_close_packet_hash,
                "row_counts": {
                    "accounting_close_formula_checks": int(len(accounting_close_formula_df)),
                    "accounting_sales_component_checks": int(len(accounting_sales_component_df)),
                    "accounting_period_drift_checks": int(len(accounting_period_drift_df)),
                    "accounting_exception_queue": int(len(accounting_exceptions_df)),
                    "tax_reporting_signoffs": int(len(tax_signoff_df)),
                    "tax_reporting_signoff_review": int(len(tax_reporting_signoff_review_df)),
                    "validation_issue_rows": int(len(accounting_validation_df)),
                },
            }
            result = execute_comp_summary(
                repo,
                query=copilot_prompt,
                ebay_rows=[],
                web_rows=[],
                spot_context=context,
                system_message=copilot_system_message,
                instruction=copilot_instruction,
            )
            copilot_text = str(result.text or "").strip()
            st.session_state["reports_copilot_raw"] = copilot_text
            copilot_review_metadata = {
                "event_type": "reports_copilot_review",
                "surface": "reports",
                "read_only": True,
                "date_range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
                "prompt_hash_sha256": copilot_prompt_hash,
                "data_scope_hash_sha256": str(copilot_data_scope.get("context_hash_sha256") or ""),
                "data_scope": copilot_data_scope,
                "context_keys": sorted(context.keys()),
                "ai_citation": dict(getattr(result, "citation", {}) or {}),
            }
            st.session_state["reports_copilot_metadata"] = copilot_review_metadata
            if hasattr(repo, "log_ai_chat_interaction"):
                try:
                    repo.log_ai_chat_interaction(
                        actor=user.username,
                        prompt=copilot_prompt,
                        intent="reports_copilot_review",
                        allowed_domains=["accounting", "reports", "sales", "orders", "inventory", "tax"],
                        citations=[
                            {
                                "table": "accounting_close_readiness_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_checks_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_period_drift_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_period_drift_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_exception_queue",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_exceptions_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "tax_reporting_signoff_review",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(tax_reporting_signoff_review_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                        ],
                        answer_preview=copilot_text,
                        denied=False,
                        elapsed_ms=int((time.perf_counter() - copilot_started) * 1000),
                        metadata=copilot_review_metadata,
                    )
                except Exception:
                    pass
            st.success("Reports copilot analysis complete.")
            st.rerun()
        except Exception as exc:
            st.error(f"Reports copilot analysis failed: {exc}")

    raw_reports_ai = str(st.session_state.get("reports_copilot_raw") or "").strip()
    if raw_reports_ai:
        with st.expander("Reports Copilot Result", expanded=False):
            _render_ai_json_sections(
                raw_reports_ai,
                [
                    ("executive_summary", "Executive Summary"),
                    ("margin_anomalies", "Margin Anomalies"),
                    ("reconciliation_findings", "Reconciliation Findings"),
                    ("tax_review_findings", "Tax Review Findings"),
                    ("recommended_exports", "Recommended Exports"),
                    ("next_actions", "Next Actions"),
                ],
            )
            st.code(raw_reports_ai, language="json")
            feedback_cols = st.columns(3)
            with feedback_cols[0]:
                if st.button("Accept Copilot Review", key="reports_copilot_accept_btn"):
                    _log_reports_ai_outcome(
                        repo,
                        actor=user.username,
                        review_type="reports_copilot_review",
                        outcome="accepted",
                        answer_text=raw_reports_ai,
                        review_metadata=st.session_state.get("reports_copilot_metadata") or {},
                    )
                    st.success("Reports Copilot review acceptance recorded.")
                    st.rerun()
            with feedback_cols[1]:
                if st.button("Copilot Needs Edits", key="reports_copilot_edit_btn"):
                    _log_reports_ai_outcome(
                        repo,
                        actor=user.username,
                        review_type="reports_copilot_review",
                        outcome="edited",
                        answer_text=raw_reports_ai,
                        review_metadata=st.session_state.get("reports_copilot_metadata") or {},
                    )
                    st.success("Reports Copilot review edit outcome recorded.")
                    st.rerun()
            with feedback_cols[2]:
                if st.button("Reject Copilot Review", key="reports_copilot_reject_btn"):
                    _log_reports_ai_outcome(
                        repo,
                        actor=user.username,
                        review_type="reports_copilot_review",
                        outcome="rejected",
                        answer_text=raw_reports_ai,
                        review_metadata=st.session_state.get("reports_copilot_metadata") or {},
                    )
                    st.success("Reports Copilot review rejection recorded.")
                    st.rerun()

    st.markdown("### AI Accountant")
    st.caption(
        "Read-only accounting review with source-table citations. Drafted recommendations require human approval before any write action."
    )
    if st.button(
        "Run AI Accountant Review",
        key="reports_accountant_review_btn",
        disabled=not load_extended_analytics,
    ):
        if not ensure_permission(user, "ai_accountant_use", "Use AI Accountant"):
            st.stop()
        try:
            accountant_started = time.perf_counter()
            accountant_context = {
                "date_range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
                "accounting_close_readiness": dict(accounting_close_summary),
                "accounting_close_packet_evidence_hash": current_accounting_close_packet_hash,
                "close_readiness_checks": (
                    accounting_close_checks_df.to_dict("records")
                    if not accounting_close_checks_df.empty
                    else []
                ),
                "accounting_close_formula_checks": (
                    accounting_close_formula_df.to_dict("records")
                    if not accounting_close_formula_df.empty
                    else []
                ),
                "accounting_sales_component_checks": (
                    accounting_sales_component_df.to_dict("records")
                    if not accounting_sales_component_df.empty
                    else []
                ),
                "accounting_return_tieout_checks": (
                    accounting_return_tieout_df.to_dict("records")
                    if not accounting_return_tieout_df.empty
                    else []
                ),
                "accounting_inventory_valuation_checks": (
                    accounting_inventory_valuation_df.to_dict("records")
                    if not accounting_inventory_valuation_df.empty
                    else []
                ),
                "accounting_fee_evidence_checks": (
                    accounting_fee_evidence_df.to_dict("records")
                    if not accounting_fee_evidence_df.empty
                    else []
                ),
                "accounting_shipping_evidence_checks": (
                    accounting_shipping_evidence_df.to_dict("records")
                    if not accounting_shipping_evidence_df.empty
                    else []
                ),
                "accounting_reconciliation_tieout_checks": (
                    accounting_reconciliation_tieout_df.to_dict("records")
                    if not accounting_reconciliation_tieout_df.empty
                    else []
                ),
                "accounting_cogs_source_checks": (
                    accounting_cogs_source_df.to_dict("records")
                    if not accounting_cogs_source_df.empty
                    else []
                ),
                "sale_fifo_cogs_evidence_rows": (
                    sale_fifo_cogs_evidence_df.head(50).to_dict("records")
                    if not sale_fifo_cogs_evidence_df.empty
                    else []
                ),
                "accounting_lot_allocation_checks": (
                    accounting_lot_allocation_df.to_dict("records")
                    if not accounting_lot_allocation_df.empty
                    else []
                ),
                "accounting_exception_queue_checks": (
                    accounting_exception_queue_checks_df.to_dict("records")
                    if not accounting_exception_queue_checks_df.empty
                    else []
                ),
                "accounting_margin_anomaly_checks": (
                    accounting_margin_anomaly_checks_df.to_dict("records")
                    if not accounting_margin_anomaly_checks_df.empty
                    else []
                ),
                "accounting_close_consistency_checks": (
                    accounting_close_consistency_checks_df.to_dict("records")
                    if not accounting_close_consistency_checks_df.empty
                    else []
                ),
                "accounting_close_packet_completeness_checks": (
                    accounting_close_packet_completeness_df.to_dict("records")
                    if not accounting_close_packet_completeness_df.empty
                    else []
                ),
                "accounting_close_packet_manifest_checks": (
                    accounting_close_packet_manifest_df.to_dict("records")
                    if not accounting_close_packet_manifest_df.empty
                    else []
                ),
                "accounting_close_packet_hash_checks": (
                    accounting_close_packet_hash_df.to_dict("records")
                    if not accounting_close_packet_hash_df.empty
                    else []
                ),
                "accounting_close_packet_evidence_hash_rows": (
                    accounting_close_packet_evidence_hash_df.to_dict("records")
                    if not accounting_close_packet_evidence_hash_df.empty
                    else []
                ),
                "accounting_period_drift_checks": (
                    accounting_period_drift_df.to_dict("records")
                    if not accounting_period_drift_df.empty
                    else []
                ),
                "accounting_period_drift_summary": {
                    "check_count": int(len(accounting_period_drift_df)),
                    "warn_count": (
                        int((accounting_period_drift_df["status"].astype(str) == "warn").sum())
                        if not accounting_period_drift_df.empty and "status" in accounting_period_drift_df.columns
                        else 0
                    ),
                },
                "accounting_exception_rows": (
                    accounting_exceptions_df.head(50).to_dict("records")
                    if not accounting_exceptions_df.empty
                    else []
                ),
                "lot_allocation_source_summary": (
                    lot_allocation_source_summary_df.to_dict("records")
                    if not lot_allocation_source_summary_df.empty
                    else []
                ),
                "sold_cogs_source_summary": (
                    cogs_source_summary_df.to_dict("records")
                    if not cogs_source_summary_df.empty
                    else []
                ),
                "cogs_margin_summary": {
                    "gross_sales_total": float(report_scalar_cache.get("cogs_gross_sales_total") or 0.0),
                    "fifo_margin_before_returns_total": float(
                        report_scalar_cache.get("cogs_fifo_margin_total") or 0.0
                    ),
                    "lot_margin_total": float(report_scalar_cache.get("cogs_lot_margin_total") or 0.0),
                    "negative_fifo_margin_rows": int(report_scalar_cache.get("cogs_negative_fifo_rows") or 0),
                },
                "profit_after_returns_summary": close_profit_summary,
                "reconciliation_flags": int(report_scalar_cache.get("reconcile_flags") or 0),
                "shipping_economics_rows": (
                    shipping_economics_df.head(30).to_dict("records")
                    if not shipping_economics_df.empty
                    else []
                ),
                "fee_reconciliation_rows": (
                    ebay_fee_reconciliation_df.head(30).to_dict("records")
                    if not ebay_fee_reconciliation_df.empty
                    else []
                ),
                "tax_review_summary": {
                    "jurisdiction": str(tax_jurisdiction or tax_default_jurisdiction or ""),
                    "tax_rate_percent": float(tax_rate_percent or 0.0),
                    "shipping_taxable": bool(tax_shipping_taxable),
                    "tax_packet_evidence_hash": current_tax_packet_hash,
                    "marketplace_scope": tax_marketplace_scope_label,
                    "facilitator_channels": sorted(facilitator_channels),
                    "tax_exempt_categories": sorted(tax_exempt_categories),
                    "tax_exception_count": int(len(tax_exceptions_df)),
                },
                "tax_profile_evidence": selected_tax_profile_context,
                "tax_reporting_signoff_rows": (
                    tax_signoff_df.head(20).to_dict("records")
                    if not tax_signoff_df.empty
                    else []
                ),
                "tax_reporting_signoff_review": (
                    tax_reporting_signoff_review_df.to_dict("records")
                    if not tax_reporting_signoff_review_df.empty
                    else []
                ),
                "accounting_close_signoff_rows": (
                    accounting_close_signoff_df.head(20).to_dict("records")
                    if not accounting_close_signoff_df.empty
                    else []
                ),
                "accounting_close_signoff_review": (
                    accounting_close_signoff_review_df.to_dict("records")
                    if not accounting_close_signoff_review_df.empty
                    else []
                ),
                "ai_review_outcome_rows": (
                    ai_review_outcomes_df.head(20).to_dict("records")
                    if not ai_review_outcomes_df.empty
                    else []
                ),
            }
            accountant_system_message = get_runtime_str(
                repo,
                "accountant_llm_system_message",
                (
                    "You are GoldenStackers' read-only AI Accountant. Cite source tables/rows, "
                    "label estimated versus actual values, and never provide tax/legal conclusions."
                ),
            ).strip()
            accountant_instruction = (
                "Return ONLY JSON with keys: `close_status`, `profit_basis_notes`, "
                "`lot_cost_findings`, `fee_shipping_findings`, `recommended_human_actions`, "
                "`unsupported_tax_or_legal_items`. Values must be concise arrays or strings. "
                "Explicitly review `accounting_period_drift_checks` for dashboard/QBO drift. "
                "Review `accounting_close_formula_checks` to confirm close arithmetic ties out before sign-off. "
                "When discussing profit, use `profit_after_returns_summary.estimated_profit_after_returns` as final estimated profit "
                "and label `cogs_margin_summary.fifo_margin_before_returns_total` as before-return margin. "
                "Review `accounting_sales_component_checks` to confirm Sales Detail fee/shipping/label components tie to COGS & Margin close totals. "
                "Review `accounting_return_tieout_checks` to confirm refund totals, return COGS reversals, and QBO adjustment staging tie out. "
                "Review `accounting_inventory_valuation_checks` to confirm stocked inventory has landed cost and inventory value ties to close readiness. "
                "Review `accounting_fee_evidence_checks` to confirm fee reconciliation/source-priority evidence ties to Sales Detail and identify fallback fee rows. "
                "Review `accounting_shipping_evidence_checks` to confirm shipping charged, label spend, and shipping delta tie across Sales Detail and Shipping Economics. "
                "Review `accounting_reconciliation_tieout_checks` to confirm marketplace reconciliation totals tie to Sales Detail, Returns, and close flags. "
                "Review `accounting_cogs_source_checks` to confirm sold COGS source totals tie to COGS & Margin and identify fallback/missing-basis COGS. "
                "When COGS evidence checks warn, review `sale_fifo_cogs_evidence_rows` to trace sale COGS back to product, lot, assignment, quantity, unit cost, total cost, and source. "
                "Review `accounting_lot_allocation_checks` to confirm Lot Allocation Source Summary ties to Lot Assignment detail and identify fallback/missing-basis assignments. "
                "Review `accounting_exception_queue_checks` to confirm exception counts and severities tie to close readiness and P0 exceptions remain blocking. "
                "Review `accounting_margin_anomaly_checks` to confirm nonpositive COGS & Margin rows tie to exception evidence and negative margin rows remain close blockers. "
                "Review `accounting_close_consistency_checks` to confirm close readiness status, blocker/warning counts, and close-check statuses are internally consistent. "
                "Review `accounting_close_packet_completeness_checks` to confirm required close-packet evidence artifacts are present before relying on sign-off evidence. "
                "Review `accounting_close_packet_manifest_checks` to confirm close-packet manifest row counts match selected export dataframes. "
                "Review `accounting_close_packet_hash_checks` to confirm close-packet CSV hashes are available for integrity review. "
                "Review `accounting_close_packet_evidence_hash_rows` to identify the current packet evidence hash for sign-off comparison. "
                    "Review `accounting_close_signoff_rows` and `accounting_close_signoff_review` for owner/date/packet evidence, packet evidence hash match, and stale or contradictory approval evidence before saying a period is sign-off ready. "
                    "Review `ai_review_outcome_rows` for accepted, edited, or rejected Copilot/AI Accountant feedback tied to prior review evidence. "
                    "Review `tax_reporting_signoff_review` for selected tax profile/jurisdiction, exception count, advisor evidence, and Tax Review Packet hash mismatches; "
                "use tax profile/sign-off evidence only to summarize configured assumptions and advisor-review status; "
                "route filing/remittance/legal conclusions to `unsupported_tax_or_legal_items`. "
                "Do not propose direct writes; draft only human-review recommendations."
            )
            accountant_prompt = "AI Accountant close-readiness review"
            accountant_prompt_hash = _stable_json_sha256(
                {
                    "query": accountant_prompt,
                    "system_message": accountant_system_message,
                    "instruction": accountant_instruction,
                }
            )
            accountant_data_scope = {
                "date_range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
                "context_keys": sorted(accountant_context.keys()),
                "context_hash_sha256": _stable_json_sha256(accountant_context),
                "tax_packet_evidence_hash": current_tax_packet_hash,
                "accounting_close_packet_evidence_hash": current_accounting_close_packet_hash,
                "row_counts": {
                    "accounting_close_readiness_checks": int(len(accounting_close_checks_df)),
                    "accounting_period_drift_checks": int(len(accounting_period_drift_df)),
                    "accounting_exception_queue": int(len(accounting_exceptions_df)),
                    "sale_fifo_cogs_evidence": int(len(sale_fifo_cogs_evidence_df)),
                    "accounting_close_signoffs": int(len(accounting_close_signoff_df)),
                    "accounting_close_signoff_review": int(len(accounting_close_signoff_review_df)),
                    "ai_review_outcomes": int(len(ai_review_outcomes_df)),
                    "tax_reporting_signoffs": int(len(tax_signoff_df)),
                    "tax_reporting_signoff_review": int(len(tax_reporting_signoff_review_df)),
                },
            }
            result = execute_comp_summary(
                repo,
                query=accountant_prompt,
                ebay_rows=[],
                web_rows=[],
                spot_context=accountant_context,
                system_message=accountant_system_message,
                instruction=accountant_instruction,
            )
            accountant_text = str(result.text or "").strip()
            st.session_state["reports_accountant_raw"] = accountant_text
            accountant_elapsed_ms = int((time.perf_counter() - accountant_started) * 1000)
            accountant_review_metadata = {
                "event_type": "ai_accountant_review",
                "surface": "reports",
                "read_only": True,
                "requires_human_approval_for_writes": True,
                "tax_legal_guardrail": "unsupported conclusions routed to human review",
                "date_range": {"from": from_date.isoformat(), "to": to_date.isoformat()},
                "prompt_hash_sha256": accountant_prompt_hash,
                "data_scope_hash_sha256": str(
                    accountant_data_scope.get("context_hash_sha256") or ""
                ),
                "data_scope": accountant_data_scope,
                "context_keys": sorted(accountant_context.keys()),
                "ai_citation": dict(getattr(result, "citation", {}) or {}),
            }
            st.session_state["reports_accountant_metadata"] = accountant_review_metadata
            if hasattr(repo, "log_ai_chat_interaction"):
                try:
                    repo.log_ai_chat_interaction(
                        actor=user.username,
                        prompt="AI Accountant close-readiness review",
                        intent="reports_ai_accountant_review",
                        allowed_domains=["accounting", "reports", "sales", "orders", "inventory", "tax"],
                        citations=[
                            {
                                "table": "accounting_close_readiness_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_checks_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_exception_queue",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_exceptions_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_period_drift_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_period_drift_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_formula_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_formula_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_sales_component_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_sales_component_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_return_tieout_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_return_tieout_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_inventory_valuation_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_inventory_valuation_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_fee_evidence_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_fee_evidence_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_shipping_evidence_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_shipping_evidence_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_reconciliation_tieout_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_reconciliation_tieout_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_cogs_source_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_cogs_source_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "sale_fifo_cogs_evidence",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(sale_fifo_cogs_evidence_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_lot_allocation_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_lot_allocation_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_exception_queue_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_exception_queue_checks_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_margin_anomaly_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_margin_anomaly_checks_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_consistency_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_consistency_checks_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_packet_completeness_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_packet_completeness_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_packet_manifest_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_packet_manifest_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_packet_hash_checks",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_packet_hash_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_packet_evidence_hash",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_packet_evidence_hash_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_signoffs",
                                "filters": "latest_audit_records",
                                "rows_considered": int(len(accounting_close_signoff_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "accounting_close_signoff_review",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(accounting_close_signoff_review_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "ai_review_outcomes",
                                "filters": "latest_audit_records",
                                "rows_considered": int(len(ai_review_outcomes_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "lot_allocation_source_summary",
                                "filters": "grouped_by=cost_source",
                                "rows_considered": int(len(lot_allocation_source_summary_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "tax_profile",
                                "filters": "selected_saved_profile",
                                "rows_considered": 1 if selected_tax_profile_context else 0,
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "tax_reporting_signoffs",
                                "filters": "latest_audit_records",
                                "rows_considered": int(len(tax_signoff_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                            {
                                "table": "tax_reporting_signoff_review",
                                "filters": f"from={from_date.isoformat()};to={to_date.isoformat()}",
                                "rows_considered": int(len(tax_reporting_signoff_review_df)),
                                "as_of_utc": datetime.now(timezone.utc).isoformat(),
                            },
                        ],
                        answer_preview=accountant_text,
                        denied=False,
                        elapsed_ms=accountant_elapsed_ms,
                        metadata=accountant_review_metadata,
                    )
                except Exception:
                    pass
            st.success("AI Accountant review complete.")
            st.rerun()
        except Exception as exc:
            st.error(f"AI Accountant review failed: {exc}")

    raw_accountant_ai = str(st.session_state.get("reports_accountant_raw") or "").strip()
    if raw_accountant_ai:
        with st.expander("AI Accountant Result", expanded=False):
            _render_ai_json_sections(
                raw_accountant_ai,
                [
                    ("close_status", "Close Status"),
                    ("profit_basis_notes", "Profit Basis Notes"),
                    ("lot_cost_findings", "Lot Cost Findings"),
                    ("fee_shipping_findings", "Fee and Shipping Findings"),
                    ("recommended_human_actions", "Recommended Human Actions"),
                    ("unsupported_tax_or_legal_items", "Unsupported Tax or Legal Items"),
                ],
            )
            st.code(raw_accountant_ai, language="json")
            feedback_cols = st.columns(3)
            with feedback_cols[0]:
                if st.button("Accept AI Accountant Review", key="reports_accountant_accept_btn"):
                    _log_reports_ai_outcome(
                        repo,
                        actor=user.username,
                        review_type="ai_accountant_review",
                        outcome="accepted",
                        answer_text=raw_accountant_ai,
                        review_metadata=st.session_state.get("reports_accountant_metadata") or {},
                    )
                    st.success("AI Accountant review acceptance recorded.")
                    st.rerun()
            with feedback_cols[1]:
                if st.button("AI Accountant Needs Edits", key="reports_accountant_edit_btn"):
                    _log_reports_ai_outcome(
                        repo,
                        actor=user.username,
                        review_type="ai_accountant_review",
                        outcome="edited",
                        answer_text=raw_accountant_ai,
                        review_metadata=st.session_state.get("reports_accountant_metadata") or {},
                    )
                    st.success("AI Accountant review edit outcome recorded.")
                    st.rerun()
            with feedback_cols[2]:
                if st.button("Reject AI Accountant Review", key="reports_accountant_reject_btn"):
                    _log_reports_ai_outcome(
                        repo,
                        actor=user.username,
                        review_type="ai_accountant_review",
                        outcome="rejected",
                        answer_text=raw_accountant_ai,
                        review_metadata=st.session_state.get("reports_accountant_metadata") or {},
                    )
                    st.success("AI Accountant review rejection recorded.")
                    st.rerun()

    reports = [
        ("Sales Detail", sales_df, "sales_detail"),
        ("Tax Summary (Estimated)", tax_summary_df, "tax_summary_estimated"),
        ("Tax by Marketplace (Estimated)", tax_by_marketplace_df, "tax_by_marketplace_estimated"),
        ("Tax Detail (Estimated)", tax_detail_df, "tax_detail_estimated"),
        ("Tax Exceptions / Advisor Review", tax_exceptions_df, "tax_exceptions_advisor_review"),
        ("Tax Reporting Sign-Off Evidence", tax_signoff_df, "tax_reporting_signoffs"),
        ("Tax Reporting Sign-Off Review", tax_reporting_signoff_review_df, "tax_reporting_signoff_review"),
        ("Inventory Snapshot", inventory_df, "inventory_snapshot"),
        ("Listing Snapshot", listings_df, "listing_snapshot"),
        ("Orders", orders_df, "orders"),
        ("Order Items", order_items_df, "order_items"),
        ("eBay Order Fee Breakdown", ebay_order_fee_breakdown_df, "ebay_order_fee_breakdown"),
        ("Returns", returns_df, "returns"),
        ("Lot Assignment", lots_df, "lot_assignment"),
        ("Lot Allocation Source Summary", lot_allocation_source_summary_df, "lot_allocation_source_summary"),
        ("Sold COGS Source Summary", cogs_source_summary_df, "cogs_source_summary"),
        ("Sale FIFO COGS Evidence", sale_fifo_cogs_evidence_df, "sale_fifo_cogs_evidence"),
        ("Inventory Movements", movements_df, "inventory_movements"),
        ("QuickBooks Sales Export", qbo_sales_df, "qbo_sales_export"),
        ("QuickBooks Refund/Adjustment Export", qbo_adjustments_df, "qbo_adjustments_export"),
        ("Reconciliation by Marketplace", reconciliation_df, "reconciliation_marketplace"),
        ("Accounting Validation Flags", accounting_validation_df, "accounting_validation_flags"),
        ("Accounting Exception Queue", accounting_exceptions_df, "accounting_exception_queue"),
        ("Accounting Close Readiness Checks", accounting_close_checks_df, "accounting_close_readiness_checks"),
        ("Accounting Close Formula Checks", accounting_close_formula_df, "accounting_close_formula_checks"),
        ("Accounting Sales Component Checks", accounting_sales_component_df, "accounting_sales_component_checks"),
        ("Accounting Return Tie-Out Checks", accounting_return_tieout_df, "accounting_return_tieout_checks"),
        (
            "Accounting Inventory Valuation Checks",
            accounting_inventory_valuation_df,
            "accounting_inventory_valuation_checks",
        ),
        ("Accounting Fee Evidence Checks", accounting_fee_evidence_df, "accounting_fee_evidence_checks"),
        (
            "Accounting Shipping Evidence Checks",
            accounting_shipping_evidence_df,
            "accounting_shipping_evidence_checks",
        ),
        (
            "Accounting Reconciliation Tie-Out Checks",
            accounting_reconciliation_tieout_df,
            "accounting_reconciliation_tieout_checks",
        ),
        ("Accounting COGS Source Checks", accounting_cogs_source_df, "accounting_cogs_source_checks"),
        ("Accounting Lot Allocation Checks", accounting_lot_allocation_df, "accounting_lot_allocation_checks"),
        (
            "Accounting Exception Queue Checks",
            accounting_exception_queue_checks_df,
            "accounting_exception_queue_checks",
        ),
        (
            "Accounting Margin Anomaly Checks",
            accounting_margin_anomaly_checks_df,
            "accounting_margin_anomaly_checks",
        ),
        (
            "Accounting Close Consistency Checks",
            accounting_close_consistency_checks_df,
            "accounting_close_consistency_checks",
        ),
        (
            "Accounting Close Packet Completeness Checks",
            accounting_close_packet_completeness_df,
            "accounting_close_packet_completeness_checks",
        ),
        (
            "Accounting Close Packet Manifest Checks",
            accounting_close_packet_manifest_df,
            "accounting_close_packet_manifest_checks",
        ),
        (
            "Accounting Close Packet Hash Checks",
            accounting_close_packet_hash_df,
            "accounting_close_packet_hash_checks",
        ),
        (
            "Accounting Close Packet Evidence Hash",
            accounting_close_packet_evidence_hash_df,
            "accounting_close_packet_evidence_hash",
        ),
        ("Accounting Close Sign-Off Evidence", accounting_close_signoff_df, "accounting_close_signoffs"),
        ("Accounting Close Sign-Off Review", accounting_close_signoff_review_df, "accounting_close_signoff_review"),
        ("AI Review Outcome Evidence", ai_review_outcomes_df, "ai_review_outcomes"),
        ("Accounting Period Drift Checks", accounting_period_drift_df, "accounting_period_drift_checks"),
        ("COGS & Margin Detail", cogs_margin_df, "cogs_margin_detail"),
        ("Margin by SKU", margin_by_sku_df, "margin_by_sku"),
        ("Margin by Marketplace", margin_by_channel_df, "margin_by_marketplace"),
        ("Margin by Period", margin_by_period_df, "margin_by_period"),
        ("Inventory Cycle Summary by SKU", inventory_cycle_summary_df, "inventory_cycle_summary"),
        ("Inventory Cycles (Rebuy/Resell)", inventory_cycles_df, "inventory_cycles"),
        ("Rebuy Cost Trend Events", rebuy_cost_trend_df, "rebuy_cost_trend"),
        ("Listing Review Activity", review_activity_df, "listing_review_activity"),
        ("Listing Review Summary", review_summary_df, "listing_review_summary"),
        (
            "Listing Format Intent vs Publish Outcome",
            listing_format_outcome_df,
            "listing_format_intent_vs_outcome",
        ),
        (
            "eBay Fee Estimate vs Actual",
            ebay_fee_reconciliation_df,
            "ebay_fee_estimate_vs_actual",
        ),
        (
            "eBay Fee Reconciliation Summary",
            ebay_fee_reconciliation_by_marketplace_df,
            "ebay_fee_reconciliation_summary",
        ),
        (
            "eBay Fee Actual Source Breakdown",
            ebay_fee_actual_source_df,
            "ebay_fee_actual_source_breakdown",
        ),
        (
            "eBay Fee Source Priority",
            ebay_fee_source_priority_df,
            "ebay_fee_source_priority",
        ),
        (
            "eBay Fee Source Priority Trend",
            ebay_fee_source_priority_trend_df,
            "ebay_fee_source_priority_trend",
        ),
        (
            "eBay Marketplace Fee Detail (Per Order/Line)",
            ebay_marketplace_fee_detail_df,
            "ebay_marketplace_fee_detail",
        ),
        (
            "eBay Marketplace Fee Summary (By Fee Type)",
            ebay_marketplace_fee_summary_df,
            "ebay_marketplace_fee_summary",
        ),
        (
            "eBay Marketplace Fee by SKU",
            ebay_marketplace_fee_by_sku_df,
            "ebay_marketplace_fee_by_sku",
        ),
        (
            "eBay Marketplace Fee by Category",
            ebay_marketplace_fee_by_category_df,
            "ebay_marketplace_fee_by_category",
        ),
        (
            "Shipping Economics Detail",
            shipping_economics_df,
            "shipping_economics_detail",
        ),
        (
            "Shipping Economics Summary",
            shipping_econ_summary_df,
            "shipping_economics_summary",
        ),
        (
            "Order Actual Economics Allocation",
            order_actual_econ_df,
            "order_actual_economics_allocation",
        ),
        (
            "Economics Intelligence Facts (Estimate vs Actual)",
            economics_intel_df,
            "economics_intelligence_facts",
        ),
        (
            "Economics Intelligence by SKU",
            economics_intel_by_sku_df,
            "economics_intelligence_by_sku",
        ),
        (
            "Economics Intelligence by Marketplace",
            economics_intel_by_marketplace_df,
            "economics_intelligence_by_marketplace",
        ),
        (
            "Economics Intelligence Alert Rows",
            economics_intel_alerts_df,
            "economics_intelligence_alert_rows",
        ),
    ]
    accounting_close_packet = _build_accounting_close_export_packet(
        reports=reports,
        close_summary=accounting_close_summary,
        from_date=from_date,
        to_date=to_date,
    )
    st.download_button(
        label="Download Accounting Close Packet ZIP",
        data=accounting_close_packet,
        file_name=f"accounting_close_packet_{from_date}_{to_date}.zip",
        mime="application/zip",
        key="reports_accounting_close_packet_zip",
    )
    tax_review_packet = _build_tax_review_export_packet(
        reports=reports,
        from_date=from_date,
        to_date=to_date,
        tax_jurisdiction=tax_jurisdiction or tax_default_jurisdiction,
        tax_rate_percent=float(tax_rate_percent),
        shipping_taxable=bool(tax_shipping_taxable),
        marketplace_scope=tax_marketplace_scope_label,
        facilitator_channels=facilitator_channels,
        tax_exempt_categories=tax_exempt_categories,
        tax_profile=selected_tax_profile_context,
    )
    st.download_button(
        label="Download Tax Review Packet ZIP",
        data=tax_review_packet,
        file_name=f"tax_review_packet_{from_date}_{to_date}.zip",
        mime="application/zip",
        key="reports_tax_review_packet_zip",
    )
    with st.expander("Tax Reporting Sign-Off Review", expanded=False):
        _render_df_with_preview(tax_reporting_signoff_review_df, hide_index=True)
    with st.expander("Record Tax Reporting Sign-Off", expanded=False):
        current_tax_packet_ref = f"tax_review_packet_{from_date}_{to_date}.zip"
        selected_profile_key = str(selected_tax_profile_context.get("profile_key") or "").strip().lower()
        st.caption(
            "Record tax sign-off only after advisor/human review of the Tax Review Packet. "
            "These outputs remain estimates and do not replace tax-advisor validation."
        )
        st.code(current_tax_packet_hash or "tax packet hash unavailable", language="text")
        can_record_tax_signoff = has_permission(user.role, "manage_settings")
        tax_signoff_context = (
            st.form("reports_tax_reporting_signoff_form")
            if hasattr(st, "form")
            else st.expander("Tax Sign-Off Fields", expanded=True)
        )
        with tax_signoff_context:
            ts1, ts2 = st.columns(2)
            with ts1:
                reports_tax_signoff_period = st.text_input(
                    "Tax Period",
                    value=_default_accounting_close_period(from_date, to_date),
                    key="reports_tax_signoff_period",
                )
                reports_tax_signoff_jurisdiction = st.text_input(
                    "Sign-Off Jurisdiction",
                    value=str(tax_jurisdiction or tax_default_jurisdiction or ""),
                    key="reports_tax_signoff_jurisdiction",
                )
                reports_tax_signoff_profile_key = st.text_input(
                    "Tax Profile Key Used",
                    value=selected_profile_key,
                    key="reports_tax_signoff_profile_key",
                )
                reports_tax_signoff_owner = st.text_input(
                    "Owner",
                    value=str(user.username or ""),
                    key="reports_tax_signoff_owner",
                )
            with ts2:
                reports_tax_signoff_status = st.selectbox(
                    "Status",
                    options=["approved", "blocked", "needs_followup"],
                    index=0 if tax_exceptions_df.empty else 2,
                    key="reports_tax_signoff_status",
                )
                reports_tax_signoff_date = st.date_input(
                    "Sign-Off Date",
                    value=utc_today(),
                    key="reports_tax_signoff_date",
                )
                reports_tax_exception_count = st.number_input(
                    "Tax Exception Count",
                    min_value=0,
                    max_value=1000000,
                    value=int(len(tax_exceptions_df)),
                    step=1,
                    key="reports_tax_signoff_exception_count",
                )
                reports_tax_packet_ref = st.text_input(
                    "Tax Packet Reference",
                    value=current_tax_packet_ref,
                    key="reports_tax_signoff_packet_ref",
                )
            reports_tax_advisor_evidence = st.text_input(
                "Advisor / Evidence Link",
                placeholder="advisor email, review ticket, or filing workpaper reference",
                key="reports_tax_signoff_advisor_evidence",
            )
            if hasattr(st, "text_area"):
                reports_tax_signoff_notes = st.text_area(
                    "Notes",
                    placeholder="Advisor comments, exemption/shipping/facilitator assumptions, or filing caveats.",
                    key="reports_tax_signoff_notes",
                )
            else:
                reports_tax_signoff_notes = st.text_input(
                    "Notes",
                    value="",
                    key="reports_tax_signoff_notes",
                )
            if hasattr(st, "form_submit_button"):
                save_tax_signoff = st.form_submit_button(
                    "Record Tax Reporting Sign-Off",
                    disabled=not can_record_tax_signoff,
                )
            else:
                save_tax_signoff = st.button(
                    "Record Tax Reporting Sign-Off",
                    key="reports_tax_signoff_submit_btn",
                    disabled=not can_record_tax_signoff,
                )
        if not can_record_tax_signoff:
            st.caption("You need `manage_settings` permission to record tax reporting sign-off evidence.")
        if save_tax_signoff:
            try:
                payload = _build_tax_reporting_signoff_payload(
                    target_env=settings.app_env,
                    tax_period=reports_tax_signoff_period,
                    jurisdiction=reports_tax_signoff_jurisdiction,
                    profile_key=reports_tax_signoff_profile_key,
                    status=reports_tax_signoff_status,
                    owner=reports_tax_signoff_owner,
                    signoff_date=reports_tax_signoff_date,
                    tax_packet_ref=reports_tax_packet_ref,
                    tax_packet_hash=current_tax_packet_hash,
                    advisor_evidence_link=reports_tax_advisor_evidence,
                    tax_exception_count=int(reports_tax_exception_count or 0),
                    notes=reports_tax_signoff_notes,
                )
                repo.record_audit_event(
                    entity_type="tax_reporting_signoff",
                    entity_id=None,
                    action="record",
                    actor=user.username,
                    changes=payload,
                )
                st.success("Tax reporting sign-off recorded.")
                st.rerun()
            except Exception as exc:
                st.error(f"Unable to record tax reporting sign-off: {exc}")
    for label, df, file_prefix in reports:
        st.markdown(f"### {label}")
        if df.empty:
            st.info("No records for this report in the selected date range.")
            continue

        st.caption(f"Rows: {int(len(df))}")
        report_context_caption = _report_context_caption(file_prefix)
        if report_context_caption:
            st.caption(report_context_caption)
        _render_df_with_preview(df)
        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                label=f"Download {label} CSV",
                data=df.to_csv(index=False).encode("utf-8"),
                file_name=f"{file_prefix}_{from_date}_{to_date}.csv",
                mime="text/csv",
            )
        with dl2:
            st.download_button(
                label=f"Download {label} XLSX",
                data=dataframe_to_xlsx_bytes(df, sheet_name=file_prefix[:31]),
                file_name=f"{file_prefix}_{from_date}_{to_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    render_workspace_feedback(
        repo=repo,
        actor=user.username,
        workspace_key="reports",
        section_title="Workspace Feedback: Reports",
    )
